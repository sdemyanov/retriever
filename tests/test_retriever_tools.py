from __future__ import annotations

import csv
import errno
import hashlib
import importlib.util
import json
import io
import mailbox
import os
import random
import re
import shutil
import sqlite3
import tempfile
import threading
import time
import types
import unittest
import zipfile
import base64
from contextlib import redirect_stderr, redirect_stdout
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from pathlib import Path
from unittest import mock


REPO_ROOT = Path(__file__).resolve().parents[1]
TOOL_PATH = REPO_ROOT / "skills" / "tool-template" / "tools.py"
BUNDLER_PATH = REPO_ROOT / "skills" / "tool-template" / "bundle_retriever_tools.py"
TOOL_TEMPLATE_PATH = REPO_ROOT / "skills" / "tool-template" / "tool-template.md"
SOURCE_HEADER_PATH = REPO_ROOT / "skills" / "tool-template" / "src" / "00_header.py"
PLUGIN_MANIFEST_PATH = REPO_ROOT / ".claude-plugin" / "plugin.json"
PING_SKILL_PATH = REPO_ROOT / "skills" / "ping" / "SKILL.md"
REGRESSION_CORPUS_ROOT = REPO_ROOT / "phase0" / "regression_corpus"

retriever_tools = None
TOOL_BYTES = b""
RANDOMIZED_FILTER_TEST_SEED = 20260419
RANDOMIZED_FILTER_FIELD_TYPES = {
    "file_name": "text",
    "review_note": "text",
    "review_score": "integer",
    "review_date": "date",
    "is_hot": "boolean",
}
RANDOMIZED_FILTER_LITERAL_POOLS = {
    "file_name": [
        "alpha.txt",
        "beta.txt",
        "gamma.txt",
        "thread.eml",
        "notes.txt",
        "zeta.txt",
    ],
    "review_note": [
        "alpha memo",
        "beta's note",
        "needs review",
        "mail thread",
        "child's attachment",
        "missing note",
    ],
    "review_score": [-1, 1, 5, 7, 9, 12],
    "review_date": [
        "2026-04-01",
        "2026-04-02",
        "2026-04-03",
        "2026-04-05",
        "2026-04-09",
    ],
    "is_hot": [True, False],
}
RANDOMIZED_FILTER_LIKE_PATTERNS = {
    "file_name": ["alpha%", "%txt", "%notes%", "thread%", "%eml"],
    "review_note": ["%memo%", "%note%", "%review%", "%thread%", "%attachment%"],
    "review_date": ["2026-04-%", "%-03", "2026-04-0_", "%-09"],
}
RANDOMIZED_FILTER_OPERATOR_CHOICES = {
    "text": ["=", "!=", "<", "<=", ">", ">=", "LIKE", "NOT LIKE", "IN", "NOT IN", "BETWEEN", "NOT BETWEEN", "IS NULL", "IS NOT NULL"],
    "date": ["=", "!=", "<", "<=", ">", ">=", "LIKE", "NOT LIKE", "IN", "NOT IN", "BETWEEN", "NOT BETWEEN", "IS NULL", "IS NOT NULL"],
    "integer": ["=", "!=", "<", "<=", ">", ">=", "IN", "NOT IN", "BETWEEN", "NOT BETWEEN", "IS NULL", "IS NOT NULL"],
    "boolean": ["=", "!=", "IS NULL", "IS NOT NULL"],
}

def load_python_module(path: Path, module_name: str):
    module = types.ModuleType(module_name)
    module.__file__ = str(path)
    module.__package__ = ""
    exec(compile(path.read_text(encoding="utf-8"), str(path), "exec"), module.__dict__)
    return module


def assert_bundled_tooling_current() -> None:
    if not TOOL_PATH.exists():
        raise AssertionError(
            f"Missing generated bundle at {TOOL_PATH}. Run ./build.sh to regenerate the bundle and checksum before running tests."
        )

    bundler = load_python_module(BUNDLER_PATH, "retriever_bundler_under_test")
    expected_text = bundler.bundle_source(TOOL_PATH.parent / "src")
    current_text = TOOL_PATH.read_text(encoding="utf-8")
    if current_text != expected_text:
        raise AssertionError(
            "Generated tools.py is stale relative to skills/tool-template/src. "
            "Run ./build.sh to regenerate the bundle and checksum before running tests."
        )

    checksum_match = re.search(
        r"^- source checksum \(SHA256\): `([0-9a-f]{64})`$",
        TOOL_TEMPLATE_PATH.read_text(encoding="utf-8"),
        re.MULTILINE,
    )
    if checksum_match is None:
        raise AssertionError(f"Could not find the canonical source checksum line in {TOOL_TEMPLATE_PATH}.")

    expected_sha = hashlib.sha256(expected_text.encode("utf-8")).hexdigest()
    if checksum_match.group(1) != expected_sha:
        raise AssertionError(
            "tool-template.md has a stale source checksum for the generated bundle. "
            "Run ./build.sh to regenerate the bundle and checksum before running tests."
        )


def assert_version_metadata_current() -> None:
    header_match = re.search(
        r'^TOOL_VERSION = "([^"]+)"$',
        SOURCE_HEADER_PATH.read_text(encoding="utf-8"),
        re.MULTILINE,
    )
    if header_match is None:
        raise AssertionError(f"Could not determine TOOL_VERSION from {SOURCE_HEADER_PATH}.")
    expected_version = header_match.group(1)

    plugin_version = json.loads(PLUGIN_MANIFEST_PATH.read_text(encoding="utf-8")).get("version")
    if plugin_version != expected_version:
        raise AssertionError(
            ".claude-plugin/plugin.json has a stale version relative to TOOL_VERSION. "
            "Run ./build.sh to synchronize version metadata before running tests."
        )

    ping_text = PING_SKILL_PATH.read_text(encoding="utf-8")
    metadata_match = re.search(r'^\s*version:\s*"([^"]+)"\s*$', ping_text, re.MULTILINE)
    body_match = re.search(r"^Version:\s*(\S+)\s*$", ping_text, re.MULTILINE)
    if metadata_match is None or body_match is None:
        raise AssertionError(f"Could not find both ping skill version markers in {PING_SKILL_PATH}.")
    if metadata_match.group(1) != expected_version or body_match.group(1) != expected_version:
        raise AssertionError(
            "skills/ping/SKILL.md has stale version text relative to TOOL_VERSION. "
            "Run ./build.sh to synchronize version metadata before running tests."
        )


def setUpModule() -> None:
    global retriever_tools, TOOL_BYTES
    assert_bundled_tooling_current()
    assert_version_metadata_current()
    retriever_tools = load_python_module(TOOL_PATH, "retriever_tools_under_test")
    TOOL_BYTES = TOOL_PATH.read_bytes()


class RetrieverToolsRegressionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory(prefix="retriever-tests-")
        self.addCleanup(self.tempdir.cleanup)
        self.root = Path(self.tempdir.name)
        self.paths = retriever_tools.workspace_paths(self.root)
        retriever_tools.ensure_layout(self.paths)

    def run_cli(self, *args: str) -> tuple[int, dict[str, object] | None, str, str]:
        stdout = io.StringIO()
        stderr = io.StringIO()
        with (
            redirect_stdout(stdout),
            redirect_stderr(stderr),
            mock.patch.object(retriever_tools.sys, "argv", ["tools.py", *args]),
        ):
            exit_code = retriever_tools.main()
        stdout_text = stdout.getvalue().strip()
        stderr_text = stderr.getvalue().strip()
        payload = None
        if stdout_text or stderr_text:
            try:
                payload = json.loads(stdout_text or stderr_text)
            except json.JSONDecodeError:
                payload = None
        return exit_code, payload, stdout_text, stderr_text

    def run_cli_raw(self, *args: str) -> tuple[int, str, str]:
        stdout = io.StringIO()
        stderr = io.StringIO()
        with (
            redirect_stdout(stdout),
            redirect_stderr(stderr),
            mock.patch.object(retriever_tools.sys, "argv", ["tools.py", *args]),
        ):
            exit_code = retriever_tools.main()
        return exit_code, stdout.getvalue().strip(), stderr.getvalue().strip()

    def run_v2_loose_ingest(self, *scan_paths: str) -> dict[str, object]:
        start_args = ["ingest-start", str(self.root), "--recursive"]
        for scan_path in scan_paths:
            start_args.extend(["--path", scan_path])
        start_exit, start_payload, _, _ = self.run_cli(*start_args)
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        assert start_payload is not None
        run_id = str(start_payload["run_id"])

        payloads: dict[str, object] = {"run_id": run_id, "start": start_payload}
        for command, payload_key in (
            ("ingest-plan-step", "plan"),
            ("ingest-prepare-step", "prepare"),
            ("ingest-commit-step", "commit"),
            ("ingest-finalize-step", "finalize"),
        ):
            exit_code, payload, _, _ = self.run_cli(command, str(self.root), "--run-id", run_id)
            self.assertEqual(exit_code, 0)
            self.assertIsNotNone(payload)
            assert payload is not None
            payloads[payload_key] = payload
        post_finalize_steps: list[dict[str, object]] = []
        while dict(payloads["finalize"])["run"]["status"] != "completed":
            exit_code, payload, _, _ = self.run_cli(
                "ingest-run-step",
                str(self.root),
                "--run-id",
                run_id,
                "--budget-seconds",
                "35",
            )
            self.assertEqual(exit_code, 0)
            self.assertIsNotNone(payload)
            assert payload is not None
            post_finalize_steps.append(payload)
            step_result = payload.get("step_result")
            if isinstance(step_result, dict) and step_result.get("step") == "finalize":
                payloads["finalize"] = step_result
            elif payload["run"]["status"] == "completed":
                payloads["finalize"] = {
                    "ok": True,
                    "implemented": True,
                    "step": "finalize",
                    "finalization_complete": True,
                    "run": payload["run"],
                }
            self.assertLessEqual(len(post_finalize_steps), 10)
        if post_finalize_steps:
            payloads["post_finalize_run_steps"] = post_finalize_steps
        return payloads

    def preview_target_file_path(self, target: dict[str, object]) -> Path:
        file_abs_path = str(target.get("file_abs_path") or target.get("abs_path") or "")
        return Path(file_abs_path.split("#", 1)[0])

    def preview_target_by_label(self, targets: list[dict[str, object]], label: str | None) -> dict[str, object]:
        for target in targets:
            if target.get("label") == label:
                return target
        raise AssertionError(f"Missing preview target with label {label!r}.")

    def create_legacy_documents_table(self, *, with_row: bool) -> None:
        connection = sqlite3.connect(self.paths["db_path"])
        try:
            connection.execute(
                """
                CREATE TABLE documents (
                  id INTEGER PRIMARY KEY,
                  rel_path TEXT NOT NULL UNIQUE,
                  file_name TEXT NOT NULL,
                  file_type TEXT,
                  file_size INTEGER,
                  page_count INTEGER,
                  author TEXT,
                  date_created TEXT,
                  date_modified TEXT,
                  title TEXT,
                  subject TEXT,
                  recipients TEXT,
                  manual_field_locks_json TEXT NOT NULL DEFAULT '[]',
                  file_hash TEXT,
                  content_hash TEXT,
                  text_status TEXT NOT NULL DEFAULT 'ok',
                  lifecycle_status TEXT NOT NULL DEFAULT 'active',
                  ingested_at TEXT,
                  last_seen_at TEXT,
                  updated_at TEXT
                )
                """
            )
            if with_row:
                connection.execute(
                    """
                    INSERT INTO documents (
                      id, rel_path, file_name, file_type, file_size, page_count, author,
                      date_created, date_modified, title, subject, recipients,
                      manual_field_locks_json, file_hash, content_hash, text_status,
                      lifecycle_status, ingested_at, last_seen_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        1,
                        "legacy.txt",
                        "legacy.txt",
                        "txt",
                        12,
                        1,
                        "Legacy Author",
                        None,
                        None,
                        "Legacy Title",
                        None,
                        None,
                        "[]",
                        "file-hash",
                        "content-hash",
                        "ok",
                        "active",
                        "2026-04-14T00:00:00Z",
                        "2026-04-14T00:00:00Z",
                        "2026-04-14T00:00:00Z",
                    ),
                )
            connection.commit()
        finally:
            connection.close()

    def create_v6_documents_table(self, *, with_row: bool) -> None:
        connection = sqlite3.connect(self.paths["db_path"])
        try:
            connection.execute(
                """
                CREATE TABLE documents (
                  id INTEGER PRIMARY KEY,
                  display_id TEXT UNIQUE,
                  parent_document_id INTEGER REFERENCES documents(id) ON DELETE CASCADE,
                  rel_path TEXT NOT NULL UNIQUE,
                  file_name TEXT NOT NULL,
                  file_type TEXT,
                  file_size INTEGER,
                  page_count INTEGER,
                  author TEXT,
                  content_type TEXT,
                  date_created TEXT,
                  date_modified TEXT,
                  title TEXT,
                  subject TEXT,
                  participants TEXT,
                  recipients TEXT,
                  manual_field_locks_json TEXT NOT NULL DEFAULT '[]',
                  file_hash TEXT,
                  content_hash TEXT,
                  text_status TEXT NOT NULL DEFAULT 'ok',
                  lifecycle_status TEXT NOT NULL DEFAULT 'active',
                  ingested_at TEXT,
                  last_seen_at TEXT,
                  updated_at TEXT,
                  display_batch INTEGER,
                  display_family_sequence INTEGER,
                  display_attachment_sequence INTEGER
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE display_id_batches (
                  batch_number INTEGER PRIMARY KEY,
                  next_family_sequence INTEGER NOT NULL DEFAULT 1,
                  created_at TEXT NOT NULL,
                  updated_at TEXT NOT NULL
                )
                """
            )
            if with_row:
                connection.execute(
                    """
                    INSERT INTO documents (
                      id, display_id, parent_document_id, rel_path, file_name, file_type, file_size,
                      page_count, author, content_type, date_created, date_modified, title, subject,
                      participants, recipients, manual_field_locks_json, file_hash, content_hash,
                      text_status, lifecycle_status, ingested_at, last_seen_at, updated_at,
                      display_batch, display_family_sequence, display_attachment_sequence
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        1,
                        "DOC001.00000001",
                        None,
                        "legacy.txt",
                        "legacy.txt",
                        "txt",
                        12,
                        1,
                        "Legacy Author",
                        "E-Doc",
                        None,
                        None,
                        "Legacy Title",
                        None,
                        None,
                        None,
                        "[]",
                        "file-hash",
                        "content-hash",
                        "ok",
                        "active",
                        "2026-04-14T00:00:00Z",
                        "2026-04-14T00:00:00Z",
                        "2026-04-14T00:00:00Z",
                        1,
                        1,
                        None,
                    ),
                )
                connection.execute(
                    """
                    INSERT INTO display_id_batches (batch_number, next_family_sequence, created_at, updated_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    (1, 2, "2026-04-14T00:00:00Z", "2026-04-14T00:00:00Z"),
                )
            connection.commit()
        finally:
            connection.close()

    def normalized_document_row(self, row: sqlite3.Row) -> dict[str, object]:
        payload = {key: row[key] for key in row.keys()}
        custodians = retriever_tools.parse_document_custodians_json(payload.get("custodians_json"))
        payload["custodians"] = custodians
        payload["custodian"] = ", ".join(custodians) if custodians else None
        return payload

    def fetch_document_row(self, rel_path: str) -> dict[str, object]:
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            row = connection.execute(
                "SELECT * FROM documents WHERE rel_path = ?",
                (rel_path,),
            ).fetchone()
            self.assertIsNotNone(row)
            assert row is not None
            return self.normalized_document_row(row)
        finally:
            connection.close()

    def fetch_document_by_id(self, document_id: int) -> dict[str, object]:
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            row = connection.execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
            self.assertIsNotNone(row)
            assert row is not None
            return self.normalized_document_row(row)
        finally:
            connection.close()

    def fetch_dataset_row(self, dataset_id: int) -> sqlite3.Row:
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            row = connection.execute("SELECT * FROM datasets WHERE id = ?", (dataset_id,)).fetchone()
            self.assertIsNotNone(row)
            return row
        finally:
            connection.close()

    def fetch_custom_field_registry_row(self, field_name: str) -> sqlite3.Row:
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            row = connection.execute(
                "SELECT * FROM custom_fields_registry WHERE field_name = ?",
                (field_name,),
            ).fetchone()
            self.assertIsNotNone(row)
            assert row is not None
            return row
        finally:
            connection.close()

    def fetch_child_rows(self, parent_document_id: int) -> list[dict[str, object]]:
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            rows = connection.execute(
                """
                SELECT *
                FROM documents
                WHERE parent_document_id = ?
                ORDER BY id ASC
                """,
                (parent_document_id,),
            ).fetchall()
            return [self.normalized_document_row(row) for row in rows]
        finally:
            connection.close()

    def fetch_occurrence_rows(self, document_id: int) -> list[sqlite3.Row]:
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            return connection.execute(
                """
                SELECT *
                FROM document_occurrences
                WHERE document_id = ?
                ORDER BY id ASC
                """,
                (document_id,),
            ).fetchall()
        finally:
            connection.close()

    def fetch_email_threading_row(self, document_id: int) -> sqlite3.Row | None:
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            return connection.execute(
                """
                SELECT *
                FROM document_email_threading
                WHERE document_id = ?
                """,
                (document_id,),
            ).fetchone()
        finally:
            connection.close()

    def count_rows(self, table_name: str) -> int:
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            row = connection.execute(f"SELECT COUNT(*) AS count FROM {table_name}").fetchone()
            self.assertIsNotNone(row)
            return int(row["count"] or 0)
        finally:
            connection.close()

    def write_email_message(
        self,
        path: Path,
        *,
        subject: str,
        body_text: str,
        body_html: str | None = None,
        author: str = "Alice Example <alice@example.com>",
        recipients: str = "Bob Example <bob@example.com>",
        cc: str | None = "Carol Example <carol@example.com>",
        date_created: str = "Tue, 14 Apr 2026 10:00:00 +0000",
        message_id: str | None = None,
        in_reply_to: str | None = None,
        references: str | None = None,
        conversation_index: str | None = None,
        conversation_topic: str | None = None,
        attachment_name: str | None = None,
        attachment_text: str | None = None,
    ) -> None:
        message = EmailMessage()
        message["From"] = author
        message["To"] = recipients
        if cc is not None:
            message["Cc"] = cc
        message["Subject"] = subject
        message["Date"] = date_created
        if message_id is not None:
            message["Message-ID"] = message_id
        if in_reply_to is not None:
            message["In-Reply-To"] = in_reply_to
        if references is not None:
            message["References"] = references
        if conversation_index is not None:
            message["Conversation-Index"] = conversation_index
        if conversation_topic is not None:
            message["Conversation-Topic"] = conversation_topic
        message.set_content(body_text)
        if body_html is not None:
            message.add_alternative(body_html, subtype="html")
        if attachment_name is not None and attachment_text is not None:
            message.add_attachment(attachment_text, subtype="plain", filename=attachment_name)
        path.write_bytes(message.as_bytes(policy=policy.default))

    def create_structured_extraction_job_version(self, job_name: str) -> int:
        sanitized_job_name = re.sub(r"[^a-zA-Z0-9_]+", "_", job_name.strip()).strip("_").lower()
        if sanitized_job_name and sanitized_job_name[0].isdigit():
            sanitized_job_name = f"job_{sanitized_job_name}"
        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            job_name,
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            sanitized_job_name,
            "--instruction",
            "Extract a stable value.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        return int(create_version_payload["job_version"]["id"])

    def setup_randomized_sql_filter_corpus(self) -> list[dict[str, object]]:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("beta body\n", encoding="utf-8")
        (self.root / "gamma.txt").write_text("gamma body\n", encoding="utf-8")
        self.write_email_message(
            self.root / "thread.eml",
            subject="Filter thread",
            body_text="Parent email body text.",
            attachment_name="notes.txt",
            attachment_text="Attachment detail.",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 4)

        for field_name, field_type in (
            ("review_score", "integer"),
            ("review_date", "date"),
            ("is_hot", "boolean"),
            ("review_note", "text"),
        ):
            add_field_exit, _, _, _ = self.run_cli("add-field", str(self.root), field_name, field_type)
            self.assertEqual(add_field_exit, 0)

        alpha_row = self.fetch_document_row("alpha.txt")
        beta_row = self.fetch_document_row("beta.txt")
        gamma_row = self.fetch_document_row("gamma.txt")
        parent_row = self.fetch_document_row("thread.eml")
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        updates = {
            int(alpha_row["id"]): (1, "2026-04-01", 1, "alpha memo"),
            int(beta_row["id"]): (5, "2026-04-02", 0, "beta's note"),
            int(gamma_row["id"]): (None, "2026-04-03", None, "needs review"),
            int(parent_row["id"]): (7, "2026-04-04", 1, "mail thread"),
            int(child_row["id"]): (9, "2026-04-05", 0, "child's attachment"),
        }

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            for document_id, (review_score, review_date, is_hot, review_note) in updates.items():
                connection.execute(
                    """
                    UPDATE documents
                    SET review_score = ?, review_date = ?, is_hot = ?, review_note = ?
                    WHERE id = ?
                    """,
                    (review_score, review_date, is_hot, review_note, document_id),
                )
            connection.commit()

            rows = connection.execute(
                f"""
                SELECT id AS doc_id, file_name, review_note, review_score, review_date, is_hot
                FROM documents
                WHERE id IN ({", ".join("?" for _ in updates)})
                ORDER BY id ASC
                """,
                tuple(updates),
            ).fetchall()
        finally:
            connection.close()

        return [dict(row) for row in rows]

    def reference_doc_ids_for_sql_filter(
        self,
        corpus_rows: list[dict[str, object]],
        expression: str,
    ) -> list[int]:
        reference_connection = sqlite3.connect(":memory:")
        try:
            reference_connection.execute(
                """
                CREATE TABLE docs (
                  doc_id INTEGER PRIMARY KEY,
                  file_name TEXT,
                  review_note TEXT,
                  review_score INTEGER,
                  review_date TEXT,
                  is_hot INTEGER
                )
                """
            )
            reference_connection.executemany(
                """
                INSERT INTO docs (doc_id, file_name, review_note, review_score, review_date, is_hot)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        int(row["doc_id"]),
                        row["file_name"],
                        row["review_note"],
                        row["review_score"],
                        row["review_date"],
                        row["is_hot"],
                    )
                    for row in corpus_rows
                ],
            )
            return [
                int(row[0])
                for row in reference_connection.execute(
                    f"SELECT doc_id FROM docs WHERE {expression} ORDER BY doc_id ASC"
                ).fetchall()
            ]
        finally:
            reference_connection.close()

    def render_sql_filter_literal(self, field_type: str, value: object) -> str:
        if field_type == "boolean":
            return "TRUE" if bool(value) else "FALSE"
        if field_type == "integer":
            return str(int(value))
        escaped = str(value).replace("'", "''")
        return f"'{escaped}'"

    def build_random_sql_filter_predicate(self, rng: random.Random) -> dict[str, object]:
        field_name = rng.choice(sorted(RANDOMIZED_FILTER_FIELD_TYPES))
        field_type = RANDOMIZED_FILTER_FIELD_TYPES[field_name]
        operator = rng.choice(RANDOMIZED_FILTER_OPERATOR_CHOICES[field_type])
        predicate: dict[str, object] = {
            "kind": "predicate",
            "field_name": field_name,
            "field_type": field_type,
            "operator": operator,
        }
        if operator in {"IS NULL", "IS NOT NULL"}:
            return predicate
        if operator in {"LIKE", "NOT LIKE"}:
            predicate["operand"] = rng.choice(RANDOMIZED_FILTER_LIKE_PATTERNS[field_name])
            return predicate
        if operator in {"IN", "NOT IN"}:
            values = RANDOMIZED_FILTER_LITERAL_POOLS[field_name]
            predicate["operand"] = [rng.choice(values) for _ in range(rng.randint(1, 4))]
            return predicate
        if operator in {"BETWEEN", "NOT BETWEEN"}:
            values = RANDOMIZED_FILTER_LITERAL_POOLS[field_name]
            predicate["operand"] = (rng.choice(values), rng.choice(values))
            return predicate
        predicate["operand"] = rng.choice(RANDOMIZED_FILTER_LITERAL_POOLS[field_name])
        return predicate

    def build_random_sql_filter_node(self, rng: random.Random, *, depth: int = 0) -> dict[str, object]:
        if depth >= 3 or (depth > 0 and rng.random() < 0.45):
            node = self.build_random_sql_filter_predicate(rng)
        else:
            branch_kind = rng.choice(["and", "or", "not", "predicate"])
            if branch_kind == "predicate":
                node = self.build_random_sql_filter_predicate(rng)
            elif branch_kind == "not":
                node = {
                    "kind": "not",
                    "child": self.build_random_sql_filter_node(rng, depth=depth + 1),
                }
            else:
                node = {
                    "kind": branch_kind,
                    "left": self.build_random_sql_filter_node(rng, depth=depth + 1),
                    "right": self.build_random_sql_filter_node(rng, depth=depth + 1),
                }
        if depth > 0 and rng.random() < 0.2:
            return {"kind": "group", "child": node}
        return node

    def render_random_sql_filter_node(self, node: dict[str, object], *, parent_precedence: int = 0) -> str:
        kind = str(node["kind"])
        if kind == "predicate":
            field_name = str(node["field_name"])
            field_type = str(node["field_type"])
            operator = str(node["operator"])
            precedence = 4
            if operator in {"IS NULL", "IS NOT NULL"}:
                text = f"{field_name} {operator}"
            elif operator in {"LIKE", "NOT LIKE"}:
                operand = self.render_sql_filter_literal(field_type, node["operand"])
                text = f"{field_name} {operator} {operand}"
            elif operator in {"IN", "NOT IN"}:
                operands = ", ".join(
                    self.render_sql_filter_literal(field_type, value)
                    for value in node["operand"]
                )
                text = f"{field_name} {operator} ({operands})"
            elif operator in {"BETWEEN", "NOT BETWEEN"}:
                left_value, right_value = node["operand"]
                text = (
                    f"{field_name} {operator} "
                    f"{self.render_sql_filter_literal(field_type, left_value)} "
                    f"AND {self.render_sql_filter_literal(field_type, right_value)}"
                )
            else:
                operand = self.render_sql_filter_literal(field_type, node["operand"])
                text = f"{field_name} {operator} {operand}"
        elif kind == "group":
            return f"({self.render_random_sql_filter_node(node['child'])})"
        elif kind == "not":
            precedence = 3
            text = f"NOT {self.render_random_sql_filter_node(node['child'], parent_precedence=precedence)}"
        else:
            operator = "AND" if kind == "and" else "OR"
            precedence = 2 if operator == "AND" else 1
            left_text = self.render_random_sql_filter_node(node["left"], parent_precedence=precedence)
            right_text = self.render_random_sql_filter_node(node["right"], parent_precedence=precedence)
            text = f"{left_text} {operator} {right_text}"
        if precedence < parent_precedence:
            return f"({text})"
        return text

    def write_fake_pst_file(self, name: str = "mailbox.pst", content: bytes = b"pst-v1") -> Path:
        path = self.root / name
        path.write_bytes(content)
        return path

    def write_csv_rows(
        self,
        path: Path,
        *,
        fieldnames: list[str],
        rows: list[dict[str, object]],
    ) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)

    def write_fake_mbox_file(self, messages: list[EmailMessage], name: str = "mailbox.mbox") -> Path:
        path = self.root / name
        if path.exists():
            path.unlink()
        archive = mailbox.mbox(str(path), create=True)
        try:
            for message in messages:
                archive.add(message)
            archive.flush()
        finally:
            archive.close()
        return path

    def build_fake_mbox_message(
        self,
        *,
        subject: str | None,
        body_text: str,
        message_id: str | None,
        author: str | None = "Alice Example <alice@example.com>",
        recipients: str | None = "Bob Example <bob@example.com>",
        date_created: str = "Tue, 14 Apr 2026 10:00:00 +0000",
        attachment_name: str | None = None,
        attachment_text: str | None = None,
    ) -> EmailMessage:
        message = EmailMessage(policy=policy.default)
        if author is not None:
            message["From"] = author
        if recipients is not None:
            message["To"] = recipients
        if subject is not None:
            message["Subject"] = subject
        message["Date"] = date_created
        if message_id is not None:
            message["Message-ID"] = message_id
        message.set_content(body_text)
        if attachment_name is not None and attachment_text is not None:
            message.add_attachment(attachment_text, subtype="plain", filename=attachment_name)
        return message

    def build_fake_pst_message(
        self,
        *,
        source_item_id: str,
        subject: str | None,
        body_text: str,
        folder_path: str = "Inbox",
        attachment_name: str | None = None,
        attachment_text: str | None = None,
        author: str | None = "Alice Example <alice@example.com>",
        recipients: str | None = "Bob Example <bob@example.com>, Carol Example <carol@example.com>",
        date_created: str | None = "2026-04-14T10:00:00Z",
        message_class: str | None = None,
        transport_headers: str | None = None,
        conversation_topic: str | None = None,
        chat_threading: dict[str, object] | None = None,
    ) -> dict[str, object]:
        attachments: list[dict[str, object]] = []
        if attachment_name is not None and attachment_text is not None:
            attachments.append(
                {
                    "file_name": attachment_name,
                    "payload": attachment_text.encode("utf-8"),
                }
            )
        return {
            "source_item_id": source_item_id,
            "folder_path": folder_path,
            "subject": subject,
            "author": author,
            "recipients": recipients,
            "date_created": date_created,
            "message_class": message_class,
            "transport_headers": transport_headers,
            "conversation_topic": conversation_topic,
            "chat_threading": dict(chat_threading or {}),
            "text_body": body_text,
            "html_body": None,
            "attachments": attachments,
        }

    def write_xls_fixture(self, path: Path) -> None:
        try:
            import xlwt
        except Exception as exc:  # pragma: no cover - test helper dependency
            self.skipTest(f"xlwt unavailable for xls fixture generation: {exc}")
        workbook = xlwt.Workbook()
        sheet1 = workbook.add_sheet("Sheet1")
        sheet1.write(0, 0, "Name")
        sheet1.write(0, 1, "Value")
        sheet1.write(1, 0, "Alpha")
        sheet1.write(1, 1, 42)
        notes = workbook.add_sheet("Notes")
        notes.write(0, 0, "Memo")
        notes.write(1, 0, "Budget approved")
        workbook.save(str(path))

    def write_xlsx_fixture(self, path: Path) -> None:
        try:
            import openpyxl
            from openpyxl.chart import BarChart, Reference
            from openpyxl.comments import Comment
            from openpyxl.workbook.defined_name import DefinedName
            from openpyxl.worksheet.datavalidation import DataValidation
        except Exception as exc:  # pragma: no cover - test helper dependency
            self.skipTest(f"openpyxl unavailable for xlsx fixture generation: {exc}")
        from datetime import datetime

        workbook = openpyxl.Workbook()
        budget = workbook.active
        budget.title = "Budget"
        budget.append(["Department", "Amount", "Quarter"])
        budget.append(["Engineering", 1200, "Q1"])
        budget.append(["Sales", 900, "Q2"])
        budget["A2"].comment = Comment("Needs review", "Sergey")
        budget["A2"].hyperlink = "https://example.com/departments/engineering"

        validation = DataValidation(type="list", formula1='"Q1,Q2,Q3,Q4"')
        budget.add_data_validation(validation)
        validation.add("C2:C10")

        chart = BarChart()
        data = Reference(budget, min_col=2, min_row=1, max_row=3)
        chart.add_data(data, titles_from_data=True)
        chart.title = "Budget Totals"
        chart.x_axis.title = "Department"
        chart.y_axis.title = "Amount"
        budget.add_chart(chart, "E2")

        notes = workbook.create_sheet("Notes")
        notes.append(["Memo"])
        notes.append(["Budget approved"])

        workbook.defined_names.add(DefinedName("DeptList", attr_text="'Budget'!$A$2:$A$3"))
        workbook.properties.creator = "Rachel Green"
        workbook.properties.lastModifiedBy = "Sergey"
        workbook.properties.title = "Quarterly Budget Workbook"
        workbook.properties.subject = "Finance Planning"
        workbook.properties.created = datetime(2026, 4, 20, 10, 0, 0)
        workbook.properties.modified = datetime(2026, 4, 20, 11, 30, 0)
        workbook.save(path)

    def write_pptx_fixture(self, path: Path) -> None:
        image_bytes = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+a9mQAAAAASUVORK5CYII="
        )
        entries = {
            "docProps/core.xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
 xmlns:dc="http://purl.org/dc/elements/1.1/"
 xmlns:dcterms="http://purl.org/dc/terms/"
 xmlns:dcmitype="http://purl.org/dc/dcmitype/"
 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>Quarterly Strategy Deck</dc:title>
  <dc:subject>Board Update</dc:subject>
  <dc:creator>Rachel Green</dc:creator>
  <dcterms:created xsi:type="dcterms:W3CDTF">2026-04-15T10:00:00Z</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">2026-04-15T11:00:00Z</dcterms:modified>
</cp:coreProperties>
""",
            "ppt/presentation.xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:presentation xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <p:sldIdLst>
    <p:sldId id="256" r:id="rId1"/>
    <p:sldId id="257" r:id="rId2"/>
  </p:sldIdLst>
</p:presentation>
""",
            "ppt/_rels/presentation.xml.rels": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide2.xml"/>
</Relationships>
""",
            "ppt/slides/slide1.xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <p:cSld>
    <p:spTree>
      <p:nvGrpSpPr><p:cNvPr id="1" name=""/></p:nvGrpSpPr>
      <p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>
      <p:sp>
        <p:nvSpPr><p:cNvPr id="2" name="Title"/><p:cNvSpPr/><p:nvPr><p:ph type="title"/></p:nvPr></p:nvSpPr>
        <p:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="4000" cy="800"/></a:xfrm></p:spPr>
        <p:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Q3 Revenue Review</a:t></a:r></a:p></p:txBody>
      </p:sp>
      <p:sp>
        <p:nvSpPr><p:cNvPr id="3" name="Body 1"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>
        <p:spPr><a:xfrm><a:off x="100" y="900"/><a:ext cx="4000" cy="1200"/></a:xfrm></p:spPr>
        <p:txBody>
          <a:bodyPr/><a:lstStyle/>
          <a:p><a:r><a:t>Revenue up 15%</a:t></a:r></a:p>
          <a:p><a:r><a:t>Pipeline remains strong</a:t></a:r></a:p>
        </p:txBody>
      </p:sp>
      <p:sp>
        <p:nvSpPr><p:cNvPr id="4" name="Body 2"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>
        <p:spPr><a:xfrm><a:off x="200" y="2300"/><a:ext cx="4000" cy="800"/></a:xfrm></p:spPr>
        <p:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Operating margin improved</a:t></a:r></a:p></p:txBody>
      </p:sp>
    </p:spTree>
  </p:cSld>
</p:sld>
""",
            "ppt/slides/_rels/slide1.xml.rels": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/notesSlide" Target="../notesSlides/notesSlide1.xml"/>
</Relationships>
""",
            "ppt/notesSlides/notesSlide1.xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:notes xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
  <p:cSld>
    <p:spTree>
      <p:nvGrpSpPr><p:cNvPr id="1" name=""/></p:nvGrpSpPr>
      <p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>
      <p:sp>
        <p:nvSpPr><p:cNvPr id="2" name="Notes"/><p:cNvSpPr/><p:nvPr><p:ph type="body"/></p:nvPr></p:nvSpPr>
        <p:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="4000" cy="800"/></a:xfrm></p:spPr>
        <p:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Emphasize retained enterprise customers.</a:t></a:r></a:p></p:txBody>
      </p:sp>
    </p:spTree>
  </p:cSld>
</p:notes>
""",
            "ppt/slides/slide2.xml": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <p:cSld>
    <p:spTree>
      <p:nvGrpSpPr><p:cNvPr id="1" name=""/></p:nvGrpSpPr>
      <p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>
      <p:sp>
        <p:nvSpPr><p:cNvPr id="2" name="Title"/><p:cNvSpPr/><p:nvPr><p:ph type="title"/></p:nvPr></p:nvSpPr>
        <p:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="4000" cy="800"/></a:xfrm></p:spPr>
        <p:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Hiring Plan</a:t></a:r></a:p></p:txBody>
      </p:sp>
      <p:sp>
        <p:nvSpPr><p:cNvPr id="3" name="Body"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>
        <p:spPr><a:xfrm><a:off x="150" y="1000"/><a:ext cx="4000" cy="1000"/></a:xfrm></p:spPr>
        <p:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Open 12 roles in engineering</a:t></a:r></a:p></p:txBody>
      </p:sp>
      <p:pic>
        <p:nvPicPr>
          <p:cNvPr id="4" name="Stark Therapeutics logo" descr="Stark Therapeutics logo"/>
          <p:cNvPicPr/>
          <p:nvPr/>
        </p:nvPicPr>
        <p:blipFill>
          <a:blip r:embed="rId1"/>
          <a:stretch><a:fillRect/></a:stretch>
        </p:blipFill>
        <p:spPr>
          <a:xfrm><a:off x="150" y="2200"/><a:ext cx="952500" cy="952500"/></a:xfrm>
          <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
        </p:spPr>
      </p:pic>
    </p:spTree>
  </p:cSld>
</p:sld>
""",
            "ppt/slides/_rels/slide2.xml.rels": """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image1.png"/>
</Relationships>
""",
        }
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, content in entries.items():
                archive.writestr(name, content)
            archive.writestr("ppt/media/image1.png", image_bytes)

    def write_tiff_fixture(self, path: Path, color: tuple[int, int, int]) -> None:
        try:
            from PIL import Image
        except Exception as exc:  # pragma: no cover - test helper dependency
            self.skipTest(f"Pillow unavailable for TIFF fixture generation: {exc}")
        image = Image.new("RGB", (8, 8), color)
        image.save(path, format="TIFF")

    def write_minimal_pdf(self, path: Path, title: str) -> None:
        path.write_bytes(
            (
                b"%PDF-1.1\n"
                b"1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n"
                b"2 0 obj<< /Type /Pages /Count 1 /Kids [3 0 R] >>endobj\n"
                b"3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 200 200] /Contents 4 0 R >>endobj\n"
                + f"4 0 obj<< /Length {len(title) + 30} >>stream\nBT /F1 12 Tf 20 100 Td ({title}) Tj ET\nendstream endobj\n".encode("latin-1")
                + b"trailer<< /Root 1 0 R >>\n%%EOF\n"
            )
        )

    def write_production_fixture(
        self,
        *,
        production_name: str = "Synthetic_Production",
        control_prefix: str = "PDX",
        loadfile_volume_prefix: str | None = None,
    ) -> Path:
        production_root = self.root / production_name
        data_dir = production_root / "DATA"
        text_dir = production_root / "TEXT" / "TEXT001"
        image_dir = production_root / "IMAGES" / "IMG001"
        native_dir = production_root / "NATIVES" / "NAT001"
        for directory in (data_dir, text_dir, image_dir, native_dir):
            directory.mkdir(parents=True, exist_ok=True)

        loadfile_root = (loadfile_volume_prefix or production_root.name).strip()

        def bates(number: int) -> str:
            return f"{control_prefix}{number:06d}"

        def loadfile_path(*parts: str) -> str:
            return ".\\" + "\\".join([loadfile_root, *parts])

        (text_dir / f"{bates(1)}.txt").write_text(
            (
                "From: Elena Steven <elena@example.com>\n"
                "To: Harry Montoro <harry@example.com>\n"
                "Date: Tue, 14 Apr 2026 10:32:00 +0000\n"
                "Subject: Attachment Handling\n\n"
                "Parent production memo\n"
                "Discuss attachment handling.\n"
            ),
            encoding="utf-8",
        )
        (text_dir / f"{bates(3)}.txt").write_text(
            (
                "From: Review Team\n"
                "Sent: 04/14/2026 09:00 AM\n\n"
                "Case status update\n"
                "Contains follow-up details.\n"
            ),
            encoding="utf-8",
        )
        (text_dir / f"{bates(4)}.txt").write_text("Native-backed production doc\nUse native preview first.\n", encoding="utf-8")

        self.write_tiff_fixture(image_dir / f"{bates(1)}.tif", (255, 0, 0))
        self.write_tiff_fixture(image_dir / f"{bates(2)}.tif", (0, 255, 0))
        self.write_tiff_fixture(image_dir / f"{bates(3)}.tif", (0, 0, 255))
        self.write_tiff_fixture(image_dir / f"{bates(5)}.tif", (128, 128, 0))
        self.write_tiff_fixture(image_dir / f"{bates(6)}.tif", (0, 128, 128))
        self.write_minimal_pdf(native_dir / f"{bates(4)}.pdf", "Native preview document")

        headers = ["Begin Bates", "End Bates", "Begin Attachment", "End Attachment", "Text Precedence", "FILE_PATH"]
        rows = [
            [bates(1), bates(2), bates(1), bates(3), loadfile_path("TEXT", "TEXT001", f"{bates(1)}.txt"), ""],
            [bates(3), bates(3), "", "", loadfile_path("TEXT", "TEXT001", f"{bates(3)}.txt"), ""],
            [bates(4), bates(4), "", "", loadfile_path("TEXT", "TEXT001", f"{bates(4)}.txt"), loadfile_path("NATIVES", "NAT001", f"{bates(4)}.pdf")],
            [bates(5), bates(6), "", "", "", ""],
        ]
        delimiter = b"\x14"
        quote = b"\xfe"

        def dat_line(fields: list[str]) -> bytes:
            return delimiter.join(quote + field.encode("latin-1") + quote for field in fields) + b"\r\n"

        (data_dir / f"{production_name}.dat").write_bytes(dat_line(headers) + b"".join(dat_line(row) for row in rows))

        opt_lines = [
            f"{bates(1)},{production_name},{loadfile_path('IMAGES', 'IMG001', f'{bates(1)}.tif')},Y,,,2",
            f"{bates(2)},{production_name},{loadfile_path('IMAGES', 'IMG001', f'{bates(2)}.tif')},,,,",
            f"{bates(3)},{production_name},{loadfile_path('IMAGES', 'IMG001', f'{bates(3)}.tif')},Y,,,1",
            f"{bates(5)},{production_name},{loadfile_path('IMAGES', 'IMG001', f'{bates(5)}.tif')},Y,,,2",
            f"{bates(6)},{production_name},{loadfile_path('IMAGES', 'IMG001', f'{bates(6)}.tif')},,,,",
        ]
        (data_dir / f"{production_name}.opt").write_text("\n".join(opt_lines) + "\n", encoding="utf-8")
        return production_root

    def test_bootstrap_migrates_legacy_schema_and_backfills_content_type(self) -> None:
        self.create_legacy_documents_table(with_row=True)

        result = retriever_tools.bootstrap(self.root)

        self.assertEqual(result["schema_version"], retriever_tools.SCHEMA_VERSION)
        self.assertEqual(result["tool_version"], retriever_tools.TOOL_VERSION)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            columns = retriever_tools.table_columns(connection, "documents")
            preview_columns = retriever_tools.table_columns(connection, "document_previews")
            self.assertIn("content_type", columns)
            self.assertIn("custodians_json", columns)
            self.assertNotIn("custodian", columns)
            self.assertIn("participants", columns)
            self.assertIn("control_number", columns)
            self.assertIn("conversation_id", columns)
            self.assertIn("conversation_assignment_mode", columns)
            self.assertIn("dataset_id", columns)
            self.assertIn("parent_document_id", columns)
            self.assertIn("child_document_kind", columns)
            self.assertIn("root_message_key", columns)
            self.assertIn("target_fragment", preview_columns)
            row = connection.execute(
                """
                SELECT content_type, custodians_json, participants, control_number, conversation_id,
                       conversation_assignment_mode, dataset_id, parent_document_id,
                       child_document_kind, root_message_key
                FROM documents
                WHERE id = 1
                """
            ).fetchone()
            conversations_table = connection.execute(
                """
                SELECT name
                FROM sqlite_master
                WHERE type = 'table' AND name = 'conversations'
                """
            ).fetchone()
            email_threading_table = connection.execute(
                """
                SELECT name
                FROM sqlite_master
                WHERE type = 'table' AND name = 'document_email_threading'
                """
            ).fetchone()
            chat_threading_table = connection.execute(
                """
                SELECT name
                FROM sqlite_master
                WHERE type = 'table' AND name = 'document_chat_threading'
                """
            ).fetchone()
            dataset_row = connection.execute(
                """
                SELECT *
                FROM datasets
                WHERE id = ?
                """,
                (row["dataset_id"],),
            ).fetchone()
            control_number_batch_row = connection.execute(
                """
                SELECT batch_number, next_family_sequence
                FROM control_number_batches
                WHERE batch_number = 1
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(row["content_type"], "E-Doc")
        self.assertEqual(retriever_tools.parse_document_custodians_json(row["custodians_json"]), [])
        self.assertIsNone(row["participants"])
        self.assertEqual(row["control_number"], "DOC001.00000001")
        self.assertIsNone(row["conversation_id"])
        self.assertEqual(row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)
        self.assertIsNotNone(row["dataset_id"])
        self.assertIsNone(row["parent_document_id"])
        self.assertIsNone(row["child_document_kind"])
        self.assertIsNone(row["root_message_key"])
        self.assertIsNotNone(conversations_table)
        self.assertIsNotNone(email_threading_table)
        self.assertIsNotNone(chat_threading_table)
        self.assertIsNotNone(dataset_row)
        self.assertEqual(dataset_row["source_kind"], retriever_tools.FILESYSTEM_SOURCE_KIND)
        self.assertEqual(dataset_row["dataset_locator"], ".")
        self.assertEqual(dataset_row["dataset_name"], self.root.name)
        self.assertIsNotNone(control_number_batch_row)
        self.assertEqual(control_number_batch_row["next_family_sequence"], 2)

        runtime = json.loads(self.paths["runtime_path"].read_text(encoding="utf-8"))
        self.assertEqual(runtime["tool_version"], retriever_tools.TOOL_VERSION)
        self.assertEqual(runtime["schema_version"], retriever_tools.SCHEMA_VERSION)
        self.assertEqual(runtime["template_sha256"], retriever_tools.sha256_file(TOOL_PATH))

    def test_entity_schema_and_parser_foundation(self) -> None:
        result = retriever_tools.bootstrap(self.root)
        self.assertEqual(result["schema_version"], retriever_tools.SCHEMA_VERSION)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            self.assertIn("dataset_source_id", retriever_tools.table_columns(connection, "document_occurrences"))
            self.assertIn("entity_hints_json", retriever_tools.table_columns(connection, "document_occurrences"))
            dataset_columns = retriever_tools.table_columns(connection, "datasets")
            for column_name in (
                "allow_auto_merge",
                "email_auto_merge",
                "handle_auto_merge",
                "phone_auto_merge",
                "name_auto_merge",
                "external_id_auto_merge_names_json",
            ):
                self.assertIn(column_name, dataset_columns)
            for table_name in (
                "entities",
                "entity_identifiers",
                "entity_resolution_keys",
                "document_entities",
                "entity_overrides",
                "entity_merge_blocks",
            ):
                self.assertTrue(retriever_tools.table_exists(connection, table_name), table_name)
        finally:
            connection.close()

        candidates = retriever_tools.parse_entity_candidates(
            "Doe, Jane <jane@example.com>; Support <support@example.com>",
            role="recipient",
        )
        self.assertEqual(len(candidates), 2)
        jane = candidates[0]
        self.assertEqual(jane["entity_type"], retriever_tools.ENTITY_TYPE_PERSON)
        jane_identifiers = {item["identifier_type"]: item for item in jane["identifiers"]}
        self.assertEqual(jane_identifiers["email"]["normalized_value"], "jane@example.com")
        self.assertEqual(jane_identifiers["name"]["normalized_full_name"], "jane doe")
        self.assertEqual(jane_identifiers["name"]["normalized_sort_name"], "doe jane")
        self.assertEqual(candidates[1]["entity_type"], retriever_tools.ENTITY_TYPE_SHARED_MAILBOX)

        mailbox_candidates = retriever_tools.parse_entity_candidates(
            (
                "Beagle Team <hello@discoverbeagle.com>; "
                "Everlaw Legalweek <legalevents@alm.com>; "
                "Slack <notification@slack.com>; "
                "Demyanov Family Team <gmail-noreply@google.com>; "
                "All Company <allcompany@maxusivanovgmail.onmicrosoft.com>"
            ),
            role="recipient",
        )
        mailbox_types = {
            next(
                identifier["normalized_value"]
                for identifier in candidate["identifiers"]
                if identifier["identifier_type"] == "email"
            ): candidate["entity_type"]
            for candidate in mailbox_candidates
        }
        self.assertEqual(mailbox_types["hello@discoverbeagle.com"], retriever_tools.ENTITY_TYPE_SHARED_MAILBOX)
        self.assertEqual(mailbox_types["legalevents@alm.com"], retriever_tools.ENTITY_TYPE_SHARED_MAILBOX)
        self.assertEqual(mailbox_types["notification@slack.com"], retriever_tools.ENTITY_TYPE_SYSTEM_MAILBOX)
        self.assertEqual(mailbox_types["gmail-noreply@google.com"], retriever_tools.ENTITY_TYPE_SYSTEM_MAILBOX)
        self.assertEqual(
            mailbox_types["allcompany@maxusivanovgmail.onmicrosoft.com"],
            retriever_tools.ENTITY_TYPE_SHARED_MAILBOX,
        )

    def test_source_custodian_inference_handles_archive_basename_clues(self) -> None:
        self.assertEqual(
            retriever_tools.infer_source_custodian(
                source_kind=retriever_tools.PST_SOURCE_KIND,
                source_rel_path="Sergey@Example.COM.pst",
            ),
            "sergey@example.com",
        )
        self.assertEqual(
            retriever_tools.infer_source_custodian(
                source_kind=retriever_tools.MBOX_SOURCE_KIND,
                source_rel_path="exports/Jane Doe Mailbox.mbox",
            ),
            "Jane Doe",
        )
        self.assertEqual(
            retriever_tools.infer_source_custodian(
                source_kind=retriever_tools.PST_SOURCE_KIND,
                source_rel_path="Legal/Acme Corp Archive.pst",
            ),
            "Acme Corp",
        )
        self.assertIsNone(
            retriever_tools.infer_source_custodian(
                source_kind=retriever_tools.PST_SOURCE_KIND,
                source_rel_path="noreply@example.com.pst",
            )
        )
        self.assertIsNone(
            retriever_tools.infer_source_custodian(
                source_kind=retriever_tools.MBOX_SOURCE_KIND,
                source_rel_path="support@example.com.mbox",
            )
        )
        self.assertEqual(
            retriever_tools.infer_source_custodian(
                source_kind=retriever_tools.MBOX_SOURCE_KIND,
                source_rel_path="mailbox.mbox",
            ),
            "mailbox",
        )

    def test_ingest_syncs_author_entities_by_email_resolution_key(self) -> None:
        self.write_email_message(
            self.root / "first.eml",
            subject="First entity thread",
            body_text="First body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            message_id="<first-entity@example.com>",
        )
        self.write_email_message(
            self.root / "second.eml",
            subject="Second entity thread",
            body_text="Second body",
            author="Alice A. <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            message_id="<second-entity@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        first_row = self.fetch_document_row("first.eml")
        second_row = self.fetch_document_row("second.eml")
        self.assertEqual(first_row["author"], "Alice Example <alice@example.com>")
        self.assertEqual(second_row["author"], "Alice Example <alice@example.com>")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            author_rows = connection.execute(
                """
                SELECT DISTINCT de.entity_id
                FROM document_entities de
                WHERE de.role = 'author'
                ORDER BY de.entity_id ASC
                """
            ).fetchall()
            self.assertEqual(len(author_rows), 1)
            author_entity_id = int(author_rows[0]["entity_id"])
            key_row = connection.execute(
                """
                SELECT *
                FROM entity_resolution_keys
                WHERE key_type = 'email' AND normalized_value = 'alice@example.com'
                """
            ).fetchone()
            self.assertIsNotNone(key_row)
            self.assertEqual(int(key_row["entity_id"]), author_entity_id)
            identifier_rows = connection.execute(
                """
                SELECT identifier_type, normalized_value
                FROM entity_identifiers
                WHERE entity_id = ?
                ORDER BY identifier_type ASC, normalized_value ASC
                """,
                (author_entity_id,),
            ).fetchall()
        finally:
            connection.close()

        identifier_pairs = {(row["identifier_type"], row["normalized_value"]) for row in identifier_rows}
        self.assertIn(("email", "alice@example.com"), identifier_pairs)
        self.assertIn(("name", "alice example"), identifier_pairs)

    def test_ingest_reuses_shared_mailbox_entity_for_author_and_participant(self) -> None:
        self.write_email_message(
            self.root / "all-company.eml",
            subject="The new All Company group is ready",
            body_text="Group ready body",
            author="All Company <allcompany@maxusivanovgmail.onmicrosoft.com>",
            recipients="Reviewer <reviewer@example.com>",
            cc=None,
            message_id="<all-company@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        document_row = self.fetch_document_row("all-company.eml")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            role_rows = connection.execute(
                """
                SELECT de.role, de.entity_id, e.entity_type, e.entity_origin, e.display_name, e.primary_email
                FROM document_entities de
                JOIN entities e ON e.id = de.entity_id
                WHERE de.document_id = ?
                  AND de.role IN ('author', 'participant')
                  AND e.primary_email = 'allcompany@maxusivanovgmail.onmicrosoft.com'
                ORDER BY de.role ASC
                """,
                (document_row["id"],),
            ).fetchall()
            key_count_row = connection.execute(
                """
                SELECT COUNT(*) AS key_count
                FROM entity_resolution_keys
                WHERE key_type = 'email'
                  AND normalized_value = 'allcompany@maxusivanovgmail.onmicrosoft.com'
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual({row["role"] for row in role_rows}, {"author", "participant"})
        self.assertEqual({int(row["entity_id"]) for row in role_rows}, {int(role_rows[0]["entity_id"])})
        self.assertEqual(
            {row["entity_type"] for row in role_rows},
            {retriever_tools.ENTITY_TYPE_SHARED_MAILBOX},
        )
        self.assertEqual(
            {row["entity_origin"] for row in role_rows},
            {retriever_tools.ENTITY_ORIGIN_IDENTIFIED},
        )
        self.assertEqual(key_count_row["key_count"], 1)

    def test_entity_rebuild_and_read_commands(self) -> None:
        self.write_email_message(
            self.root / "first.eml",
            subject="First entity command thread",
            body_text="First body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            message_id="<first-entity-command@example.com>",
        )
        self.write_email_message(
            self.root / "second.eml",
            subject="Second entity command thread",
            body_text="Second body",
            author="Alice A. <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            message_id="<second-entity-command@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        exit_code, rebuild_payload, _, _ = self.run_cli(
            "rebuild-entities",
            str(self.root),
            "--batch-size",
            "1",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(rebuild_payload)
        assert rebuild_payload is not None
        self.assertEqual(rebuild_payload["status"], "ok")
        self.assertEqual(rebuild_payload["mode"], "full")
        self.assertEqual(rebuild_payload["documents_scanned"], 2)
        self.assertEqual(rebuild_payload["documents_synced"], 2)
        self.assertGreaterEqual(rebuild_payload["auto_links_created"], 4)
        self.assertGreaterEqual(rebuild_payload["active_entity_count"], 2)

        exit_code, list_payload, _, _ = self.run_cli(
            "list-entities",
            str(self.root),
            "--query",
            "alice@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        assert list_payload is not None
        alice_entities = [
            entity
            for entity in list_payload["entities"]
            if entity["primary_email"] == "alice@example.com"
        ]
        self.assertEqual(len(alice_entities), 1)
        alice_entity = alice_entities[0]
        self.assertEqual(alice_entity["label"], "Alice Example <alice@example.com>")
        self.assertEqual(alice_entity["document_count"], 2)
        self.assertIn("author", alice_entity["roles"])

        exit_code, show_payload, _, _ = self.run_cli(
            "show-entity",
            str(self.root),
            str(alice_entity["id"]),
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(show_payload)
        assert show_payload is not None
        self.assertEqual(show_payload["entity"]["primary_email"], "alice@example.com")
        role_counts = {item["role"]: item["document_count"] for item in show_payload["role_counts"]}
        self.assertEqual(role_counts["author"], 2)
        self.assertEqual(
            {
                (identifier["identifier_type"], identifier["normalized_value"])
                for identifier in show_payload["identifiers"]
            }
            & {("email", "alice@example.com"), ("name", "alice example")},
            {("email", "alice@example.com"), ("name", "alice example")},
        )
        author_links = [item for item in show_payload["documents"] if item["role"] == "author"]
        self.assertEqual(len(author_links), 2)

        exit_code, entity_filter_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "author = 'alice@example.com'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(entity_filter_payload)
        assert entity_filter_payload is not None
        self.assertEqual(entity_filter_payload["total_hits"], 2)

        exit_code, entity_id_filter_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            f"author_entity_id = {alice_entity['id']}",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(entity_id_filter_payload)
        assert entity_id_filter_payload is not None
        self.assertEqual(entity_id_filter_payload["total_hits"], 2)

        exit_code, raw_author_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "raw_author LIKE '%Alice A%'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(raw_author_payload)
        assert raw_author_payload is not None
        self.assertEqual(raw_author_payload["total_hits"], 1)
        self.assertEqual(raw_author_payload["results"][0]["file_name"], "second.eml")

        exit_code, inventory_payload, _, _ = self.run_cli(
            "list-entity-role-inventory",
            str(self.root),
            "--role",
            "author",
            "--filter",
            "author = 'alice@example.com'",
            "--examples",
            "1",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(inventory_payload)
        assert inventory_payload is not None
        author_inventory_rows = [
            row
            for row in inventory_payload["rows"]
            if row["entity_id"] == alice_entity["id"] and row["role"] == "author"
        ]
        self.assertEqual(len(author_inventory_rows), 1)
        self.assertEqual(author_inventory_rows[0]["document_count"], 2)
        self.assertEqual(len(author_inventory_rows[0]["examples"]), 1)

    def test_resumable_entity_rebuild_full_run_completes(self) -> None:
        self.write_email_message(
            self.root / "first.eml",
            subject="First resumable entity rebuild",
            body_text="First body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            message_id="<first-resumable-entity@example.com>",
        )
        self.write_email_message(
            self.root / "second.eml",
            subject="Second resumable entity rebuild",
            body_text="Second body",
            author="Alice A. <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            message_id="<second-resumable-entity@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        start_exit, start_payload, _, _ = self.run_cli(
            "rebuild-entities-start",
            str(self.root),
            "--batch-size",
            "1",
            "--budget-seconds",
            "35",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        assert start_payload is not None
        self.assertTrue(start_payload["created"])
        self.assertEqual(start_payload["mode"], "full")
        self.assertEqual(start_payload["phase"], "resetting")
        run_id = str(start_payload["run_id"])
        self.assertIn("rebuild-entities-run-step", start_payload["next_recommended_commands"][0])

        legacy_exit, legacy_payload, _, _ = self.run_cli("rebuild-entities", str(self.root))
        self.assertEqual(legacy_exit, 2)
        self.assertIsNotNone(legacy_payload)
        assert legacy_payload is not None
        self.assertEqual(legacy_payload["error"], "active_entity_rebuild_run")

        final_payload: dict[str, object] | None = None
        executed_steps: list[str] = []
        for _ in range(10):
            step_exit, step_payload, _, _ = self.run_cli(
                "rebuild-entities-run-step",
                str(self.root),
                "--run-id",
                run_id,
                "--budget-seconds",
                "35",
            )
            self.assertEqual(step_exit, 0)
            self.assertIsNotNone(step_payload)
            assert step_payload is not None
            executed_steps.extend(str(step) for step in step_payload["executed_steps"])
            final_payload = step_payload
            if step_payload["run"]["status"] == "completed":
                break
        self.assertIsNotNone(final_payload)
        assert final_payload is not None
        self.assertEqual(final_payload["run"]["status"], "completed")
        self.assertIn("reset", executed_steps)
        self.assertIn("plan", executed_steps)
        self.assertIn("rebuild", executed_steps)
        self.assertEqual(final_payload["run"]["counts"]["work_items"]["committed"], 2)
        self.assertEqual(final_payload["run"]["counts"]["work_items"]["failed"], 0)
        self.assertEqual(final_payload["run"]["progress"]["documents_synced"], 2)
        self.assertGreaterEqual(final_payload["run"]["progress"]["auto_links_created"], 4)
        self.assertGreaterEqual(final_payload["run"]["progress"]["reset_counts"]["auto_document_links_deleted"], 4)

        status_exit, status_payload, _, _ = self.run_cli(
            "rebuild-entities-status",
            str(self.root),
            "--run-id",
            run_id,
        )
        self.assertEqual(status_exit, 0)
        self.assertIsNotNone(status_payload)
        assert status_payload is not None
        self.assertEqual(status_payload["status"], "completed")

        list_exit, list_payload, _, _ = self.run_cli(
            "list-entities",
            str(self.root),
            "--query",
            "alice@example.com",
        )
        self.assertEqual(list_exit, 0)
        self.assertIsNotNone(list_payload)
        assert list_payload is not None
        alice_entities = [
            entity
            for entity in list_payload["entities"]
            if entity["primary_email"] == "alice@example.com"
        ]
        self.assertEqual(len(alice_entities), 1)
        self.assertEqual(alice_entities[0]["document_count"], 2)

    def test_resumable_entity_rebuild_selected_documents(self) -> None:
        self.write_email_message(
            self.root / "first.eml",
            subject="Selected entity rebuild one",
            body_text="First body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            message_id="<selected-entity-one@example.com>",
        )
        self.write_email_message(
            self.root / "second.eml",
            subject="Selected entity rebuild two",
            body_text="Second body",
            author="Carol Example <carol@example.com>",
            recipients="Dan Example <dan@example.com>",
            message_id="<selected-entity-two@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        document_row = self.fetch_document_row("first.eml")

        start_exit, start_payload, _, _ = self.run_cli(
            "rebuild-entities-start",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
            "--batch-size",
            "1",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        assert start_payload is not None
        self.assertEqual(start_payload["mode"], "selected")
        self.assertEqual(start_payload["phase"], "planning")
        run_id = str(start_payload["run_id"])

        step_exit, step_payload, _, _ = self.run_cli(
            "rebuild-entities-run-step",
            str(self.root),
            "--run-id",
            run_id,
        )
        self.assertEqual(step_exit, 0)
        self.assertIsNotNone(step_payload)
        assert step_payload is not None
        self.assertEqual(step_payload["run"]["status"], "completed")
        self.assertNotIn("reset", step_payload["executed_steps"])
        self.assertIn("plan", step_payload["executed_steps"])
        self.assertIn("rebuild", step_payload["executed_steps"])
        self.assertEqual(step_payload["run"]["counts"]["work_items"]["committed"], 1)
        self.assertEqual(step_payload["run"]["progress"]["documents_synced"], 1)
        self.assertEqual(step_payload["run"]["progress"]["reset_counts"]["auto_document_links_deleted"], 0)

    def test_list_entities_paginates_sorts_and_seeds_slash_entity_browse(self) -> None:
        for name, email in [
            ("Alice Example", "alice@people.test"),
            ("Bob Example", "bob@people.test"),
            ("Carol Example", "carol@people.test"),
            ("Dave Example", "dave@people.test"),
        ]:
            self.write_email_message(
                self.root / f"{email}.eml",
                subject=f"{name} note",
                body_text=f"{name} body",
                author=f"{name} <{email}>",
                recipients="Reviewer <reviewer@review.test>",
                message_id=f"<{email}>",
            )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 4)

        exit_code, first_payload, _, _ = self.run_cli(
            "list-entities",
            str(self.root),
            "--query",
            "people.test",
            "--limit",
            "2",
            "--sort",
            "display_name",
            "--order",
            "asc",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(first_payload)
        assert first_payload is not None
        self.assertEqual(first_payload["offset"], 0)
        self.assertEqual(first_payload["limit"], 2)
        self.assertEqual(first_payload["total_hits"], 4)
        self.assertEqual(
            [entity["primary_email"] for entity in first_payload["entities"]],
            ["alice@people.test", "bob@people.test"],
        )

        exit_code, second_payload, _, _ = self.run_cli(
            "list-entities",
            str(self.root),
            "--query",
            "people.test",
            "--limit",
            "2",
            "--offset",
            "2",
            "--sort",
            "display_name",
            "--order",
            "asc",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(second_payload)
        assert second_payload is not None
        self.assertEqual(second_payload["offset"], 2)
        self.assertEqual(
            [entity["primary_email"] for entity in second_payload["entities"]],
            ["carol@people.test", "dave@people.test"],
        )

        state = retriever_tools.read_session_state(self.paths)
        self.assertEqual(state["browse_mode"], retriever_tools.BROWSE_MODE_ENTITIES)
        self.assertEqual(state["display"][retriever_tools.BROWSE_MODE_ENTITIES]["page_size"], 2)
        self.assertEqual(
            state["browsing"][retriever_tools.BROWSE_MODE_ENTITIES]["sort"],
            [["display_name", "asc"]],
        )

        page_exit, page_stdout, _ = self.run_cli_raw("slash", str(self.root), "/page first")
        self.assertEqual(page_exit, 0)
        self.assertIn("alice@people.test", page_stdout)
        self.assertIn("bob@people.test", page_stdout)
        self.assertNotIn("carol@people.test", page_stdout)

        page_size_exit, page_size_stdout, _ = self.run_cli_raw("slash", str(self.root), "/page-size 3")
        self.assertEqual(page_size_exit, 0)
        self.assertIn("alice@people.test", page_size_stdout)
        self.assertIn("bob@people.test", page_size_stdout)
        self.assertIn("carol@people.test", page_size_stdout)
        self.assertNotIn("dave@people.test", page_size_stdout)

        sort_exit, sort_stdout, _ = self.run_cli_raw("slash", str(self.root), "/sort primary_email desc")
        self.assertEqual(sort_exit, 0)
        self.assertIn("dave@people.test", sort_stdout)
        self.assertIn("carol@people.test", sort_stdout)
        self.assertNotIn("alice@people.test", sort_stdout)

        reset_sort_exit, reset_sort_stdout, _ = self.run_cli_raw("slash", str(self.root), "/sort display_name asc")
        self.assertEqual(reset_sort_exit, 0)
        self.assertIn("alice@people.test", reset_sort_stdout)
        self.assertIn("bob@people.test", reset_sort_stdout)

        reset_size_exit, reset_size_stdout, _ = self.run_cli_raw("slash", str(self.root), "/page-size 2")
        self.assertEqual(reset_size_exit, 0)
        self.assertIn("alice@people.test", reset_size_stdout)
        self.assertIn("bob@people.test", reset_size_stdout)
        self.assertNotIn("carol@people.test", reset_size_stdout)

        next_exit, next_stdout, _ = self.run_cli_raw("slash", str(self.root), "/next")
        self.assertEqual(next_exit, 0)
        self.assertIn("carol@people.test", next_stdout)
        self.assertIn("dave@people.test", next_stdout)
        self.assertNotIn("alice@people.test", next_stdout)

        columns_exit, columns_stdout, _ = self.run_cli_raw(
            "slash",
            str(self.root),
            "/columns set label,primary_email,entity_status",
        )
        self.assertEqual(columns_exit, 0)
        self.assertIn("| label | primary_email | entity_status |", columns_stdout)

        documents_exit, _, _ = self.run_cli_raw("slash", str(self.root), "/documents")
        self.assertEqual(documents_exit, 0)
        self.assertEqual(
            retriever_tools.read_session_state(self.paths)["browse_mode"],
            retriever_tools.BROWSE_MODE_DOCUMENTS,
        )

        entities_exit, entities_stdout, _ = self.run_cli_raw("slash", str(self.root), "/entities")
        self.assertEqual(entities_exit, 0)
        self.assertIn("| label | primary_email | entity_status |", entities_stdout)
        self.assertIn("carol@people.test", entities_stdout)
        self.assertIn("dave@people.test", entities_stdout)
        self.assertNotIn("alice@people.test", entities_stdout)

    def test_entity_similar_block_and_merge_commands(self) -> None:
        self.write_email_message(
            self.root / "old.eml",
            subject="Old Jane",
            body_text="Old Jane body",
            author="Jane Doe <jane.old@example.com>",
            recipients="Reviewer <reviewer@example.com>",
            cc=None,
            message_id="<old-jane@example.com>",
        )
        self.write_email_message(
            self.root / "new.eml",
            subject="New Jane",
            body_text="New Jane body",
            author="Jane Doe <jane.new@example.org>",
            recipients="Reviewer <reviewer@example.com>",
            cc=None,
            message_id="<new-jane@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        exit_code, list_payload, _, _ = self.run_cli("list-entities", str(self.root), "--query", "Jane Doe")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        jane_entities = {
            entity["primary_email"]: entity
            for entity in list_payload["entities"]
            if entity["primary_email"] in {"jane.old@example.com", "jane.new@example.org"}
        }
        self.assertEqual(set(jane_entities), {"jane.old@example.com", "jane.new@example.org"})
        source_id = int(jane_entities["jane.old@example.com"]["id"])
        target_id = int(jane_entities["jane.new@example.org"]["id"])

        exit_code, similar_payload, _, _ = self.run_cli("similar-entities", str(self.root), str(source_id))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(similar_payload)
        assert similar_payload is not None
        similar_ids = [int(item["entity"]["id"]) for item in similar_payload["suggestions"]]
        self.assertIn(target_id, similar_ids)
        target_suggestion = next(item for item in similar_payload["suggestions"] if int(item["entity"]["id"]) == target_id)
        self.assertIn("exact_full_name", {reason["kind"] for reason in target_suggestion["reasons"]})

        exit_code, block_payload, _, _ = self.run_cli(
            "block-entity-merge",
            str(self.root),
            str(source_id),
            str(target_id),
            "--reason",
            "different Jane records",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(block_payload)
        self.assertTrue(block_payload["created"])

        exit_code, similar_payload, _, _ = self.run_cli("similar-entities", str(self.root), str(source_id))
        self.assertEqual(exit_code, 0)
        self.assertNotIn(target_id, [int(item["entity"]["id"]) for item in similar_payload["suggestions"]])

        blocked_exit, blocked_payload, _, _ = self.run_cli(
            "merge-entities",
            str(self.root),
            str(source_id),
            str(target_id),
        )
        self.assertEqual(blocked_exit, 2)
        self.assertIsNotNone(blocked_payload)
        self.assertIn("merge block", blocked_payload["error"])

        exit_code, merge_payload, _, _ = self.run_cli(
            "merge-entities",
            str(self.root),
            str(source_id),
            str(target_id),
            "--force",
            "--reason",
            "manual cleanup",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(merge_payload)
        self.assertEqual(merge_payload["source_entity_id"], source_id)
        self.assertEqual(merge_payload["target_entity_id"], target_id)
        self.assertEqual(sorted(merge_payload["affected_document_ids"]), sorted([self.fetch_document_row("old.eml")["id"], self.fetch_document_row("new.eml")["id"]]))

        exit_code, show_payload, _, _ = self.run_cli("show-entity", str(self.root), str(target_id))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(show_payload)
        merged_identifiers = {
            (identifier["identifier_type"], identifier["normalized_value"])
            for identifier in show_payload["identifiers"]
        }
        self.assertIn(("email", "jane.old@example.com"), merged_identifiers)
        self.assertIn(("email", "jane.new@example.org"), merged_identifiers)

        exit_code, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            f"author_entity_id = {target_id}",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(search_payload)
        self.assertEqual(search_payload["total_hits"], 2)

        exit_code, source_search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            f"author_entity_id = {source_id}",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(source_search_payload)
        self.assertEqual(source_search_payload["total_hits"], 0)

    def test_ignore_entity_hides_links_but_preserves_raw_metadata_after_rebuild(self) -> None:
        self.write_email_message(
            self.root / "artifact.eml",
            subject="Artifact",
            body_text="Artifact body",
            author="Alice Example <alice@example.com>",
            recipients="Parser Artifact <artifact@example.com>",
            cc=None,
            message_id="<artifact@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        exit_code, list_payload, _, _ = self.run_cli("list-entities", str(self.root), "--query", "artifact@example.com")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        artifact_entities = [
            entity
            for entity in list_payload["entities"]
            if entity["primary_email"] == "artifact@example.com"
        ]
        self.assertEqual(len(artifact_entities), 1)
        artifact_id = int(artifact_entities[0]["id"])

        exit_code, ignore_payload, _, _ = self.run_cli(
            "ignore-entity",
            str(self.root),
            str(artifact_id),
            "--reason",
            "parser artifact",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(ignore_payload)
        self.assertGreaterEqual(ignore_payload["document_links_deleted"], 1)
        self.assertGreaterEqual(ignore_payload["override_count"], 1)

        exit_code, hidden_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "recipient = 'artifact@example.com'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(hidden_payload)
        self.assertEqual(hidden_payload["total_hits"], 0)

        exit_code, raw_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "raw_recipient LIKE '%artifact@example.com%'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(raw_payload)
        self.assertEqual(raw_payload["total_hits"], 1)

        exit_code, rebuild_payload, _, _ = self.run_cli("rebuild-entities", str(self.root))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(rebuild_payload)

        exit_code, hidden_after_rebuild_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "recipient = 'artifact@example.com'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(hidden_after_rebuild_payload)
        self.assertEqual(hidden_after_rebuild_payload["total_hits"], 0)

        exit_code, active_list_payload, _, _ = self.run_cli("list-entities", str(self.root), "--query", "artifact@example.com")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(active_list_payload)
        self.assertEqual(active_list_payload["entities"], [])

        exit_code, ignored_list_payload, _, _ = self.run_cli(
            "list-entities",
            str(self.root),
            "--query",
            "artifact@example.com",
            "--include-ignored",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(ignored_list_payload)
        self.assertEqual(ignored_list_payload["entities"][0]["canonical_status"], retriever_tools.ENTITY_STATUS_IGNORED)

    def test_split_entity_moves_document_link_and_survives_rebuild(self) -> None:
        self.write_email_message(
            self.root / "one.eml",
            subject="Shared one",
            body_text="Shared one body",
            author="Shared Identity <shared@example.com>",
            recipients="Reviewer <reviewer@example.com>",
            cc=None,
            message_id="<shared-one@example.com>",
        )
        self.write_email_message(
            self.root / "two.eml",
            subject="Shared two",
            body_text="Shared two body",
            author="Shared Identity <shared@example.com>",
            recipients="Reviewer <reviewer@example.com>",
            cc=None,
            message_id="<shared-two@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        one_row = self.fetch_document_row("one.eml")
        two_row = self.fetch_document_row("two.eml")

        exit_code, list_payload, _, _ = self.run_cli("list-entities", str(self.root), "--query", "shared@example.com")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        shared_entities = [
            entity
            for entity in list_payload["entities"]
            if entity["primary_email"] == "shared@example.com"
        ]
        self.assertEqual(len(shared_entities), 1)
        source_id = int(shared_entities[0]["id"])
        self.assertEqual(shared_entities[0]["document_count"], 2)

        exit_code, split_payload, _, _ = self.run_cli(
            "split-entity",
            str(self.root),
            str(source_id),
            "--doc-id",
            str(two_row["id"]),
            "--role",
            "author",
            "--display-name",
            "Split Identity",
            "--reason",
            "separate person",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(split_payload)
        target_id = int(split_payload["target_entity_id"])
        self.assertTrue(split_payload["created_target"])
        self.assertEqual(split_payload["moved_document_links"], 1)
        self.assertGreaterEqual(split_payload["overrides_created"], 1)

        exit_code, source_search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            f"author_entity_id = {source_id}",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(source_search_payload)
        self.assertEqual(source_search_payload["total_hits"], 1)
        self.assertEqual(source_search_payload["results"][0]["id"], one_row["id"])

        exit_code, target_search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            f"author_entity_id = {target_id}",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(target_search_payload)
        self.assertEqual(target_search_payload["total_hits"], 1)
        self.assertEqual(target_search_payload["results"][0]["id"], two_row["id"])
        self.assertEqual(self.fetch_document_row("two.eml")["author"], "Split Identity")

        blocked_exit, blocked_payload, _, _ = self.run_cli(
            "merge-entities",
            str(self.root),
            str(target_id),
            str(source_id),
        )
        self.assertEqual(blocked_exit, 2)
        self.assertIsNotNone(blocked_payload)
        self.assertIn("merge block", blocked_payload["error"])

        rebuild_exit, rebuild_payload, _, _ = self.run_cli(
            "rebuild-entities",
            str(self.root),
            "--doc-id",
            str(two_row["id"]),
        )
        self.assertEqual(rebuild_exit, 0)
        self.assertIsNotNone(rebuild_payload)

        exit_code, target_after_rebuild_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            f"author_entity_id = {target_id}",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(target_after_rebuild_payload)
        self.assertEqual(target_after_rebuild_payload["total_hits"], 1)
        self.assertEqual(target_after_rebuild_payload["results"][0]["id"], two_row["id"])
        self.assertEqual(self.fetch_document_row("two.eml")["author"], "Split Identity")

    def test_assign_and_unassign_entity_update_caches_and_survive_rebuild(self) -> None:
        self.write_email_message(
            self.root / "assign.eml",
            subject="Assign",
            body_text="Assign body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            cc=None,
            message_id="<assign@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        document_row = self.fetch_document_row("assign.eml")

        exit_code, list_payload, _, _ = self.run_cli("list-entities", str(self.root), "--query", "alice@example.com")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        alice_id = next(
            int(entity["id"])
            for entity in list_payload["entities"]
            if entity["primary_email"] == "alice@example.com"
        )

        exit_code, assign_payload, _, _ = self.run_cli(
            "assign-entity",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
            "--role",
            "custodian",
            "--entity-id",
            str(alice_id),
            "--reason",
            "manual custodian",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(assign_payload)
        self.assertTrue(assign_payload["created"])
        self.assertEqual(self.fetch_document_row("assign.eml")["custodians"], ["Alice Example <alice@example.com>"])

        exit_code, custodian_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "custodian = 'alice@example.com'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(custodian_payload)
        self.assertEqual(custodian_payload["total_hits"], 1)

        exit_code, unassign_payload, _, _ = self.run_cli(
            "unassign-entity",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
            "--role",
            "author",
            "--entity-id",
            str(alice_id),
            "--reason",
            "not the author",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(unassign_payload)
        self.assertEqual(unassign_payload["auto_links_removed"], 1)
        self.assertGreaterEqual(unassign_payload["overrides_created"], 1)
        self.assertIsNone(self.fetch_document_row("assign.eml")["author"])

        exit_code, author_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "author = 'alice@example.com'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(author_payload)
        self.assertEqual(author_payload["total_hits"], 0)

        exit_code, raw_author_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "raw_author LIKE '%alice@example.com%'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(raw_author_payload)
        self.assertEqual(raw_author_payload["total_hits"], 1)

        rebuild_exit, rebuild_payload, _, _ = self.run_cli(
            "rebuild-entities",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
        )
        self.assertEqual(rebuild_exit, 0)
        self.assertIsNotNone(rebuild_payload)

        exit_code, author_after_rebuild_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "author = 'alice@example.com'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(author_after_rebuild_payload)
        self.assertEqual(author_after_rebuild_payload["total_hits"], 0)
        self.assertEqual(self.fetch_document_row("assign.eml")["custodians"], ["Alice Example <alice@example.com>"])

    def test_create_and_edit_entity_commands_update_manual_links(self) -> None:
        self.write_email_message(
            self.root / "manual-person.eml",
            subject="Manual person",
            body_text="Manual person body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            cc=None,
            message_id="<manual-person@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        document_row = self.fetch_document_row("manual-person.eml")

        exit_code, create_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--entity-type",
            "person",
            "--display-name",
            "Manual Person",
            "--email",
            "manual@example.com",
            "--name",
            "Manual Person",
            "--notes",
            "seed profile",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(create_payload)
        assert create_payload is not None
        entity_id = int(create_payload["entity_id"])
        self.assertEqual(create_payload["entity"]["label"], "Manual Person <manual@example.com>")
        self.assertEqual(create_payload["created_identifier_count"], 2)
        self.assertEqual(create_payload["created_resolution_keys"], 1)

        exit_code, assign_payload, _, _ = self.run_cli(
            "assign-entity",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
            "--role",
            "custodian",
            "--entity-id",
            str(entity_id),
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(assign_payload)
        self.assertEqual(self.fetch_document_row("manual-person.eml")["custodians"], ["Manual Person <manual@example.com>"])

        exit_code, edit_payload, _, _ = self.run_cli(
            "edit-entity",
            str(self.root),
            str(entity_id),
            "--display-name",
            "Manual Renamed",
            "--add-email",
            "renamed@example.com",
            "--add-phone",
            "+1 (212) 555-0100",
            "--notes",
            "updated profile",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(edit_payload)
        assert edit_payload is not None
        self.assertEqual(edit_payload["entity"]["label"], "Manual Renamed <manual@example.com>")
        self.assertEqual(edit_payload["created_identifier_count"], 2)
        self.assertEqual(edit_payload["created_resolution_keys"], 1)
        self.assertEqual(edit_payload["affected_document_ids"], [document_row["id"]])
        self.assertEqual(self.fetch_document_row("manual-person.eml")["custodians"], ["Manual Renamed <manual@example.com>"])

        exit_code, custodian_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "custodian = 'renamed@example.com'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(custodian_payload)
        self.assertEqual(custodian_payload["total_hits"], 1)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            entity_row = connection.execute("SELECT notes, display_name_source FROM entities WHERE id = ?", (entity_id,)).fetchone()
            self.assertIsNotNone(entity_row)
            assert entity_row is not None
            self.assertEqual(entity_row["notes"], "updated profile")
            self.assertEqual(entity_row["display_name_source"], retriever_tools.ENTITY_DISPLAY_SOURCE_MANUAL)
        finally:
            connection.close()

    def test_dataset_policy_commands_update_source_backed_policy(self) -> None:
        (self.root / "policy.txt").write_text("Dataset policy source document", encoding="utf-8")
        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        exit_code, list_payload, _, _ = self.run_cli("list-datasets", str(self.root))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        assert list_payload is not None
        dataset = next(item for item in list_payload["datasets"] if item["source_kind"] == retriever_tools.FILESYSTEM_SOURCE_KIND)
        dataset_id = int(dataset["id"])
        self.assertTrue(dataset["merge_policy"]["source_backed"])
        self.assertTrue(dataset["merge_policy"]["email_auto_merge"])
        self.assertFalse(dataset["merge_policy"]["name_auto_merge"])

        exit_code, show_payload, _, _ = self.run_cli(
            "show-dataset-policy",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(show_payload)
        assert show_payload is not None
        self.assertEqual(show_payload["merge_policy"]["dataset_id"], dataset_id)
        self.assertTrue(show_payload["merge_policy"]["allow_auto_merge"])

        exit_code, set_payload, _, _ = self.run_cli(
            "set-dataset-policy",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
            "--email-auto-merge",
            "false",
            "--phone-auto-merge",
            "true",
            "--name-auto-merge",
            "true",
            "--external-id-auto-merge-name",
            "Employee ID",
            "--external-id-auto-merge-name",
            "HR-ID",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(set_payload)
        assert set_payload is not None
        self.assertTrue(set_payload["changed"])
        self.assertTrue(set_payload["rebuild_recommended"])
        self.assertIn("rebuild-entities", set_payload["rebuild_command"])
        self.assertTrue(set_payload["before_merge_policy"]["email_auto_merge"])
        self.assertFalse(set_payload["merge_policy"]["email_auto_merge"])
        self.assertTrue(set_payload["merge_policy"]["phone_auto_merge"])
        self.assertTrue(set_payload["merge_policy"]["name_auto_merge"])
        self.assertEqual(set_payload["merge_policy"]["external_id_auto_merge_names"], ["employee_id", "hr_id"])

        exit_code, show_after_payload, _, _ = self.run_cli(
            "show-dataset-policy",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(show_after_payload)
        self.assertEqual(show_after_payload["merge_policy"], set_payload["merge_policy"])

        exit_code, create_dataset_payload, _, _ = self.run_cli("create-dataset", str(self.root), "Review Set")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(create_dataset_payload)
        rejected_exit, rejected_payload, _, _ = self.run_cli(
            "set-dataset-policy",
            str(self.root),
            "--dataset-name",
            "Review Set",
            "--email-auto-merge",
            "false",
        )
        self.assertEqual(rejected_exit, 2)
        self.assertIsNotNone(rejected_payload)
        self.assertIn("source-backed", rejected_payload["error"])

    def test_rebuild_after_email_policy_demotion_preserves_manual_state(self) -> None:
        self.write_email_message(
            self.root / "policy-one.eml",
            subject="Policy one",
            body_text="Policy one body",
            author="Policy Person <policy@example.com>",
            recipients="Shared Recipient <recipient@example.com>",
            cc=None,
            message_id="<policy-one@example.com>",
        )
        self.write_email_message(
            self.root / "policy-two.eml",
            subject="Policy two",
            body_text="Policy two body",
            author="Policy Alias <policy@example.com>",
            recipients="Shared Recipient <recipient@example.com>",
            cc=None,
            message_id="<policy-two@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        one_row = self.fetch_document_row("policy-one.eml")
        two_row = self.fetch_document_row("policy-two.eml")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            dataset_row = connection.execute(
                """
                SELECT id
                FROM datasets
                WHERE source_kind = ?
                ORDER BY id ASC
                LIMIT 1
                """,
                (retriever_tools.FILESYSTEM_SOURCE_KIND,),
            ).fetchone()
            self.assertIsNotNone(dataset_row)
            assert dataset_row is not None
            dataset_id = int(dataset_row["id"])
            author_entity_rows = connection.execute(
                """
                SELECT DISTINCT de.entity_id
                FROM document_entities de
                JOIN entity_identifiers ei ON ei.entity_id = de.entity_id
                WHERE de.role = 'author'
                  AND ei.identifier_type = 'email'
                  AND ei.normalized_value = 'policy@example.com'
                ORDER BY de.entity_id ASC
                """
            ).fetchall()
            key_row = connection.execute(
                """
                SELECT *
                FROM entity_resolution_keys
                WHERE key_type = 'email'
                  AND normalized_value = 'policy@example.com'
                """
            ).fetchone()
        finally:
            connection.close()
        self.assertEqual(len(author_entity_rows), 1)
        original_author_entity_id = int(author_entity_rows[0]["entity_id"])
        self.assertIsNotNone(key_row)

        exit_code, manual_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--display-name",
            "Manual Keeper",
            "--email",
            "keeper@example.com",
            "--name",
            "Manual Keeper",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(manual_payload)
        assert manual_payload is not None
        manual_entity_id = int(manual_payload["entity_id"])

        exit_code, assign_payload, _, _ = self.run_cli(
            "assign-entity",
            str(self.root),
            "--doc-id",
            str(one_row["id"]),
            "--role",
            "custodian",
            "--entity-id",
            str(manual_entity_id),
            "--reason",
            "manual keeper",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(assign_payload)

        exit_code, recipient_list_payload, _, _ = self.run_cli(
            "list-entities",
            str(self.root),
            "--query",
            "recipient@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(recipient_list_payload)
        assert recipient_list_payload is not None
        recipient_entity_id = next(
            int(entity["id"])
            for entity in recipient_list_payload["entities"]
            if entity["primary_email"] == "recipient@example.com"
        )
        exit_code, unassign_payload, _, _ = self.run_cli(
            "unassign-entity",
            str(self.root),
            "--doc-id",
            str(one_row["id"]),
            "--role",
            "recipient",
            "--entity-id",
            str(recipient_entity_id),
            "--reason",
            "doc one should not show recipient",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(unassign_payload)
        self.assertGreaterEqual(unassign_payload["overrides_created"], 1)

        exit_code, block_a_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--display-name",
            "Blocked A",
            "--email",
            "blocked.a@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(block_a_payload)
        exit_code, block_b_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--display-name",
            "Blocked B",
            "--email",
            "blocked.b@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(block_b_payload)
        block_a_id = int(block_a_payload["entity_id"])
        block_b_id = int(block_b_payload["entity_id"])
        exit_code, block_payload, _, _ = self.run_cli(
            "block-entity-merge",
            str(self.root),
            str(block_a_id),
            str(block_b_id),
            "--reason",
            "dogfood block",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(block_payload)
        self.assertTrue(block_payload["created"])

        exit_code, policy_payload, _, _ = self.run_cli(
            "set-dataset-policy",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
            "--email-auto-merge",
            "false",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(policy_payload)
        self.assertTrue(policy_payload["rebuild_recommended"])

        exit_code, rebuild_payload, _, _ = self.run_cli("rebuild-entities", str(self.root))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(rebuild_payload)
        assert rebuild_payload is not None
        self.assertGreaterEqual(rebuild_payload["auto_resolution_keys_deleted"], 1)
        self.assertGreaterEqual(rebuild_payload["auto_identifiers_deleted"], 1)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            demoted_key_count = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM entity_resolution_keys
                    WHERE key_type = 'email'
                      AND normalized_value = 'policy@example.com'
                    """
                ).fetchone()[0]
                or 0
            )
            author_entity_rows_after = connection.execute(
                """
                SELECT de.document_id, de.entity_id
                FROM document_entities de
                JOIN entity_identifiers ei ON ei.entity_id = de.entity_id
                WHERE de.role = 'author'
                  AND ei.identifier_type = 'email'
                  AND ei.normalized_value = 'policy@example.com'
                ORDER BY de.document_id ASC, de.entity_id ASC
                """
            ).fetchall()
            manual_identifier_row = connection.execute(
                """
                SELECT *
                FROM entity_identifiers
                WHERE entity_id = ?
                  AND identifier_type = 'email'
                  AND normalized_value = 'keeper@example.com'
                  AND source_kind = 'manual'
                """,
                (manual_entity_id,),
            ).fetchone()
            manual_resolution_key_row = connection.execute(
                """
                SELECT *
                FROM entity_resolution_keys
                WHERE entity_id = ?
                  AND key_type = 'email'
                  AND normalized_value = 'keeper@example.com'
                """,
                (manual_entity_id,),
            ).fetchone()
            manual_assignment_row = connection.execute(
                """
                SELECT *
                FROM document_entities
                WHERE document_id = ?
                  AND entity_id = ?
                  AND role = 'custodian'
                  AND assignment_mode = 'manual'
                """,
                (one_row["id"], manual_entity_id),
            ).fetchone()
            override_row = connection.execute(
                """
                SELECT *
                FROM entity_overrides
                WHERE scope_type = 'document'
                  AND scope_id = ?
                  AND role = 'recipient'
                  AND override_effect = 'remove'
                """,
                (one_row["id"],),
            ).fetchone()
            merge_block_count = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM entity_merge_blocks
                    WHERE left_entity_id = MIN(?, ?)
                      AND right_entity_id = MAX(?, ?)
                    """,
                    (block_a_id, block_b_id, block_a_id, block_b_id),
                ).fetchone()[0]
                or 0
            )
        finally:
            connection.close()

        self.assertEqual(demoted_key_count, 0)
        self.assertEqual(len(author_entity_rows_after), 2)
        self.assertEqual(
            {int(row["document_id"]) for row in author_entity_rows_after},
            {one_row["id"], two_row["id"]},
        )
        self.assertNotEqual(
            int(author_entity_rows_after[0]["entity_id"]),
            int(author_entity_rows_after[1]["entity_id"]),
        )
        self.assertNotIn(
            original_author_entity_id,
            {int(row["entity_id"]) for row in author_entity_rows_after},
        )
        self.assertIsNotNone(manual_identifier_row)
        self.assertIsNotNone(manual_resolution_key_row)
        self.assertIsNotNone(manual_assignment_row)
        self.assertIsNotNone(override_row)
        self.assertEqual(merge_block_count, 1)
        self.assertEqual(self.fetch_document_row("policy-one.eml")["custodians"], ["Manual Keeper <keeper@example.com>"])

        exit_code, recipient_search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            "recipient = 'recipient@example.com'",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(recipient_search_payload)
        self.assertEqual(recipient_search_payload["total_hits"], 1)
        self.assertEqual(recipient_search_payload["results"][0]["id"], two_row["id"])

    def test_rebuild_after_source_backed_dataset_delete_preserves_manual_state(self) -> None:
        self.write_email_message(
            self.root / "delete-one.eml",
            subject="Delete one",
            body_text="Delete one body",
            author="Delete Person <delete.person@example.com>",
            recipients="Delete Recipient <delete.recipient@example.com>",
            cc=None,
            message_id="<delete-one@example.com>",
        )
        self.write_email_message(
            self.root / "delete-two.eml",
            subject="Delete two",
            body_text="Delete two body",
            author="Delete Alias <delete.person@example.com>",
            recipients="Delete Recipient <delete.recipient@example.com>",
            cc=None,
            message_id="<delete-two@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        one_row = self.fetch_document_row("delete-one.eml")
        two_row = self.fetch_document_row("delete-two.eml")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            dataset_row = connection.execute(
                """
                SELECT id
                FROM datasets
                WHERE source_kind = ?
                ORDER BY id ASC
                LIMIT 1
                """,
                (retriever_tools.FILESYSTEM_SOURCE_KIND,),
            ).fetchone()
            self.assertIsNotNone(dataset_row)
            assert dataset_row is not None
            dataset_id = int(dataset_row["id"])
            source_resolution_key_count = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM entity_resolution_keys
                    WHERE key_type = 'email'
                      AND normalized_value IN ('delete.person@example.com', 'delete.recipient@example.com')
                    """
                ).fetchone()[0]
                or 0
            )
            source_auto_link_count = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM document_entities
                    WHERE assignment_mode = 'auto'
                      AND document_id IN (?, ?)
                    """,
                    (one_row["id"], two_row["id"]),
                ).fetchone()[0]
                or 0
            )
        finally:
            connection.close()
        self.assertGreaterEqual(source_resolution_key_count, 2)
        self.assertGreaterEqual(source_auto_link_count, 4)

        exit_code, manual_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--display-name",
            "Deleted Source Keeper",
            "--email",
            "deleted.keeper@example.com",
            "--name",
            "Deleted Source Keeper",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(manual_payload)
        assert manual_payload is not None
        manual_entity_id = int(manual_payload["entity_id"])

        exit_code, assign_payload, _, _ = self.run_cli(
            "assign-entity",
            str(self.root),
            "--doc-id",
            str(one_row["id"]),
            "--role",
            "custodian",
            "--entity-id",
            str(manual_entity_id),
            "--reason",
            "manual keeper",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(assign_payload)

        exit_code, recipient_list_payload, _, _ = self.run_cli(
            "list-entities",
            str(self.root),
            "--query",
            "delete.recipient@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(recipient_list_payload)
        assert recipient_list_payload is not None
        recipient_entity_id = next(
            int(entity["id"])
            for entity in recipient_list_payload["entities"]
            if entity["primary_email"] == "delete.recipient@example.com"
        )
        exit_code, unassign_payload, _, _ = self.run_cli(
            "unassign-entity",
            str(self.root),
            "--doc-id",
            str(one_row["id"]),
            "--role",
            "recipient",
            "--entity-id",
            str(recipient_entity_id),
            "--reason",
            "preserve override while deleting source",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(unassign_payload)
        self.assertGreaterEqual(unassign_payload["overrides_created"], 1)

        exit_code, block_a_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--display-name",
            "Delete Block A",
            "--email",
            "delete.block.a@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(block_a_payload)
        exit_code, block_b_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--display-name",
            "Delete Block B",
            "--email",
            "delete.block.b@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(block_b_payload)
        block_a_id = int(block_a_payload["entity_id"])
        block_b_id = int(block_b_payload["entity_id"])
        exit_code, block_payload, _, _ = self.run_cli(
            "block-entity-merge",
            str(self.root),
            str(block_a_id),
            str(block_b_id),
            "--reason",
            "preserve block while deleting source",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(block_payload)
        self.assertTrue(block_payload["created"])

        exit_code, delete_payload, _, _ = self.run_cli(
            "delete-dataset",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(delete_payload)
        assert delete_payload is not None
        self.assertEqual(
            sorted(delete_payload["documents_without_dataset_memberships"]),
            sorted([one_row["id"], two_row["id"]]),
        )
        self.assertEqual(retriever_tools.search(self.root, "", None, None, None, 1, 20)["total_hits"], 0)

        exit_code, rebuild_payload, _, _ = self.run_cli("rebuild-entities", str(self.root))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(rebuild_payload)
        assert rebuild_payload is not None
        self.assertEqual(rebuild_payload["status"], "ok")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            remaining_source_auto_links = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM document_entities
                    WHERE assignment_mode = 'auto'
                      AND document_id IN (?, ?)
                    """,
                    (one_row["id"], two_row["id"]),
                ).fetchone()[0]
                or 0
            )
            remaining_source_resolution_keys = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM entity_resolution_keys
                    WHERE key_type = 'email'
                      AND normalized_value IN ('delete.person@example.com', 'delete.recipient@example.com')
                    """
                ).fetchone()[0]
                or 0
            )
            remaining_source_auto_identifiers = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM entity_identifiers
                    WHERE source_kind = 'auto'
                      AND identifier_type = 'email'
                      AND normalized_value IN ('delete.person@example.com', 'delete.recipient@example.com')
                    """
                ).fetchone()[0]
                or 0
            )
            manual_identifier_row = connection.execute(
                """
                SELECT *
                FROM entity_identifiers
                WHERE entity_id = ?
                  AND identifier_type = 'email'
                  AND normalized_value = 'deleted.keeper@example.com'
                  AND source_kind = 'manual'
                """,
                (manual_entity_id,),
            ).fetchone()
            manual_assignment_row = connection.execute(
                """
                SELECT *
                FROM document_entities
                WHERE document_id = ?
                  AND entity_id = ?
                  AND role = 'custodian'
                  AND assignment_mode = 'manual'
                """,
                (one_row["id"], manual_entity_id),
            ).fetchone()
            override_row = connection.execute(
                """
                SELECT *
                FROM entity_overrides
                WHERE scope_type = 'document'
                  AND scope_id = ?
                  AND role = 'recipient'
                  AND override_effect = 'remove'
                """,
                (one_row["id"],),
            ).fetchone()
            merge_block_count = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM entity_merge_blocks
                    WHERE left_entity_id = MIN(?, ?)
                      AND right_entity_id = MAX(?, ?)
                    """,
                    (block_a_id, block_b_id, block_a_id, block_b_id),
                ).fetchone()[0]
                or 0
            )
        finally:
            connection.close()

        self.assertEqual(remaining_source_auto_links, 0)
        self.assertEqual(remaining_source_resolution_keys, 0)
        self.assertEqual(remaining_source_auto_identifiers, 0)
        self.assertIsNotNone(manual_identifier_row)
        self.assertIsNotNone(manual_assignment_row)
        self.assertIsNotNone(override_row)
        self.assertEqual(merge_block_count, 1)
        self.assertIsNone(self.fetch_document_row("delete-one.eml")["author"])
        self.assertIsNone(self.fetch_document_row("delete-two.eml")["author"])
        self.assertEqual(self.fetch_document_row("delete-one.eml")["custodians"], ["Deleted Source Keeper <deleted.keeper@example.com>"])

        exit_code, source_entity_payload, _, _ = self.run_cli(
            "list-entities",
            str(self.root),
            "--query",
            "delete.person@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(source_entity_payload)
        self.assertEqual(source_entity_payload["entities"], [])

    def test_bootstrap_initializes_processing_schema_and_job_crud(self) -> None:
        result = retriever_tools.bootstrap(self.root)
        self.assertEqual(result["schema_version"], retriever_tools.SCHEMA_VERSION)

        exit_code, create_job_payload, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Contract Metadata",
            "structured_extraction",
            "--description",
            "Extract key contract facts",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(create_job_payload)
        self.assertEqual(create_job_payload["job"]["job_name"], "contract_metadata")
        self.assertEqual(create_job_payload["job"]["job_kind"], "structured_extraction")

        exit_code, add_output_payload, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "contract_metadata",
            "Governing Law",
            "--value-type",
            "text",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(add_output_payload)
        self.assertEqual(add_output_payload["job_output"]["output_name"], "governing_law")

        exit_code, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "contract_metadata",
            "--provider",
            "openai",
            "--model",
            "gpt-5.4",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Extract the governing law field.",
            "--response-schema-json",
            "{\"type\":\"object\",\"properties\":{\"governing_law\":{\"type\":\"string\"}}}",
            "--parameters-json",
            "{\"temperature\":0}",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(create_version_payload)
        self.assertEqual(create_version_payload["job_version"]["version"], 1)
        self.assertEqual(create_version_payload["job_version"]["capability"], "text_structured")
        self.assertEqual(create_version_payload["job_version"]["provider"], "openai")
        self.assertEqual(create_version_payload["job_version"]["parameters"], {"temperature": 0})
        self.assertEqual(
            create_version_payload["job_version"]["response_schema"]["type"],
            "object",
        )

        exit_code, list_jobs_payload, _, _ = self.run_cli("list-jobs", str(self.root))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_jobs_payload)
        self.assertEqual(len(list_jobs_payload["jobs"]), 1)
        self.assertEqual(list_jobs_payload["jobs"][0]["latest_job_version"]["version"], 1)
        self.assertEqual(list_jobs_payload["jobs"][0]["outputs"][0]["output_name"], "governing_law")

        exit_code, versions_payload, _, _ = self.run_cli("list-job-versions", str(self.root), "contract_metadata")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(versions_payload)
        self.assertEqual(len(versions_payload["job_versions"]), 1)
        self.assertEqual(versions_payload["job_versions"][0]["display_name"], "contract_metadata v1")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            document_columns = retriever_tools.table_columns(connection, "documents")
            self.assertIn("source_text_revision_id", document_columns)
            self.assertIn("active_search_text_revision_id", document_columns)
            self.assertIn("active_text_source_kind", document_columns)
            self.assertIn("active_text_language", document_columns)
            self.assertIn("active_text_quality_score", document_columns)
            self.assertIn("capability", retriever_tools.table_columns(connection, "job_versions"))
            run_item_columns = retriever_tools.table_columns(connection, "run_items")
            self.assertIn("claimed_by", run_item_columns)
            self.assertIn("claimed_at", run_item_columns)
            self.assertIn("last_heartbeat_at", run_item_columns)

            expected_tables = {
                "jobs",
                "job_outputs",
                "job_versions",
                "runs",
                "run_snapshot_documents",
                "run_items",
                "attempts",
                "results",
                "result_outputs",
                "text_revisions",
                "text_revision_segments",
                "embedding_vectors",
                "publications",
                "text_revision_activation_events",
            }
            actual_tables = {
                row["name"]
                for row in connection.execute(
                    """
                    SELECT name
                    FROM sqlite_master
                    WHERE type = 'table'
                    """
                ).fetchall()
            }
            self.assertTrue(expected_tables.issubset(actual_tables))

            job_row = connection.execute(
                "SELECT id FROM jobs WHERE job_name = ?",
                ("contract_metadata",),
            ).fetchone()
            self.assertIsNotNone(job_row)
            job_version_row = connection.execute(
                "SELECT id FROM job_versions WHERE job_id = ? AND version = 1",
                (job_row["id"],),
            ).fetchone()
            self.assertIsNotNone(job_version_row)

            document_cursor = connection.execute(
                """
                INSERT INTO documents (rel_path, file_name, updated_at)
                VALUES (?, ?, ?)
                """,
                ("sample.txt", "sample.txt", "2026-04-17T00:00:00Z"),
            )
            document_id = int(document_cursor.lastrowid)
            text_revision_cursor = connection.execute(
                """
                INSERT INTO text_revisions (
                  document_id,
                  revision_kind,
                  language,
                  parent_revision_id,
                  created_by_job_version_id,
                  storage_rel_path,
                  content_hash,
                  char_count,
                  token_estimate,
                  quality_score,
                  provider_metadata_json,
                  created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    document_id,
                    "source_extract",
                    "en",
                    None,
                    None,
                    None,
                    retriever_tools.sha256_text("alpha"),
                    5,
                    1,
                    None,
                    "{}",
                    "2026-04-17T00:00:00Z",
                ),
            )
            input_revision_id = int(text_revision_cursor.lastrowid)
            input_identity = retriever_tools.build_text_revision_input_identity(input_revision_id)

            connection.execute(
                """
                INSERT INTO results (
                  run_id,
                  document_id,
                  job_version_id,
                  input_revision_id,
                  input_identity,
                  raw_output_json,
                  normalized_output_json,
                  created_text_revision_id,
                  provider_metadata_json,
                  created_at,
                  retracted_at,
                  retraction_reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    None,
                    document_id,
                    int(job_version_row["id"]),
                    input_revision_id,
                    input_identity,
                    "{\"governing_law\":\"Delaware\"}",
                    "{\"governing_law\":\"Delaware\"}",
                    None,
                    "{}",
                    "2026-04-17T00:00:00Z",
                    None,
                    None,
                ),
            )
            with self.assertRaises(sqlite3.IntegrityError):
                connection.execute(
                    """
                    INSERT INTO results (
                      run_id,
                      document_id,
                      job_version_id,
                      input_revision_id,
                      input_identity,
                      raw_output_json,
                      normalized_output_json,
                      created_text_revision_id,
                      provider_metadata_json,
                      created_at,
                      retracted_at,
                      retraction_reason
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        None,
                        document_id,
                        int(job_version_row["id"]),
                        input_revision_id,
                        input_identity,
                        "{\"governing_law\":\"Delaware\"}",
                        "{\"governing_law\":\"Delaware\"}",
                        None,
                        "{}",
                        "2026-04-17T00:00:01Z",
                        None,
                        None,
                    ),
                )

            connection.execute(
                """
                UPDATE results
                SET retracted_at = ?, retraction_reason = ?
                WHERE document_id = ? AND job_version_id = ? AND input_identity = ?
                """,
                (
                    "2026-04-17T00:00:02Z",
                    "Superseded during testing",
                    document_id,
                    int(job_version_row["id"]),
                    input_identity,
                ),
            )
            connection.execute(
                """
                INSERT INTO results (
                  run_id,
                  document_id,
                  job_version_id,
                  input_revision_id,
                  input_identity,
                  raw_output_json,
                  normalized_output_json,
                  created_text_revision_id,
                  provider_metadata_json,
                  created_at,
                  retracted_at,
                  retraction_reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    None,
                    document_id,
                    int(job_version_row["id"]),
                    input_revision_id,
                    input_identity,
                    "{\"governing_law\":\"New York\"}",
                    "{\"governing_law\":\"New York\"}",
                    None,
                    "{}",
                    "2026-04-17T00:00:03Z",
                    None,
                    None,
                ),
            )
            connection.commit()

            active_count_row = connection.execute(
                """
                SELECT COUNT(*) AS count
                FROM results
                WHERE document_id = ? AND job_version_id = ? AND input_identity = ? AND retracted_at IS NULL
                """,
                (document_id, int(job_version_row["id"]), input_identity),
            ).fetchone()
            self.assertEqual(active_count_row["count"], 1)
        finally:
            connection.close()

    def test_workspace_init_command_bootstraps_and_reports_ready_status(self) -> None:
        runtime_paths = retriever_tools.plugin_runtime_paths(root=self.root)
        self.assertIsNotNone(runtime_paths)

        with mock.patch.object(
            retriever_tools,
            "ensure_plugin_runtime",
            return_value={
                "status": "pass",
                "detail": "Shared plugin runtime and pinned requirements are ready.",
                "venv_created": True,
                "requirements_installed": True,
                "requirements_version": retriever_tools.REQUIREMENTS_VERSION,
                "python_executable": str(runtime_paths["venv_python_path"]),
            },
        ) as ensure_plugin_runtime:
            exit_code, payload, _, _ = self.run_cli("workspace", "init", str(self.root))

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["action"], "init")
        self.assertEqual(payload["status"], "ready")
        self.assertEqual(payload["initialization"]["status"], "initialized")
        self.assertEqual(payload["status_report"]["workspace"]["state"], "initialized")
        self.assertEqual(payload["tool_update"]["status"], "updated-runtime")
        self.assertEqual(payload["runtime_init"]["status"], "pass")
        self.assertEqual(payload["status_report"]["plugin_runtime"]["plugin_root"], str(REPO_ROOT))
        self.assertTrue(self.paths["runtime_path"].exists())
        ensure_plugin_runtime.assert_called_once_with(
            runtime_paths,
            install_requirements=True,
            force_requirements_install=False,
            reason="init",
        )

    def test_plugin_runtime_paths_live_under_plugin_root_not_workspace(self) -> None:
        runtime_paths = retriever_tools.plugin_runtime_paths(root=self.root)
        self.assertIsNotNone(runtime_paths)
        self.assertEqual(runtime_paths["plugin_root"], REPO_ROOT)
        self.assertIn(REPO_ROOT, runtime_paths["runtime_root"].parents)
        self.assertNotIn(self.root.resolve(), runtime_paths["runtime_root"].parents)

    def test_load_dependency_auto_installs_plugin_runtime_and_retries_import(self) -> None:
        retriever_tools.set_active_workspace_root(self.root)
        retriever_tools.pypff = retriever_tools._UNLOADED_DEPENDENCY
        runtime_paths = retriever_tools.plugin_runtime_paths(root=self.root)
        self.assertIsNotNone(runtime_paths)
        dummy_module = types.SimpleNamespace(file=lambda: None)

        with (
            mock.patch.object(retriever_tools, "activate_plugin_site_packages", return_value=False) as activate_site_packages,
            mock.patch.object(
                retriever_tools,
                "ensure_plugin_runtime",
                return_value={"status": "pass", "requirements_installed": True},
            ) as ensure_plugin_runtime,
            mock.patch.object(
                retriever_tools.importlib,
                "import_module",
                side_effect=[ImportError("missing pypff"), dummy_module],
            ) as import_module,
        ):
            loaded = retriever_tools.load_dependency("pypff", allow_auto_install=True)

        self.assertIs(loaded, dummy_module)
        self.assertIs(retriever_tools.pypff, dummy_module)
        self.assertGreaterEqual(activate_site_packages.call_count, 1)
        ensure_plugin_runtime.assert_called_once_with(
            runtime_paths,
            install_requirements=True,
            force_requirements_install=True,
            reason="dependency:pypff",
        )
        self.assertEqual(import_module.call_count, 2)

    def test_ingest_seeds_text_revisions_and_create_run_freezes_family_snapshot(self) -> None:
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Run planning",
            body_text="Parent body text for the email.",
            attachment_name="notes.txt",
            attachment_text="confidential attachment detail",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        parent_row = self.fetch_document_row("thread.eml")
        child_rows = self.fetch_child_rows(parent_row["id"])
        self.assertEqual(len(child_rows), 1)
        child_row = child_rows[0]

        self.assertIsNotNone(parent_row["source_text_revision_id"])
        self.assertEqual(parent_row["source_text_revision_id"], parent_row["active_search_text_revision_id"])
        self.assertIsNotNone(child_row["source_text_revision_id"])
        self.assertEqual(child_row["source_text_revision_id"], child_row["active_search_text_revision_id"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            parent_revision_row = connection.execute(
                "SELECT * FROM text_revisions WHERE id = ?",
                (parent_row["source_text_revision_id"],),
            ).fetchone()
            child_revision_row = connection.execute(
                "SELECT * FROM text_revisions WHERE id = ?",
                (child_row["source_text_revision_id"],),
            ).fetchone()
            self.assertIsNotNone(parent_revision_row)
            self.assertIsNotNone(child_revision_row)
            parent_revision_text = retriever_tools.read_text_revision_body(
                self.paths,
                parent_revision_row["storage_rel_path"],
            )
            child_revision_text = retriever_tools.read_text_revision_body(
                self.paths,
                child_revision_row["storage_rel_path"],
            )
        finally:
            connection.close()

        self.assertIn("Parent body text for the email.", parent_revision_text)
        self.assertIn("confidential attachment detail", child_revision_text)

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Issue Tags",
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "issue_tags",
            "primary_issue",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        version_exit, version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "issue_tags",
            "--provider",
            "openai",
            "--model",
            "gpt-5.4",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Extract the main issue tag.",
            "--parameters-json",
            "{\"temperature\":0}",
        )
        self.assertEqual(version_exit, 0)
        self.assertIsNotNone(version_payload)
        self.assertEqual(version_payload["job_version"]["version"], 1)

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-name",
            "issue_tags",
            "--job-version",
            "1",
            "--keyword",
            "confidential",
            "--family-mode",
            "with_family",
            "--limit",
            "1",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_payload = create_run_payload["run"]
        self.assertEqual(run_payload["planned_count"], 2)
        self.assertEqual(len(run_payload["documents"]), 2)

        snapshot_by_document_id = {
            int(item["document_id"]): item
            for item in run_payload["documents"]
        }
        self.assertEqual(set(snapshot_by_document_id), {int(parent_row["id"]), int(child_row["id"])})
        self.assertEqual(
            snapshot_by_document_id[int(child_row["id"])]["pinned_input_revision_id"],
            int(child_row["source_text_revision_id"]),
        )
        self.assertEqual(
            snapshot_by_document_id[int(child_row["id"])]["pinned_input_identity"],
            retriever_tools.build_text_revision_input_identity(int(child_row["source_text_revision_id"])),
        )
        self.assertEqual(
            snapshot_by_document_id[int(child_row["id"])]["inclusion_reason"]["direct_reasons"][0]["type"],
            "keyword",
        )
        self.assertEqual(
            snapshot_by_document_id[int(parent_row["id"])]["inclusion_reason"]["family_seed_document_ids"],
            [int(child_row["id"])],
        )

        first_run_id = int(run_payload["id"])
        second_run_exit, second_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-name",
            "issue_tags",
            "--job-version",
            "1",
            "--from-run-id",
            str(first_run_id),
        )
        self.assertEqual(second_run_exit, 0)
        self.assertIsNotNone(second_run_payload)
        second_snapshot = {
            int(item["document_id"]): item
            for item in second_run_payload["run"]["documents"]
        }
        self.assertEqual(set(second_snapshot), set(snapshot_by_document_id))
        self.assertEqual(
            second_snapshot[int(parent_row["id"])]["pinned_input_revision_id"],
            snapshot_by_document_id[int(parent_row["id"])]["pinned_input_revision_id"],
        )
        self.assertEqual(
            second_snapshot[int(child_row["id"])]["pinned_input_identity"],
            snapshot_by_document_id[int(child_row["id"])]["pinned_input_identity"],
        )

    def test_activate_text_revision_rebuilds_search_chunks_and_records_event(self) -> None:
        note_path = self.root / "note.txt"
        note_path.write_text("sourcealphaonly original searchable text", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("note.txt")
        source_revision_id = int(document_row["source_text_revision_id"])
        replacement_text = "promotedbetaonly replacement searchable text"

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute("BEGIN")
            try:
                promoted_revision_id = retriever_tools.create_text_revision_row(
                    connection,
                    self.paths,
                    document_id=int(document_row["id"]),
                    revision_kind="ocr",
                    text_content=replacement_text,
                    language="en",
                    parent_revision_id=source_revision_id,
                    created_by_job_version_id=None,
                    quality_score=0.95,
                    provider_metadata={"provider": "test"},
                )
                connection.commit()
            except Exception:
                connection.rollback()
                raise
        finally:
            connection.close()

        before_activation = retriever_tools.search(self.root, "promotedbetaonly", None, None, None, 1, 20)
        self.assertEqual(before_activation["total_hits"], 0)
        original_search = retriever_tools.search(self.root, "sourcealphaonly", None, None, None, 1, 20)
        self.assertEqual(original_search["total_hits"], 1)

        activate_exit, activate_payload, _, _ = self.run_cli(
            "activate-text-revision",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
            "--text-revision-id",
            str(promoted_revision_id),
        )
        self.assertEqual(activate_exit, 0)
        self.assertIsNotNone(activate_payload)
        self.assertEqual(activate_payload["activation_policy"], "manual")
        self.assertEqual(activate_payload["text_revision"]["id"], promoted_revision_id)
        self.assertTrue(activate_payload["text_revision"]["is_active_search_revision"])

        list_exit, list_payload, _, _ = self.run_cli(
            "list-text-revisions",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
        )
        self.assertEqual(list_exit, 0)
        self.assertIsNotNone(list_payload)
        listed_by_id = {int(item["id"]): item for item in list_payload["text_revisions"]}
        self.assertEqual(set(listed_by_id), {source_revision_id, int(promoted_revision_id)})
        self.assertTrue(listed_by_id[int(promoted_revision_id)]["is_active_search_revision"])
        self.assertTrue(listed_by_id[source_revision_id]["is_source_revision"])

        updated_row = self.fetch_document_by_id(int(document_row["id"]))
        self.assertEqual(updated_row["source_text_revision_id"], source_revision_id)
        self.assertEqual(updated_row["active_search_text_revision_id"], promoted_revision_id)
        self.assertEqual(updated_row["active_text_source_kind"], "ocr")
        self.assertEqual(updated_row["content_hash"], retriever_tools.sha256_text(replacement_text))

        promoted_search = retriever_tools.search(self.root, "promotedbetaonly", None, None, None, 1, 20)
        self.assertEqual(promoted_search["total_hits"], 1)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            chunk_texts = [
                row["text_content"]
                for row in connection.execute(
                    """
                    SELECT text_content
                    FROM document_chunks
                    WHERE document_id = ?
                    ORDER BY chunk_index ASC
                    """,
                    (document_row["id"],),
                ).fetchall()
            ]
            self.assertEqual(chunk_texts, [replacement_text])
            activation_row = connection.execute(
                """
                SELECT *
                FROM text_revision_activation_events
                WHERE document_id = ?
                ORDER BY id DESC
                LIMIT 1
                """,
                (document_row["id"],),
            ).fetchone()
            self.assertIsNotNone(activation_row)
            self.assertEqual(activation_row["text_revision_id"], promoted_revision_id)
            self.assertEqual(activation_row["activation_policy"], "manual")
        finally:
            connection.close()

    def test_execute_run_reuses_results_and_publishes_bound_outputs(self) -> None:
        note_path = self.root / "contract.txt"
        note_path.write_text("This contract mentions Delaware and an automatic renewal clause.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("contract.txt")

        add_field_exit, _, _, _ = self.run_cli("add-field", str(self.root), "governing_law", "text")
        self.assertEqual(add_field_exit, 0)
        add_bool_field_exit, _, _, _ = self.run_cli("add-field", str(self.root), "auto_renewal", "boolean")
        self.assertEqual(add_bool_field_exit, 0)

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Contract Metadata",
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        add_governing_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "contract_metadata",
            "governing_law",
            "--value-type",
            "text",
            "--bind-custom-field",
            "governing_law",
        )
        self.assertEqual(add_governing_output_exit, 0)
        add_renewal_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "contract_metadata",
            "auto_renewal",
            "--value-type",
            "boolean",
            "--bind-custom-field",
            "auto_renewal",
        )
        self.assertEqual(add_renewal_output_exit, 0)

        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "contract_metadata",
            "--provider",
            "static_json",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Return contract metadata.",
            "--parameters-json",
            "{\"output_values\":{\"governing_law\":\"Delaware\",\"auto_renewal\":true}}",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        first_run_id = int(create_run_payload["run"]["id"])

        execute_run_exit, execute_run_payload, _, _ = self.run_cli(
            "execute-run",
            str(self.root),
            "--run-id",
            str(first_run_id),
        )
        self.assertEqual(execute_run_exit, 0)
        self.assertIsNotNone(execute_run_payload)
        self.assertEqual(execute_run_payload["run"]["status"], "completed")
        self.assertEqual(execute_run_payload["run"]["completed_count"], 1)
        self.assertEqual(execute_run_payload["run"]["skipped_count"], 0)
        self.assertEqual(len(execute_run_payload["run_items"]), 1)
        self.assertEqual(execute_run_payload["run_items"][0]["status"], "completed")
        self.assertEqual(len(execute_run_payload["results"]), 1)
        first_result = execute_run_payload["results"][0]
        outputs_by_name = {item["output_name"]: item for item in first_result["outputs"]}
        self.assertEqual(outputs_by_name["governing_law"]["output_value"], "Delaware")
        self.assertEqual(outputs_by_name["auto_renewal"]["output_value"], True)

        list_results_exit, list_results_payload, _, _ = self.run_cli(
            "list-results",
            str(self.root),
            "--run-id",
            str(first_run_id),
        )
        self.assertEqual(list_results_exit, 0)
        self.assertIsNotNone(list_results_payload)
        self.assertEqual(len(list_results_payload["results"]), 1)

        publish_exit, publish_payload, _, _ = self.run_cli(
            "publish-run-results",
            str(self.root),
            "--run-id",
            str(first_run_id),
        )
        self.assertEqual(publish_exit, 0)
        self.assertIsNotNone(publish_payload)
        self.assertEqual(publish_payload["published_count"], 2)
        updated_document_row = self.fetch_document_by_id(int(document_row["id"]))
        self.assertEqual(updated_document_row["governing_law"], "Delaware")
        self.assertEqual(updated_document_row["auto_renewal"], 1)

        second_run_exit, second_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--from-run-id",
            str(first_run_id),
        )
        self.assertEqual(second_run_exit, 0)
        self.assertIsNotNone(second_run_payload)
        second_run_id = int(second_run_payload["run"]["id"])

        second_execute_exit, second_execute_payload, _, _ = self.run_cli(
            "execute-run",
            str(self.root),
            "--run-id",
            str(second_run_id),
        )
        self.assertEqual(second_execute_exit, 0)
        self.assertIsNotNone(second_execute_payload)
        self.assertEqual(second_execute_payload["run"]["status"], "completed")
        self.assertEqual(second_execute_payload["run"]["completed_count"], 0)
        self.assertEqual(second_execute_payload["run"]["skipped_count"], 1)
        self.assertEqual(len(second_execute_payload["results"]), 1)
        self.assertEqual(second_execute_payload["results"][0]["id"], first_result["id"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            result_count = connection.execute(
                """
                SELECT COUNT(*) AS count
                FROM results
                WHERE document_id = ? AND job_version_id = ? AND retracted_at IS NULL
                """,
                (document_row["id"], job_version_id),
            ).fetchone()["count"]
            self.assertEqual(result_count, 1)
            publication_count = connection.execute(
                """
                SELECT COUNT(*) AS count
                FROM publications
                WHERE document_id = ?
                """,
                (document_row["id"],),
            ).fetchone()["count"]
            self.assertEqual(publication_count, 2)
            reused_run_item = connection.execute(
                """
                SELECT *
                FROM run_items
                WHERE run_id = ?
                ORDER BY id ASC
                LIMIT 1
                """,
                (second_run_id,),
            ).fetchone()
            self.assertIsNotNone(reused_run_item)
            self.assertEqual(reused_run_item["status"], "skipped")
            self.assertEqual(reused_run_item["result_id"], first_result["id"])
        finally:
            connection.close()

    def test_claim_complete_translation_run_creates_derived_text_revision(self) -> None:
        note_path = self.root / "memo.txt"
        note_path.write_text("Original English memo text.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("memo.txt")
        source_revision_id = int(document_row["source_text_revision_id"])

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Translate ES",
            "translation",
        )
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "translate_es",
            "--provider",
            "static_text",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Translate to Spanish.",
            "--parameters-json",
            "{\"target_language\":\"es\",\"translated_text\":\"ES::{text}\"}",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "translator-es",
            "--limit",
            "1",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        self.assertEqual(len(claim_payload["run_items"]), 1)
        run_item_id = int(claim_payload["run_items"][0]["id"])

        context_exit, context_payload, _, _ = self.run_cli(
            "get-run-item-context",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
        )
        self.assertEqual(context_exit, 0)
        self.assertIsNotNone(context_payload)
        self.assertEqual(context_payload["context"]["job_version"]["capability"], "text_translation")
        self.assertEqual(
            context_payload["context"]["input"]["inline_text"],
            "Original English memo text.",
        )
        self.assertEqual(context_payload["context"]["execution"]["target_language"], "es")

        completion_template = context_payload["context"]["execution"]["completion_template"]
        raw_output = dict(completion_template["raw_output_json"])
        raw_output["translated_text"] = "ES::Original English memo text."
        normalized_output = dict(completion_template["normalized_output_json"])
        normalized_output["translated_text"] = "ES::Original English memo text."
        created_text_revision = dict(completion_template["created_text_revision_json"])
        created_text_revision["text_content"] = "ES::Original English memo text."

        complete_exit, complete_payload, _, _ = self.run_cli(
            "complete-run-item",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
            "--claimed-by",
            "translator-es",
            "--raw-output-json",
            json.dumps(raw_output),
            "--normalized-output-json",
            json.dumps(normalized_output),
            "--created-text-revision-json",
            json.dumps(created_text_revision),
        )
        self.assertEqual(complete_exit, 0)
        self.assertIsNotNone(complete_payload)
        self.assertFalse(complete_payload["idempotent"])
        self.assertEqual(complete_payload["run"]["status"], "completed")
        result_payload = complete_payload["result"]
        self.assertIsNotNone(result_payload["created_text_revision_id"])

        updated_row = self.fetch_document_by_id(int(document_row["id"]))
        self.assertEqual(updated_row["active_search_text_revision_id"], source_revision_id)
        self.assertEqual(updated_row["source_text_revision_id"], source_revision_id)

        revision_exit, revision_payload, _, _ = self.run_cli(
            "list-text-revisions",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
        )
        self.assertEqual(revision_exit, 0)
        self.assertIsNotNone(revision_payload)
        revisions_by_id = {int(item["id"]): item for item in revision_payload["text_revisions"]}
        translated_revision = revisions_by_id[int(result_payload["created_text_revision_id"])]
        self.assertEqual(translated_revision["revision_kind"], "translation")
        self.assertEqual(translated_revision["parent_revision_id"], source_revision_id)
        self.assertFalse(translated_revision["is_active_search_revision"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            translated_row = connection.execute(
                "SELECT * FROM text_revisions WHERE id = ?",
                (result_payload["created_text_revision_id"],),
            ).fetchone()
            self.assertIsNotNone(translated_row)
            translated_text = retriever_tools.read_text_revision_body(
                self.paths,
                translated_row["storage_rel_path"],
            )
            self.assertEqual(translated_text, "ES::Original English memo text.")
        finally:
            connection.close()

    def test_claim_complete_translation_run_with_always_activation_promotes_revision(self) -> None:
        note_path = self.root / "memo-activate.txt"
        note_path.write_text("Original English memo text.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("memo-activate.txt")
        source_revision_id = int(document_row["source_text_revision_id"])

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Translate RU",
            "translation",
        )
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "translate_ru",
            "--provider",
            "static_text",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Translate to Russian.",
            "--parameters-json",
            "{\"target_language\":\"ru\",\"translated_text\":\"RU::{text}\"}",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--doc-id",
            str(document_row["id"]),
            "--activation-policy",
            "always",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        self.assertEqual(create_run_payload["run"]["activation_policy"], "always")
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "translator-ru",
            "--limit",
            "1",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        run_item_id = int(claim_payload["run_items"][0]["id"])

        context_exit, context_payload, _, _ = self.run_cli(
            "get-run-item-context",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
        )
        self.assertEqual(context_exit, 0)
        self.assertIsNotNone(context_payload)
        completion_template = context_payload["context"]["execution"]["completion_template"]
        raw_output = dict(completion_template["raw_output_json"])
        raw_output["translated_text"] = "RU::Original English memo text."
        normalized_output = dict(completion_template["normalized_output_json"])
        normalized_output["translated_text"] = "RU::Original English memo text."
        created_text_revision = dict(completion_template["created_text_revision_json"])
        created_text_revision["text_content"] = "RU::Original English memo text."

        complete_exit, complete_payload, _, _ = self.run_cli(
            "complete-run-item",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
            "--claimed-by",
            "translator-ru",
            "--raw-output-json",
            json.dumps(raw_output),
            "--normalized-output-json",
            json.dumps(normalized_output),
            "--created-text-revision-json",
            json.dumps(created_text_revision),
        )
        self.assertEqual(complete_exit, 0)
        self.assertIsNotNone(complete_payload)
        self.assertIn("activation", complete_payload)
        result_payload = complete_payload["result"]
        translated_revision_id = int(result_payload["created_text_revision_id"])
        self.assertEqual(complete_payload["activation"]["text_revision"]["id"], translated_revision_id)
        self.assertEqual(complete_payload["activation"]["activation_policy"], "always")

        updated_row = self.fetch_document_by_id(int(document_row["id"]))
        self.assertEqual(updated_row["source_text_revision_id"], source_revision_id)
        self.assertEqual(updated_row["active_search_text_revision_id"], translated_revision_id)
        self.assertEqual(updated_row["active_text_source_kind"], "translation")

        revision_exit, revision_payload, _, _ = self.run_cli(
            "list-text-revisions",
            str(self.root),
            "--doc-id",
            str(document_row["id"]),
        )
        self.assertEqual(revision_exit, 0)
        self.assertIsNotNone(revision_payload)
        revisions_by_id = {int(item["id"]): item for item in revision_payload["text_revisions"]}
        self.assertTrue(revisions_by_id[translated_revision_id]["is_active_search_revision"])

        repeat_complete_exit, repeat_complete_payload, _, _ = self.run_cli(
            "complete-run-item",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
            "--claimed-by",
            "translator-ru",
            "--raw-output-json",
            json.dumps(raw_output),
            "--normalized-output-json",
            json.dumps(normalized_output),
            "--created-text-revision-json",
            json.dumps(created_text_revision),
        )
        self.assertEqual(repeat_complete_exit, 0)
        self.assertIsNotNone(repeat_complete_payload)
        self.assertTrue(repeat_complete_payload["idempotent"])
        self.assertEqual(repeat_complete_payload["result"]["id"], result_payload["id"])

    def test_create_run_rejects_always_activation_for_structured_extraction(self) -> None:
        note_path = self.root / "contract-activation.txt"
        note_path.write_text("Contract text.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("contract-activation.txt")

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Extract Contract Metadata",
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "extract_contract_metadata",
            "--instruction",
            "Extract metadata fields from the contract.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, stderr_text = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--doc-id",
            str(document_row["id"]),
            "--activation-policy",
            "always",
        )
        self.assertNotEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        self.assertIn("only supported", create_run_payload["error"])

    def test_translation_run_item_context_includes_execution_template(self) -> None:
        note_path = self.root / "translation.txt"
        note_path.write_text("Translate this sentence.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("translation.txt")

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Translate FR",
            "translation",
        )
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "translate_fr",
            "--instruction",
            "Translate to French and preserve meaning.",
            "--parameters-json",
            "{\"target_language\":\"fr\"}",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        self.assertEqual(create_version_payload["job_version"]["capability"], "text_translation")
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "translator-a",
            "--limit",
            "1",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        self.assertEqual(len(claim_payload["run_items"]), 1)
        run_item_id = int(claim_payload["run_items"][0]["id"])

        context_exit, context_payload, _, _ = self.run_cli(
            "get-run-item-context",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
        )
        self.assertEqual(context_exit, 0)
        self.assertIsNotNone(context_payload)
        self.assertEqual(context_payload["context"]["job_version"]["capability"], "text_translation")
        self.assertEqual(context_payload["context"]["execution"]["capability"], "text_translation")
        self.assertEqual(context_payload["context"]["execution"]["target_language"], "fr")
        self.assertIn(
            "Translate the entire input text into fr.",
            context_payload["context"]["execution"]["task_prompt"],
        )
        self.assertEqual(
            context_payload["context"]["execution"]["completion_template"]["created_text_revision_json"]["language"],
            "fr",
        )
        self.assertEqual(
            context_payload["context"]["execution"]["completion_template"]["created_text_revision_json"]["revision_kind"],
            "translation",
        )

    def test_claim_complete_run_item_flow_is_idempotent(self) -> None:
        note_path = self.root / "contract.txt"
        note_path.write_text("Governing law is Delaware.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("contract.txt")

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Contract Metadata",
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "contract_metadata",
            "governing_law",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "contract_metadata",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Extract the governing law field.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        self.assertEqual(create_version_payload["job_version"]["capability"], "text_structured")
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "worker-a",
            "--limit",
            "1",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        self.assertEqual(len(claim_payload["run_items"]), 1)
        run_item_id = int(claim_payload["run_items"][0]["id"])
        self.assertEqual(claim_payload["run_items"][0]["claimed_by"], "worker-a")

        context_exit, context_payload, _, _ = self.run_cli(
            "get-run-item-context",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
        )
        self.assertEqual(context_exit, 0)
        self.assertIsNotNone(context_payload)
        self.assertEqual(context_payload["context"]["job_version"]["capability"], "text_structured")
        self.assertEqual(
            context_payload["context"]["input"]["inline_text"],
            "Governing law is Delaware.",
        )
        self.assertIn("governing_law", context_payload["context"]["response_schema"]["properties"])
        self.assertEqual(context_payload["context"]["execution"]["capability"], "text_structured")
        self.assertIn(
            "Return only a JSON object that matches response_schema exactly.",
            context_payload["context"]["execution"]["task_prompt"],
        )
        self.assertEqual(
            context_payload["context"]["execution"]["output_defaults"]["governing_law"],
            "",
        )
        self.assertEqual(
            context_payload["context"]["execution"]["completion_template"]["output_values_json"]["governing_law"],
            "<final governing_law value>",
        )

        completion_template = context_payload["context"]["execution"]["completion_template"]
        raw_output = dict(completion_template["raw_output_json"])
        raw_output["governing_law"] = "Delaware"
        normalized_output = dict(completion_template["normalized_output_json"])
        normalized_output["governing_law"] = "Delaware"
        output_values = dict(completion_template["output_values_json"])
        output_values["governing_law"] = "Delaware"

        complete_exit, complete_payload, _, _ = self.run_cli(
            "complete-run-item",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
            "--claimed-by",
            "worker-a",
            "--raw-output-json",
            json.dumps(raw_output),
            "--normalized-output-json",
            json.dumps(normalized_output),
            "--output-values-json",
            json.dumps(output_values),
        )
        self.assertEqual(complete_exit, 0)
        self.assertIsNotNone(complete_payload)
        self.assertFalse(complete_payload["idempotent"])
        self.assertEqual(complete_payload["run"]["status"], "completed")
        self.assertEqual(
            complete_payload["result"]["outputs"][0]["output_value"],
            "Delaware",
        )

        repeat_complete_exit, repeat_complete_payload, _, _ = self.run_cli(
            "complete-run-item",
            str(self.root),
            "--run-item-id",
            str(run_item_id),
            "--claimed-by",
            "worker-a",
            "--raw-output-json",
            json.dumps(raw_output),
            "--normalized-output-json",
            json.dumps(normalized_output),
            "--output-values-json",
            json.dumps(output_values),
        )
        self.assertEqual(repeat_complete_exit, 0)
        self.assertIsNotNone(repeat_complete_payload)
        self.assertTrue(repeat_complete_payload["idempotent"])
        self.assertEqual(repeat_complete_payload["result"]["id"], complete_payload["result"]["id"])

        status_exit, status_payload, _, _ = self.run_cli(
            "run-status",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(status_exit, 0)
        self.assertIsNotNone(status_payload)
        self.assertEqual(status_payload["run"]["status"], "completed")
        self.assertEqual(status_payload["run"]["run_item_counts"]["completed"], 1)

    def test_prepare_run_batch_returns_contexts_and_worker_hints(self) -> None:
        note_path = self.root / "batch-contract.txt"
        note_path.write_text("Governing law is New York.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("batch-contract.txt")

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Batch Contract Metadata",
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "batch_contract_metadata",
            "governing_law",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "batch_contract_metadata",
            "--instruction",
            "Extract the governing law field.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "worker-loop",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(prepare_payload["worker"]["claimed_by"], "worker-loop")
        self.assertEqual(prepare_payload["worker"]["next_action"], "process_batch")
        self.assertEqual(prepare_payload["worker"]["recommended_execution_mode"], "inline")
        self.assertEqual(prepare_payload["worker"]["prepared_batch_size"], 1)
        self.assertEqual(len(prepare_payload["batch"]), 1)
        batch_entry = prepare_payload["batch"][0]
        self.assertEqual(batch_entry["run_item"]["claimed_by"], "worker-loop")
        self.assertEqual(batch_entry["context"]["job_version"]["capability"], "text_structured")
        self.assertEqual(
            batch_entry["context"]["input"]["inline_text"],
            "Governing law is New York.",
        )
        self.assertEqual(
            batch_entry["context"]["execution"]["completion_template"]["output_values_json"]["governing_law"],
            "<final governing_law value>",
        )

    def test_prepare_run_batch_registers_background_worker_and_tracks_task(self) -> None:
        note_path = self.root / "background-contract.txt"
        note_path.write_text("The governing law is Delaware.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("background-contract.txt")

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Background Contract Metadata",
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "background_contract_metadata",
            "governing_law",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "background_contract_metadata",
            "--instruction",
            "Extract the governing law field.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "background-worker",
            "--launch-mode",
            "background",
            "--worker-task-id",
            "task-123",
            "--max-batches",
            "1",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(prepare_payload["worker"]["next_action"], "process_batch")
        self.assertEqual(prepare_payload["worker_record"]["launch_mode"], "background")
        self.assertEqual(prepare_payload["worker_record"]["worker_task_id"], "task-123")
        self.assertEqual(prepare_payload["worker_record"]["max_batches"], 1)
        self.assertEqual(prepare_payload["worker_record"]["batches_prepared"], 1)
        self.assertEqual(prepare_payload["worker_record"]["status"], "active")

        status_exit, status_payload, _, _ = self.run_cli(
            "run-status",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(status_exit, 0)
        self.assertIsNotNone(status_payload)
        self.assertEqual(status_payload["run"]["workers"][0]["claimed_by"], "background-worker")
        self.assertEqual(status_payload["run"]["workers"][0]["worker_task_id"], "task-123")
        self.assertEqual(status_payload["run"]["supervision"]["background_worker_count"], 1)

    def test_run_status_supervision_recommends_wakeup_and_bounded_worker_count(self) -> None:
        for index in range(12):
            (self.root / f"background-{index:02d}.txt").write_text(
                f"Document {index} governing law is Delaware.",
                encoding="utf-8",
            )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 12)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            document_rows = connection.execute("SELECT id FROM documents ORDER BY id ASC").fetchall()
        finally:
            connection.close()
        document_ids = [int(row["id"]) for row in document_rows]

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Long Background Contract Metadata",
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "long_background_contract_metadata",
            "governing_law",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "long_background_contract_metadata",
            "--instruction",
            "Extract the governing law field.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_args = [
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id IN ({', '.join(str(document_id) for document_id in document_ids)})",
        ]
        create_run_exit, create_run_payload, _, _ = self.run_cli(*create_run_args)
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        status_exit, status_payload, _, _ = self.run_cli(
            "run-status",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(status_exit, 0)
        self.assertIsNotNone(status_payload)
        supervision = status_payload["run"]["supervision"]
        self.assertEqual(supervision["recommended_action"], "spawn_background_worker")
        self.assertTrue(supervision["continuation_needed"])
        self.assertTrue(supervision["should_schedule_wakeup"])
        self.assertEqual(supervision["wake_interval_seconds"], 60)
        self.assertEqual(supervision["wakeup_reason"], "pending_work")
        self.assertEqual(supervision["max_parallel_workers"], 4)
        self.assertEqual(supervision["suggested_worker_count"], 3)
        self.assertEqual(supervision["spawn_additional_worker_count"], 3)

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "background-supervisor-worker",
            "--launch-mode",
            "background",
            "--worker-task-id",
            "task-supervisor-1",
            "--max-batches",
            "1",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)

        after_status_exit, after_status_payload, _, _ = self.run_cli(
            "run-status",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(after_status_exit, 0)
        self.assertIsNotNone(after_status_payload)
        after_supervision = after_status_payload["run"]["supervision"]
        self.assertEqual(after_supervision["background_worker_count"], 1)
        self.assertTrue(after_supervision["should_schedule_wakeup"])
        self.assertEqual(after_supervision["wake_interval_seconds"], 60)
        self.assertEqual(after_supervision["wakeup_reason"], "workers_active")
        self.assertGreaterEqual(after_supervision["spawn_additional_worker_count"], 1)
        self.assertLessEqual(after_supervision["spawn_additional_worker_count"], 3)

    def test_prepare_run_batch_handoffs_after_worker_batch_budget(self) -> None:
        (self.root / "handoff-a.txt").write_text("Document A governing law is New York.", encoding="utf-8")
        (self.root / "handoff-b.txt").write_text("Document B governing law is California.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        first_doc = self.fetch_document_row("handoff-a.txt")
        second_doc = self.fetch_document_row("handoff-b.txt")

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Handoff Test", "structured_extraction")
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "handoff_test",
            "governing_law",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "handoff_test",
            "--instruction",
            "Extract the governing law field.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id IN ({first_doc['id']}, {second_doc['id']})",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "handoff-worker",
            "--launch-mode",
            "background",
            "--worker-task-id",
            "task-handoff",
            "--max-batches",
            "1",
            "--limit",
            "1",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(len(prepare_payload["batch"]), 1)

        batch_entry = prepare_payload["batch"][0]
        complete_exit, complete_payload, _, _ = self.run_cli(
            "complete-run-item",
            str(self.root),
            "--run-item-id",
            str(batch_entry["run_item"]["id"]),
            "--claimed-by",
            "handoff-worker",
            "--raw-output-json",
            json.dumps({"governing_law": "New York"}),
            "--normalized-output-json",
            json.dumps({"governing_law": "New York"}),
            "--output-values-json",
            json.dumps({"governing_law": "New York"}),
        )
        self.assertEqual(complete_exit, 0)
        self.assertIsNotNone(complete_payload)

        handoff_exit, handoff_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "handoff-worker",
            "--launch-mode",
            "background",
            "--worker-task-id",
            "task-handoff",
            "--max-batches",
            "1",
        )
        self.assertEqual(handoff_exit, 0)
        self.assertIsNotNone(handoff_payload)
        self.assertEqual(handoff_payload["batch"], [])
        self.assertEqual(handoff_payload["worker"]["next_action"], "handoff")
        self.assertEqual(handoff_payload["worker"]["stop_reason"], "max_batches_reached")
        self.assertTrue(handoff_payload["worker"]["should_exit_after_batch"])

        finish_exit, finish_payload, _, _ = self.run_cli(
            "finish-run-worker",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "handoff-worker",
            "--worker-status",
            "stopped",
            "--summary-json",
            json.dumps({"reason": "handoff", "processed": 1}),
        )
        self.assertEqual(finish_exit, 0)
        self.assertIsNotNone(finish_payload)
        self.assertEqual(finish_payload["worker"]["status"], "stopped")
        self.assertEqual(finish_payload["worker"]["summary"]["reason"], "handoff")

        status_exit, status_payload, _, _ = self.run_cli(
            "run-status",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(status_exit, 0)
        self.assertIsNotNone(status_payload)
        self.assertEqual(status_payload["run"]["supervision"]["active_worker_count"], 0)
        self.assertTrue(status_payload["run"]["supervision"]["continuation_needed"])

    def test_claim_run_items_reclaims_stale_running_items(self) -> None:
        note_path = self.root / "stale.txt"
        note_path.write_text("Counterparty is Acme.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("stale.txt")

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Counterparty", "structured_extraction")
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "counterparty",
            "counterparty_name",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "counterparty",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Extract the counterparty.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        first_claim_exit, first_claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "worker-a",
            "--launch-mode",
            "background",
            "--worker-task-id",
            "task-stale-a",
            "--limit",
            "1",
        )
        self.assertEqual(first_claim_exit, 0)
        self.assertIsNotNone(first_claim_payload)
        run_item_id = int(first_claim_payload["run_items"][0]["id"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute(
                "UPDATE run_items SET last_heartbeat_at = ? WHERE id = ?",
                ("2000-01-01T00:00:00Z", run_item_id),
            )
            connection.commit()
        finally:
            connection.close()

        second_claim_exit, second_claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "worker-b",
            "--launch-mode",
            "background",
            "--worker-task-id",
            "task-stale-b",
            "--limit",
            "1",
            "--stale-seconds",
            "60",
        )
        self.assertEqual(second_claim_exit, 0)
        self.assertIsNotNone(second_claim_payload)
        self.assertEqual(len(second_claim_payload["run_items"]), 1)
        self.assertEqual(int(second_claim_payload["run_items"][0]["id"]), run_item_id)
        self.assertEqual(second_claim_payload["run_items"][0]["claimed_by"], "worker-b")

        status_exit, status_payload, _, _ = self.run_cli(
            "run-status",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(status_exit, 0)
        self.assertIsNotNone(status_payload)
        workers_by_claimed_by = {
            worker_payload["claimed_by"]: worker_payload for worker_payload in status_payload["run"]["workers"]
        }
        self.assertEqual(workers_by_claimed_by["worker-a"]["status"], "orphaned")
        self.assertEqual(workers_by_claimed_by["worker-b"]["status"], "active")

    def test_cancel_run_skips_pending_items_and_blocks_new_claims(self) -> None:
        (self.root / "a.txt").write_text("Alpha text.", encoding="utf-8")
        (self.root / "b.txt").write_text("Beta text.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        first_doc = self.fetch_document_row("a.txt")
        second_doc = self.fetch_document_row("b.txt")

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Cancel Test", "structured_extraction")
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "cancel_test",
            "summary_text",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "cancel_test",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Summarize the document.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id IN ({first_doc['id']}, {second_doc['id']})",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "worker-a",
            "--limit",
            "1",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        self.assertEqual(len(claim_payload["run_items"]), 1)

        cancel_exit, cancel_payload, _, _ = self.run_cli(
            "cancel-run",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(cancel_exit, 0)
        self.assertIsNotNone(cancel_payload)
        self.assertEqual(cancel_payload["run"]["status"], "canceled")
        self.assertEqual(cancel_payload["canceled_pending_items"], 1)

        post_cancel_claim_exit, post_cancel_claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "worker-b",
            "--limit",
            "1",
        )
        self.assertEqual(post_cancel_claim_exit, 0)
        self.assertIsNotNone(post_cancel_claim_payload)
        self.assertEqual(post_cancel_claim_payload["run"]["status"], "canceled")
        self.assertEqual(post_cancel_claim_payload["run_items"], [])

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "worker-c",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(prepare_payload["batch"], [])
        self.assertEqual(prepare_payload["worker"]["next_action"], "stop")
        self.assertEqual(prepare_payload["worker"]["stop_reason"], "canceled")

    def test_cancel_run_force_returns_background_worker_task_ids(self) -> None:
        note_path = self.root / "force-cancel.txt"
        note_path.write_text("Summary this text.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("force-cancel.txt")

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Force Cancel", "structured_extraction")
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "force_cancel",
            "summary_text",
            "--value-type",
            "text",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "force_cancel",
            "--instruction",
            "Summarize the document.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "force-worker",
            "--launch-mode",
            "background",
            "--worker-task-id",
            "task-force-123",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)

        cancel_exit, cancel_payload, _, _ = self.run_cli(
            "cancel-run",
            str(self.root),
            "--run-id",
            str(run_id),
            "--force",
        )
        self.assertEqual(cancel_exit, 0)
        self.assertIsNotNone(cancel_payload)
        self.assertTrue(cancel_payload["force_stop_requested"])
        self.assertEqual(cancel_payload["force_stop_task_ids"], ["task-force-123"])
        self.assertEqual(cancel_payload["run"]["status"], "canceled")
        self.assertEqual(cancel_payload["run"]["workers"][0]["status"], "canceled")
        self.assertIsNotNone(cancel_payload["run"]["workers"][0]["cancel_requested_at"])
        self.assertEqual(cancel_payload["run"]["supervision"]["force_stop_task_ids"], ["task-force-123"])

    def test_ocr_page_run_items_finalize_into_document_result(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        image_only_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )
        self.assertEqual(image_only_row["text_status"], "empty")

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Production OCR", "ocr")
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "production_ocr",
            "--instruction",
            "OCR each page image and preserve reading order.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        self.assertEqual(create_version_payload["job_version"]["capability"], "vision_ocr")
        self.assertEqual(create_version_payload["job_version"]["input_basis"], "source_parts")
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {image_only_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "ocr-worker",
            "--limit",
            "10",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        claimed_items = claim_payload["run_items"]
        self.assertEqual(len(claimed_items), 2)
        self.assertTrue(all(item["item_kind"] == "page" for item in claimed_items))
        self.assertEqual([item["page_number"] for item in claimed_items], [1, 2])

        first_item_id = int(claimed_items[0]["id"])
        context_exit, context_payload, _, _ = self.run_cli(
            "get-run-item-context",
            str(self.root),
            "--run-item-id",
            str(first_item_id),
        )
        self.assertEqual(context_exit, 0)
        self.assertIsNotNone(context_payload)
        self.assertEqual(context_payload["context"]["input"]["kind"], "ocr_page_image")
        self.assertEqual(context_payload["context"]["input"]["page_number"], 1)
        self.assertTrue(context_payload["context"]["input"]["artifact_path"].endswith("PDX000005.tif"))

        for item in claimed_items:
            page_number = int(item["page_number"])
            complete_exit, complete_payload, _, _ = self.run_cli(
                "complete-run-item",
                str(self.root),
                "--run-item-id",
                str(item["id"]),
                "--claimed-by",
                "ocr-worker",
                "--page-text",
                f"OCR page {page_number}",
            )
            self.assertEqual(complete_exit, 0)
            self.assertIsNotNone(complete_payload)
            self.assertIn("ocr_page_output", complete_payload)
            self.assertEqual(complete_payload["ocr_page_output"]["page_number"], page_number)
            self.assertEqual(complete_payload["ocr_page_output"]["text_content"], f"OCR page {page_number}")

        finalize_exit, finalize_payload, _, _ = self.run_cli(
            "finalize-ocr-run",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(finalize_exit, 0)
        self.assertIsNotNone(finalize_payload)
        self.assertEqual(finalize_payload["run"]["status"], "completed")
        self.assertEqual(len(finalize_payload["results"]), 1)
        result_payload = finalize_payload["results"][0]
        self.assertIsNotNone(result_payload["created_text_revision_id"])

        revisions_exit, revisions_payload, _, _ = self.run_cli(
            "list-text-revisions",
            str(self.root),
            "--doc-id",
            str(image_only_row["id"]),
        )
        self.assertEqual(revisions_exit, 0)
        self.assertIsNotNone(revisions_payload)
        revisions_by_id = {int(item["id"]): item for item in revisions_payload["text_revisions"]}
        ocr_revision = revisions_by_id[int(result_payload["created_text_revision_id"])]
        self.assertEqual(ocr_revision["revision_kind"], "ocr")
        self.assertFalse(ocr_revision["is_active_search_revision"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            text_revision_row = connection.execute(
                "SELECT * FROM text_revisions WHERE id = ?",
                (result_payload["created_text_revision_id"],),
            ).fetchone()
            self.assertIsNotNone(text_revision_row)
            merged_text = retriever_tools.read_text_revision_body(
                self.paths,
                text_revision_row["storage_rel_path"],
            )
            self.assertEqual(merged_text, "OCR page 1\n\nOCR page 2")

            page_item_rows = connection.execute(
                """
                SELECT page_number, result_id, status
                FROM run_items
                WHERE run_id = ?
                ORDER BY page_number ASC, id ASC
                """,
                (run_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual([int(row["page_number"]) for row in page_item_rows], [1, 2])
        self.assertTrue(all(str(row["status"]) == "completed" for row in page_item_rows))
        self.assertTrue(all(int(row["result_id"]) == int(result_payload["id"]) for row in page_item_rows))

    def test_prepare_run_batch_requests_ocr_finalization_after_completed_pages(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        image_only_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Queued OCR", "ocr")
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "queued_ocr",
            "--instruction",
            "OCR each page image.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {image_only_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "ocr-loop",
            "--limit",
            "10",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(prepare_payload["worker"]["next_action"], "process_batch")
        self.assertEqual(len(prepare_payload["batch"]), 2)

        for batch_entry in prepare_payload["batch"]:
            page_number = int(batch_entry["run_item"]["page_number"])
            complete_exit, complete_payload, _, _ = self.run_cli(
                "complete-run-item",
                str(self.root),
                "--run-item-id",
                str(batch_entry["run_item"]["id"]),
                "--claimed-by",
                "ocr-loop",
                "--page-text",
                f"OCR batch page {page_number}",
            )
            self.assertEqual(complete_exit, 0)
            self.assertIsNotNone(complete_payload)

        final_prepare_exit, final_prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "ocr-loop",
        )
        self.assertEqual(final_prepare_exit, 0)
        self.assertIsNotNone(final_prepare_payload)
        self.assertEqual(final_prepare_payload["batch"], [])
        self.assertTrue(final_prepare_payload["worker"]["needs_ocr_finalization"])
        self.assertEqual(final_prepare_payload["worker"]["next_action"], "finalize_ocr")

    def test_finalize_ocr_run_with_always_activation_promotes_revision(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        image_only_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Queued OCR Auto Activate", "ocr")
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "queued_ocr_auto_activate",
            "--instruction",
            "OCR each page image.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--doc-id",
            str(image_only_row["id"]),
            "--activation-policy",
            "always",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        self.assertEqual(create_run_payload["run"]["activation_policy"], "always")
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "ocr-activate-worker",
            "--limit",
            "10",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        for item in claim_payload["run_items"]:
            page_number = int(item["page_number"])
            complete_exit, complete_payload, _, _ = self.run_cli(
                "complete-run-item",
                str(self.root),
                "--run-item-id",
                str(item["id"]),
                "--claimed-by",
                "ocr-activate-worker",
                "--page-text",
                f"OCR auto page {page_number}",
            )
            self.assertEqual(complete_exit, 0)
            self.assertIsNotNone(complete_payload)

        finalize_exit, finalize_payload, _, _ = self.run_cli(
            "finalize-ocr-run",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(finalize_exit, 0)
        self.assertIsNotNone(finalize_payload)
        self.assertEqual(len(finalize_payload["activations"]), 1)
        result_payload = finalize_payload["results"][0]
        created_revision_id = int(result_payload["created_text_revision_id"])
        self.assertEqual(finalize_payload["activations"][0]["text_revision"]["id"], created_revision_id)
        self.assertEqual(finalize_payload["activations"][0]["activation_policy"], "always")

        updated_row = self.fetch_document_by_id(int(image_only_row["id"]))
        self.assertEqual(updated_row["active_search_text_revision_id"], created_revision_id)
        self.assertEqual(updated_row["active_text_source_kind"], "ocr")

    def test_image_description_page_run_items_finalize_into_document_result(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        image_only_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )
        self.assertEqual(image_only_row["text_status"], "empty")

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Production Image Description",
            "image_description",
        )
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "production_image_description",
            "--instruction",
            "Describe each page so image-only productions become searchable.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        self.assertEqual(create_version_payload["job_version"]["capability"], "vision_description")
        self.assertEqual(create_version_payload["job_version"]["input_basis"], "source_parts")
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {image_only_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "image-description-worker",
            "--limit",
            "10",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        claimed_items = claim_payload["run_items"]
        self.assertEqual(len(claimed_items), 2)
        self.assertTrue(all(item["item_kind"] == "page" for item in claimed_items))

        first_item_id = int(claimed_items[0]["id"])
        context_exit, context_payload, _, _ = self.run_cli(
            "get-run-item-context",
            str(self.root),
            "--run-item-id",
            str(first_item_id),
        )
        self.assertEqual(context_exit, 0)
        self.assertIsNotNone(context_payload)
        self.assertEqual(context_payload["context"]["input"]["kind"], "image_description_page_image")
        self.assertEqual(context_payload["context"]["execution"]["capability"], "vision_description")

        for item in claimed_items:
            page_number = int(item["page_number"])
            complete_exit, complete_payload, _, _ = self.run_cli(
                "complete-run-item",
                str(self.root),
                "--run-item-id",
                str(item["id"]),
                "--claimed-by",
                "image-description-worker",
                "--page-text",
                f"Image description page {page_number}",
            )
            self.assertEqual(complete_exit, 0)
            self.assertIsNotNone(complete_payload)
            self.assertIn("image_description_page_output", complete_payload)
            self.assertEqual(
                complete_payload["image_description_page_output"]["text_content"],
                f"Image description page {page_number}",
            )

        finalize_exit, finalize_payload, _, _ = self.run_cli(
            "finalize-image-description-run",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(finalize_exit, 0)
        self.assertIsNotNone(finalize_payload)
        self.assertEqual(finalize_payload["run"]["status"], "completed")
        self.assertEqual(len(finalize_payload["results"]), 1)
        result_payload = finalize_payload["results"][0]
        self.assertIsNotNone(result_payload["created_text_revision_id"])

        revisions_exit, revisions_payload, _, _ = self.run_cli(
            "list-text-revisions",
            str(self.root),
            "--doc-id",
            str(image_only_row["id"]),
        )
        self.assertEqual(revisions_exit, 0)
        self.assertIsNotNone(revisions_payload)
        revisions_by_id = {int(item["id"]): item for item in revisions_payload["text_revisions"]}
        description_revision = revisions_by_id[int(result_payload["created_text_revision_id"])]
        self.assertEqual(description_revision["revision_kind"], "image_description")
        self.assertFalse(description_revision["is_active_search_revision"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            text_revision_row = connection.execute(
                "SELECT * FROM text_revisions WHERE id = ?",
                (result_payload["created_text_revision_id"],),
            ).fetchone()
            self.assertIsNotNone(text_revision_row)
            merged_text = retriever_tools.read_text_revision_body(
                self.paths,
                text_revision_row["storage_rel_path"],
            )
            self.assertEqual(
                merged_text,
                "[IMAGE DESCRIPTION - PAGE 1]\nImage description page 1\n\n"
                "[IMAGE DESCRIPTION - PAGE 2]\nImage description page 2",
            )
        finally:
            connection.close()

    def test_finalize_image_description_run_with_always_activation_promotes_revision(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        image_only_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Queued Image Description Auto Activate",
            "image_description",
        )
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "queued_image_description_auto_activate",
            "--instruction",
            "Describe each page image in search-friendly prose.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--doc-id",
            str(image_only_row["id"]),
            "--activation-policy",
            "always",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        self.assertEqual(create_run_payload["run"]["activation_policy"], "always")
        run_id = int(create_run_payload["run"]["id"])

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "image-description-activate-worker",
            "--limit",
            "10",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        for item in claim_payload["run_items"]:
            page_number = int(item["page_number"])
            complete_exit, complete_payload, _, _ = self.run_cli(
                "complete-run-item",
                str(self.root),
                "--run-item-id",
                str(item["id"]),
                "--claimed-by",
                "image-description-activate-worker",
                "--page-text",
                f"Image description auto page {page_number}",
            )
            self.assertEqual(complete_exit, 0)
            self.assertIsNotNone(complete_payload)

        finalize_exit, finalize_payload, _, _ = self.run_cli(
            "finalize-image-description-run",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(finalize_exit, 0)
        self.assertIsNotNone(finalize_payload)
        self.assertEqual(len(finalize_payload["activations"]), 1)
        result_payload = finalize_payload["results"][0]
        created_revision_id = int(result_payload["created_text_revision_id"])
        self.assertEqual(finalize_payload["activations"][0]["text_revision"]["id"], created_revision_id)
        self.assertEqual(finalize_payload["activations"][0]["activation_policy"], "always")

        updated_row = self.fetch_document_by_id(int(image_only_row["id"]))
        self.assertEqual(updated_row["active_search_text_revision_id"], created_revision_id)
        self.assertEqual(updated_row["active_text_source_kind"], "image_description")

    def test_prepare_run_batch_requests_image_description_finalization_after_completed_pages(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        image_only_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )

        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Queued Image Description",
            "image_description",
        )
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "queued_image_description",
            "--instruction",
            "Describe each page image in search-friendly prose.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {image_only_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "image-description-loop",
            "--limit",
            "10",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(prepare_payload["worker"]["next_action"], "process_batch")
        self.assertEqual(len(prepare_payload["batch"]), 2)

        for batch_entry in prepare_payload["batch"]:
            page_number = int(batch_entry["run_item"]["page_number"])
            complete_exit, complete_payload, _, _ = self.run_cli(
                "complete-run-item",
                str(self.root),
                "--run-item-id",
                str(batch_entry["run_item"]["id"]),
                "--claimed-by",
                "image-description-loop",
                "--page-text",
                f"Batch image description page {page_number}",
            )
            self.assertEqual(complete_exit, 0)
            self.assertIsNotNone(complete_payload)

        final_prepare_exit, final_prepare_payload, _, _ = self.run_cli(
            "prepare-run-batch",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "image-description-loop",
        )
        self.assertEqual(final_prepare_exit, 0)
        self.assertIsNotNone(final_prepare_payload)
        self.assertEqual(final_prepare_payload["batch"], [])
        self.assertTrue(final_prepare_payload["worker"]["needs_image_description_finalization"])
        self.assertEqual(final_prepare_payload["worker"]["next_action"], "finalize_image_description")

    def test_ocr_source_parts_uses_native_pdf_parts_for_production_docs(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        native_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000004.logical"
        )

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Native PDF OCR", "ocr")
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "native_pdf_ocr",
            "--instruction",
            "OCR the native production PDF.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        self.assertEqual(create_version_payload["job_version"]["input_basis"], "source_parts")
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {native_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])
        self.assertEqual(create_run_payload["run"]["planned_count"], 1)

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(run_id),
            "--claimed-by",
            "native-ocr-worker",
            "--limit",
            "10",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        self.assertEqual(len(claim_payload["run_items"]), 1)
        run_item = claim_payload["run_items"][0]
        self.assertEqual(run_item["item_kind"], "page")
        self.assertEqual(run_item["page_number"], 1)
        self.assertTrue(
            str(run_item["input_artifact_rel_path"]).startswith(
                f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/jobs/ocr/run-{run_id}/doc-{native_row['id']}/page-0001"
            )
        )
        self.assertTrue(str(run_item["input_artifact_rel_path"]).endswith(".png"))

        context_exit, context_payload, _, _ = self.run_cli(
            "get-run-item-context",
            str(self.root),
            "--run-item-id",
            str(run_item["id"]),
        )
        self.assertEqual(context_exit, 0)
        self.assertIsNotNone(context_payload)
        self.assertEqual(context_payload["context"]["input"]["kind"], "ocr_page_image")
        self.assertEqual(context_payload["context"]["input"]["page_number"], 1)
        self.assertTrue(str(context_payload["context"]["input"]["artifact_rel_path"]).endswith(".png"))
        self.assertTrue(str(context_payload["context"]["input"]["artifact_path"]).endswith(".png"))

    def test_ocr_run_freezes_artifacts_at_create_run_and_reuses_them_from_prior_run(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        image_only_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            source_part_rows = connection.execute(
                """
                SELECT rel_source_path
                FROM document_source_parts
                WHERE document_id = ?
                  AND part_kind = 'image'
                ORDER BY ordinal ASC, id ASC
                """,
                (image_only_row["id"],),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(len(source_part_rows), 2)
        source_part_paths = [str(row["rel_source_path"]) for row in source_part_rows]

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Frozen OCR", "ocr")
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "frozen_ocr",
            "--instruction",
            "OCR the document pages.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        self.assertEqual(create_version_payload["job_version"]["input_basis"], "source_parts")
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {image_only_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        first_run_id = int(create_run_payload["run"]["id"])
        self.assertEqual(create_run_payload["run"]["planned_count"], 2)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            first_run_item_rows = connection.execute(
                """
                SELECT page_number, input_artifact_rel_path
                FROM run_items
                WHERE run_id = ?
                ORDER BY page_number ASC, id ASC
                """,
                (first_run_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual([int(row["page_number"]) for row in first_run_item_rows], [1, 2])
        first_run_artifact_paths = [str(row["input_artifact_rel_path"]) for row in first_run_item_rows]
        self.assertTrue(
            all(
                path.startswith(
                    f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/jobs/ocr/run-{first_run_id}/doc-{image_only_row['id']}/"
                )
                for path in first_run_artifact_paths
            )
        )
        self.assertTrue(all(path not in source_part_paths for path in first_run_artifact_paths))

        self.write_tiff_fixture(self.root / source_part_paths[0], (17, 34, 51))

        claim_exit, claim_payload, _, _ = self.run_cli(
            "claim-run-items",
            str(self.root),
            "--run-id",
            str(first_run_id),
            "--claimed-by",
            "freeze-worker",
            "--limit",
            "10",
        )
        self.assertEqual(claim_exit, 0)
        self.assertIsNotNone(claim_payload)
        self.assertEqual(len(claim_payload["run_items"]), 2)
        first_run_item_id = int(claim_payload["run_items"][0]["id"])

        context_exit, context_payload, _, _ = self.run_cli(
            "get-run-item-context",
            str(self.root),
            "--run-item-id",
            str(first_run_item_id),
        )
        self.assertEqual(context_exit, 0)
        self.assertIsNotNone(context_payload)
        self.assertEqual(
            context_payload["context"]["input"]["artifact_rel_path"],
            first_run_artifact_paths[0],
        )

        second_run_exit, second_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--from-run-id",
            str(first_run_id),
        )
        self.assertEqual(second_run_exit, 0)
        self.assertIsNotNone(second_run_payload)
        second_run_id = int(second_run_payload["run"]["id"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            second_run_item_rows = connection.execute(
                """
                SELECT page_number, input_artifact_rel_path
                FROM run_items
                WHERE run_id = ?
                ORDER BY page_number ASC, id ASC
                """,
                (second_run_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual([int(row["page_number"]) for row in second_run_item_rows], [1, 2])
        self.assertEqual(
            [str(row["input_artifact_rel_path"]) for row in second_run_item_rows],
            first_run_artifact_paths,
        )

    def test_execute_openai_structured_extraction_run_uses_provider_api(self) -> None:
        note_path = self.root / "party.txt"
        note_path.write_text("Counterparty is Acme Corp.", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        document_row = self.fetch_document_row("party.txt")

        add_field_exit, _, _, _ = self.run_cli("add-field", str(self.root), "counterparty_name", "text")
        self.assertEqual(add_field_exit, 0)
        create_job_exit, _, _, _ = self.run_cli(
            "create-job",
            str(self.root),
            "Counterparty Extract",
            "structured_extraction",
        )
        self.assertEqual(create_job_exit, 0)
        add_output_exit, _, _, _ = self.run_cli(
            "add-job-output",
            str(self.root),
            "counterparty_extract",
            "counterparty_name",
            "--value-type",
            "text",
            "--bind-custom-field",
            "counterparty_name",
        )
        self.assertEqual(add_output_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "counterparty_extract",
            "--provider",
            "openai_responses",
            "--model",
            "gpt-5.4",
            "--input-basis",
            "active_search_text",
            "--instruction",
            "Extract the counterparty.",
            "--parameters-json",
            "{\"temperature\":0}",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--filter",
            f"id = {document_row['id']}",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_id = int(create_run_payload["run"]["id"])

        fake_response = {
            "id": "resp_test_123",
            "status": "completed",
            "output_text": "{\"counterparty_name\":\"Acme Corp\"}",
            "usage": {"input_tokens": 17, "output_tokens": 9},
        }
        with mock.patch.object(retriever_tools, "call_openai_responses_api", return_value=fake_response):
            execute_run_exit, execute_run_payload, _, _ = self.run_cli(
                "execute-run",
                str(self.root),
                "--run-id",
                str(run_id),
            )
        self.assertEqual(execute_run_exit, 0)
        self.assertIsNotNone(execute_run_payload)
        self.assertEqual(execute_run_payload["run"]["status"], "completed")
        self.assertEqual(len(execute_run_payload["results"]), 1)
        result_payload = execute_run_payload["results"][0]
        outputs_by_name = {item["output_name"]: item for item in result_payload["outputs"]}
        self.assertEqual(outputs_by_name["counterparty_name"]["output_value"], "Acme Corp")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            attempt_row = connection.execute(
                """
                SELECT *
                FROM attempts
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(attempt_row)
            self.assertEqual(attempt_row["provider_request_id"], "resp_test_123")
            self.assertEqual(attempt_row["input_tokens"], 17)
            self.assertEqual(attempt_row["output_tokens"], 9)
        finally:
            connection.close()

        publish_exit, publish_payload, _, _ = self.run_cli(
            "publish-run-results",
            str(self.root),
            "--run-id",
            str(run_id),
        )
        self.assertEqual(publish_exit, 0)
        self.assertIsNotNone(publish_payload)
        updated_row = self.fetch_document_by_id(int(document_row["id"]))
        self.assertEqual(updated_row["counterparty_name"], "Acme Corp")

    def test_create_run_select_from_scope_snapshots_current_scope(self) -> None:
        (self.root / "alpha.txt").write_text("scopealpha only\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("scopebeta only\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        alpha_row = self.fetch_document_row("alpha.txt")

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Scope Snapshot", "structured_extraction")
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "scope_snapshot",
            "--instruction",
            "Extract a stable value.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        scope_exit, scope_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "scopealpha")
        self.assertEqual(scope_exit, 0)
        self.assertIsNotNone(scope_payload)
        self.assertEqual(scope_payload["total_hits"], 1)

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--select-from-scope",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_payload = create_run_payload["run"]
        self.assertEqual(run_payload["planned_count"], 1)
        self.assertEqual(run_payload["selector"]["keyword"], "scopealpha")
        self.assertEqual([item["document_id"] for item in run_payload["documents"]], [alpha_row["id"]])

        shifted_scope_exit, shifted_scope_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "scopebeta")
        self.assertEqual(shifted_scope_exit, 0)
        self.assertIsNotNone(shifted_scope_payload)
        self.assertEqual(shifted_scope_payload["total_hits"], 1)

        stored_run_exit, stored_run_payload, _, _ = self.run_cli(
            "get-run",
            str(self.root),
            "--run-id",
            str(run_payload["id"]),
        )
        self.assertEqual(stored_run_exit, 0)
        self.assertIsNotNone(stored_run_payload)
        self.assertEqual(stored_run_payload["run"]["selector"]["keyword"], "scopealpha")
        self.assertEqual([item["document_id"] for item in stored_run_payload["run"]["documents"]], [alpha_row["id"]])

    def test_create_run_select_from_scope_and_narrows_with_explicit_dataset(self) -> None:
        (self.root / "alpha.txt").write_text("scope dataset alpha body\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("scope dataset beta body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        alpha_row = self.fetch_document_row("alpha.txt")
        beta_row = self.fetch_document_row("beta.txt")

        create_scope_dataset_exit, _, _, _ = self.run_cli("create-dataset", str(self.root), "Scope Set")
        self.assertEqual(create_scope_dataset_exit, 0)
        create_narrow_dataset_exit, _, _, _ = self.run_cli("create-dataset", str(self.root), "Narrow Set")
        self.assertEqual(create_narrow_dataset_exit, 0)

        add_scope_set_exit, _, _, _ = self.run_cli(
            "add-to-dataset",
            str(self.root),
            "--dataset-name",
            "Scope Set",
            "--doc-id",
            str(alpha_row["id"]),
            "--doc-id",
            str(beta_row["id"]),
        )
        self.assertEqual(add_scope_set_exit, 0)
        add_narrow_set_exit, _, _, _ = self.run_cli(
            "add-to-dataset",
            str(self.root),
            "--dataset-name",
            "Narrow Set",
            "--doc-id",
            str(beta_row["id"]),
        )
        self.assertEqual(add_narrow_set_exit, 0)

        create_job_exit, _, _, _ = self.run_cli("create-job", str(self.root), "Scope Dataset Run", "structured_extraction")
        self.assertEqual(create_job_exit, 0)
        create_version_exit, create_version_payload, _, _ = self.run_cli(
            "create-job-version",
            str(self.root),
            "scope_dataset_run",
            "--instruction",
            "Extract a stable value.",
        )
        self.assertEqual(create_version_exit, 0)
        self.assertIsNotNone(create_version_payload)
        job_version_id = int(create_version_payload["job_version"]["id"])

        scope_exit, scope_payload, _, _ = self.run_cli("slash", str(self.root), "/dataset", "Scope Set")
        self.assertEqual(scope_exit, 0)
        self.assertIsNotNone(scope_payload)
        self.assertEqual(scope_payload["total_hits"], 2)

        create_run_exit, create_run_payload, _, _ = self.run_cli(
            "create-run",
            str(self.root),
            "--job-version-id",
            str(job_version_id),
            "--select-from-scope",
            "--dataset",
            "Narrow Set",
        )
        self.assertEqual(create_run_exit, 0)
        self.assertIsNotNone(create_run_payload)
        run_payload = create_run_payload["run"]
        self.assertEqual(run_payload["planned_count"], 1)
        self.assertEqual([item["document_id"] for item in run_payload["documents"]], [beta_row["id"]])
        self.assertIn("all_of", run_payload["selector"])
        self.assertEqual(len(run_payload["selector"]["all_of"]), 2)
        self.assertEqual(run_payload["selector"]["all_of"][0]["dataset"][0]["name"], "Scope Set")
        self.assertEqual(run_payload["selector"]["all_of"][1]["dataset"][0]["name"], "Narrow Set")

    def test_filter_errors_match_search_create_run_and_export_archive(self) -> None:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        job_version_id = self.create_structured_extraction_job_version("Filter Parity")

        for expression in ("authr = 'Smith'", "is_attachment > 5"):
            with self.subTest(expression=expression):
                search_exit, search_payload, _, _ = self.run_cli(
                    "search",
                    str(self.root),
                    "--filter",
                    expression,
                )
                self.assertEqual(search_exit, 2)
                self.assertIsNotNone(search_payload)

                create_run_exit, create_run_payload, _, _ = self.run_cli(
                    "create-run",
                    str(self.root),
                    "--job-version-id",
                    str(job_version_id),
                    "--filter",
                    expression,
                )
                self.assertEqual(create_run_exit, 2)
                self.assertIsNotNone(create_run_payload)
                self.assertEqual(create_run_payload["error"], search_payload["error"])

                archive_exit, archive_payload, _, _ = self.run_cli(
                    "export-archive",
                    str(self.root),
                    "filter-error-parity.zip",
                    "--filter",
                    expression,
                )
                self.assertEqual(archive_exit, 2)
                self.assertIsNotNone(archive_payload)
                self.assertEqual(archive_payload["error"], search_payload["error"])

    def test_filter_surface_parity_matches_search_create_run_and_export_archive(self) -> None:
        corpus_rows = self.setup_randomized_sql_filter_corpus()
        job_version_id = self.create_structured_extraction_job_version("Surface Filter Parity")
        expressions = [
            "review_score BETWEEN 1 AND 7 AND NOT is_hot = FALSE",
            "(file_name IN ('alpha.txt', 'notes.txt') OR review_note LIKE '%review%') AND review_date IS NOT NULL",
            "review_score IS NULL OR review_note = 'beta''s note'",
            "NOT (review_note LIKE '%thread%' OR review_score > 8)",
            "file_name NOT LIKE '%eml' AND (review_score >= 5 OR is_hot IS NULL)",
        ]

        for index, expression in enumerate(expressions, start=1):
            expected_doc_ids = self.reference_doc_ids_for_sql_filter(corpus_rows, expression)

            with self.subTest(expression=expression):
                search_exit, search_payload, _, _ = self.run_cli(
                    "search",
                    str(self.root),
                    "--filter",
                    expression,
                )
                self.assertEqual(search_exit, 0)
                self.assertIsNotNone(search_payload)
                search_doc_ids = sorted(int(item["id"]) for item in search_payload["results"])

                create_run_exit, create_run_payload, _, _ = self.run_cli(
                    "create-run",
                    str(self.root),
                    "--job-version-id",
                    str(job_version_id),
                    "--filter",
                    expression,
                )
                self.assertEqual(create_run_exit, 0)
                self.assertIsNotNone(create_run_payload)
                create_run_doc_ids = sorted(
                    int(item["document_id"]) for item in create_run_payload["run"]["documents"]
                )

                archive_exit, archive_payload, _, _ = self.run_cli(
                    "export-archive",
                    str(self.root),
                    f"filter-surface-parity-{index}.zip",
                    "--filter",
                    expression,
                )
                self.assertEqual(archive_exit, 0)
                self.assertIsNotNone(archive_payload)
                archive_doc_ids = sorted(
                    int(item["document_id"]) for item in archive_payload["documents"]
                )

                self.assertEqual(search_doc_ids, expected_doc_ids)
                self.assertEqual(create_run_doc_ids, expected_doc_ids)
                self.assertEqual(archive_doc_ids, expected_doc_ids)

    def test_connect_db_falls_back_to_delete_when_wal_fails(self) -> None:
        class FakeCursor:
            def __init__(self, row):
                self._row = row

            def fetchone(self):
                return self._row

        class FakeConnection:
            def __init__(self) -> None:
                self.row_factory = None
                self.commands: list[str] = []

            def execute(self, statement: str):
                self.commands.append(statement)
                if statement == "PRAGMA journal_mode = WAL":
                    raise sqlite3.OperationalError("wal unsupported")
                if statement == "PRAGMA journal_mode = DELETE":
                    return FakeCursor(["delete"])
                return FakeCursor(None)

            def close(self) -> None:
                return None

        fake_connection = FakeConnection()
        with mock.patch.object(retriever_tools.sqlite3, "connect", return_value=fake_connection):
            connection = retriever_tools.connect_db(self.paths["db_path"])

        self.assertIs(connection, fake_connection)
        self.assertIn("PRAGMA journal_mode = WAL", fake_connection.commands)
        self.assertIn("PRAGMA journal_mode = DELETE", fake_connection.commands)

    def test_workspace_paths_include_ingest_lock_and_tmp_dirs(self) -> None:
        self.assertEqual(self.paths["tmp_dir"], self.root / ".retriever" / "tmp")
        self.assertEqual(self.paths["ingest_tmp_dir"], self.root / ".retriever" / "tmp" / "ingest")
        self.assertEqual(self.paths["locks_dir"], self.root / ".retriever" / "locks")
        self.assertEqual(self.paths["ingest_lock_path"], self.root / ".retriever" / "locks" / "ingest.lock")

    def test_workspace_ingest_session_sweeps_stale_tmp_dirs_and_cleans_current_dir(self) -> None:
        stale_dir = self.paths["ingest_tmp_dir"] / "stale-session"
        stale_dir.mkdir(parents=True, exist_ok=True)
        (stale_dir / "payload.txt").write_text("stale\n", encoding="utf-8")

        with retriever_tools.workspace_ingest_session(self.paths, command_name="ingest") as session:
            self.assertEqual(session["stale_tmp_dirs_removed"], 1)
            self.assertFalse(stale_dir.exists())
            session_dir = Path(session["tmp_dir"])
            self.assertTrue(session_dir.exists())
            self.assertEqual(session_dir.parent, self.paths["ingest_tmp_dir"])
            self.assertTrue(self.paths["ingest_lock_path"].exists())

        self.assertFalse(session_dir.exists())

    def test_workspace_ingest_session_warns_when_stale_tmp_dir_cannot_be_removed(self) -> None:
        stale_dir = self.paths["ingest_tmp_dir"] / "stale-session"
        stale_dir.mkdir(parents=True, exist_ok=True)
        (stale_dir / "payload.txt").write_text("stale\n", encoding="utf-8")
        real_remove_directory_tree = retriever_tools.remove_directory_tree

        def remove_directory_tree(path: Path) -> bool:
            if Path(path) == stale_dir:
                raise PermissionError("sandbox denied delete")
            return real_remove_directory_tree(path)

        with mock.patch.object(retriever_tools, "remove_directory_tree", side_effect=remove_directory_tree):
            with retriever_tools.workspace_ingest_session(self.paths, command_name="ingest") as session:
                self.assertEqual(session["stale_tmp_dirs_removed"], 0)
                self.assertEqual(session["stale_tmp_dirs_failed"], 1)
                self.assertIn("PermissionError", session["warnings"][0])
                self.assertTrue(Path(session["tmp_dir"]).exists())

        self.assertTrue(stale_dir.exists())

    def test_acquire_workspace_ingest_lock_raises_clear_error_on_lock_conflict(self) -> None:
        blocking_error = BlockingIOError(errno.EAGAIN, "already locked")
        with mock.patch.object(retriever_tools, "acquire_os_file_lock", side_effect=blocking_error):
            with self.assertRaisesRegex(
                retriever_tools.RetrieverError,
                "Another ingest is already running in this workspace",
            ):
                retriever_tools.acquire_workspace_ingest_lock(self.paths)

    def test_iter_prepared_loose_file_items_preserves_input_order_with_multiple_workers(self) -> None:
        items = [
            {
                "path": self.root / "alpha.txt",
                "rel_path": "alpha.txt",
                "file_type": "txt",
                "file_hash": "alpha",
            },
            {
                "path": self.root / "beta.txt",
                "rel_path": "beta.txt",
                "file_type": "txt",
                "file_hash": "beta",
            },
        ]

        def fake_prepare(item: dict[str, object]) -> dict[str, object]:
            if item["rel_path"] == "alpha.txt":
                time.sleep(0.05)
            prepared = dict(item)
            prepared["prepare_ms"] = 1.0
            prepared["prepare_error"] = None
            prepared["extracted_payload"] = {"text_content": item["rel_path"]}
            prepared["attachments"] = []
            return prepared

        with mock.patch.dict(os.environ, {"RETRIEVER_INGEST_WORKERS": "2"}):
            with mock.patch.object(retriever_tools, "prepare_loose_file_item", side_effect=fake_prepare):
                prepared = list(retriever_tools.iter_prepared_loose_file_items(items))

        self.assertEqual([item["rel_path"] for item, _ in prepared], ["alpha.txt", "beta.txt"])

    def test_iter_prepared_container_message_items_preserves_input_order_with_multiple_workers(self) -> None:
        def raw_messages():
            yield {"source_item_id": "first"}
            yield {"source_item_id": "second"}

        def fake_normalize(source_rel_path: str, raw_message: dict[str, object]) -> dict[str, object]:
            self.assertEqual(source_rel_path, "mailbox.mbox")
            if raw_message["source_item_id"] == "first":
                time.sleep(0.05)
            return {
                "rel_path": f"_retriever/logical/mbox/{raw_message['source_item_id']}.eml",
                "file_name": f"{raw_message['source_item_id']}.eml",
                "file_hash": str(raw_message["source_item_id"]),
                "source_item_id": str(raw_message["source_item_id"]),
                "source_folder_path": None,
                "extracted": {
                    "text_content": str(raw_message["source_item_id"]),
                    "preview_artifacts": [],
                    "attachments": [],
                },
            }

        with mock.patch.object(retriever_tools, "ingest_container_prepare_worker_count", return_value=2):
            prepared = list(
                retriever_tools.iter_prepared_container_message_items(
                    source_kind=retriever_tools.MBOX_SOURCE_KIND,
                    source_rel_path="mailbox.mbox",
                    raw_messages=raw_messages(),
                    normalize_message=fake_normalize,
                )
            )

        self.assertEqual(
            [item["source_item_id"] for item, _ in prepared],
            ["first", "second"],
        )
        self.assertFalse(prepared[0][0]["skip"])
        self.assertTrue(prepared[0][0]["prepared_chunks"])

    def test_refresh_prepared_loose_file_item_if_stale_reprepares_with_new_contents(self) -> None:
        path = self.root / "alpha.txt"
        path.write_text("first version\n", encoding="utf-8")
        item = retriever_tools.refresh_ingest_item_filesystem_facts(
            {
                "path": path,
                "rel_path": "alpha.txt",
                "file_type": "txt",
            }
        )
        prepared = retriever_tools.prepare_loose_file_item(item)

        time.sleep(0.02)
        path.write_text("second version\nwith more text\n", encoding="utf-8")

        refreshed, did_refresh = retriever_tools.refresh_prepared_loose_file_item_if_stale(prepared)

        self.assertTrue(did_refresh)
        self.assertNotEqual(prepared["file_hash"], refreshed["file_hash"])
        self.assertIn("second version", refreshed["extracted_payload"]["text_content"])
        self.assertTrue(refreshed["prepared_chunks"])
        self.assertIn("second version", refreshed["prepared_chunks"][0]["text_content"])

    def test_iter_prepared_loose_file_items_spills_large_payloads_to_temp_dir(self) -> None:
        items = [
            {
                "path": self.root / "alpha.txt",
                "rel_path": "alpha.txt",
                "file_type": "txt",
                "file_hash": "alpha",
            }
        ]
        staging_root = self.paths["ingest_tmp_dir"] / "spill-test"

        def fake_prepare(item: dict[str, object]) -> dict[str, object]:
            prepared = dict(item)
            prepared["prepare_ms"] = 1.0
            prepared["prepare_chunk_ms"] = 0.1
            prepared["prepare_error"] = None
            prepared["extracted_payload"] = {
                "text_content": "x" * 4096,
                "preview_artifacts": [{"file_name": "preview.html", "preview_type": "html", "content": "x" * 4096}],
            }
            prepared["attachments"] = [
                {
                    "file_name": "payload.bin",
                    "payload": b"x" * 4096,
                    "file_hash": "payload-hash",
                }
            ]
            prepared["prepared_chunks"] = [
                {
                    "chunk_index": 0,
                    "char_start": 0,
                    "char_end": 32,
                    "token_estimate": 8,
                    "text_content": "x" * 32,
                }
            ]
            return prepared

        with mock.patch.object(retriever_tools, "prepare_loose_file_item", side_effect=fake_prepare):
            with mock.patch.object(retriever_tools, "ingest_prepare_worker_count", return_value=1):
                with mock.patch.object(retriever_tools, "ingest_prepare_queue_capacity", return_value=1):
                    with mock.patch.object(retriever_tools, "ingest_prepare_queue_max_bytes", return_value=1):
                        with mock.patch.object(retriever_tools, "ingest_prepare_spill_threshold_bytes", return_value=1):
                            prepared = list(retriever_tools.iter_prepared_loose_file_items(items, staging_root))

        self.assertEqual(prepared[0][0]["rel_path"], "alpha.txt")
        spill_dir = staging_root / "prepared-loose"
        self.assertTrue(spill_dir.exists())
        self.assertEqual(list(spill_dir.iterdir()), [])

    def test_bootstrap_recovers_zero_byte_sqlite_artifacts(self) -> None:
        self.paths["db_path"].touch()
        Path(f"{self.paths['db_path']}-journal").write_text("stale\n", encoding="utf-8")

        result = retriever_tools.bootstrap(self.root)

        self.assertEqual(result["schema_version"], retriever_tools.SCHEMA_VERSION)
        self.assertIn("recovered_sqlite_artifacts", result)
        self.assertIn(str(self.paths["db_path"]), result["recovered_sqlite_artifacts"])
        self.assertGreater(self.paths["db_path"].stat().st_size, 0)

    def test_bootstrap_then_ingest_creates_email_attachment_children_and_search_context(self) -> None:
        self.create_legacy_documents_table(with_row=False)
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Upgrade test",
            body_text="Hello team,\nThis is the email body.",
            attachment_name="notes.txt",
            attachment_text="confidential attachment detail",
        )

        retriever_tools.bootstrap(self.root)
        result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(result["new"], 1)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(result["scanned_files"], 1)
        self.assertEqual(result["workspace_parent_documents"], 1)
        self.assertEqual(result["workspace_attachment_children"], 1)
        self.assertEqual(result["workspace_documents_total"], 2)

        parent_row = self.fetch_document_row("thread.eml")
        self.assertEqual(parent_row["content_type"], "Email")
        self.assertEqual(parent_row["control_number"], "DOC001.00000001")
        self.assertIsNotNone(parent_row["participants"])
        self.assertIn("alice@example.com", parent_row["participants"])
        self.assertIn("bob@example.com", parent_row["participants"])
        self.assertIn("carol@example.com", parent_row["participants"])

        child_rows = self.fetch_child_rows(parent_row["id"])
        self.assertEqual(len(child_rows), 1)
        child_row = child_rows[0]
        self.assertEqual(child_row["file_name"], "notes.txt")
        self.assertEqual(child_row["control_number"], "DOC001.00000001.001")
        self.assertEqual(child_row["parent_document_id"], parent_row["id"])
        self.assertTrue(
            str(child_row["rel_path"]).startswith(
                f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/previews/thread.eml/attachments/"
            )
        )
        self.assertEqual(child_row["content_type"], "E-Doc")

        parent_search = retriever_tools.search(self.root, "Upgrade test", None, None, None, 1, 20)
        parent_result = next(item for item in parent_search["results"] if item["id"] == parent_row["id"])
        self.assertEqual(parent_result["control_number"], parent_row["control_number"])
        self.assertEqual(parent_result["attachment_count"], 1)
        self.assertEqual(parent_result["attachments"][0]["control_number"], child_row["control_number"])
        self.assertEqual(parent_result["attachments"][0]["file_name"], "notes.txt")

        attachment_search = retriever_tools.search(self.root, "confidential attachment detail", None, None, None, 1, 20)
        attachment_result = next(item for item in attachment_search["results"] if item["id"] == child_row["id"])
        self.assertEqual(attachment_result["control_number"], child_row["control_number"])
        self.assertEqual(attachment_result["parent"]["control_number"], parent_row["control_number"])
        self.assertEqual(attachment_result["parent"]["file_name"], "thread.eml")

        attachments_only = retriever_tools.search(
            self.root,
            "",
            [["is_attachment", "eq", "true"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(attachments_only["total_hits"], 1)
        self.assertEqual(attachments_only["results"][0]["id"], child_row["id"])

        parents_with_attachments = retriever_tools.search(
            self.root,
            "",
            [["has_attachments", "eq", "true"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(parents_with_attachments["total_hits"], 1)
        self.assertEqual(parents_with_attachments["results"][0]["id"], parent_row["id"])

    def test_email_preview_promotes_calendar_attachments_into_invite_cards(self) -> None:
        email_path = self.root / "thread.eml"
        message = EmailMessage()
        message["From"] = "Sergey Demyanov <sergey@discoverbeagle.com>"
        message["To"] = "Max Faleev <max@discoverbeagle.com>"
        message["Subject"] = "Discuss Relativity"
        message["Date"] = "Thu, 01 Jun 2023 00:31:26 +0000"
        message["Message-ID"] = "<invite@example.com>"
        message.set_content("Invitation from Google Calendar")
        message.add_attachment(
            "\r\n".join(
                [
                    "BEGIN:VCALENDAR",
                    "VERSION:2.0",
                    "METHOD:REQUEST",
                    "BEGIN:VEVENT",
                    "SUMMARY:Discuss Relativity",
                    "DTSTART;TZID=America/New_York:20230601T150000",
                    "DTEND;TZID=America/New_York:20230601T153000",
                    "ORGANIZER;CN=Sergey Demyanov:mailto:sergey@discoverbeagle.com",
                    "ATTENDEE;CN=Max Faleev:mailto:max@discoverbeagle.com",
                    "STATUS:CONFIRMED",
                    "X-GOOGLE-CONFERENCE:https://meet.google.com/fps-qara-aie",
                    "END:VEVENT",
                    "END:VCALENDAR",
                    "",
                ]
            ),
            subtype="calendar",
            filename="invite.ics",
        )
        message.add_attachment("Agenda note", subtype="plain", filename="notes.txt")
        email_path.write_bytes(message.as_bytes(policy=policy.default))

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)

        parent_row = self.fetch_document_row("thread.eml")
        child_rows = self.fetch_child_rows(parent_row["id"])
        self.assertEqual(len(child_rows), 1)
        child_row = child_rows[0]
        self.assertEqual(child_row["file_name"], "notes.txt")

        search_result = retriever_tools.search(self.root, "Invitation from Google Calendar", None, None, None, 1, 20)
        parent_result = next(item for item in search_result["results"] if item["id"] == parent_row["id"])
        self.assertEqual(parent_result["attachment_count"], 1)
        self.assertEqual(parent_result["attachments"][0]["file_name"], "notes.txt")

        invite_search = retriever_tools.search(self.root, "qara", None, None, None, 1, 20)
        self.assertEqual(invite_search["total_hits"], 1)
        self.assertEqual(invite_search["results"][0]["id"], parent_row["id"])

        attachments_only = retriever_tools.search(
            self.root,
            "",
            [["is_attachment", "eq", "true"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(attachments_only["total_hits"], 1)
        self.assertEqual(attachments_only["results"][0]["id"], child_row["id"])

        message_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(parent_result["preview_targets"], "message")
        ).read_text(encoding="utf-8")
        self.assertEqual([target.get("label") for target in parent_result["preview_targets"]], ["message"])
        self.assertIn("Calendar invite", message_preview_html)
        self.assertIn("Discuss Relativity", message_preview_html)
        self.assertIn("Jun 1, 2023 3:00 PM - 3:30 PM EDT", message_preview_html)
        self.assertIn("Sergey Demyanov &lt;sergey@discoverbeagle.com&gt;", message_preview_html)
        self.assertIn("Max Faleev &lt;max@discoverbeagle.com&gt;", message_preview_html)
        self.assertIn("meet.google.com/fps-qara-aie", message_preview_html)
        self.assertIn("notes.txt", message_preview_html)
        self.assertNotIn(">invite.ics<", message_preview_html)

    def test_ingest_email_inline_html_image_stays_in_preview_without_attachment_child(self) -> None:
        email_path = self.root / "inline.eml"
        png_pixel = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+a9mQAAAAASUVORK5CYII="
        )
        encoded_png = base64.b64encode(png_pixel).decode("ascii")
        boundary = "boundary-inline-ingest"
        email_path.write_bytes(
            (
                "From: Alice Example <alice@example.com>\r\n"
                "To: Bob Example <bob@example.com>\r\n"
                "Subject: Inline icon test\r\n"
                "Date: Tue, 14 Apr 2026 10:00:00 +0000\r\n"
                "MIME-Version: 1.0\r\n"
                f'Content-Type: multipart/related; boundary="{boundary}"\r\n'
                "\r\n"
                f"--{boundary}\r\n"
                "Content-Type: text/html; charset=utf-8\r\n"
                "\r\n"
                '<html><body><p>Hello team.</p><img src="cid:logo-icon" alt="logo"/></body></html>\r\n'
                f"--{boundary}\r\n"
                "Content-Type: image/png\r\n"
                "Content-Transfer-Encoding: base64\r\n"
                "Content-ID: <logo-icon>\r\n"
                'Content-Disposition: inline; filename="logo.png"\r\n'
                "\r\n"
                f"{encoded_png}\r\n"
                f"--{boundary}--\r\n"
            ).encode("ascii")
        )

        retriever_tools.bootstrap(self.root)
        result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(result["new"], 1)
        self.assertEqual(result["failed"], 0)
        self.assertEqual(result["workspace_parent_documents"], 1)
        self.assertEqual(result["workspace_attachment_children"], 0)
        self.assertEqual(result["workspace_documents_total"], 1)

        parent_row = self.fetch_document_row("inline.eml")
        self.assertEqual(len(self.fetch_child_rows(parent_row["id"])), 0)

        browse_results = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        parent_result = next(item for item in browse_results["results"] if item["id"] == parent_row["id"])
        self.assertEqual(parent_result["attachment_count"], 0)

        preview_html = Path(parent_result["preview_abs_path"]).read_text(encoding="utf-8")
        self.assertIn("data:image/png;base64,", preview_html)
        self.assertNotIn('src="cid:logo-icon"', preview_html)

    def test_ingest_email_preview_moves_attachment_lists_out_of_titles_and_into_links(self) -> None:
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Energy Balance Revenue Summary and Back Up Attachments: notes.txt; revenue.txt",
            body_text="Hello team,\nThis is the email body.",
            attachment_name="notes.txt",
            attachment_text="confidential attachment detail",
        )

        retriever_tools.bootstrap(self.root)
        result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(result["new"], 1)
        self.assertEqual(result["failed"], 0)

        parent_row = self.fetch_document_row("thread.eml")
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        self.assertEqual(parent_row["title"], "Energy Balance Revenue Summary")
        self.assertEqual(parent_row["subject"], "Energy Balance Revenue Summary")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            parent_preview_target = retriever_tools.default_preview_target(self.paths, parent_row, connection)
            child_preview_target = retriever_tools.default_preview_target(self.paths, child_row, connection)
            parent_preview_targets = retriever_tools.collect_preview_targets(
                self.paths,
                int(parent_row["id"]),
                str(parent_row["rel_path"]),
                connection,
            )
            parent_preview_abs_path = str(
                parent_preview_target.get("file_abs_path") or parent_preview_target.get("abs_path") or ""
            ).split("#", 1)[0]
            child_preview_abs_path = str(
                child_preview_target.get("file_abs_path") or child_preview_target.get("abs_path") or ""
            ).split("#", 1)[0]
        finally:
            connection.close()

        self.assertEqual([target.get("label") for target in parent_preview_targets], ["message"])
        preview_path = self.preview_target_file_path(
            self.preview_target_by_label(parent_preview_targets, "message")
        )
        preview_html = preview_path.read_text(encoding="utf-8")

        self.assertIn("<title>Energy Balance Revenue Summary</title>", preview_html)
        self.assertIn('class="gmail-thread-title">Energy Balance Revenue Summary</h1>', preview_html)
        self.assertNotIn("Back Up Attachments:", preview_html)
        self.assertIn("<h2>Attachments</h2>", preview_html)
        self.assertIn(">notes.txt<", preview_html)
        self.assertRegex(preview_html, r'href="[^"]*notes\.txt"')

        with mock.patch.object(retriever_tools, "pypff", object()):
            doctor_result = retriever_tools.doctor(self.root, quick=False)
        self.assertEqual(doctor_result["workspace_inventory"]["parent_documents"], 1)
        self.assertEqual(doctor_result["workspace_inventory"]["attachment_children"], 1)
        self.assertEqual(doctor_result["workspace_inventory"]["documents_total"], 2)
        self.assertIn(doctor_result["sqlite_journal_mode"], {"wal", "delete"})

        with self.assertRaises(retriever_tools.RetrieverError) as context:
            retriever_tools.search(self.root, "", None, "is_attachment", None, 1, 20)
        self.assertIn("virtual filter field", str(context.exception))

    def test_email_preview_styles_attachment_section_heading(self) -> None:
        preview_html = retriever_tools.build_email_thread_preview_html(
            thread_title="Energy Balance Revenue Summary",
            page_title="Energy Balance Revenue Summary",
            documents=[
                {
                    "id": 1,
                    "author": "Alice Example <alice@example.com>",
                    "recipients": "Bob Example <bob@example.com>",
                    "date_created": "2026-04-15T09:00:00Z",
                    "content_type": "Email",
                    "text_content": "Hello team.",
                }
            ],
            attachment_links_by_document_id={
                1: [
                    {
                        "href": "attachments/notes.txt",
                        "label": "notes.txt",
                        "detail": "text/plain",
                    }
                ]
            },
        )

        self.assertIn('<section class="retriever-attachments"><h2>Attachments</h2><ul>', preview_html)
        self.assertIn(".retriever-attachments h2", preview_html)
        self.assertNotIn(".retriever-attachments h3", preview_html)

    def test_report_style_colon_sections_are_not_extracted_as_participants(self) -> None:
        report_text = "\n".join(
            [
                "INTELLECTUAL PROPERTY AUDIT REPORT",
                "",
                "TO:",
                "",
                "Amanda Foster, General Counsel",
                "",
                "Dr. Patricia Liu, Chief Regulatory Officer",
                "",
                "FROM:",
                "",
                "Dr. Helen Chang, Director of Intellectual Property",
                "",
                "DATE:",
                "",
                "March 15, 2023",
                "",
                "January 15, 2023: Routine patent maintenance review identified discrepancies.",
                "February 3, 2023: Stanford University Technology Licensing Office inquiry.",
                "MIT Licensing Obligations: Three patents involving manufacturing processes.",
                "Due Diligence Exposure: Potential acquirers will demand full disclosure.",
            ]
        )

        email_headers = retriever_tools.extract_email_like_headers(report_text)

        self.assertEqual(email_headers, {})
        self.assertIsNone(retriever_tools.extract_chat_transcript_metadata(report_text))
        self.assertIsNone(retriever_tools.extract_chat_participants(report_text))
        self.assertEqual(
            retriever_tools.determine_content_type(
                Path("audit.docx"),
                report_text,
                email_headers=email_headers,
                chat_metadata=None,
                explicit_content_type="E-Doc",
            ),
            "E-Doc",
        )
        self.assertEqual(
            retriever_tools.extract_chat_participants(
                "Alice Example: Hi Bob.\nBob Example: Hi Alice.\nAlice Example: Thanks."
            ),
            "Alice Example, Bob Example",
        )

    def test_reingest_clears_stale_conversation_when_document_becomes_e_doc(self) -> None:
        memo_path = self.root / "audit.txt"
        memo_path.write_text("Plain audit report.\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["failed"], 0)

        row = self.fetch_document_row("audit.txt")
        self.assertEqual(row["content_type"], "E-Doc")
        self.assertIsNone(row["conversation_id"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            conversation_id = retriever_tools.upsert_conversation_row(
                connection,
                source_kind=retriever_tools.EMAIL_CONVERSATION_SOURCE_KIND,
                source_locator=retriever_tools.filesystem_dataset_locator(),
                conversation_key="stale-audit-thread",
                conversation_type="email",
                display_name="Plain audit report",
            )
            connection.execute(
                """
                UPDATE documents
                SET content_type = 'Email',
                    canonical_kind = 'email',
                    conversation_id = ?,
                    conversation_assignment_mode = ?,
                    participants = 'January 15, 2023'
                WHERE id = ?
                """,
                (
                    conversation_id,
                    retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO,
                    int(row["id"]),
                ),
            )
            connection.commit()
        finally:
            connection.close()

        memo_path.write_text("Plain audit report, revised.\n", encoding="utf-8")
        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_ingest["failed"], 0)

        updated_row = self.fetch_document_row("audit.txt")
        self.assertEqual(updated_row["content_type"], "E-Doc")
        self.assertEqual(updated_row["canonical_kind"], "document")
        self.assertIsNone(updated_row["conversation_id"])
        self.assertEqual(updated_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)
        self.assertIsNone(updated_row["participants"])

    def test_reingest_refreshes_unchanged_file_when_current_extract_differs_from_stored_metadata(self) -> None:
        memo_path = self.root / "ip-audit.txt"
        memo_path.write_text(
            "\n".join(
                [
                    "INTELLECTUAL PROPERTY AUDIT REPORT",
                    "TO:",
                    "Amanda Foster, General Counsel",
                    "FROM:",
                    "Dr. Helen Chang, Director of Intellectual Property",
                    "DATE:",
                    "March 15, 2023",
                    "RE:",
                    "Critical Patent Encumbrance Issues",
                    "January 15, 2023: Routine patent maintenance review identified discrepancies.",
                    "MIT Licensing Obligations: Three patents involving manufacturing processes.",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["failed"], 0)

        row = self.fetch_document_row("ip-audit.txt")
        self.assertEqual(row["content_type"], "E-Doc")
        self.assertIsNone(row["conversation_id"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            conversation_id = retriever_tools.upsert_conversation_row(
                connection,
                source_kind=retriever_tools.EMAIL_CONVERSATION_SOURCE_KIND,
                source_locator=retriever_tools.filesystem_dataset_locator(),
                conversation_key="stale-ip-audit-thread",
                conversation_type="email",
                display_name="INTELLECTUAL PROPERTY AUDIT REPORT",
            )
            connection.execute(
                """
                UPDATE documents
                SET content_type = 'Email',
                    canonical_kind = 'email',
                    conversation_id = ?,
                    conversation_assignment_mode = ?,
                    participants = 'January 15, 2023, MIT Licensing Obligations'
                WHERE id = ?
                """,
                (
                    conversation_id,
                    retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO,
                    int(row["id"]),
                ),
            )
            connection.execute(
                """
                UPDATE document_occurrences
                SET extracted_content_type = 'Email',
                    extracted_participants = 'January 15, 2023, MIT Licensing Obligations'
                WHERE document_id = ?
                """,
                (int(row["id"]),),
            )
            connection.commit()
        finally:
            connection.close()

        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_ingest["failed"], 0)
        self.assertEqual(second_ingest["updated"], 1)
        self.assertEqual(second_ingest["skipped"], 0)

        updated_row = self.fetch_document_row("ip-audit.txt")
        self.assertEqual(updated_row["content_type"], "E-Doc")
        self.assertEqual(updated_row["canonical_kind"], "document")
        self.assertIsNone(updated_row["conversation_id"])
        self.assertIsNone(updated_row["participants"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            occurrence_row = connection.execute(
                """
                SELECT extracted_content_type, extracted_participants
                FROM document_occurrences
                WHERE document_id = ?
                """,
                (int(row["id"]),),
            ).fetchone()
            self.assertIsNotNone(occurrence_row)
            assert occurrence_row is not None
            self.assertEqual(occurrence_row["extracted_content_type"], "E-Doc")
            self.assertIsNone(occurrence_row["extracted_participants"])
        finally:
            connection.close()

    def test_scoped_ingest_only_marks_missing_documents_inside_scope(self) -> None:
        raw_dir = self.root / "raw"
        other_dir = self.root / "other"
        raw_dir.mkdir()
        other_dir.mkdir()
        (raw_dir / "remove-me.txt").write_text("raw body\n", encoding="utf-8")
        (other_dir / "keep-me.txt").write_text("outside body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["new"], 2)
        self.assertEqual(first_ingest["failed"], 0)

        (raw_dir / "remove-me.txt").unlink()
        scoped_ingest = retriever_tools.ingest(
            self.root,
            recursive=True,
            raw_file_types=None,
            raw_paths=["raw"],
        )

        self.assertEqual(scoped_ingest["scan_paths"], ["raw"])
        self.assertEqual(scoped_ingest["missing"], 1)
        self.assertEqual(self.fetch_document_row("raw/remove-me.txt")["lifecycle_status"], "missing")
        self.assertEqual(self.fetch_document_row("other/keep-me.txt")["lifecycle_status"], "active")

    def test_scoped_ingest_does_not_rename_from_outside_scope_by_hash(self) -> None:
        raw_dir = self.root / "raw"
        other_dir = self.root / "other"
        raw_dir.mkdir()
        other_dir.mkdir()
        (other_dir / "original.txt").write_text("same body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["new"], 1)

        (raw_dir / "copy.txt").write_text("same body\n", encoding="utf-8")
        scoped_ingest = retriever_tools.ingest(
            self.root,
            recursive=True,
            raw_file_types=None,
            raw_paths=["raw"],
        )

        self.assertEqual(scoped_ingest["new"], 1)
        self.assertEqual(scoped_ingest["renamed"], 0)
        self.assertEqual(scoped_ingest["missing"], 0)
        self.assertEqual(self.fetch_document_row("other/original.txt")["lifecycle_status"], "active")
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            copied_occurrence = connection.execute(
                "SELECT * FROM document_occurrences WHERE rel_path = ?",
                ("raw/copy.txt",),
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(copied_occurrence)

    def test_scoped_ingest_rejects_paths_outside_workspace_or_state_dir(self) -> None:
        outside_path = self.root.parent / f"{self.root.name}-outside.txt"
        outside_path.write_text("outside\n", encoding="utf-8")
        self.addCleanup(lambda: outside_path.exists() and outside_path.unlink())

        with self.assertRaisesRegex(retriever_tools.RetrieverError, "inside the workspace root"):
            retriever_tools.ingest(self.root, recursive=True, raw_file_types=None, raw_paths=[str(outside_path)])
        with self.assertRaisesRegex(retriever_tools.RetrieverError, ".retriever"):
            retriever_tools.ingest(self.root, recursive=True, raw_file_types=None, raw_paths=[".retriever"])

    def test_ingest_cli_accepts_scoped_path(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        (raw_dir / "sample.txt").write_text("scoped body\n", encoding="utf-8")

        exit_code, payload, _, _ = self.run_cli(
            "ingest",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["new"], 1)
        self.assertEqual(payload["scan_paths"], ["raw"])

    def test_ingest_v2_start_status_and_cancel_foundation(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
            "--budget-seconds",
            "35",
        )

        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        self.assertTrue(start_payload["ok"])
        self.assertTrue(start_payload["created"])
        self.assertEqual(start_payload["status"], "planning")
        self.assertEqual(start_payload["phase"], "planning")
        self.assertEqual(start_payload["scope"], ["raw"])
        self.assertEqual(start_payload["budget_recommendation_seconds"], 35)
        self.assertIn("ingest-plan-step", " ".join(start_payload["next_recommended_commands"]))
        run_id = str(start_payload["run_id"])

        status_exit, status_payload, _, _ = self.run_cli(
            "ingest-status",
            str(self.root),
            "--run-id",
            run_id,
        )
        self.assertEqual(status_exit, 0)
        self.assertIsNotNone(status_payload)
        self.assertEqual(status_payload["run_id"], run_id)
        self.assertEqual(status_payload["counts"]["work_items"]["pending"], 0)
        self.assertEqual(status_payload["artifacts"]["orphan_pending_sweep"], 0)
        self.assertFalse(status_payload["entity"]["graph_stale"])

        cancel_exit, cancel_payload, _, _ = self.run_cli(
            "ingest-cancel",
            str(self.root),
            "--run-id",
            run_id,
        )
        self.assertEqual(cancel_exit, 0)
        self.assertIsNotNone(cancel_payload)
        self.assertTrue(cancel_payload["cancel_requested"])
        self.assertEqual(cancel_payload["status"], "canceled")
        self.assertEqual(cancel_payload["phase"], "canceled")

    def test_ingest_v2_run_step_reports_no_run(self) -> None:
        exit_code, payload, _, _ = self.run_cli("ingest-run-step", str(self.root))

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertTrue(payload["implemented"])
        self.assertFalse(payload["executed"])
        self.assertEqual(payload["executed_steps"], [])
        self.assertEqual(payload["reason"], "no_ingest_run")
        self.assertEqual(payload["run"]["status"], "none")

    def test_ingest_v2_plan_step_creates_loose_file_work_items(self) -> None:
        raw_dir = self.root / "raw"
        nested_dir = raw_dir / "nested"
        other_dir = self.root / "other"
        nested_dir.mkdir(parents=True)
        other_dir.mkdir()
        (raw_dir / "alpha.txt").write_text("alpha body\n", encoding="utf-8")
        (nested_dir / "beta.md").write_text("# beta\n", encoding="utf-8")
        (raw_dir / "mailbox.pst").write_bytes(b"not a real pst")
        (other_dir / "outside.txt").write_text("outside body\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        plan_exit, plan_payload, _, _ = self.run_cli(
            "ingest-plan-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )

        self.assertEqual(plan_exit, 0)
        self.assertIsNotNone(plan_payload)
        self.assertTrue(plan_payload["implemented"])
        self.assertFalse(plan_payload["more_planning_remaining"])
        self.assertEqual(plan_payload["planned_loose_files"], 2)
        self.assertEqual(plan_payload["cursor"]["skipped_container_files"], 0)
        self.assertEqual(plan_payload["cursor"]["scanned_pst_source_rel_paths"], ["raw/mailbox.pst"])
        self.assertEqual(len(plan_payload["cursor"]["pst_failures"]), 1)
        self.assertEqual(plan_payload["timings"]["work_item_insert_ms"]["count"], 3)
        self.assertGreaterEqual(plan_payload["timings"]["cursor_save_ms"]["count"], 1)
        self.assertIsInstance(plan_payload["timings"]["status_payload_ms"], float)
        run_payload = plan_payload["run"]
        self.assertEqual(run_payload["phase"], "preparing")
        self.assertEqual(run_payload["status"], "preparing")
        self.assertEqual(run_payload["counts"]["work_items"]["pending"], 2)
        self.assertEqual(run_payload["counts"]["by_unit_type"]["loose_file"]["pending"], 2)
        self.assertIn("ingest-prepare-step", " ".join(run_payload["next_recommended_commands"]))
        self.assertNotIn("ingest-plan-step", " ".join(run_payload["next_recommended_commands"]))

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            rows = connection.execute(
                """
                SELECT rel_path, unit_type, source_kind, source_key, commit_order, payload_json
                FROM ingest_work_items
                WHERE run_id = ?
                ORDER BY commit_order ASC
                """,
                (run_id,),
            ).fetchall()
        finally:
            connection.close()
        self.assertEqual([row["rel_path"] for row in rows], ["raw/alpha.txt", "raw/nested/beta.md"])
        self.assertTrue(all(row["unit_type"] == "loose_file" for row in rows))
        self.assertTrue(all(row["source_kind"] == retriever_tools.FILESYSTEM_SOURCE_KIND for row in rows))
        self.assertEqual([int(row["commit_order"]) for row in rows], [1, 2])
        payloads = [json.loads(row["payload_json"]) for row in rows]
        self.assertEqual(payloads[0]["file_type"], "txt")
        self.assertEqual(payloads[1]["file_type"], "md")
        self.assertNotIn("file_hash", payloads[0])

        second_plan_exit, second_plan_payload, _, _ = self.run_cli(
            "ingest-plan-step",
            str(self.root),
            "--run-id",
            run_id,
        )
        self.assertEqual(second_plan_exit, 0)
        self.assertIsNotNone(second_plan_payload)
        self.assertEqual(second_plan_payload["processed_paths"], 0)
        self.assertEqual(second_plan_payload["planned_loose_files"], 0)
        self.assertEqual(second_plan_payload["run"]["counts"]["work_items"]["pending"], 2)

    def test_ingest_v2_run_step_chains_recommended_steps_within_budget(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        (raw_dir / "alpha.txt").write_text("alpha runner body\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        self.assertIn("ingest-run-step", start_payload["next_recommended_commands"][0])
        run_id = str(start_payload["run_id"])

        with mock.patch.object(retriever_tools, "apply_schema", wraps=retriever_tools.apply_schema) as apply_schema_mock:
            exit_code, payload, _, _ = self.run_cli(
                "ingest-run-step",
                str(self.root),
                "--run-id",
                run_id,
                "--budget-seconds",
                "35",
            )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertTrue(payload["executed"])
        self.assertEqual(payload["selected_step"], "plan")
        self.assertEqual(payload["executed_steps"], ["plan", "prepare", "commit", "finalize"])
        self.assertEqual(payload["reason"], "run_terminal")
        self.assertEqual(payload["run"]["status"], "completed")
        self.assertEqual(apply_schema_mock.call_count, 1)
        self.assertGreaterEqual(payload["timings"]["schema_ms"], 0.0)
        self.assertEqual(payload["timings"]["inner_step_ms"]["count"], 4)
        completed_row = self.fetch_document_row("raw/alpha.txt")
        self.assertEqual(completed_row["control_number"], "DOC001.00000001")

        final_exit, final_payload, _, _ = self.run_cli("ingest-run-step", str(self.root), "--run-id", run_id)
        self.assertEqual(final_exit, 0)
        self.assertIsNotNone(final_payload)
        self.assertFalse(final_payload["executed"])
        self.assertEqual(final_payload["reason"], "run_terminal")
        self.assertEqual(final_payload["run"]["status"], "completed")

    def test_ingest_v2_mbox_creates_message_rows_with_source_context(self) -> None:
        mbox_path = self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="V2 MBOX Parent",
                    body_text="V2 parent message body",
                    message_id="<v2-mbox-msg-001@example.com>",
                    attachment_name="notes.txt",
                    attachment_text="v2 mbox attachment body",
                ),
                self.build_fake_mbox_message(
                    subject="V2 MBOX Sibling",
                    body_text="V2 sibling body text",
                    message_id="<v2-mbox-msg-002@example.com>",
                    date_created="Tue, 14 Apr 2026 10:05:00 +0000",
                ),
            ]
        )

        payloads = self.run_v2_loose_ingest()
        run_id = str(payloads["run_id"])
        commit_payload = dict(payloads["commit"])
        finalize_payload = dict(payloads["finalize"])

        self.assertEqual(commit_payload["failed"], 0)
        self.assertEqual(commit_payload["committed"], 3)
        self.assertEqual(commit_payload["actions"], {"new": 2, "finalized": 1})
        self.assertEqual(finalize_payload["run"]["status"], "completed")
        self.assertEqual(finalize_payload["run"]["counts"]["by_unit_type"]["mbox_message"]["committed"], 2)
        self.assertEqual(finalize_payload["run"]["counts"]["by_unit_type"]["mbox_source_finalizer"]["committed"], 1)

        parent_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<v2-mbox-msg-001@example.com>")
        sibling_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<v2-mbox-msg-002@example.com>")
        parent_row = self.fetch_document_row(parent_rel_path)
        sibling_row = self.fetch_document_row(sibling_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        dataset_row = self.fetch_dataset_row(int(parent_row["dataset_id"]))

        self.assertEqual(parent_row["source_kind"], retriever_tools.MBOX_SOURCE_KIND)
        self.assertEqual(parent_row["source_rel_path"], "mailbox.mbox")
        self.assertEqual(parent_row["source_item_id"], "<v2-mbox-msg-001@example.com>")
        self.assertEqual(parent_row["file_type"], "mbox")
        self.assertEqual(parent_row["content_type"], "Email")
        self.assertEqual(parent_row["custodian"], "mailbox")
        self.assertEqual(sibling_row["dataset_id"], parent_row["dataset_id"])
        self.assertEqual(child_row["dataset_id"], parent_row["dataset_id"])
        self.assertEqual(dataset_row["source_kind"], retriever_tools.MBOX_SOURCE_KIND)
        self.assertEqual(dataset_row["dataset_locator"], "mailbox.mbox")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            container_row = connection.execute(
                "SELECT * FROM container_sources WHERE source_kind = ? AND source_rel_path = ?",
                (retriever_tools.MBOX_SOURCE_KIND, "mailbox.mbox"),
            ).fetchone()
            work_items = connection.execute(
                """
                SELECT unit_type, status, affected_document_ids_json, artifact_manifest_json
                FROM ingest_work_items
                WHERE run_id = ?
                ORDER BY commit_order ASC
                """,
                (run_id,),
            ).fetchall()
            cursor_row = connection.execute(
                """
                SELECT cursor_json
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'committing'
                  AND cursor_key = 'loose_file_commit'
                """,
                (run_id,),
            ).fetchone()
        finally:
            connection.close()

        self.assertIsNotNone(container_row)
        self.assertEqual(container_row["dataset_id"], parent_row["dataset_id"])
        self.assertEqual(container_row["message_count"], 2)
        self.assertEqual(container_row["file_size"], mbox_path.stat().st_size)
        self.assertIsNotNone(container_row["last_scan_completed_at"])
        self.assertEqual(
            [row["unit_type"] for row in work_items],
            ["mbox_message", "mbox_message", "mbox_source_finalizer", "conversation_preview", "conversation_preview"],
        )
        self.assertTrue(all(row["status"] == "committed" for row in work_items))
        self.assertTrue(json.loads(work_items[0]["affected_document_ids_json"]))
        self.assertEqual(json.loads(work_items[2]["artifact_manifest_json"])["commit_action"], "finalized")
        self.assertIsNotNone(cursor_row)
        mbox_stats = json.loads(cursor_row["cursor_json"])["mbox_stats"]
        self.assertEqual(mbox_stats["mbox_messages_created"], 2)
        self.assertEqual(mbox_stats["mbox_sources_finalized"], 1)

    def test_ingest_v2_mbox_reingest_retires_removed_messages(self) -> None:
        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="V2 Original MBOX Parent",
                    body_text="Parent v1",
                    message_id="<v2-mbox-retire-001@example.com>",
                    attachment_name="notes.txt",
                    attachment_text="stable attachment body",
                ),
                self.build_fake_mbox_message(
                    subject="V2 Removed later",
                    body_text="Remove me",
                    message_id="<v2-mbox-retire-002@example.com>",
                    date_created="Tue, 14 Apr 2026 10:05:00 +0000",
                ),
            ]
        )
        first_payloads = self.run_v2_loose_ingest()
        self.assertEqual(dict(first_payloads["finalize"])["run"]["status"], "completed")

        parent_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<v2-mbox-retire-001@example.com>")
        removed_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<v2-mbox-retire-002@example.com>")
        parent_row = self.fetch_document_row(parent_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        retriever_tools.set_field(self.root, parent_row["id"], "title", "Manual V2 MBOX Title")

        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="V2 Updated MBOX Parent",
                    body_text="Parent v2",
                    message_id="<v2-mbox-retire-001@example.com>",
                    attachment_name="notes.txt",
                    attachment_text="stable attachment body",
                )
            ]
        )
        second_payloads = self.run_v2_loose_ingest()
        second_commit = dict(second_payloads["commit"])
        self.assertEqual(second_commit["failed"], 0)
        self.assertEqual(second_commit["actions"], {"updated": 1, "finalized": 1})

        updated_parent = self.fetch_document_row(parent_rel_path)
        updated_child = self.fetch_child_rows(updated_parent["id"])[0]
        retired_row = self.fetch_document_row(removed_rel_path)
        self.assertEqual(updated_parent["control_number"], parent_row["control_number"])
        self.assertEqual(updated_parent["title"], "Manual V2 MBOX Title")
        self.assertEqual(updated_child["id"], child_row["id"])
        self.assertEqual(updated_child["control_number"], child_row["control_number"])
        self.assertEqual(retired_row["lifecycle_status"], "deleted")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            cursor_row = connection.execute(
                """
                SELECT cursor_json
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'committing'
                  AND cursor_key = 'loose_file_commit'
                """,
                (str(second_payloads["run_id"]),),
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(cursor_row)
        mbox_stats = json.loads(cursor_row["cursor_json"])["mbox_stats"]
        self.assertEqual(mbox_stats["mbox_messages_updated"], 1)
        self.assertEqual(mbox_stats["mbox_messages_deleted"], 1)

    def test_ingest_v2_pst_creates_message_rows_with_source_context(self) -> None:
        pst_path = self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="v2-pst-msg-001",
                subject="V2 PST Parent",
                body_text="V2 PST parent message body",
                folder_path="Inbox",
                recipients=None,
                transport_headers="\n".join(
                    [
                        "To: Bob Example <bob@example.com>",
                        "Cc: Carol Example <carol@example.com>",
                        "",
                    ]
                ),
                attachment_name="notes.txt",
                attachment_text="v2 pst attachment body",
            ),
            self.build_fake_pst_message(
                source_item_id="v2-pst-msg-002",
                subject="V2 PST Sibling",
                body_text="V2 PST sibling body text",
                folder_path="Sent Items",
            ),
        ]

        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            payloads = self.run_v2_loose_ingest()
        run_id = str(payloads["run_id"])
        commit_payload = dict(payloads["commit"])
        finalize_payload = dict(payloads["finalize"])

        self.assertEqual(commit_payload["failed"], 0)
        self.assertEqual(commit_payload["committed"], 3)
        self.assertEqual(commit_payload["actions"], {"new": 2, "finalized": 1})
        self.assertEqual(finalize_payload["run"]["status"], "completed")
        self.assertEqual(finalize_payload["run"]["counts"]["by_unit_type"]["pst_message"]["committed"], 2)
        self.assertEqual(finalize_payload["run"]["counts"]["by_unit_type"]["pst_source_finalizer"]["committed"], 1)

        parent_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "v2-pst-msg-001")
        sibling_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "v2-pst-msg-002")
        parent_row = self.fetch_document_row(parent_rel_path)
        sibling_row = self.fetch_document_row(sibling_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        dataset_row = self.fetch_dataset_row(int(parent_row["dataset_id"]))

        self.assertEqual(parent_row["source_kind"], retriever_tools.PST_SOURCE_KIND)
        self.assertEqual(parent_row["source_rel_path"], "mailbox.pst")
        self.assertEqual(parent_row["source_item_id"], "v2-pst-msg-001")
        self.assertEqual(parent_row["source_folder_path"], "Inbox")
        self.assertEqual(parent_row["file_type"], "pst")
        self.assertEqual(parent_row["content_type"], "Email")
        self.assertEqual(parent_row["custodian"], "mailbox")
        self.assertEqual(
            parent_row["recipients"],
            "Bob Example <bob@example.com>, Carol Example <carol@example.com>",
        )
        self.assertEqual(sibling_row["source_folder_path"], "Sent Items")
        self.assertEqual(sibling_row["dataset_id"], parent_row["dataset_id"])
        self.assertEqual(child_row["dataset_id"], parent_row["dataset_id"])
        self.assertEqual(dataset_row["source_kind"], retriever_tools.PST_SOURCE_KIND)
        self.assertEqual(dataset_row["dataset_locator"], "mailbox.pst")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            container_row = connection.execute(
                "SELECT * FROM container_sources WHERE source_kind = ? AND source_rel_path = ?",
                (retriever_tools.PST_SOURCE_KIND, "mailbox.pst"),
            ).fetchone()
            work_items = connection.execute(
                """
                SELECT unit_type, status, affected_document_ids_json, artifact_manifest_json
                FROM ingest_work_items
                WHERE run_id = ?
                ORDER BY commit_order ASC
                """,
                (run_id,),
            ).fetchall()
            cursor_row = connection.execute(
                """
                SELECT cursor_json
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'committing'
                  AND cursor_key = 'loose_file_commit'
                """,
                (run_id,),
            ).fetchone()
        finally:
            connection.close()

        self.assertIsNotNone(container_row)
        self.assertEqual(container_row["dataset_id"], parent_row["dataset_id"])
        self.assertEqual(container_row["message_count"], 2)
        self.assertEqual(container_row["file_size"], pst_path.stat().st_size)
        self.assertIsNotNone(container_row["last_scan_completed_at"])
        self.assertEqual(
            [row["unit_type"] for row in work_items],
            ["pst_message", "pst_message", "pst_source_finalizer", "conversation_preview", "conversation_preview"],
        )
        self.assertTrue(all(row["status"] == "committed" for row in work_items))
        self.assertTrue(json.loads(work_items[0]["affected_document_ids_json"]))
        self.assertEqual(json.loads(work_items[2]["artifact_manifest_json"])["commit_action"], "finalized")
        self.assertIsNotNone(cursor_row)
        pst_stats = json.loads(cursor_row["cursor_json"])["pst_stats"]
        self.assertEqual(pst_stats["pst_messages_created"], 2)
        self.assertEqual(pst_stats["pst_sources_finalized"], 1)

    def test_ingest_v2_pst_reingest_retires_removed_messages(self) -> None:
        pst_path = self.write_fake_pst_file(content=b"pst-v2-a")
        first_messages = [
            self.build_fake_pst_message(
                source_item_id="v2-pst-retire-001",
                subject="V2 Original PST Parent",
                body_text="Parent v1",
                attachment_name="notes.txt",
                attachment_text="stable attachment body",
            ),
            self.build_fake_pst_message(
                source_item_id="v2-pst-retire-002",
                subject="V2 Removed later",
                body_text="Remove me",
            ),
        ]
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(first_messages)):
            first_payloads = self.run_v2_loose_ingest()
        self.assertEqual(dict(first_payloads["finalize"])["run"]["status"], "completed")

        parent_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "v2-pst-retire-001")
        removed_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "v2-pst-retire-002")
        parent_row = self.fetch_document_row(parent_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        retriever_tools.set_field(self.root, parent_row["id"], "title", "Manual V2 PST Title")

        pst_path.write_bytes(b"pst-v2-b")
        second_messages = [
            self.build_fake_pst_message(
                source_item_id="v2-pst-retire-001",
                subject="V2 Updated PST Parent",
                body_text="Parent v2",
                attachment_name="notes.txt",
                attachment_text="stable attachment body",
            )
        ]
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(second_messages)):
            second_payloads = self.run_v2_loose_ingest()
        second_commit = dict(second_payloads["commit"])
        self.assertEqual(second_commit["failed"], 0)
        self.assertEqual(second_commit["actions"], {"updated": 1, "finalized": 1})

        updated_parent = self.fetch_document_row(parent_rel_path)
        updated_child = self.fetch_child_rows(updated_parent["id"])[0]
        retired_row = self.fetch_document_row(removed_rel_path)
        self.assertEqual(updated_parent["control_number"], parent_row["control_number"])
        self.assertEqual(updated_parent["title"], "Manual V2 PST Title")
        self.assertEqual(updated_child["id"], child_row["id"])
        self.assertEqual(updated_child["control_number"], child_row["control_number"])
        self.assertEqual(retired_row["lifecycle_status"], "deleted")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            container_row = connection.execute(
                "SELECT * FROM container_sources WHERE source_kind = ? AND source_rel_path = ?",
                (retriever_tools.PST_SOURCE_KIND, "mailbox.pst"),
            ).fetchone()
            cursor_row = connection.execute(
                """
                SELECT cursor_json
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'committing'
                  AND cursor_key = 'loose_file_commit'
                """,
                (str(second_payloads["run_id"]),),
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(container_row)
        self.assertEqual(container_row["message_count"], 1)
        self.assertIsNotNone(cursor_row)
        pst_stats = json.loads(cursor_row["cursor_json"])["pst_stats"]
        self.assertEqual(pst_stats["pst_messages_updated"], 1)
        self.assertEqual(pst_stats["pst_messages_deleted"], 1)

    def test_ingest_v2_slack_export_creates_day_and_thread_documents(self) -> None:
        export_root = self.root / "data" / "slack"
        export_root.mkdir(parents=True)
        (export_root / "users.json").write_text(
            json.dumps(
                [
                    {
                        "id": "U04SERGEY1",
                        "name": "sergey",
                        "profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                        },
                    },
                    {
                        "id": "U04MAX0001",
                        "name": "maksim",
                        "profile": {
                            "real_name": "Maksim Faleev",
                            "display_name": "Maksim",
                        },
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (export_root / "channels.json").write_text(
            json.dumps(
                [
                    {
                        "id": "C04GENERAL1",
                        "name": "general",
                        "is_channel": True,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        channel_dir = export_root / "general"
        channel_dir.mkdir()
        thread_ts = "1671235434.237949"
        (channel_dir / "2022-12-16.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Kickoff thread",
                        "user": "U04SERGEY1",
                        "ts": thread_ts,
                        "thread_ts": thread_ts,
                        "reply_count": 1,
                    },
                    {
                        "type": "message",
                        "text": "Standalone channel update",
                        "user": "U04MAX0001",
                        "ts": "1671235834.237949",
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (channel_dir / "2022-12-17.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Following up on kickoff",
                        "user": "U04MAX0001",
                        "ts": "1671321834.237949",
                        "thread_ts": thread_ts,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        payloads = self.run_v2_loose_ingest("data/slack")
        run_id = str(payloads["run_id"])
        plan_payload = dict(payloads["plan"])
        commit_payload = dict(payloads["commit"])
        finalize_payload = dict(payloads["finalize"])

        self.assertEqual(plan_payload["cursor"]["planned_slack_export_roots"], ["data/slack"])
        self.assertEqual(plan_payload["cursor"]["planned_slack_conversations"], 1)
        self.assertEqual(plan_payload["cursor"]["planned_slack_day_documents"], 2)
        self.assertEqual(commit_payload["failed"], 0)
        self.assertEqual(commit_payload["committed"], 1)
        self.assertEqual(commit_payload["actions"], {"committed": 1})
        self.assertEqual(finalize_payload["run"]["status"], "completed")
        self.assertEqual(finalize_payload["run"]["counts"]["by_unit_type"]["slack_conversation"]["committed"], 1)

        day_one_row = self.fetch_document_row("data/slack/general/2022-12-16.json")
        day_two_row = self.fetch_document_row("data/slack/general/2022-12-17.json")
        child_rel_path = retriever_tools.slack_reply_thread_rel_path("C04GENERAL1", thread_ts)
        child_row = self.fetch_document_row(child_rel_path)
        dataset_row = self.fetch_dataset_row(int(day_one_row["dataset_id"]))

        self.assertEqual(day_one_row["source_kind"], retriever_tools.SLACK_EXPORT_SOURCE_KIND)
        self.assertEqual(day_two_row["source_kind"], retriever_tools.SLACK_EXPORT_SOURCE_KIND)
        self.assertEqual(child_row["source_kind"], retriever_tools.SLACK_EXPORT_SOURCE_KIND)
        self.assertEqual(child_row["parent_document_id"], day_one_row["id"])
        self.assertEqual(child_row["child_document_kind"], retriever_tools.CHILD_DOCUMENT_KIND_REPLY_THREAD)
        self.assertEqual(child_row["source_rel_path"], "data/slack/general/2022-12-16.json")
        self.assertEqual(child_row["source_item_id"], thread_ts)
        self.assertEqual(child_row["root_message_key"], f"C04GENERAL1:{thread_ts}")
        self.assertEqual(day_two_row["conversation_id"], day_one_row["conversation_id"])
        self.assertEqual(child_row["conversation_id"], day_one_row["conversation_id"])
        self.assertEqual(dataset_row["source_kind"], retriever_tools.SLACK_EXPORT_SOURCE_KIND)
        self.assertEqual(dataset_row["dataset_locator"], "data/slack")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            work_item = connection.execute(
                """
                SELECT affected_document_ids_json, affected_conversation_keys_json, artifact_manifest_json
                FROM ingest_work_items
                WHERE run_id = ?
                  AND unit_type = 'slack_conversation'
                """,
                (run_id,),
            ).fetchone()
            source_parts = connection.execute(
                """
                SELECT part_kind, rel_source_path
                FROM document_source_parts
                WHERE document_id = ?
                ORDER BY ordinal ASC, id ASC
                """,
                (child_row["id"],),
            ).fetchall()
        finally:
            connection.close()

        self.assertIsNotNone(work_item)
        self.assertEqual(len(json.loads(work_item["affected_document_ids_json"])), 3)
        self.assertEqual(json.loads(work_item["affected_conversation_keys_json"]), ["data/slack:C04GENERAL1"])
        self.assertEqual(json.loads(work_item["artifact_manifest_json"])["new"], 3)
        self.assertEqual(
            [(row["part_kind"], row["rel_source_path"]) for row in source_parts],
            [
                ("slack_thread_root_day", "data/slack/general/2022-12-16.json"),
                ("slack_thread_reply_day", "data/slack/general/2022-12-17.json"),
            ],
        )

    def test_ingest_v2_gmail_export_preserves_sidecar_enrichment(self) -> None:
        export_root = self.root / "gmail-filtered"
        export_root.mkdir()

        mbox_path = export_root / "Filtered.mbox"
        archive = mailbox.mbox(str(mbox_path), create=True)
        try:
            archive.add(
                self.build_fake_mbox_message(
                    subject="Drive sharing update",
                    body_text="Email body that references a linked Drive document.",
                    message_id="<gmail-v2-filter-001@example.com>",
                    author="Sender Example <sender@example.com>",
                    recipients="Receiver Example <receiver@example.com>",
                )
            )
            archive.flush()
        finally:
            archive.close()

        with (export_root / "Filtered-metadata.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "Rfc822MessageId",
                    "GmailMessageId",
                    "Account",
                    "Labels",
                    "Subject",
                    "From",
                    "To",
                    "DateSent",
                    "DateReceived",
                    "ThreadedMessageCount",
                ],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "Rfc822MessageId": "gmail-v2-filter-001@example.com",
                    "GmailMessageId": "1767000000000000042",
                    "Account": "owner@example.com",
                    "Labels": "^INBOX,projectalpha",
                    "Subject": "Drive sharing update",
                    "From": "sender@example.com Sender Example",
                    "To": "receiver@example.com Receiver Example",
                    "DateSent": "2026-04-14T10:00:00Z",
                    "DateReceived": "2026-04-14T10:00:05Z",
                    "ThreadedMessageCount": "1",
                }
            )

        with (export_root / "Filtered-drive-links.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=["Account", "Rfc822MessageId", "GmailMessageId", "DriveUrl", "DriveItemId"],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "Account": "owner@example.com",
                    "Rfc822MessageId": "gmail-v2-filter-001@example.com",
                    "GmailMessageId": "1767000000000000042",
                    "DriveUrl": "https://docs.google.com/document/d/drive-doc-001/edit",
                    "DriveItemId": "drive-doc-001",
                }
            )

        drive_export_dir = export_root / "Filtered_Drive_Link_Export_0"
        drive_export_dir.mkdir()
        drive_file = drive_export_dir / "Linked notes_drive-doc-001.txt"
        drive_file.write_text("Drive export body text.\n", encoding="utf-8")
        (export_root / "Filtered_Drive_Link_Export-metadata.xml").write_text(
            """<?xml version='1.0' encoding='UTF-8' standalone='yes'?>
<Root DataInterchangeType='Update' Description='Test'>
  <Batch name='New Batch'>
    <Documents>
      <Document DocID='drive-doc-001'>
        <Tags>
          <Tag TagName='#Author' TagDataType='Text' TagValue='owner@example.com'/>
          <Tag TagName='#Title' TagDataType='Text' TagValue='Linked notes from metadata'/>
        </Tags>
        <Files>
          <File FileType='Native'>
            <ExternalFile FileName='Linked notes_drive-doc-001.txt' FileSize='24' Hash='abc123'/>
          </File>
        </Files>
      </Document>
    </Documents>
  </Batch>
</Root>
""",
            encoding="utf-8",
        )

        payloads = self.run_v2_loose_ingest("gmail-filtered")
        run_id = str(payloads["run_id"])
        plan_payload = dict(payloads["plan"])
        finalize_payload = dict(payloads["finalize"])

        self.assertEqual(finalize_payload["run"]["status"], "completed")
        self.assertEqual(plan_payload["cursor"]["special_source_counts"]["gmail_export_roots"], 1)
        self.assertEqual(plan_payload["cursor"]["planned_mbox_sources"], ["gmail-filtered/Filtered.mbox"])
        self.assertEqual(plan_payload["cursor"]["planned_mbox_messages"], 1)
        self.assertEqual(finalize_payload["run"]["counts"]["by_unit_type"]["mbox_message"]["committed"], 1)
        self.assertEqual(finalize_payload["run"]["counts"]["by_unit_type"]["mbox_source_finalizer"]["committed"], 1)

        email_rel_path = retriever_tools.mbox_message_rel_path(
            "gmail-filtered/Filtered.mbox",
            "<gmail-v2-filter-001@example.com>",
        )
        email_row = self.fetch_document_row(email_rel_path)
        child_rows = self.fetch_child_rows(int(email_row["id"]))
        self.assertEqual(len(child_rows), 1)
        self.assertEqual(child_rows[0]["title"], "Linked notes from metadata")
        self.assertEqual(child_rows[0]["source_kind"], retriever_tools.EMAIL_ATTACHMENT_SOURCE_KIND)

        label_search = retriever_tools.search(self.root, "projectalpha", None, None, None, 1, 20)
        self.assertEqual(label_search["total_hits"], 1)
        self.assertEqual(label_search["results"][0]["id"], email_row["id"])
        self.assertEqual(label_search["results"][0]["attachment_count"], 1)

        linked_title_search = retriever_tools.search(self.root, "Linked notes from metadata", None, None, None, 1, 20)
        self.assertEqual(linked_title_search["total_hits"], 2)
        returned_ids = {item["id"] for item in linked_title_search["results"]}
        self.assertIn(email_row["id"], returned_ids)
        self.assertIn(child_rows[0]["id"], returned_ids)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            parent_rel_paths = {
                str(row["rel_path"])
                for row in connection.execute(
                    """
                    SELECT rel_path
                    FROM documents
                    WHERE parent_document_id IS NULL
                    ORDER BY rel_path ASC
                    """
                ).fetchall()
            }
            work_items = connection.execute(
                """
                SELECT artifact_manifest_json
                FROM ingest_work_items
                WHERE run_id = ?
                ORDER BY commit_order ASC
                """,
                (run_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertNotIn(
            "gmail-filtered/Filtered_Drive_Link_Export_0/Linked notes_drive-doc-001.txt",
            parent_rel_paths,
        )
        self.assertNotIn("gmail-filtered/Filtered-metadata.csv", parent_rel_paths)
        self.assertNotIn("gmail-filtered/Filtered-drive-links.csv", parent_rel_paths)
        self.assertNotIn("gmail-filtered/Filtered_Drive_Link_Export-metadata.xml", parent_rel_paths)
        manifests = [json.loads(row["artifact_manifest_json"]) for row in work_items]
        self.assertTrue(any(manifest.get("source_plan_kind") == "gmail" for manifest in manifests))

    def test_ingest_v2_gmail_mbox_planning_saves_partial_source_cursor(self) -> None:
        export_root = self.root / "gmail-bulk"
        export_root.mkdir()

        message_count = retriever_tools.INGEST_V2_MBOX_PLAN_BATCH_SIZE + 5
        mbox_path = export_root / "Bulk.mbox"
        archive = mailbox.mbox(str(mbox_path), create=True)
        try:
            for index in range(message_count):
                archive.add(
                    self.build_fake_mbox_message(
                        subject=f"Bulk Gmail message {index:03d}",
                        body_text=f"Bulk Gmail body {index:03d}",
                        message_id=f"<gmail-v2-bulk-{index:03d}@example.com>",
                        author="Sender Example <sender@example.com>",
                        recipients="Receiver Example <receiver@example.com>",
                    )
                )
            archive.flush()
        finally:
            archive.close()

        with (export_root / "Bulk-metadata.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "Rfc822MessageId",
                    "GmailMessageId",
                    "Account",
                    "Labels",
                    "Subject",
                    "From",
                    "To",
                    "DateSent",
                    "DateReceived",
                    "ThreadedMessageCount",
                ],
            )
            writer.writeheader()
            for index in range(message_count):
                writer.writerow(
                    {
                        "Rfc822MessageId": f"gmail-v2-bulk-{index:03d}@example.com",
                        "GmailMessageId": f"1767000000001{index:04d}",
                        "Account": "owner@example.com",
                        "Labels": "^INBOX,bulklabel",
                        "Subject": f"Bulk Gmail message {index:03d}",
                        "From": "sender@example.com Sender Example",
                        "To": "receiver@example.com Receiver Example",
                        "DateSent": "2026-04-14T10:00:00Z",
                        "DateReceived": "2026-04-14T10:00:05Z",
                        "ThreadedMessageCount": "1",
                    }
                )

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "gmail-bulk",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        assert start_payload is not None
        run_id = str(start_payload["run_id"])

        first_plan_exit, first_plan_payload, _, _ = self.run_cli(
            "ingest-plan-step",
            str(self.root),
            "--run-id",
            run_id,
        )
        self.assertEqual(first_plan_exit, 0)
        self.assertIsNotNone(first_plan_payload)
        assert first_plan_payload is not None
        self.assertEqual(
            first_plan_payload["cursor"]["planned_mbox_messages"],
            retriever_tools.INGEST_V2_MBOX_PLAN_BATCH_SIZE,
        )
        self.assertEqual(first_plan_payload["cursor"]["current_mbox_source"], "gmail-bulk/Bulk.mbox")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            cursor_row = connection.execute(
                """
                SELECT cursor_json
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'planning'
                  AND cursor_key = 'loose_file_scan'
                """,
                (run_id,),
            ).fetchone()
            first_batch_count = connection.execute(
                """
                SELECT COUNT(*)
                FROM ingest_work_items
                WHERE run_id = ?
                  AND unit_type = 'mbox_message'
                """,
                (run_id,),
            ).fetchone()[0]
        finally:
            connection.close()

        self.assertIsNotNone(cursor_row)
        cursor_json = json.loads(cursor_row["cursor_json"])
        current_source = cursor_json["current_mbox_source"]
        self.assertEqual(current_source["next_message_index"], retriever_tools.INGEST_V2_MBOX_PLAN_BATCH_SIZE)
        self.assertEqual(current_source["planned_message_count"], retriever_tools.INGEST_V2_MBOX_PLAN_BATCH_SIZE)
        self.assertEqual(first_batch_count, retriever_tools.INGEST_V2_MBOX_PLAN_BATCH_SIZE)

        second_plan_exit, second_plan_payload, _, _ = self.run_cli(
            "ingest-plan-step",
            str(self.root),
            "--run-id",
            run_id,
        )
        self.assertEqual(second_plan_exit, 0)
        self.assertIsNotNone(second_plan_payload)
        assert second_plan_payload is not None
        self.assertEqual(second_plan_payload["run"]["phase"], "preparing")
        self.assertEqual(second_plan_payload["cursor"]["planned_mbox_messages"], message_count)
        self.assertEqual(second_plan_payload["cursor"]["planned_mbox_sources"], ["gmail-bulk/Bulk.mbox"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            unit_counts = {
                str(row["unit_type"]): int(row["count"])
                for row in connection.execute(
                    """
                    SELECT unit_type, COUNT(*) AS count
                    FROM ingest_work_items
                    WHERE run_id = ?
                    GROUP BY unit_type
                    """,
                    (run_id,),
                ).fetchall()
            }
            finalizer_row = connection.execute(
                """
                SELECT payload_json
                FROM ingest_work_items
                WHERE run_id = ?
                  AND unit_type = 'mbox_source_finalizer'
                """,
                (run_id,),
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(unit_counts, {"mbox_message": message_count, "mbox_source_finalizer": 1})
        self.assertIsNotNone(finalizer_row)
        self.assertEqual(json.loads(finalizer_row["payload_json"])["message_count"], message_count)

    def test_ingest_v2_auto_routes_production_root(self) -> None:
        self.write_production_fixture()

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        run_exit, run_payload, _, _ = self.run_cli(
            "ingest-run-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(run_exit, 0)
        self.assertIsNotNone(run_payload)
        self.assertTrue(run_payload["executed"])
        self.assertEqual(run_payload["run"]["status"], "completed")
        self.assertEqual(run_payload["run"]["counts"]["by_unit_type"]["production_row"]["committed"], 4)

        parent_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000001.logical"
        )
        child_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000003.logical"
        )
        native_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000004.logical"
        )
        self.assertEqual(parent_row["source_kind"], retriever_tools.PRODUCTION_SOURCE_KIND)
        self.assertEqual(parent_row["content_type"], "Email")
        self.assertEqual(child_row["parent_document_id"], parent_row["id"])
        self.assertEqual(native_row["file_name"], "PDX000004.pdf")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            loose_artifact_rows = connection.execute(
                """
                SELECT rel_path
                FROM documents
                WHERE rel_path LIKE 'Synthetic_Production/TEXT/%'
                   OR rel_path LIKE 'Synthetic_Production/IMAGES/%'
                   OR rel_path LIKE 'Synthetic_Production/NATIVES/%'
                """
            ).fetchall()
            cursor_row = connection.execute(
                """
                SELECT cursor_json
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'finalizing'
                  AND cursor_key = 'loose_file_finalize'
                """,
                (run_id,),
            ).fetchone()
        finally:
            connection.close()
        self.assertEqual(loose_artifact_rows, [])
        self.assertIsNotNone(cursor_row)
        cursor_payload = json.loads(cursor_row["cursor_json"])
        self.assertEqual(cursor_payload["production_finalized_roots"], ["Synthetic_Production"])
        self.assertEqual(cursor_payload["production_stats"]["families_reconstructed"], 1)

    def test_ingest_v2_generates_all_production_preview_pages_in_batches(self) -> None:
        production_name = "Batched_Production"
        production_root = self.root / production_name
        data_dir = production_root / "DATA"
        text_dir = production_root / "TEXT" / "TEXT001"
        image_dir = production_root / "IMAGES" / "IMG001"
        for directory in (data_dir, text_dir, image_dir):
            directory.mkdir(parents=True, exist_ok=True)

        def bates(number: int) -> str:
            return f"BP{number:06d}"

        def loadfile_path(*parts: str) -> str:
            return ".\\" + "\\".join([production_name, *parts])

        page_count = retriever_tools.INGEST_V2_PRODUCTION_PREVIEW_BATCH_SIZE + 3
        (text_dir / f"{bates(1)}.txt").write_text("Large image-backed production row\n", encoding="utf-8")
        for index in range(1, page_count + 1):
            self.write_tiff_fixture(
                image_dir / f"{bates(index)}.tif",
                ((index * 11) % 255, (index * 17) % 255, (index * 23) % 255),
            )

        delimiter = b"\x14"
        quote = b"\xfe"

        def dat_line(fields: list[str]) -> bytes:
            return delimiter.join(quote + field.encode("latin-1") + quote for field in fields) + b"\r\n"

        headers = ["Begin Bates", "End Bates", "Begin Attachment", "End Attachment", "Text Precedence", "FILE_PATH"]
        row = [
            bates(1),
            bates(page_count),
            "",
            "",
            loadfile_path("TEXT", "TEXT001", f"{bates(1)}.txt"),
            "",
        ]
        (data_dir / f"{production_name}.dat").write_bytes(dat_line(headers) + dat_line(row))
        opt_lines = [
            f"{bates(1)},{production_name},{loadfile_path('IMAGES', 'IMG001', f'{bates(1)}.tif')},Y,,,{page_count}",
            *[
                f"{bates(index)},{production_name},{loadfile_path('IMAGES', 'IMG001', f'{bates(index)}.tif')},,,,"
                for index in range(2, page_count + 1)
            ],
        ]
        (data_dir / f"{production_name}.opt").write_text("\n".join(opt_lines) + "\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli("ingest-start", str(self.root), "--recursive")
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])
        run_exit, run_payload, _, _ = self.run_cli(
            "ingest-run-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(run_exit, 0)
        self.assertIsNotNone(run_payload)
        self.assertEqual(run_payload["run"]["status"], "completed")
        by_unit_type = run_payload["run"]["counts"]["by_unit_type"]
        self.assertEqual(by_unit_type["production_row"]["committed"], 1)
        self.assertEqual(by_unit_type["production_preview_batch"]["committed"], 2)

        document_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/{production_name}/documents/{bates(1)}.logical"
        )
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            preview_rows = connection.execute(
                """
                SELECT rel_preview_path, preview_type, ordinal
                FROM document_previews
                WHERE document_id = ?
                ORDER BY ordinal ASC, id ASC
                """,
                (document_row["id"],),
            ).fetchall()
        finally:
            connection.close()

        html_rows = [row for row in preview_rows if row["preview_type"] == "html"]
        image_rows = [row for row in preview_rows if row["preview_type"] == "image"]
        self.assertEqual(len(html_rows), 1)
        self.assertEqual(len(image_rows), page_count)
        preview_html = (self.paths["state_dir"] / html_rows[0]["rel_preview_path"]).read_text(encoding="utf-8")
        self.assertEqual(preview_html.count("<figure>"), page_count)
        self.assertNotIn("Preview shows the first", preview_html)
        self.assertNotIn("data:image/png;base64,", preview_html)
        for row in image_rows:
            self.assertTrue((self.paths["state_dir"] / row["rel_preview_path"]).exists())

    def test_ingest_v2_production_rerun_retires_missing_loadfile_rows(self) -> None:
        production_root = self.write_production_fixture()

        first_exit, first_payload, _, _ = self.run_cli("ingest-start", str(self.root), "--recursive")
        self.assertEqual(first_exit, 0)
        self.assertIsNotNone(first_payload)
        first_run_id = str(first_payload["run_id"])
        first_run_exit, first_run_payload, _, _ = self.run_cli(
            "ingest-run-step",
            str(self.root),
            "--run-id",
            first_run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(first_run_exit, 0)
        self.assertIsNotNone(first_run_payload)
        self.assertEqual(first_run_payload["run"]["status"], "completed")

        data_path = production_root / "DATA" / "Synthetic_Production.dat"
        lines = data_path.read_bytes().splitlines()
        filtered = [line for line in lines if b"PDX000005" not in line]
        data_path.write_bytes(b"\r\n".join(filtered) + b"\r\n")

        second_exit, second_payload, _, _ = self.run_cli("ingest-start", str(self.root), "--recursive")
        self.assertEqual(second_exit, 0)
        self.assertIsNotNone(second_payload)
        second_run_id = str(second_payload["run_id"])
        second_run_exit, second_run_payload, _, _ = self.run_cli(
            "ingest-run-step",
            str(self.root),
            "--run-id",
            second_run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(second_run_exit, 0)
        self.assertIsNotNone(second_run_payload)
        self.assertEqual(second_run_payload["run"]["status"], "completed")
        self.assertEqual(second_run_payload["run"]["counts"]["by_unit_type"]["production_row"]["committed"], 3)

        retired_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )
        self.assertEqual(retired_row["lifecycle_status"], "deleted")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            cursor_row = connection.execute(
                """
                SELECT cursor_json
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'finalizing'
                  AND cursor_key = 'loose_file_finalize'
                """,
                (second_run_id,),
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(cursor_row)
        self.assertEqual(json.loads(cursor_row["cursor_json"])["production_stats"]["retired"], 1)

    def test_ingest_v2_prepare_step_prepares_loose_files_without_document_writes(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        alpha_text = "alpha body\n"
        beta_text = "# beta\n\nbody\n"
        (raw_dir / "alpha.txt").write_text(alpha_text, encoding="utf-8")
        (raw_dir / "beta.md").write_text(beta_text, encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        plan_exit, plan_payload, _, _ = self.run_cli(
            "ingest-plan-step",
            str(self.root),
            "--run-id",
            run_id,
        )
        self.assertEqual(plan_exit, 0)
        self.assertIsNotNone(plan_payload)
        self.assertEqual(plan_payload["run"]["phase"], "preparing")

        prepare_exit, prepare_payload, _, _ = self.run_cli(
            "ingest-prepare-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertTrue(prepare_payload["implemented"])
        self.assertEqual(prepare_payload["claimed"], 2)
        self.assertEqual(prepare_payload["prepared"], 2)
        self.assertEqual(prepare_payload["deferred_timeout"], 0)
        self.assertFalse(prepare_payload["more_prepare_remaining"])
        self.assertTrue(prepare_payload["advanced_to_commit"])
        self.assertEqual(prepare_payload["timings"]["prepare_ms"]["count"], 2)
        self.assertEqual(prepare_payload["timings"]["hash_ms"]["count"], 2)
        self.assertEqual(prepare_payload["timings"]["extract_ms"]["count"], 2)
        self.assertEqual(prepare_payload["timings"]["chunk_ms"]["count"], 2)
        self.assertEqual(prepare_payload["timings"]["prepared_serialize_ms"]["count"], 2)
        self.assertEqual(prepare_payload["timings"]["prepared_write_ms"]["count"], 1)
        self.assertGreater(prepare_payload["timings"]["prepared_payload_bytes"], 0)
        self.assertIsInstance(prepare_payload["timings"]["status_payload_ms"], float)
        run_payload = prepare_payload["run"]
        self.assertEqual(run_payload["phase"], "committing")
        self.assertEqual(run_payload["status"], "committing")
        self.assertEqual(run_payload["counts"]["work_items"]["prepared"], 2)
        self.assertIn("ingest-commit-step", " ".join(run_payload["next_recommended_commands"]))

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            document_count = int(connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] or 0)
            rows = connection.execute(
                """
                SELECT wi.rel_path, wi.status, pi.payload_kind, pi.payload_json, pi.source_fingerprint_json
                FROM ingest_work_items wi
                JOIN ingest_prepared_items pi ON pi.work_item_id = wi.id
                WHERE wi.run_id = ?
                ORDER BY wi.commit_order ASC
                """,
                (run_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(document_count, 0)
        self.assertEqual([row["rel_path"] for row in rows], ["raw/alpha.txt", "raw/beta.md"])
        self.assertTrue(all(row["status"] == "prepared" for row in rows))
        self.assertTrue(all(row["payload_kind"] == "loose_file" for row in rows))
        first_payload = json.loads(rows[0]["payload_json"])["prepared_item"]
        first_fingerprint = json.loads(rows[0]["source_fingerprint_json"])
        self.assertEqual(first_payload["rel_path"], "raw/alpha.txt")
        self.assertIn("alpha body", first_payload["extracted_payload"]["text_content"])
        self.assertTrue(first_payload["prepared_chunks"])
        self.assertEqual(
            first_fingerprint["hash"],
            hashlib.sha256(alpha_text.encode("utf-8")).hexdigest(),
        )

    def test_ingest_v2_prepare_step_uses_parallel_workers(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        for index in range(4):
            (raw_dir / f"doc-{index}.txt").write_text(f"parallel body {index}\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        plan_exit, plan_payload, _, _ = self.run_cli("ingest-plan-step", str(self.root), "--run-id", run_id)
        self.assertEqual(plan_exit, 0)
        self.assertIsNotNone(plan_payload)

        original_prepare = retriever_tools.ingest_v2_prepare_loose_file_item
        barrier = threading.Barrier(4)
        thread_ids: set[int] = set()
        thread_ids_lock = threading.Lock()

        def wrapped_prepare(root: Path, work_item_row, *, deadline: float):
            with thread_ids_lock:
                thread_ids.add(threading.get_ident())
            barrier.wait(timeout=5)
            return original_prepare(root, work_item_row, deadline=deadline)

        with (
            mock.patch.object(retriever_tools, "INGEST_V2_PREPARE_BATCH_SIZE", 4),
            mock.patch.object(retriever_tools, "ingest_prepare_worker_count", return_value=4),
            mock.patch.object(retriever_tools, "ingest_container_prepare_worker_count", return_value=1),
            mock.patch.object(retriever_tools, "ingest_v2_prepare_loose_file_item", side_effect=wrapped_prepare),
        ):
            prepare_exit, prepare_payload, _, _ = self.run_cli(
                "ingest-prepare-step",
                str(self.root),
                "--run-id",
                run_id,
                "--budget-seconds",
                "35",
            )

        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(prepare_payload["claimed"], 4)
        self.assertEqual(prepare_payload["prepared"], 4)
        self.assertEqual(prepare_payload["claim_limit"], 4)
        self.assertEqual(prepare_payload["prepare_workers"], 4)
        self.assertEqual(len(thread_ids), 4)

    def test_ingest_v2_commit_step_commits_prepared_loose_files(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        (raw_dir / "alpha.txt").write_text("alpha body\n", encoding="utf-8")
        (raw_dir / "beta.md").write_text("# beta\n\nbody\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        plan_exit, plan_payload, _, _ = self.run_cli("ingest-plan-step", str(self.root), "--run-id", run_id)
        self.assertEqual(plan_exit, 0)
        self.assertIsNotNone(plan_payload)
        prepare_exit, prepare_payload, _, _ = self.run_cli("ingest-prepare-step", str(self.root), "--run-id", run_id)
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(prepare_payload["run"]["phase"], "committing")

        commit_exit, commit_payload, _, _ = self.run_cli(
            "ingest-commit-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(commit_exit, 0)
        self.assertIsNotNone(commit_payload)
        self.assertTrue(commit_payload["implemented"])
        self.assertEqual(commit_payload["committed"], 2)
        self.assertEqual(commit_payload["failed"], 0)
        self.assertEqual(commit_payload["actions"], {"new": 2})
        self.assertFalse(commit_payload["more_commit_remaining"])
        self.assertTrue(commit_payload["advanced_to_finalize"])
        run_payload = commit_payload["run"]
        self.assertEqual(run_payload["phase"], "finalizing")
        self.assertEqual(run_payload["status"], "finalizing")
        self.assertEqual(run_payload["counts"]["work_items"]["committed"], 2)
        self.assertIn("ingest-finalize-step", " ".join(run_payload["next_recommended_commands"]))

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            documents = connection.execute(
                """
                SELECT rel_path, control_number
                FROM documents
                ORDER BY rel_path ASC
                """
            ).fetchall()
            work_items = connection.execute(
                """
                SELECT status, affected_document_ids_json, artifact_manifest_json
                FROM ingest_work_items
                WHERE run_id = ?
                ORDER BY commit_order ASC
                """,
                (run_id,),
            ).fetchall()
            chunk_count = int(connection.execute("SELECT COUNT(*) FROM document_chunks").fetchone()[0] or 0)
            cursor_row = connection.execute(
                """
                SELECT cursor_json, status
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'committing'
                  AND cursor_key = 'loose_file_commit'
                """,
                (run_id,),
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual([row["rel_path"] for row in documents], ["raw/alpha.txt", "raw/beta.md"])
        self.assertEqual([row["control_number"] for row in documents], ["DOC001.00000001", "DOC001.00000002"])
        self.assertTrue(all(row["status"] == "committed" for row in work_items))
        self.assertTrue(all(json.loads(row["affected_document_ids_json"]) for row in work_items))
        self.assertTrue(all(json.loads(row["artifact_manifest_json"])["commit_action"] == "new" for row in work_items))
        self.assertGreaterEqual(chunk_count, 2)
        self.assertIsNotNone(cursor_row)
        self.assertEqual(cursor_row["status"], "complete")
        cursor_payload = json.loads(cursor_row["cursor_json"])
        self.assertEqual(cursor_payload["actions"], {"new": 2})

    def test_ingest_v2_interleaves_prepare_and_commit_batches(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        for index in range(3):
            (raw_dir / f"doc-{index}.txt").write_text(f"body {index}\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        plan_exit, plan_payload, _, _ = self.run_cli("ingest-plan-step", str(self.root), "--run-id", run_id)
        self.assertEqual(plan_exit, 0)
        self.assertIsNotNone(plan_payload)

        with (
            mock.patch.object(retriever_tools, "INGEST_V2_PREPARE_BATCH_SIZE", 2),
            mock.patch.object(retriever_tools, "INGEST_V2_PREPARED_COMMIT_BATCH_TARGET", 2),
            mock.patch.object(retriever_tools, "ingest_prepare_worker_count", return_value=1),
            mock.patch.object(retriever_tools, "ingest_container_prepare_worker_count", return_value=1),
        ):
            first_prepare_exit, first_prepare_payload, _, _ = self.run_cli(
                "ingest-prepare-step",
                str(self.root),
                "--run-id",
                run_id,
            )
            self.assertEqual(first_prepare_exit, 0)
            self.assertIsNotNone(first_prepare_payload)
            self.assertEqual(first_prepare_payload["prepared"], 2)
            self.assertTrue(first_prepare_payload["advanced_to_commit"])
            self.assertTrue(first_prepare_payload["more_prepare_remaining"])
            self.assertEqual(first_prepare_payload["run"]["phase"], "committing")
            self.assertEqual(first_prepare_payload["run"]["counts"]["work_items"]["pending"], 1)
            self.assertEqual(first_prepare_payload["run"]["counts"]["work_items"]["prepared"], 2)

            first_commit_exit, first_commit_payload, _, _ = self.run_cli(
                "ingest-commit-step",
                str(self.root),
                "--run-id",
                run_id,
            )
            self.assertEqual(first_commit_exit, 0)
            self.assertIsNotNone(first_commit_payload)
            self.assertEqual(first_commit_payload["committed"], 2)
            self.assertFalse(first_commit_payload["advanced_to_finalize"])
            self.assertTrue(first_commit_payload["advanced_to_prepare"])
            self.assertEqual(first_commit_payload["run"]["phase"], "preparing")
            self.assertEqual(first_commit_payload["run"]["counts"]["work_items"]["pending"], 1)
            self.assertEqual(first_commit_payload["run"]["counts"]["work_items"]["committed"], 2)

            second_prepare_exit, second_prepare_payload, _, _ = self.run_cli(
                "ingest-prepare-step",
                str(self.root),
                "--run-id",
                run_id,
            )
            self.assertEqual(second_prepare_exit, 0)
            self.assertIsNotNone(second_prepare_payload)
            self.assertEqual(second_prepare_payload["prepared"], 1)
            self.assertTrue(second_prepare_payload["advanced_to_commit"])
            self.assertEqual(second_prepare_payload["run"]["phase"], "committing")

            second_commit_exit, second_commit_payload, _, _ = self.run_cli(
                "ingest-commit-step",
                str(self.root),
                "--run-id",
                run_id,
            )
            self.assertEqual(second_commit_exit, 0)
            self.assertIsNotNone(second_commit_payload)
            self.assertEqual(second_commit_payload["committed"], 1)
            self.assertTrue(second_commit_payload["advanced_to_finalize"])
            self.assertFalse(second_commit_payload["advanced_to_prepare"])
            self.assertEqual(second_commit_payload["run"]["phase"], "finalizing")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            committed_count = int(
                connection.execute(
                    "SELECT COUNT(*) FROM ingest_work_items WHERE run_id = ? AND status = 'committed'",
                    (run_id,),
                ).fetchone()[0]
                or 0
            )
            document_count = int(connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] or 0)
        finally:
            connection.close()
        self.assertEqual(committed_count, 3)
        self.assertEqual(document_count, 3)

    def test_ingest_v2_run_step_reclaims_stale_prepare_lease(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        (raw_dir / "alpha.txt").write_text("alpha body\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        plan_exit, plan_payload, _, _ = self.run_cli("ingest-plan-step", str(self.root), "--run-id", run_id)
        self.assertEqual(plan_exit, 0)
        self.assertIsNotNone(plan_payload)
        self.assertEqual(plan_payload["run"]["phase"], "preparing")

        old_timestamp = "2000-01-01T00:00:00Z"
        future_timestamp = "2999-01-01T00:00:00Z"
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute(
                """
                UPDATE ingest_work_items
                SET status = 'leased',
                    lease_owner = 'dead-prepare-worker',
                    lease_expires_at = ?,
                    updated_at = ?
                WHERE run_id = ?
                """,
                (future_timestamp, old_timestamp, run_id),
            )
            connection.commit()
        finally:
            connection.close()

        step_exit, step_payload, _, _ = self.run_cli(
            "ingest-run-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(step_exit, 0)
        self.assertIsNotNone(step_payload)
        self.assertEqual(step_payload["stale_reclaimed"], {"prepare": 1, "commit": 0})
        self.assertEqual(step_payload["run"]["status"], "completed")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            document_count = int(connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] or 0)
            committed_count = int(
                connection.execute(
                    "SELECT COUNT(*) FROM ingest_work_items WHERE run_id = ? AND status = 'committed'",
                    (run_id,),
                ).fetchone()[0]
                or 0
            )
        finally:
            connection.close()
        self.assertEqual(document_count, 1)
        self.assertEqual(committed_count, 1)

    def test_ingest_v2_run_step_reclaims_stale_commit_lease(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        (raw_dir / "alpha.txt").write_text("alpha body\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        plan_exit, plan_payload, _, _ = self.run_cli("ingest-plan-step", str(self.root), "--run-id", run_id)
        self.assertEqual(plan_exit, 0)
        self.assertIsNotNone(plan_payload)
        prepare_exit, prepare_payload, _, _ = self.run_cli("ingest-prepare-step", str(self.root), "--run-id", run_id)
        self.assertEqual(prepare_exit, 0)
        self.assertIsNotNone(prepare_payload)
        self.assertEqual(prepare_payload["run"]["phase"], "committing")

        old_timestamp = "2000-01-01T00:00:00Z"
        future_timestamp = "2999-01-01T00:00:00Z"
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute(
                """
                UPDATE ingest_work_items
                SET status = 'committing',
                    lease_owner = 'dead-commit-worker',
                    lease_expires_at = ?,
                    updated_at = ?
                WHERE run_id = ?
                """,
                (future_timestamp, old_timestamp, run_id),
            )
            connection.execute(
                """
                UPDATE ingest_runs
                SET committer_lease_owner = 'dead-commit-worker',
                    committer_lease_expires_at = ?,
                    committer_heartbeat_at = ?,
                    last_heartbeat_at = ?
                WHERE run_id = ?
                """,
                (future_timestamp, old_timestamp, old_timestamp, run_id),
            )
            connection.commit()
        finally:
            connection.close()

        step_exit, step_payload, _, _ = self.run_cli(
            "ingest-run-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(step_exit, 0)
        self.assertIsNotNone(step_payload)
        self.assertEqual(step_payload["stale_reclaimed"], {"prepare": 0, "commit": 1})
        self.assertEqual(step_payload["run"]["status"], "completed")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            document_count = int(connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] or 0)
            committed_count = int(
                connection.execute(
                    "SELECT COUNT(*) FROM ingest_work_items WHERE run_id = ? AND status = 'committed'",
                    (run_id,),
                ).fetchone()[0]
                or 0
            )
        finally:
            connection.close()
        self.assertEqual(document_count, 1)
        self.assertEqual(committed_count, 1)

    def test_ingest_v2_finalize_step_completes_and_marks_missing_loose_files(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        stale_path = raw_dir / "stale.txt"
        stale_path.write_text("stale body\n", encoding="utf-8")

        legacy_exit, legacy_payload, _, _ = self.run_cli(
            "ingest",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(legacy_exit, 0)
        self.assertIsNotNone(legacy_payload)
        self.assertEqual(legacy_payload["new"], 1)

        stale_path.unlink()
        (raw_dir / "alpha.txt").write_text("alpha body\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        for command in ("ingest-plan-step", "ingest-prepare-step", "ingest-commit-step"):
            exit_code, payload, _, _ = self.run_cli(command, str(self.root), "--run-id", run_id)
            self.assertEqual(exit_code, 0)
            self.assertIsNotNone(payload)

        finalize_exit, finalize_payload, _, _ = self.run_cli(
            "ingest-finalize-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(finalize_exit, 0)
        self.assertIsNotNone(finalize_payload)
        self.assertTrue(finalize_payload["implemented"])
        self.assertTrue(finalize_payload["finalization_complete"])
        self.assertEqual(finalize_payload["stages_completed"], ["missing", "conversations", "prune", "complete"])
        self.assertEqual(finalize_payload["cursor"]["filesystem_missing"], 1)
        run_payload = finalize_payload["run"]
        self.assertEqual(run_payload["phase"], "completed")
        self.assertEqual(run_payload["status"], "completed")
        self.assertFalse(finalize_payload["more_work_remaining"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            occurrence_rows = connection.execute(
                """
                SELECT rel_path, lifecycle_status
                FROM document_occurrences
                WHERE rel_path IN ('raw/alpha.txt', 'raw/stale.txt')
                ORDER BY rel_path ASC
                """
            ).fetchall()
            cursor_row = connection.execute(
                """
                SELECT cursor_json, status
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'finalizing'
                  AND cursor_key = 'loose_file_finalize'
                """,
                (run_id,),
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(
            [(row["rel_path"], row["lifecycle_status"]) for row in occurrence_rows],
            [("raw/alpha.txt", "active"), ("raw/stale.txt", "missing")],
        )
        self.assertIsNotNone(cursor_row)
        self.assertEqual(cursor_row["status"], "complete")
        self.assertEqual(json.loads(cursor_row["cursor_json"])["stage"], "complete")

    def test_ingest_v2_finalize_records_conversation_preview_permission_error(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        self.write_email_message(
            raw_dir / "alpha.eml",
            subject="Preview blocked",
            body_text="preview body\n",
            message_id="<preview-blocked@example.com>",
        )

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        for command in ("ingest-plan-step", "ingest-prepare-step", "ingest-commit-step"):
            exit_code, payload, _, _ = self.run_cli(command, str(self.root), "--run-id", run_id)
            self.assertEqual(exit_code, 0)
            self.assertIsNotNone(payload)

        finalize_exit, finalize_payload, _, _ = self.run_cli(
            "ingest-finalize-step",
            str(self.root),
            "--run-id",
            run_id,
            "--budget-seconds",
            "35",
        )
        self.assertEqual(finalize_exit, 0)
        self.assertIsNotNone(finalize_payload)
        self.assertFalse(finalize_payload["finalization_complete"])
        self.assertEqual(finalize_payload["run"]["phase"], "preparing")
        self.assertEqual(finalize_payload["run"]["counts"]["by_unit_type"]["conversation_preview"]["pending"], 1)

        with mock.patch.object(
            retriever_tools,
            "refresh_conversation_previews",
            side_effect=PermissionError("blocked preview write"),
        ):
            run_exit, run_payload, _, _ = self.run_cli(
                "ingest-run-step",
                str(self.root),
                "--run-id",
                run_id,
                "--budget-seconds",
                "35",
            )

        self.assertEqual(run_exit, 0)
        self.assertIsNotNone(run_payload)
        self.assertEqual(run_payload["run"]["status"], "completed")
        self.assertEqual(run_payload["run"]["counts"]["by_unit_type"]["conversation_preview"]["failed"], 1)
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            failed_row = connection.execute(
                """
                SELECT last_error
                FROM ingest_work_items
                WHERE run_id = ?
                  AND unit_type = 'conversation_preview'
                """,
                (run_id,),
            ).fetchone()
            cursor_row = connection.execute(
                """
                SELECT cursor_json
                FROM ingest_phase_cursors
                WHERE run_id = ?
                  AND phase = 'finalizing'
                  AND cursor_key = 'loose_file_finalize'
                """,
                (run_id,),
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(failed_row)
        self.assertIn("PermissionError", failed_row["last_error"])
        self.assertIsNotNone(cursor_row)
        self.assertEqual(json.loads(cursor_row["cursor_json"])["conversation_preview_failures"], 1)

    def test_ingest_v2_reingest_skips_unchanged_loose_file(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        alpha_path = raw_dir / "alpha.txt"
        alpha_path.write_text("alpha body\n", encoding="utf-8")

        first_payloads = self.run_v2_loose_ingest("raw")
        self.assertEqual(first_payloads["commit"]["actions"], {"new": 1})
        original_row = self.fetch_document_row("raw/alpha.txt")

        second_payloads = self.run_v2_loose_ingest("raw")
        self.assertEqual(second_payloads["commit"]["actions"], {"skipped": 1})
        self.assertEqual(second_payloads["commit"]["committed"], 1)
        self.assertTrue(second_payloads["finalize"]["finalization_complete"])

        skipped_row = self.fetch_document_row("raw/alpha.txt")
        self.assertEqual(skipped_row["id"], original_row["id"])
        self.assertEqual(skipped_row["control_number"], original_row["control_number"])
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            document_count = int(connection.execute("SELECT COUNT(*) FROM documents").fetchone()[0] or 0)
            artifact_row = connection.execute(
                """
                SELECT artifact_manifest_json
                FROM ingest_work_items
                WHERE run_id = ?
                """,
                (second_payloads["run_id"],),
            ).fetchone()
        finally:
            connection.close()
        self.assertEqual(document_count, 1)
        self.assertIsNotNone(artifact_row)
        self.assertEqual(json.loads(artifact_row["artifact_manifest_json"])["commit_action"], "skipped")

    def test_ingest_v2_reingest_updates_modified_loose_file(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        alpha_path = raw_dir / "alpha.txt"
        alpha_path.write_text("original v2 body\n", encoding="utf-8")

        first_payloads = self.run_v2_loose_ingest("raw")
        self.assertEqual(first_payloads["commit"]["actions"], {"new": 1})
        original_row = self.fetch_document_row("raw/alpha.txt")

        alpha_path.write_text("updated v2 body with searchable marker\n", encoding="utf-8")
        second_payloads = self.run_v2_loose_ingest("raw")
        self.assertEqual(second_payloads["commit"]["actions"], {"updated": 1})
        self.assertTrue(second_payloads["finalize"]["finalization_complete"])

        updated_row = self.fetch_document_row("raw/alpha.txt")
        self.assertEqual(updated_row["id"], original_row["id"])
        self.assertEqual(updated_row["control_number"], original_row["control_number"])
        search_payload = retriever_tools.search(self.root, "searchable marker", None, None, None, 1, 20)
        self.assertEqual(search_payload["total_hits"], 1)
        self.assertEqual(search_payload["results"][0]["id"], updated_row["id"])

    def test_ingest_v2_rename_consumption_is_durable_across_commit_steps(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        original_path = raw_dir / "original.txt"
        body = "same durable rename body\n"
        original_path.write_text(body, encoding="utf-8")

        first_payloads = self.run_v2_loose_ingest("raw")
        self.assertEqual(first_payloads["commit"]["actions"], {"new": 1})
        original_row = self.fetch_document_row("raw/original.txt")

        original_path.unlink()
        (raw_dir / "renamed-a.txt").write_text(body, encoding="utf-8")
        (raw_dir / "renamed-b.txt").write_text(body, encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])
        for command in ("ingest-plan-step", "ingest-prepare-step"):
            exit_code, payload, _, _ = self.run_cli(command, str(self.root), "--run-id", run_id)
            self.assertEqual(exit_code, 0)
            self.assertIsNotNone(payload)

        first_commit_exit, first_commit_payload, _, _ = self.run_cli(
            "ingest-commit-step",
            str(self.root),
            "--run-id",
            run_id,
            "--max-items",
            "1",
        )
        self.assertEqual(first_commit_exit, 0)
        self.assertIsNotNone(first_commit_payload)
        self.assertEqual(first_commit_payload["actions"], {"renamed": 1})
        self.assertTrue(first_commit_payload["more_commit_remaining"])

        second_commit_exit, second_commit_payload, _, _ = self.run_cli(
            "ingest-commit-step",
            str(self.root),
            "--run-id",
            run_id,
            "--max-items",
            "1",
        )
        self.assertEqual(second_commit_exit, 0)
        self.assertIsNotNone(second_commit_payload)
        self.assertEqual(second_commit_payload["actions"], {"new": 1})
        self.assertTrue(second_commit_payload["advanced_to_finalize"])

        finalize_exit, finalize_payload, _, _ = self.run_cli("ingest-finalize-step", str(self.root), "--run-id", run_id)
        self.assertEqual(finalize_exit, 0)
        self.assertIsNotNone(finalize_payload)
        self.assertTrue(finalize_payload["finalization_complete"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            rename_rows = connection.execute(
                """
                SELECT source_document_id, source_occurrence_id
                FROM ingest_rename_consumptions
                WHERE run_id = ?
                """,
                (run_id,),
            ).fetchall()
            occurrence_rows = connection.execute(
                """
                SELECT rel_path, lifecycle_status
                FROM document_occurrences
                WHERE document_id = ?
                ORDER BY rel_path ASC
                """,
                (original_row["id"],),
            ).fetchall()
            work_item_actions = [
                json.loads(row["artifact_manifest_json"])["commit_action"]
                for row in connection.execute(
                    """
                    SELECT artifact_manifest_json
                    FROM ingest_work_items
                    WHERE run_id = ?
                    ORDER BY commit_order ASC
                    """,
                    (run_id,),
                ).fetchall()
            ]
        finally:
            connection.close()

        self.assertEqual(len(rename_rows), 1)
        self.assertEqual(int(rename_rows[0]["source_document_id"]), original_row["id"])
        self.assertEqual(work_item_actions, ["renamed", "new"])
        self.assertEqual(
            [(row["rel_path"], row["lifecycle_status"]) for row in occurrence_rows],
            [("raw/renamed-a.txt", "active"), ("raw/renamed-b.txt", "active")],
        )

    def test_ingest_v2_commit_reprepares_when_source_changes_after_prepare(self) -> None:
        raw_dir = self.root / "raw"
        raw_dir.mkdir()
        alpha_path = raw_dir / "alpha.txt"
        alpha_path.write_text("old prepared body\n", encoding="utf-8")

        start_exit, start_payload, _, _ = self.run_cli(
            "ingest-start",
            str(self.root),
            "--recursive",
            "--path",
            "raw",
        )
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])
        for command in ("ingest-plan-step", "ingest-prepare-step"):
            exit_code, payload, _, _ = self.run_cli(command, str(self.root), "--run-id", run_id)
            self.assertEqual(exit_code, 0)
            self.assertIsNotNone(payload)

        alpha_path.write_text("new prepared body with fallback marker\n", encoding="utf-8")
        commit_exit, commit_payload, _, _ = self.run_cli("ingest-commit-step", str(self.root), "--run-id", run_id)
        self.assertEqual(commit_exit, 0)
        self.assertIsNotNone(commit_payload)
        self.assertEqual(commit_payload["actions"], {"new": 1})
        self.assertEqual(commit_payload["freshness_fallbacks"], 1)
        finalize_exit, finalize_payload, _, _ = self.run_cli("ingest-finalize-step", str(self.root), "--run-id", run_id)
        self.assertEqual(finalize_exit, 0)
        self.assertIsNotNone(finalize_payload)
        self.assertTrue(finalize_payload["finalization_complete"])

        row = self.fetch_document_row("raw/alpha.txt")
        search_payload = retriever_tools.search(self.root, "fallback marker", None, None, None, 1, 20)
        self.assertEqual(search_payload["total_hits"], 1)
        self.assertEqual(search_payload["results"][0]["id"], row["id"])
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            artifact_row = connection.execute(
                """
                SELECT artifact_manifest_json
                FROM ingest_work_items
                WHERE run_id = ?
                """,
                (run_id,),
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(artifact_row)
        self.assertTrue(json.loads(artifact_row["artifact_manifest_json"])["freshness_fallback"])

    def test_ingest_v2_rejects_budget_above_hard_cap(self) -> None:
        exit_code, payload, _, stderr = self.run_cli(
            "ingest-start",
            str(self.root),
            "--budget-seconds",
            "41",
        )

        self.assertEqual(exit_code, 2)
        self.assertIsNotNone(payload)
        self.assertIn("budget-seconds", payload["error"])
        self.assertIn("40", payload["error"])
        self.assertTrue(stderr)

    def test_active_ingest_v2_run_blocks_conflicting_commands(self) -> None:
        start_exit, start_payload, _, _ = self.run_cli("ingest-start", str(self.root))
        self.assertEqual(start_exit, 0)
        self.assertIsNotNone(start_payload)
        run_id = str(start_payload["run_id"])

        second_exit, second_payload, _, _ = self.run_cli("ingest-start", str(self.root))
        self.assertEqual(second_exit, 2)
        self.assertIsNotNone(second_payload)
        self.assertEqual(second_payload["error"], "active_ingest_run")
        self.assertEqual(second_payload["active_run_id"], run_id)
        self.assertIn("ingest-status", second_payload["status_command"])
        self.assertIn("ingest-cancel", second_payload["cancel_command"])

        entity_exit, entity_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--display-name",
            "Alice Example",
        )
        self.assertEqual(entity_exit, 2)
        self.assertIsNotNone(entity_payload)
        self.assertEqual(entity_payload["error"], "active_ingest_run")
        self.assertEqual(entity_payload["active_run_id"], run_id)

        legacy_ingest_exit, legacy_ingest_payload, _, _ = self.run_cli("ingest", str(self.root))
        self.assertEqual(legacy_ingest_exit, 2)
        self.assertIsNotNone(legacy_ingest_payload)
        self.assertEqual(legacy_ingest_payload["error"], "active_ingest_run")
        self.assertEqual(legacy_ingest_payload["active_run_id"], run_id)

    def test_non_attachment_child_documents_are_not_treated_as_attachments(self) -> None:
        parent_path = self.root / "parent.txt"
        parent_path.write_text("parent body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        parent_row = self.fetch_document_row("parent.txt")
        self.assertIsNotNone(parent_row["dataset_id"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            child_control_number = retriever_tools.format_control_number(
                int(parent_row["control_number_batch"]),
                int(parent_row["control_number_family_sequence"]),
                retriever_tools.next_attachment_sequence(connection, int(parent_row["id"])),
            )
            child_id = retriever_tools.upsert_document_row(
                connection,
                "_retriever/logical/slack/general/threads/1700000000.000001.slackthread",
                None,
                None,
                {
                    "text_content": "Root message\nReply one",
                    "content_type": "Chat",
                    "title": "general thread",
                    "participants": "Alice Example; Bob Example",
                    "text_status": "ok",
                    "page_count": None,
                    "author": None,
                    "custodian": None,
                    "date_created": "2026-04-14T10:00:00Z",
                    "date_modified": "2026-04-14T11:00:00Z",
                    "subject": None,
                    "recipients": None,
                },
                file_name="1700000000.000001.slackthread",
                parent_document_id=int(parent_row["id"]),
                child_document_kind=retriever_tools.CHILD_DOCUMENT_KIND_REPLY_THREAD,
                control_number=child_control_number,
                dataset_id=int(parent_row["dataset_id"]),
                control_number_batch=int(parent_row["control_number_batch"]),
                control_number_family_sequence=int(parent_row["control_number_family_sequence"]),
                control_number_attachment_sequence=1,
                source_kind=retriever_tools.SLACK_EXPORT_SOURCE_KIND,
                source_rel_path="data/slack/general/2026-04-14.json",
                source_item_id="1700000000.000001",
                root_message_key="general:1700000000.000001",
            )
            retriever_tools.ensure_dataset_document_membership(
                connection,
                dataset_id=int(parent_row["dataset_id"]),
                document_id=child_id,
                dataset_source_id=None,
            )
            connection.commit()
            inventory = retriever_tools.document_inventory_counts(connection)
        finally:
            connection.close()

        child_row = self.fetch_document_by_id(child_id)
        self.assertEqual(child_row["child_document_kind"], retriever_tools.CHILD_DOCUMENT_KIND_REPLY_THREAD)
        self.assertEqual(child_row["source_kind"], retriever_tools.SLACK_EXPORT_SOURCE_KIND)
        self.assertEqual(inventory["parent_documents"], 1)
        self.assertEqual(inventory["attachment_children"], 0)
        self.assertEqual(inventory["documents_total"], 2)

        browse_results = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        parent_result = next(item for item in browse_results["results"] if item["id"] == parent_row["id"])
        child_result = next(item for item in browse_results["results"] if item["id"] == child_id)
        self.assertEqual(parent_result["attachment_count"], 0)
        self.assertEqual(child_result["child_document_kind"], retriever_tools.CHILD_DOCUMENT_KIND_REPLY_THREAD)
        self.assertEqual(child_result["parent"]["control_number"], parent_row["control_number"])

        attachments_only = retriever_tools.search(
            self.root,
            "",
            [["is_attachment", "eq", "true"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(attachments_only["total_hits"], 0)

        parents_with_attachments = retriever_tools.search(
            self.root,
            "",
            [["has_attachments", "eq", "true"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(parents_with_attachments["total_hits"], 0)

    def test_doctor_probes_pst_backend_when_unloaded(self) -> None:
        retriever_tools.bootstrap(self.root)

        with (
            mock.patch.object(retriever_tools, "pypff", retriever_tools._UNLOADED_DEPENDENCY),
            mock.patch.object(retriever_tools, "load_dependency", return_value=object()) as load_dependency,
        ):
            doctor_result = retriever_tools.doctor(self.root, quick=True)

        self.assertEqual(doctor_result["overall"], "pass")
        self.assertEqual(doctor_result["pst_backend"]["status"], "pass")
        self.assertIn("PST backend import succeeded", doctor_result["pst_backend"]["detail"])
        self.assertIn(doctor_result["plugin_runtime"]["status"], {"missing", "partial", "pass"})
        load_dependency.assert_called_once_with("pypff", allow_auto_install=False)

    def test_workspace_status_reports_registry_drift_without_repairing_it(self) -> None:
        retriever_tools.bootstrap(self.root)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute("ALTER TABLE documents ADD COLUMN issue_tag TEXT")
            connection.commit()
        finally:
            connection.close()

        exit_code, payload, _, _ = self.run_cli("workspace", "status", str(self.root))

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["action"], "status")
        self.assertEqual(payload["custom_field_registry"]["missing_registry"], ["issue_tag"])
        self.assertEqual(payload["custom_field_registry"]["repaired_registry"], [])
        self.assertNotIn("schema_apply", payload)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            registry_status = retriever_tools.inspect_custom_fields_registry(connection)
        finally:
            connection.close()

        self.assertEqual(registry_status["missing_registry"], ["issue_tag"])

    def test_doctor_reports_optional_pst_backend_failure_without_failing_runtime(self) -> None:
        retriever_tools.bootstrap(self.root)

        with (
            mock.patch.object(retriever_tools, "pypff", retriever_tools._UNLOADED_DEPENDENCY),
            mock.patch.object(retriever_tools, "load_dependency", return_value=None) as load_dependency,
        ):
            doctor_result = retriever_tools.doctor(self.root, quick=True)

        self.assertEqual(doctor_result["overall"], "pass")
        self.assertEqual(doctor_result["pst_backend"]["status"], "fail")
        self.assertIn("libpff-python", doctor_result["pst_backend"]["detail"])
        self.assertIn(doctor_result["plugin_runtime"]["status"], {"missing", "partial", "pass"})
        load_dependency.assert_called_once_with("pypff", allow_auto_install=False)

    def test_inspect_pst_properties_surfaces_chat_scope_candidates(self) -> None:
        pst_path = self.write_fake_pst_file()

        class FakeRecordEntry:
            def __init__(self, *, entry_type: int, value_type: int, data: bytes, property_name: str | None = None):
                self.entry_type = entry_type
                self.value_type = value_type
                self.data = data
                self.property_name = property_name

        class FakeRecordSet:
            def __init__(self, entries: list[object]):
                self.entries = entries

        class FakeMessage:
            def __init__(
                self,
                *,
                entry_identifier: str,
                message_class: str,
                subject: str,
                body: str,
                sender_name: str,
                sender_email_address: str,
                display_to: str = "",
                delivery_time: str = "2024-04-23T23:50:00Z",
                record_sets: list[object] | None = None,
            ):
                self.entry_identifier = entry_identifier
                self.message_class = message_class
                self.subject = subject
                self.body = body
                self.sender_name = sender_name
                self.sender_email_address = sender_email_address
                self.display_to = display_to
                self.delivery_time = delivery_time
                self.attachments = []
                self.record_sets = record_sets or []

        class FakeFolder:
            def __init__(
                self,
                name: str,
                *,
                messages: list[object] | None = None,
                sub_folders: list[object] | None = None,
            ):
                self.name = name
                self.messages = messages or []
                self.sub_folders = sub_folders or []

        class FakePstFile:
            def __init__(self, root_folder: object):
                self.root_folder = root_folder

            def open(self, raw_path: str) -> None:
                self.opened_path = raw_path

            def get_root_folder(self) -> object:
                return self.root_folder

            def close(self) -> None:
                return None

        teams_thread_id = "19:meeting_abc@thread.v2"
        teams_channel_id = "19:channel_xyz@thread.skype"
        root_folder = FakeFolder(
            "Top of Personal Folders",
            sub_folders=[
                FakeFolder(
                    "TeamsMessagesData",
                    messages=[
                        FakeMessage(
                            entry_identifier="pst-chat-001",
                            message_class="IPM.Note.Microsoft.Conversation",
                            subject="Hey",
                            body="Hey",
                            sender_name="Sergey Demyanov",
                            sender_email_address="sergey@example.com",
                            record_sets=[
                                FakeRecordSet(
                                    [
                                        FakeRecordEntry(
                                            entry_type=0x8001,
                                            value_type=0x001F,
                                            data=(teams_thread_id + "\x00").encode("utf-16-le"),
                                            property_name="TeamsThreadId",
                                        )
                                    ]
                                )
                            ],
                        ),
                        FakeMessage(
                            entry_identifier="pst-chat-002",
                            message_class="IPM.Note.Microsoft.Conversation",
                            subject="test",
                            body="test",
                            sender_name="Sergey Demyanov",
                            sender_email_address="sergey@example.com",
                            record_sets=[
                                FakeRecordSet(
                                    [
                                        FakeRecordEntry(
                                            entry_type=0x8002,
                                            value_type=0x001F,
                                            data=(teams_channel_id + "\x00").encode("utf-16-le"),
                                            property_name="TeamsChannelId",
                                        )
                                    ]
                                )
                            ],
                        ),
                    ],
                ),
                FakeFolder(
                    "Inbox",
                    messages=[
                        FakeMessage(
                            entry_identifier="pst-email-001",
                            message_class="IPM.Note",
                            subject="Email subject",
                            body="Email body",
                            sender_name="Alice Example",
                            sender_email_address="alice@example.com",
                            display_to="Bob Example <bob@example.com>",
                        )
                    ],
                ),
            ],
        )
        fake_pypff = types.SimpleNamespace(file=lambda: FakePstFile(root_folder))

        with mock.patch.object(retriever_tools, "pypff", fake_pypff):
            exit_code, payload, _, _ = self.run_cli(
                "inspect-pst-properties",
                str(self.root),
                str(pst_path.name),
                "--message-kind",
                "chat",
                "--limit",
                "1",
                "--max-record-entries",
                "4",
            )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["pst_rel_path"], "mailbox.pst")
        self.assertEqual(payload["message_kind"], "chat")
        self.assertEqual(payload["scanned"], 3)
        self.assertEqual(payload["matched"], 2)
        self.assertEqual(payload["returned"], 1)
        self.assertTrue(payload["truncated"])
        first_message = payload["messages"][0]
        self.assertEqual(first_message["source_item_id"], "pst-chat-001")
        self.assertEqual(first_message["message_kind"], "chat")
        self.assertEqual(first_message["candidate_scope_values"], [teams_thread_id])
        self.assertEqual(first_message["interesting_properties"][0]["property_name"], "TeamsThreadId")
        self.assertEqual(first_message["interesting_properties"][0]["decoded_value"], teams_thread_id)

        with mock.patch.object(retriever_tools, "pypff", fake_pypff):
            filtered_exit, filtered_payload, _, _ = self.run_cli(
                "inspect-pst-properties",
                str(self.root),
                str(pst_path.name),
                "--message-kind",
                "chat",
                "--source-item-id",
                "pst-chat-002",
            )

        self.assertEqual(filtered_exit, 0)
        self.assertIsNotNone(filtered_payload)
        self.assertEqual(filtered_payload["matched"], 1)
        self.assertEqual(filtered_payload["returned"], 1)
        self.assertEqual(filtered_payload["messages"][0]["source_item_id"], "pst-chat-002")
        self.assertEqual(filtered_payload["messages"][0]["candidate_scope_values"], [teams_channel_id])

    def test_ingest_supports_ics_calendar_files(self) -> None:
        calendar_path = self.root / "invite.ics"
        calendar_path.write_text(
            "\n".join(
                [
                    "BEGIN:VCALENDAR",
                    "VERSION:2.0",
                    "BEGIN:VEVENT",
                    "SUMMARY:Board Meeting",
                    "DESCRIPTION:Discuss launch plans",
                    "END:VEVENT",
                    "END:VCALENDAR",
                    "",
                ]
            ),
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("invite.ics")
        self.assertEqual(row["content_type"], "Calendar")
        self.assertEqual(row["text_status"], "ok")

        search_result = retriever_tools.search(self.root, "Board Meeting", None, None, None, 1, 20)
        self.assertEqual(search_result["results"][0]["file_name"], "invite.ics")
        self.assertEqual(search_result["results"][0]["preview_rel_path"], "invite.ics")

    def test_ingest_supports_png_native_preview_only(self) -> None:
        image_path = self.root / "sample.png"
        image_path.write_bytes(
            bytes.fromhex(
                "89504E470D0A1A0A0000000D49484452000000010000000108060000001F15C489"
                "0000000D49444154789C6360000000020001E221BC330000000049454E44AE426082"
            )
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("sample.png")
        self.assertEqual(row["content_type"], "Image")
        self.assertEqual(row["text_status"], "empty")

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        png_result = next(item for item in browse_result["results"] if item["file_name"] == "sample.png")
        self.assertEqual(png_result["preview_rel_path"], "sample.png")
        self.assertEqual(png_result["preview_targets"][0]["preview_type"], "native")

    def test_ingest_supports_additional_native_preview_images(self) -> None:
        image_files = {
            "sample.bmp": b"bmp bytes",
            "sample.gif": b"gif bytes",
            "sample.jpeg": b"jpeg bytes",
            "sample.jpg": b"jpg bytes",
            "sample.tif": b"tif bytes",
            "sample.tiff": b"tiff bytes",
            "sample.webp": b"webp bytes",
        }
        for file_name, payload in image_files.items():
            (self.root / file_name).write_bytes(payload)

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], len(image_files))
        self.assertEqual(ingest_result["failed"], 0)

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 50)
        for file_name in image_files:
            row = self.fetch_document_row(file_name)
            self.assertEqual(row["content_type"], "Image")
            self.assertEqual(row["text_status"], "empty")
            result = next(item for item in browse_result["results"] if item["file_name"] == file_name)
            self.assertEqual(result["preview_rel_path"], file_name)
            self.assertEqual(result["preview_targets"][0]["preview_type"], "native")

    def test_image_path_data_url_preserves_compact_tiff_modes_before_rgb_fallback(self) -> None:
        try:
            from PIL import Image
        except Exception as exc:  # pragma: no cover - test helper dependency
            self.skipTest(f"Pillow unavailable for TIFF data-url test: {exc}")

        image_path = self.root / "monochrome.tif"
        image = Image.new("1", (256, 256))
        for y in range(256):
            for x in range(256):
                image.putpixel((x, y), (x + y) % 2)
        image.save(image_path, format="TIFF")

        data_url = retriever_tools.image_path_data_url(image_path)
        self.assertIsNotNone(data_url)
        self.assertTrue(str(data_url).startswith("data:image/png;base64,"))
        generated_png_bytes = base64.b64decode(str(data_url).split(",", 1)[1])

        rgb_buffer = io.BytesIO()
        with Image.open(image_path) as reopened:
            reopened.convert("RGB").save(rgb_buffer, format="PNG", optimize=True)
        self.assertLess(len(generated_png_bytes), len(rgb_buffer.getvalue()))

    def test_build_production_extracted_payload_can_limit_preview_images(self) -> None:
        try:
            from PIL import Image
        except Exception as exc:  # pragma: no cover - test helper dependency
            self.skipTest(f"Pillow unavailable for production preview image limit test: {exc}")

        image_paths: list[Path] = []
        for index in range(3):
            image_path = self.root / f"page-{index + 1}.tif"
            Image.new("RGB", (32, 32), color=(index * 40, 20, 120)).save(image_path, format="TIFF")
            image_paths.append(image_path)

        payload = retriever_tools.build_production_extracted_payload(
            self.root,
            production_name="Synthetic",
            control_number="PDX000001",
            begin_bates="PDX000001",
            end_bates="PDX000003",
            begin_attachment=None,
            end_attachment=None,
            text_path=None,
            image_paths=image_paths,
            native_path=None,
            preview_image_limit=1,
            preview_image_max_dimension=16,
        )

        self.assertEqual(payload["page_count"], 3)
        preview_content = payload["preview_artifacts"][0]["content"]
        self.assertEqual(preview_content.count("<figure>"), 1)
        self.assertIn("Preview shows the first 1 of 3 produced pages", preview_content)

    def test_ingest_supports_rtf_text_extraction(self) -> None:
        rtf_path = self.root / "memo.rtf"
        rtf_path.write_text(
            r"{\rtf1\ansi\deff0 {\fonttbl {\f0 Arial;}}\f0 Contract Review Memorandum\par Budget approved for Q2.\par}",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("memo.rtf")
        self.assertEqual(row["content_type"], "E-Doc")
        self.assertEqual(row["text_status"], "ok")
        self.assertEqual(row["title"], "Contract Review Memorandum")

        search_result = retriever_tools.search(self.root, "Budget approved", None, None, None, 1, 20)
        self.assertEqual(search_result["results"][0]["file_name"], "memo.rtf")
        self.assertTrue(search_result["results"][0]["preview_rel_path"].endswith(".html"))
        self.assertEqual(search_result["results"][0]["preview_targets"][0]["preview_type"], "html")
        preview_html = Path(search_result["results"][0]["preview_targets"][0]["abs_path"]).read_text(encoding="utf-8")
        self.assertIn("<title>Contract Review Memorandum</title>", preview_html)
        self.assertIn("<h1>Contract Review Memorandum</h1>", preview_html)

    def test_ingest_prefers_native_preview_for_chat_like_rtf_files(self) -> None:
        rtf_path = self.root / "chat-thread.rtf"
        rtf_path.write_text(
            (
                r"{\rtf1\ansi\deff0 {\fonttbl {\f0 Arial;}}\f0 "
                r"[2026-04-15 09:00] Alice Example: Kickoff thread for launch planning.\par "
                r"[2026-04-15 09:05] Bob Example: I'll draft the update.\par "
                r"[2026-04-15 09:07] Alice Example: Great, thanks.\par}"
            ),
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("chat-thread.rtf")
        self.assertEqual(row["content_type"], "Chat")

        search_result = retriever_tools.search(self.root, "draft the update", None, None, None, 1, 20)
        result = search_result["results"][0]
        self.assertEqual(result["preview_rel_path"], "chat-thread.rtf")
        self.assertEqual(result["preview_targets"][0]["preview_type"], "native")
        self.assertGreaterEqual(len(result["preview_targets"]), 2)
        self.assertEqual(result["preview_targets"][1]["preview_type"], "html")
        preview_html = Path(result["preview_targets"][1]["abs_path"]).read_text(encoding="utf-8")
        self.assertIn("<title>Kickoff thread for launch planning.</title>", preview_html)
        self.assertIn("<h1>Kickoff thread for launch planning.</h1>", preview_html)
        self.assertIn("Alice Example", preview_html)
        self.assertIn("Bob Example", preview_html)

    def test_ingest_supports_chat_transcript_metadata_for_text_files(self) -> None:
        transcript_path = self.root / "chat-thread.txt"
        transcript_path.write_text(
            "\n".join(
                [
                    "[2026-04-15 09:00] Alice Example: Kickoff thread for launch planning.",
                    "[2026-04-15 09:05] Bob Example: I'll draft the update.",
                    "[2026-04-15 09:07] Alice Example: Great, thanks.",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("chat-thread.txt")
        self.assertEqual(row["content_type"], "Chat")
        self.assertIsNone(row["author"])
        self.assertEqual(row["participants"], "Alice Example, Bob Example")
        self.assertEqual(row["date_created"], "2026-04-15T09:00:00Z")
        self.assertEqual(row["date_modified"], "2026-04-15T09:07:00Z")
        self.assertEqual(row["title"], "Kickoff thread for launch planning.")
        self.assertIsNone(row["subject"])
        self.assertIsNone(row["recipients"])

        search_result = retriever_tools.search(self.root, "draft the update", None, None, None, 1, 20)
        result = search_result["results"][0]
        self.assertEqual(result["file_name"], "chat-thread.txt")
        self.assertTrue(result["preview_rel_path"].endswith(".html"))
        self.assertEqual(result["preview_targets"][0]["preview_type"], "html")
        preview_html = Path(result["preview_targets"][0]["abs_path"]).read_text(encoding="utf-8")
        self.assertIn("<title>Kickoff thread for launch planning.</title>", preview_html)
        self.assertIn("<h1>Kickoff thread for launch planning.</h1>", preview_html)
        self.assertIn("Alice Example", preview_html)
        self.assertIn("Bob Example", preview_html)
        self.assertIn("Kickoff thread for launch planning.", preview_html)
        self.assertIn("Full transcript", preview_html)

    def test_ingest_renders_slack_json_as_named_chat_messages(self) -> None:
        users_path = self.root / "users.json"
        users_path.write_text(
            json.dumps(
                [
                    {
                        "id": "U04SERGEY1",
                        "name": "sergey",
                        "profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                            "first_name": "Sergey",
                            "color": "9f69e7",
                        },
                    },
                    {
                        "id": "U04MAX0001",
                        "name": "maksim",
                        "profile": {
                            "real_name": "Maksim Faleev",
                            "display_name": "Maksim",
                            "first_name": "Maksim",
                            "color": "2eb67d",
                        },
                    },
                ],
                ensure_ascii=True,
            ),
            encoding="utf-8",
        )
        channel_dir = self.root / "general"
        channel_dir.mkdir()
        slack_day_path = channel_dir / "2022-12-16.json"
        slack_day_path.write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "<@U04MAX0001> can we sync?",
                        "user": "U04SERGEY1",
                        "ts": "1671235434.237949",
                    },
                    {
                        "type": "message",
                        "text": "Let's sync at 10:05 :partying_face::christmas_tree:",
                        "user": "U04MAX0001",
                        "ts": "1671235734.237949",
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("general/2022-12-16.json")
        self.assertEqual(row["content_type"], "Chat")
        self.assertIsNone(row["author"])
        self.assertEqual(row["participants"], "Sergey Demyanov, Maksim Faleev")
        self.assertEqual(row["date_created"], "2022-12-17T00:03:54Z")
        self.assertEqual(row["date_modified"], "2022-12-17T00:08:54Z")
        self.assertEqual(row["title"], "#general - Dec 16, 2022")

        search_result = retriever_tools.search(self.root, "sync", None, None, None, 1, 20)
        result = next(item for item in search_result["results"] if item["file_name"] == "2022-12-16.json")
        self.assertTrue(result["preview_rel_path"].endswith(".html"))
        self.assertEqual(result["preview_targets"][0]["preview_type"], "html")
        preview_html = Path(result["preview_targets"][0]["abs_path"]).read_text(encoding="utf-8")
        self.assertIn("Sergey Demyanov", preview_html)
        self.assertIn("Maksim Faleev", preview_html)
        self.assertIn("@Maksim can we sync?", preview_html)
        self.assertIn("sync at 10:05 🥳🎄", preview_html)
        self.assertIn("chat-avatar-svg", preview_html)
        self.assertIn('aria-label="Sergey Demyanov"', preview_html)
        self.assertIn("#general - Dec 16, 2022", preview_html)
        self.assertIn("Dec 17, 2022 12:03 AM UTC", preview_html)
        self.assertIn("[Dec 17, 2022 12:03 AM UTC]", preview_html)
        self.assertNotIn("<th>Author</th>", preview_html)
        self.assertNotIn("&quot;type&quot;", preview_html)

    def test_ingest_routes_slack_export_root_into_dedicated_dataset(self) -> None:
        export_root = self.root / "data" / "slack"
        export_root.mkdir(parents=True)
        (export_root / "users.json").write_text(
            json.dumps(
                [
                    {
                        "id": "U04SERGEY1",
                        "name": "sergey",
                        "profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                        },
                    },
                    {
                        "id": "U04MAX0001",
                        "name": "maksim",
                        "profile": {
                            "real_name": "Maksim Faleev",
                            "display_name": "Maksim",
                        },
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (export_root / "channels.json").write_text(
            json.dumps(
                [
                    {
                        "id": "C04GENERAL1",
                        "name": "general",
                        "is_channel": True,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (export_root / "canvases.json").write_text("[]\n", encoding="utf-8")
        channel_dir = export_root / "general"
        channel_dir.mkdir()
        (channel_dir / "2022-12-16.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "<@U04MAX0001> can we sync?",
                        "user": "U04SERGEY1",
                        "ts": "1671235434.237949",
                    },
                    {
                        "type": "message",
                        "text": "Let's sync at 10:05",
                        "user": "U04MAX0001",
                        "ts": "1671235734.237949",
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (self.root / "alpha.txt").write_text("ordinary workspace file\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["slack_exports_detected"], 1)
        self.assertEqual(ingest_result["slack_day_documents_scanned"], 1)
        self.assertEqual(ingest_result["slack_documents_created"], 1)

        day_row = self.fetch_document_row("data/slack/general/2022-12-16.json")
        self.assertEqual(day_row["source_kind"], retriever_tools.SLACK_EXPORT_SOURCE_KIND)
        self.assertIsNotNone(day_row["conversation_id"])
        self.assertIsNotNone(day_row["dataset_id"])
        self.assertEqual(day_row["title"], "#general - Dec 16, 2022")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            metadata_rows = connection.execute(
                """
                SELECT rel_path
                FROM documents
                WHERE rel_path IN (?, ?, ?)
                ORDER BY rel_path ASC
                """,
                (
                    "data/slack/users.json",
                    "data/slack/channels.json",
                    "data/slack/canvases.json",
                ),
            ).fetchall()
            conversation_row = connection.execute(
                """
                SELECT c.conversation_key, c.conversation_type, c.display_name
                FROM conversations c
                JOIN documents d ON d.conversation_id = c.id
                WHERE d.rel_path = ?
                """,
                ("data/slack/general/2022-12-16.json",),
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(metadata_rows, [])
        self.assertIsNotNone(conversation_row)
        self.assertEqual(conversation_row["conversation_key"], "C04GENERAL1")
        self.assertEqual(conversation_row["conversation_type"], "public_channel")
        self.assertEqual(conversation_row["display_name"], "#general")

        dataset_payload = retriever_tools.list_datasets(self.root)
        slack_dataset = next(
            item for item in dataset_payload["datasets"] if item["source_kind"] == retriever_tools.SLACK_EXPORT_SOURCE_KIND
        )
        filesystem_dataset = next(
            item for item in dataset_payload["datasets"] if item["source_kind"] == retriever_tools.FILESYSTEM_SOURCE_KIND
        )
        self.assertEqual(slack_dataset["dataset_locator"], "data/slack")
        self.assertEqual(slack_dataset["document_count"], 1)
        self.assertEqual(filesystem_dataset["document_count"], 1)

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_result["total_hits"], 2)
        day_result = next(item for item in browse_result["results"] if item["id"] == day_row["id"])
        self.assertEqual(day_result["dataset_name"], slack_dataset["dataset_name"])
        self.assertEqual(day_result["conversation_id"], day_row["conversation_id"])
        self.assertEqual(
            [target.get("label") for target in day_result["preview_targets"]],
            ["message", "conversation"],
        )
        self.assertEqual(
            day_result["preview_rel_path"],
            self.preview_target_by_label(day_result["preview_targets"], "message")["rel_path"],
        )
        day_preview_html = Path(str(day_result["preview_abs_path"]).split("#", 1)[0]).read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', day_preview_html)
        self.assertIn("sync", day_preview_html)

        search_result = retriever_tools.search(self.root, "sync", None, None, None, 1, 20)
        self.assertEqual(search_result["results"][0]["id"], day_row["id"])

    def test_slack_user_hints_store_email_handle_and_merge_by_default(self) -> None:
        export_root = self.root / "data" / "slack"
        export_root.mkdir(parents=True)
        (export_root / "users.json").write_text(
            json.dumps(
                [
                    {
                        "id": "U04SERGEY1",
                        "name": "sergey",
                        "profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                            "email": "sergey@example.com",
                        },
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (export_root / "channels.json").write_text(
            json.dumps(
                [
                    {
                        "id": "C04GENERAL1",
                        "name": "general",
                        "is_channel": True,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        channel_dir = export_root / "general"
        channel_dir.mkdir()
        (channel_dir / "2022-12-16.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "First Slack identity clue",
                        "user": "U04SERGEY1",
                        "user_profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                        },
                        "ts": "1671235434.237949",
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (channel_dir / "2022-12-17.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Second Slack identity clue",
                        "user": "U04SERGEY1",
                        "user_profile": {
                            "real_name": "Sergey D.",
                            "display_name": "SD",
                        },
                        "ts": "1671321834.237949",
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["slack_documents_created"], 2)

        day_one_row = self.fetch_document_row("data/slack/general/2022-12-16.json")
        day_two_row = self.fetch_document_row("data/slack/general/2022-12-17.json")
        self.assertEqual(day_one_row["participants"], "Sergey Demyanov <sergey@example.com>")
        self.assertEqual(day_two_row["participants"], "Sergey Demyanov <sergey@example.com>")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            participant_rows = connection.execute(
                """
                SELECT d.rel_path, de.entity_id, e.display_name, e.entity_origin
                FROM document_entities de
                JOIN documents d ON d.id = de.document_id
                JOIN entities e ON e.id = de.entity_id
                WHERE de.role = 'participant'
                  AND d.id IN (?, ?)
                ORDER BY d.rel_path ASC
                """,
                (day_one_row["id"], day_two_row["id"]),
            ).fetchall()
            hint_row = connection.execute(
                """
                SELECT entity_hints_json
                FROM document_occurrences
                WHERE document_id = ?
                """,
                (day_one_row["id"],),
            ).fetchone()
            external_identifier_rows = connection.execute(
                """
                SELECT entity_id, identifier_name, identifier_scope, normalized_value
                FROM entity_identifiers
                WHERE identifier_type = 'external_id'
                  AND identifier_name = 'slack_user_id'
                  AND identifier_scope = 'data/slack'
                  AND normalized_value = 'u04sergey1'
                ORDER BY entity_id ASC
                """
            ).fetchall()
            email_identifier_rows = connection.execute(
                """
                SELECT entity_id, normalized_value
                FROM entity_identifiers
                WHERE identifier_type = 'email'
                  AND normalized_value = 'sergey@example.com'
                ORDER BY entity_id ASC
                """
            ).fetchall()
            handle_identifier_rows = connection.execute(
                """
                SELECT entity_id, provider, provider_scope, normalized_value
                FROM entity_identifiers
                WHERE identifier_type = 'handle'
                  AND provider = 'slack'
                  AND provider_scope = 'data/slack'
                  AND normalized_value = 'sergey'
                ORDER BY entity_id ASC
                """
            ).fetchall()
            external_key_count = connection.execute(
                """
                SELECT COUNT(*) AS row_count
                FROM entity_resolution_keys
                WHERE key_type = 'external_id'
                  AND identifier_name = 'slack_user_id'
                  AND identifier_scope = 'data/slack'
                  AND normalized_value = 'u04sergey1'
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(len(participant_rows), 2)
        self.assertEqual(
            {row["entity_id"] for row in participant_rows},
            {participant_rows[0]["entity_id"]},
        )
        self.assertEqual({row["entity_origin"] for row in participant_rows}, {retriever_tools.ENTITY_ORIGIN_IDENTIFIED})
        self.assertIsNotNone(hint_row)
        hint_payload = json.loads(hint_row["entity_hints_json"])
        hint_identifiers = hint_payload["participants"][0]["identifiers"]
        self.assertEqual(
            {
                (
                    identifier["identifier_type"],
                    identifier.get("identifier_name"),
                    identifier.get("identifier_scope"),
                    identifier.get("provider"),
                    identifier.get("provider_scope"),
                    identifier["normalized_value"],
                )
                for identifier in hint_identifiers
            },
            {
                ("external_id", "slack_user_id", "data/slack", None, None, "u04sergey1"),
                ("email", None, None, None, None, "sergey@example.com"),
                ("handle", None, None, "slack", "data/slack", "sergey"),
            },
        )
        self.assertEqual(len(external_identifier_rows), 1)
        self.assertEqual(len(email_identifier_rows), 1)
        self.assertEqual(len(handle_identifier_rows), 1)
        self.assertEqual(
            {
                int(external_identifier_rows[0]["entity_id"]),
                int(email_identifier_rows[0]["entity_id"]),
                int(handle_identifier_rows[0]["entity_id"]),
            },
            {int(participant_rows[0]["entity_id"])},
        )
        self.assertEqual(int(external_key_count["row_count"] or 0), 1)

    def test_slack_profile_email_merges_with_existing_manual_entity_on_ingest(self) -> None:
        export_root = self.root / "data" / "slack"
        export_root.mkdir(parents=True)
        (export_root / "users.json").write_text(
            json.dumps(
                [
                    {
                        "id": "U04SERGEY1",
                        "name": "sergey",
                        "profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                            "email": "sergey@example.com",
                        },
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (export_root / "channels.json").write_text(
            json.dumps(
                [
                    {
                        "id": "C04GENERAL1",
                        "name": "general",
                        "is_channel": True,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        channel_dir = export_root / "general"
        channel_dir.mkdir()
        (channel_dir / "2022-12-16.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Slack profile should attach to the known person.",
                        "user": "U04SERGEY1",
                        "ts": "1671235434.237949",
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        exit_code, create_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--entity-type",
            "person",
            "--display-name",
            "Known Sergey",
            "--email",
            "sergey@example.com",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(create_payload)
        manual_entity_id = int(create_payload["entity_id"])

        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)
        day_row = self.fetch_document_row("data/slack/general/2022-12-16.json")
        self.assertEqual(day_row["participants"], "Known Sergey <sergey@example.com>")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            participant_row = connection.execute(
                """
                SELECT de.entity_id, de.assignment_mode
                FROM document_entities de
                WHERE de.document_id = ?
                  AND de.role = 'participant'
                """,
                (day_row["id"],),
            ).fetchone()
            slack_identifier_rows = connection.execute(
                """
                SELECT identifier_type, identifier_name, identifier_scope, provider, provider_scope, normalized_value
                FROM entity_identifiers
                WHERE entity_id = ?
                  AND (
                    identifier_type = 'email'
                    OR identifier_type = 'handle'
                    OR identifier_name = 'slack_user_id'
                  )
                ORDER BY identifier_type ASC, normalized_value ASC
                """,
                (manual_entity_id,),
            ).fetchall()
        finally:
            connection.close()

        self.assertIsNotNone(participant_row)
        self.assertEqual(int(participant_row["entity_id"]), manual_entity_id)
        self.assertEqual(participant_row["assignment_mode"], "auto")
        self.assertEqual(
            {
                (
                    row["identifier_type"],
                    row["identifier_name"],
                    row["identifier_scope"],
                    row["provider"],
                    row["provider_scope"],
                    row["normalized_value"],
                )
                for row in slack_identifier_rows
            },
            {
                ("email", None, None, None, None, "sergey@example.com"),
                ("external_id", "slack_user_id", "data/slack", None, None, "u04sergey1"),
                ("handle", None, None, "slack", "data/slack", "sergey"),
            },
        )

    def test_ingest_creates_slack_reply_thread_child_documents(self) -> None:
        export_root = self.root / "data" / "slack"
        export_root.mkdir(parents=True)
        (export_root / "users.json").write_text(
            json.dumps(
                [
                    {
                        "id": "U04SERGEY1",
                        "name": "sergey",
                        "profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                        },
                    },
                    {
                        "id": "U04MAX0001",
                        "name": "maksim",
                        "profile": {
                            "real_name": "Maksim Faleev",
                            "display_name": "Maksim",
                        },
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (export_root / "channels.json").write_text(
            json.dumps(
                [
                    {
                        "id": "C04GENERAL1",
                        "name": "general",
                        "is_channel": True,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        channel_dir = export_root / "general"
        channel_dir.mkdir()
        thread_ts = "1671235434.237949"
        (channel_dir / "2022-12-16.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Kickoff thread",
                        "user": "U04SERGEY1",
                        "ts": thread_ts,
                        "thread_ts": thread_ts,
                        "reply_count": 1,
                    },
                    {
                        "type": "message",
                        "text": "Standalone channel update",
                        "user": "U04MAX0001",
                        "ts": "1671235834.237949",
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (channel_dir / "2022-12-17.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Following up on kickoff",
                        "user": "U04MAX0001",
                        "ts": "1671321834.237949",
                        "thread_ts": thread_ts,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 3)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["slack_exports_detected"], 1)
        self.assertEqual(ingest_result["slack_day_documents_scanned"], 2)
        self.assertEqual(ingest_result["slack_documents_created"], 3)

        day_one_row = self.fetch_document_row("data/slack/general/2022-12-16.json")
        day_two_row = self.fetch_document_row("data/slack/general/2022-12-17.json")
        child_rel_path = retriever_tools.slack_reply_thread_rel_path("C04GENERAL1", thread_ts)
        child_row = self.fetch_document_row(child_rel_path)

        self.assertEqual(child_row["parent_document_id"], day_one_row["id"])
        self.assertEqual(child_row["child_document_kind"], retriever_tools.CHILD_DOCUMENT_KIND_REPLY_THREAD)
        self.assertEqual(child_row["source_kind"], retriever_tools.SLACK_EXPORT_SOURCE_KIND)
        self.assertEqual(child_row["source_rel_path"], "data/slack/general/2022-12-16.json")
        self.assertEqual(child_row["source_item_id"], thread_ts)
        self.assertEqual(child_row["root_message_key"], f"C04GENERAL1:{thread_ts}")
        self.assertEqual(child_row["conversation_id"], day_one_row["conversation_id"])
        self.assertEqual(day_two_row["conversation_id"], day_one_row["conversation_id"])
        self.assertEqual(day_two_row["text_status"], "empty")

        reply_search = retriever_tools.search(self.root, "Following up on kickoff", None, None, None, 1, 20)
        self.assertEqual(reply_search["total_hits"], 1)
        self.assertEqual(reply_search["results"][0]["id"], child_row["id"])

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        day_one_result = next(item for item in browse_result["results"] if item["id"] == day_one_row["id"])
        day_two_result = next(item for item in browse_result["results"] if item["id"] == day_two_row["id"])
        self.assertEqual(day_one_result["attachment_count"], 0)
        self.assertEqual(day_one_result["child_document_count"], 1)
        self.assertEqual(day_one_result["child_documents"][0]["id"], child_row["id"])
        self.assertEqual(day_one_result["child_documents"][0]["child_document_kind"], retriever_tools.CHILD_DOCUMENT_KIND_REPLY_THREAD)
        self.assertEqual(day_two_result["child_document_count"], 0)
        self.assertEqual(
            [target.get("label") for target in day_one_result["preview_targets"]],
            ["message", "conversation"],
        )
        self.assertEqual(
            [target.get("label") for target in reply_search["results"][0]["preview_targets"]],
            ["message", "conversation"],
        )
        self.assertEqual(
            day_one_result["preview_rel_path"],
            self.preview_target_by_label(day_one_result["preview_targets"], "message")["rel_path"],
        )
        self.assertEqual(
            reply_search["results"][0]["preview_rel_path"],
            self.preview_target_by_label(reply_search["results"][0]["preview_targets"], "message")["rel_path"],
        )
        self.assertEqual(
            self.preview_target_file_path(
                self.preview_target_by_label(day_one_result["preview_targets"], "conversation")
            ),
            self.preview_target_file_path(
                self.preview_target_by_label(reply_search["results"][0]["preview_targets"], "conversation")
            ),
        )
        slack_conversation_preview = self.preview_target_file_path(
            self.preview_target_by_label(day_one_result["preview_targets"], "conversation")
        )
        reply_conversation_preview = self.preview_target_file_path(
            self.preview_target_by_label(reply_search["results"][0]["preview_targets"], "conversation")
        )
        slack_entry_path = self.preview_target_file_path(day_one_result["preview_targets"][0])
        slack_entry_html = slack_entry_path.read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', slack_entry_html)
        self.assertIn("Kickoff thread", slack_entry_html)
        self.assertNotIn("Contents", slack_entry_html)
        self.assertNotIn('class="conversation-nav-segment"', slack_entry_html)
        self.assertNotIn("window.location.replace", slack_entry_html)
        slack_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(day_one_result["preview_targets"], "conversation")
        ).read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', slack_preview_html)
        self.assertIn("Kickoff thread", slack_preview_html)
        self.assertIn("Following up on kickoff", slack_preview_html)
        self.assertTrue(slack_conversation_preview.exists())
        self.assertTrue(reply_conversation_preview.exists())

        dataset_payload = retriever_tools.list_datasets(self.root)
        slack_dataset = next(
            item for item in dataset_payload["datasets"] if item["source_kind"] == retriever_tools.SLACK_EXPORT_SOURCE_KIND
        )
        self.assertEqual(slack_dataset["document_count"], 3)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            source_parts = connection.execute(
                """
                SELECT part_kind, rel_source_path, ordinal
                FROM document_source_parts
                WHERE document_id = ?
                ORDER BY ordinal ASC, id ASC
                """,
                (child_row["id"],),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(
            [(row["part_kind"], row["rel_source_path"]) for row in source_parts],
            [
                ("slack_thread_root_day", "data/slack/general/2022-12-16.json"),
                ("slack_thread_reply_day", "data/slack/general/2022-12-17.json"),
            ],
        )

    def test_ingest_slack_rerun_adds_late_day_file_to_existing_conversation(self) -> None:
        export_root = self.root / "data" / "slack"
        export_root.mkdir(parents=True)
        (export_root / "users.json").write_text(
            json.dumps(
                [
                    {
                        "id": "U04SERGEY1",
                        "name": "sergey",
                        "profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                        },
                    },
                    {
                        "id": "U04MAX0001",
                        "name": "maksim",
                        "profile": {
                            "real_name": "Maksim Faleev",
                            "display_name": "Maksim",
                        },
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (export_root / "channels.json").write_text(
            json.dumps(
                [
                    {
                        "id": "C04GENERAL1",
                        "name": "general",
                        "is_channel": True,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        channel_dir = export_root / "general"
        channel_dir.mkdir()
        thread_ts = "1671235434.237949"
        (channel_dir / "2022-12-16.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Kickoff thread",
                        "user": "U04SERGEY1",
                        "ts": thread_ts,
                        "thread_ts": thread_ts,
                        "reply_count": 1,
                    },
                    {
                        "type": "message",
                        "text": "Standalone channel update",
                        "user": "U04MAX0001",
                        "ts": "1671235834.237949",
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        first_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_result["new"], 2)
        self.assertEqual(first_result["failed"], 0)
        self.assertEqual(first_result["slack_documents_created"], 2)

        day_one_row = self.fetch_document_row("data/slack/general/2022-12-16.json")
        child_rel_path = retriever_tools.slack_reply_thread_rel_path("C04GENERAL1", thread_ts)
        child_row = self.fetch_document_row(child_rel_path)
        first_child_id = child_row["id"]
        first_child_control_number = child_row["control_number"]

        (channel_dir / "2022-12-17.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Following up on kickoff",
                        "user": "U04MAX0001",
                        "ts": "1671321834.237949",
                        "thread_ts": thread_ts,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        second_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_result["new"], 1)
        self.assertEqual(second_result["updated"], 2)
        self.assertEqual(second_result["failed"], 0)
        self.assertEqual(second_result["slack_exports_detected"], 1)
        self.assertEqual(second_result["slack_day_documents_scanned"], 2)
        self.assertEqual(second_result["slack_documents_created"], 1)
        self.assertEqual(second_result["slack_documents_updated"], 2)

        updated_day_one_row = self.fetch_document_row("data/slack/general/2022-12-16.json")
        day_two_row = self.fetch_document_row("data/slack/general/2022-12-17.json")
        updated_child_row = self.fetch_document_row(child_rel_path)

        self.assertEqual(updated_day_one_row["conversation_id"], day_one_row["conversation_id"])
        self.assertEqual(day_two_row["conversation_id"], day_one_row["conversation_id"])
        self.assertEqual(updated_child_row["conversation_id"], day_one_row["conversation_id"])
        self.assertEqual(updated_child_row["id"], first_child_id)
        self.assertEqual(updated_child_row["control_number"], first_child_control_number)
        self.assertEqual(updated_child_row["parent_document_id"], updated_day_one_row["id"])
        self.assertEqual(day_two_row["text_status"], "empty")

        reply_search = retriever_tools.search(self.root, "Following up on kickoff", None, None, None, 1, 20)
        self.assertEqual(reply_search["total_hits"], 1)
        self.assertEqual(reply_search["results"][0]["id"], updated_child_row["id"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            source_parts = connection.execute(
                """
                SELECT part_kind, rel_source_path, ordinal
                FROM document_source_parts
                WHERE document_id = ?
                ORDER BY ordinal ASC, id ASC
                """,
                (updated_child_row["id"],),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(
            [(row["part_kind"], row["rel_source_path"]) for row in source_parts],
            [
                ("slack_thread_root_day", "data/slack/general/2022-12-16.json"),
                ("slack_thread_reply_day", "data/slack/general/2022-12-17.json"),
            ],
        )

    def test_ingest_groups_loose_eml_messages_into_one_conversation(self) -> None:
        root_path = self.root / "root.eml"
        reply_path = self.root / "reply.eml"
        self.write_email_message(
            root_path,
            subject="Status Update",
            body_text="Root message body",
            message_id="<root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            reply_path,
            subject="Re: Status Update",
            body_text="Reply message body",
            body_html=(
                '<div class="reply-rich-email">'
                "<p><strong>Reply message body</strong></p>"
                '<table role="presentation"><tr><td>Rendered from the original HTML body.</td></tr></table>'
                "</div>"
            ),
            message_id="<reply@example.com>",
            in_reply_to="<root@example.com>",
            references="<root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["email_conversations"], 1)

        root_row = self.fetch_document_row("root.eml")
        reply_row = self.fetch_document_row("reply.eml")
        self.assertIsNotNone(root_row["conversation_id"])
        self.assertEqual(root_row["conversation_id"], reply_row["conversation_id"])
        self.assertEqual(root_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)
        self.assertEqual(reply_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            signal_rows = connection.execute(
                """
                SELECT document_id, message_id, in_reply_to, references_json, normalized_subject
                FROM document_email_threading
                ORDER BY document_id ASC
                """
            ).fetchall()
            conversation_row = connection.execute(
                """
                SELECT source_kind, source_locator, conversation_type, display_name
                FROM conversations
                WHERE id = ?
                """,
                (root_row["conversation_id"],),
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(len(signal_rows), 2)
        signals_by_message_id = {row["message_id"]: row for row in signal_rows}
        self.assertEqual(set(signals_by_message_id), {"root@example.com", "reply@example.com"})
        self.assertEqual(signals_by_message_id["root@example.com"]["in_reply_to"], None)
        self.assertEqual(json.loads(signals_by_message_id["root@example.com"]["references_json"]), [])
        self.assertEqual(signals_by_message_id["root@example.com"]["normalized_subject"], "status update")
        self.assertEqual(signals_by_message_id["reply@example.com"]["in_reply_to"], "root@example.com")
        self.assertEqual(json.loads(signals_by_message_id["reply@example.com"]["references_json"]), ["root@example.com"])
        self.assertEqual(signals_by_message_id["reply@example.com"]["normalized_subject"], "status update")
        self.assertEqual(conversation_row["source_kind"], retriever_tools.EMAIL_CONVERSATION_SOURCE_KIND)
        self.assertEqual(conversation_row["source_locator"], ".")
        self.assertEqual(conversation_row["conversation_type"], "email")
        self.assertEqual(conversation_row["display_name"], "Status Update")

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        root_result = next(item for item in browse_result["results"] if item["id"] == root_row["id"])
        search_result = retriever_tools.search(self.root, "Reply message body", None, None, None, 1, 20)
        reply_result = next(item for item in search_result["results"] if item["id"] == reply_row["id"])
        self.assertEqual(
            reply_result["preview_rel_path"],
            self.preview_target_by_label(reply_result["preview_targets"], "message")["rel_path"],
        )
        self.assertIsNone(reply_result["preview_target_fragment"])
        self.assertEqual(len(reply_result["preview_targets"]), 2)
        self.assertEqual(reply_result["preview_targets"][0]["label"], "message")
        self.assertEqual(reply_result["preview_targets"][1]["label"], "segment")
        self.assertFalse(any(target.get("label") == "entry" for target in reply_result["preview_targets"]))
        self.assertFalse(any(target.get("label") == "contents" for target in reply_result["preview_targets"]))
        self.assertEqual(self.preview_target_by_label(reply_result["preview_targets"], "segment")["target_fragment"], f"doc-{reply_row['id']}")
        self.assertEqual(self.preview_target_by_label(reply_result["preview_targets"], "message")["label"], "message")
        self.assertEqual(
            self.preview_target_file_path(self.preview_target_by_label(root_result["preview_targets"], "segment")),
            self.preview_target_file_path(self.preview_target_by_label(reply_result["preview_targets"], "segment")),
        )
        root_message_preview_path = self.preview_target_file_path(
            self.preview_target_by_label(root_result["preview_targets"], "message")
        )
        root_message_preview_html = root_message_preview_path.read_text(encoding="utf-8")
        message_preview_path = self.preview_target_file_path(
            self.preview_target_by_label(reply_result["preview_targets"], "message")
        )
        message_preview_html = message_preview_path.read_text(encoding="utf-8")
        segment_preview_path = self.preview_target_file_path(
            self.preview_target_by_label(reply_result["preview_targets"], "segment")
        )
        expected_thread_href = retriever_tools.urllib_request.pathname2url(
            os.path.relpath(str(segment_preview_path), start=str(message_preview_path.parent))
        )
        expected_entry_path = (
            self.paths["state_dir"]
            / retriever_tools.conversation_preview_entry_rel_path(root_row["conversation_id"], reply_row["id"])
        )
        conversation_preview_dir = (
            self.paths["state_dir"]
            / retriever_tools.conversation_preview_base_path(root_row["conversation_id"])
        )
        self.assertFalse(expected_entry_path.exists())
        self.assertTrue((conversation_preview_dir / "conversation.html").exists())
        self.assertFalse((conversation_preview_dir / "index.html").exists())
        self.assertEqual(list(conversation_preview_dir.glob("segment-*.html")), [])
        self.assertIn("<title>Status Update (2/2 in thread)</title>", message_preview_html)
        self.assertIn('class="gmail-thread-title-link"', message_preview_html)
        self.assertIn('class="gmail-thread-title-meta">(2/2 in thread)</span>', message_preview_html)
        self.assertIn(f'href="{expected_thread_href}"', message_preview_html)
        self.assertIn("2 messages", message_preview_html)
        self.assertIn("Created Apr 14, 2026 10:00 AM UTC", message_preview_html)
        self.assertIn("Last modified Apr 14, 2026 11:00 AM UTC", message_preview_html)
        self.assertIn("Viewing message 2 of 2", message_preview_html)
        self.assertNotIn("data-retriever-email-body-source", message_preview_html)
        self.assertNotIn("data-retriever-email-body-source", root_message_preview_html)
        self.assertIn("<title>Status Update (1/2 in thread)</title>", root_message_preview_html)
        self.assertIn("Viewing message 1 of 2", root_message_preview_html)
        self.assertIn("Root message body", root_message_preview_html)
        self.assertNotIn("Reply message body", root_message_preview_html)
        self.assertEqual(root_message_preview_html.count('class="gmail-message-card'), 1)
        self.assertIn("Root message body", message_preview_html)
        self.assertIn("Reply message body", message_preview_html)
        self.assertEqual(message_preview_html.count('class="gmail-message-card'), 2)
        self.assertIn('class="gmail-message-card gmail-message-card--selected"', message_preview_html)
        self.assertIn('class="reply-rich-email"', message_preview_html)
        segment_html = segment_preview_path.read_text(encoding="utf-8")
        self.assertIn(f'id="doc-{root_row["id"]}"', segment_html)
        self.assertIn(f'id="doc-{reply_row["id"]}"', segment_html)
        self.assertIn("Root message body", segment_html)
        self.assertIn("Reply message body", segment_html)
        self.assertIn('class="reply-rich-email"', segment_html)
        self.assertIn('class="gmail-thread-messages"', segment_html)
        self.assertIn('class="gmail-message-card"', segment_html)
        self.assertIn("to Bob Example &lt;bob@example.com&gt;, Carol Example &lt;carol@example.com&gt;", segment_html)
        self.assertNotIn("<dt>Author</dt>", segment_html)
        self.assertNotIn("<dt>Recipients</dt>", segment_html)
        self.assertNotIn("&lt;div class=&quot;reply-rich-email&quot;&gt;", segment_html)
        self.assertTrue(segment_preview_path.name == "conversation.html")
        self.assertTrue(
            self.preview_target_file_path(self.preview_target_by_label(root_result["preview_targets"], "message")).exists()
        )
        self.assertTrue(
            self.preview_target_file_path(self.preview_target_by_label(reply_result["preview_targets"], "message")).exists()
        )

    def test_email_message_preview_renders_thread_up_to_selected_message_only(self) -> None:
        self.write_email_message(
            self.root / "root.eml",
            subject="Design Review",
            body_text="Root thread message",
            message_id="<root@example.com>",
            date_created="Tue, 14 Apr 2026 09:00:00 +0000",
        )
        self.write_email_message(
            self.root / "middle.eml",
            subject="Re: Design Review",
            body_text="Middle thread message",
            message_id="<middle@example.com>",
            in_reply_to="<root@example.com>",
            references="<root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            self.root / "final.eml",
            subject="Re: Design Review",
            body_text="Final thread message",
            message_id="<final@example.com>",
            in_reply_to="<middle@example.com>",
            references="<root@example.com> <middle@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["email_conversations"], 1)

        search_result = retriever_tools.search(self.root, "Middle thread message", None, None, None, 1, 20)
        middle_result = next(item for item in search_result["results"] if item["rel_path"] == "middle.eml")
        middle_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(middle_result["preview_targets"], "message")
        ).read_text(encoding="utf-8")

        self.assertIn("<title>Design Review (2/3 in thread)</title>", middle_preview_html)
        self.assertIn("Viewing message 2 of 3", middle_preview_html)
        self.assertIn("Root thread message", middle_preview_html)
        self.assertIn("Middle thread message", middle_preview_html)
        self.assertNotIn("Final thread message", middle_preview_html)
        self.assertEqual(middle_preview_html.count('class="gmail-message-card'), 2)

    def test_thread_preview_omits_repeated_quoted_history_from_rich_reply_bodies(self) -> None:
        root_path = self.root / "root.eml"
        reply_path = self.root / "reply.eml"
        self.write_email_message(
            root_path,
            subject="Status Update",
            body_text="Root message body",
            message_id="<root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            reply_path,
            subject="Re: Status Update",
            body_text=(
                "Fresh reply body\n\n"
                "On Tue, 14 Apr 2026 at 10:00 AM Root Sender <root@example.com> wrote:\n"
                "> Root message body"
            ),
            body_html=(
                '<div class="reply-rich-email">'
                "<p><strong>Fresh reply body</strong></p>"
                '<div class="gmail_quote">'
                '<div class="gmail_attr">On Tue, 14 Apr 2026 at 10:00 AM Root Sender &lt;root@example.com&gt; wrote:</div>'
                "<blockquote>Root message body</blockquote>"
                "</div>"
                "</div>"
            ),
            message_id="<reply@example.com>",
            in_reply_to="<root@example.com>",
            references="<root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        search_result = retriever_tools.search(self.root, "Fresh reply body", None, None, None, 1, 20)
        reply_result = next(item for item in search_result["results"] if item["rel_path"] == "reply.eml")

        message_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(reply_result["preview_targets"], "message")
        ).read_text(encoding="utf-8")
        segment_html = self.preview_target_file_path(
            self.preview_target_by_label(reply_result["preview_targets"], "segment")
        ).read_text(encoding="utf-8")

        self.assertIn("Fresh reply body", message_preview_html)
        self.assertIn("Root message body", message_preview_html)
        self.assertEqual(message_preview_html.count('class="gmail-message-card'), 2)
        self.assertIn("Fresh reply body", segment_html)
        self.assertEqual(segment_html.count("Root message body"), 1)
        self.assertNotIn('class="gmail_quote"', segment_html)

    def test_refresh_conversation_previews_backfills_missing_email_message_preview(self) -> None:
        root_path = self.root / "root.eml"
        reply_path = self.root / "reply.eml"
        self.write_email_message(
            root_path,
            subject="Status Update",
            body_text="Root message body",
            message_id="<root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            reply_path,
            subject="Re: Status Update",
            body_text="Reply message body",
            message_id="<reply@example.com>",
            in_reply_to="<root@example.com>",
            references="<root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        root_row = self.fetch_document_row("root.eml")
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            message_preview_row = connection.execute(
                """
                SELECT rel_preview_path
                FROM document_previews
                WHERE document_id = ? AND label = 'message'
                ORDER BY ordinal ASC, id ASC
                LIMIT 1
                """,
                (root_row["id"],),
            ).fetchone()
            self.assertIsNotNone(message_preview_row)
            legacy_message_preview_path = self.paths["state_dir"] / str(message_preview_row["rel_preview_path"])
            if legacy_message_preview_path.exists():
                legacy_message_preview_path.unlink()
            connection.execute(
                "DELETE FROM document_previews WHERE document_id = ? AND label = 'message'",
                (root_row["id"],),
            )
            connection.commit()

            legacy_search = retriever_tools.search(self.root, "Root message body", None, None, None, 1, 20)
            legacy_result = next(item for item in legacy_search["results"] if item["id"] == root_row["id"])
            self.assertEqual(
                legacy_result["preview_rel_path"],
                self.preview_target_by_label(legacy_result["preview_targets"], "segment")["rel_path"],
            )
            self.assertFalse(any(target.get("label") == "message" for target in legacy_result["preview_targets"]))

            refreshed = retriever_tools.refresh_conversation_previews(
                connection,
                self.paths,
                [root_row["conversation_id"]],
            )
            connection.commit()
        finally:
            connection.close()

        self.assertEqual(refreshed, 1)
        refreshed_search = retriever_tools.search(self.root, "Root message body", None, None, None, 1, 20)
        refreshed_result = next(item for item in refreshed_search["results"] if item["id"] == root_row["id"])
        self.assertEqual(
            refreshed_result["preview_rel_path"],
            self.preview_target_by_label(refreshed_result["preview_targets"], "message")["rel_path"],
        )
        self.assertEqual(refreshed_result["preview_targets"][0]["label"], "message")
        refreshed_message_preview_path = self.preview_target_file_path(
            self.preview_target_by_label(refreshed_result["preview_targets"], "message")
        )
        self.assertTrue(refreshed_message_preview_path.exists())
        refreshed_message_preview_html = refreshed_message_preview_path.read_text(encoding="utf-8")
        self.assertIn("Root message body", refreshed_message_preview_html)
        self.assertIn('class="gmail-thread-title-link"', refreshed_message_preview_html)

    def test_rebuild_conversations_reassigns_refreshes_and_prunes_empty_preview_dirs(self) -> None:
        root_path = self.root / "root.eml"
        reply_path = self.root / "reply.eml"
        self.write_email_message(
            root_path,
            subject="Rebuild Conversation",
            body_text="Root rebuild body",
            message_id="<rebuild-root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            reply_path,
            subject="Re: Rebuild Conversation",
            body_text="Reply rebuild body",
            message_id="<rebuild-reply@example.com>",
            in_reply_to="<rebuild-root@example.com>",
            references="<rebuild-root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        stale_dir = self.paths["state_dir"] / "previews" / "conversations" / "conversation-99999999"
        stale_dir.mkdir(parents=True)
        exit_code, payload, _, _ = self.run_cli("rebuild-conversations", str(self.root), "--batch-size", "1")

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["target_conversations"], 1)
        self.assertEqual(payload["refreshed_conversations"], 1)
        self.assertGreaterEqual(payload["empty_conversation_preview_dirs_pruned"], 1)
        self.assertFalse(stale_dir.exists())

    def test_single_large_email_conversation_skips_redundant_conversation_artifacts(self) -> None:
        single_path = self.root / "single.eml"
        huge_body = "A" * (retriever_tools.CONVERSATION_PREVIEW_MAX_CHARS + 1024)
        self.write_email_message(
            single_path,
            subject="Large Single Message",
            body_text=huge_body,
            message_id="<single@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["email_conversations"], 1)

        single_row = self.fetch_document_row("single.eml")
        search_result = retriever_tools.search(self.root, "Large Single Message", None, None, None, 1, 20)
        single_result = next(item for item in search_result["results"] if item["id"] == single_row["id"])

        self.assertEqual(single_result["preview_rel_path"], single_result["preview_targets"][0]["rel_path"])
        self.assertEqual([target.get("label") for target in single_result["preview_targets"]], ["message"])

        conversation_preview_dir = (
            self.paths["state_dir"]
            / retriever_tools.conversation_preview_base_path(single_row["conversation_id"])
        )
        self.assertFalse((conversation_preview_dir / "conversation.html").exists())
        self.assertFalse((conversation_preview_dir / "index.html").exists())
        self.assertEqual(list(conversation_preview_dir.glob("segment-*.html")), [])

        conversations_payload = retriever_tools.run_slash_command(self.root, "/conversations")
        self.assertEqual(conversations_payload["total_hits"], 1)
        conversation_result = conversations_payload["results"][0]
        self.assertFalse(conversation_result["preview_rel_path"].endswith("/conversation.html"))
        self.assertTrue(Path(str(conversation_result["preview_abs_path"]).split("#", 1)[0]).exists())

    def test_ingest_groups_loose_eml_messages_without_thread_headers_using_heuristics(self) -> None:
        first_path = self.root / "first.eml"
        second_path = self.root / "second.eml"
        self.write_email_message(
            first_path,
            subject="Project Sync",
            body_text="First loose message",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            cc=None,
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
            message_id=None,
            in_reply_to=None,
            references=None,
            conversation_index=None,
            conversation_topic=None,
        )
        self.write_email_message(
            second_path,
            subject="Re: Project Sync",
            body_text="Second loose message",
            author="Bob Example <bob@example.com>",
            recipients="Alice Example <alice@example.com>",
            cc=None,
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
            message_id=None,
            in_reply_to=None,
            references=None,
            conversation_index=None,
            conversation_topic=None,
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["email_conversations"], 1)

        first_row = self.fetch_document_row("first.eml")
        second_row = self.fetch_document_row("second.eml")
        self.assertIsNotNone(first_row["conversation_id"])
        self.assertEqual(first_row["conversation_id"], second_row["conversation_id"])

    def test_manual_conversation_assignment_mode_prevents_root_overwrite_and_retargets_auto_reply(self) -> None:
        root_path = self.root / "root.eml"
        reply_path = self.root / "reply.eml"
        self.write_email_message(
            root_path,
            subject="Deal Review",
            body_text="Root message body",
            message_id="<deal-root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            reply_path,
            subject="Re: Deal Review",
            body_text="Reply message body",
            message_id="<deal-reply@example.com>",
            in_reply_to="<deal-root@example.com>",
            references="<deal-root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["failed"], 0)

        root_row = self.fetch_document_row("root.eml")
        reply_row = self.fetch_document_row("reply.eml")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute("BEGIN")
            manual_conversation_id = retriever_tools.upsert_conversation_row(
                connection,
                source_kind=retriever_tools.EMAIL_CONVERSATION_SOURCE_KIND,
                source_locator=".",
                conversation_key="manual:deal-review",
                conversation_type="email",
                display_name="Deal Review (Manual)",
            )
            connection.execute(
                """
                UPDATE documents
                SET conversation_id = ?, conversation_assignment_mode = ?
                WHERE id = ?
                """,
                (
                    manual_conversation_id,
                    retriever_tools.CONVERSATION_ASSIGNMENT_MODE_MANUAL,
                    root_row["id"],
                ),
            )
            connection.commit()
        finally:
            connection.close()

        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_ingest["failed"], 0)

        updated_root = self.fetch_document_row("root.eml")
        updated_reply = self.fetch_document_row("reply.eml")
        self.assertEqual(updated_root["conversation_id"], manual_conversation_id)
        self.assertEqual(
            updated_root["conversation_assignment_mode"],
            retriever_tools.CONVERSATION_ASSIGNMENT_MODE_MANUAL,
        )
        self.assertEqual(updated_reply["conversation_id"], manual_conversation_id)
        self.assertEqual(
            updated_reply["conversation_assignment_mode"],
            retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO,
        )

    def test_split_and_clear_conversation_assignment_cli_for_loose_eml(self) -> None:
        root_path = self.root / "root.eml"
        reply_path = self.root / "reply.eml"
        self.write_email_message(
            root_path,
            subject="Status Update",
            body_text="Root message body",
            message_id="<root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            reply_path,
            subject="Re: Status Update",
            body_text="Reply message body",
            message_id="<reply@example.com>",
            in_reply_to="<root@example.com>",
            references="<root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        root_row = self.fetch_document_row("root.eml")
        reply_row = self.fetch_document_row("reply.eml")
        self.assertEqual(root_row["conversation_id"], reply_row["conversation_id"])

        split_exit, split_payload, _, _ = self.run_cli(
            "split-from-conversation",
            str(self.root),
            "--doc-id",
            str(reply_row["id"]),
        )
        self.assertEqual(split_exit, 0)
        self.assertIsNotNone(split_payload)
        self.assertEqual(split_payload["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_MANUAL)

        split_root_row = self.fetch_document_row("root.eml")
        split_reply_row = self.fetch_document_row("reply.eml")
        self.assertNotEqual(split_root_row["conversation_id"], split_reply_row["conversation_id"])
        self.assertEqual(split_root_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)
        self.assertEqual(split_reply_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_MANUAL)

        clear_exit, clear_payload, _, _ = self.run_cli(
            "clear-conversation-assignment",
            str(self.root),
            "--doc-id",
            str(reply_row["id"]),
        )
        self.assertEqual(clear_exit, 0)
        self.assertIsNotNone(clear_payload)
        self.assertEqual(clear_payload["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)

        cleared_root_row = self.fetch_document_row("root.eml")
        cleared_reply_row = self.fetch_document_row("reply.eml")
        self.assertEqual(cleared_root_row["conversation_id"], cleared_reply_row["conversation_id"])
        self.assertEqual(cleared_reply_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)

    def test_merge_into_conversation_cli_retargets_auto_email_replies(self) -> None:
        root_path = self.root / "root.eml"
        reply_path = self.root / "reply.eml"
        target_path = self.root / "target.eml"
        self.write_email_message(
            root_path,
            subject="Deal Review",
            body_text="Root message body",
            message_id="<deal-root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            reply_path,
            subject="Re: Deal Review",
            body_text="Reply message body",
            message_id="<deal-reply@example.com>",
            in_reply_to="<deal-root@example.com>",
            references="<deal-root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )
        self.write_email_message(
            target_path,
            subject="Separate Matter",
            body_text="Target message body",
            message_id="<target@example.com>",
            date_created="Tue, 14 Apr 2026 12:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        root_row = self.fetch_document_row("root.eml")
        reply_row = self.fetch_document_row("reply.eml")
        target_row = self.fetch_document_row("target.eml")
        self.assertNotEqual(root_row["conversation_id"], target_row["conversation_id"])
        self.assertEqual(root_row["conversation_id"], reply_row["conversation_id"])

        merge_exit, merge_payload, _, _ = self.run_cli(
            "merge-into-conversation",
            str(self.root),
            "--doc-id",
            str(root_row["id"]),
            "--target-doc-id",
            str(target_row["id"]),
        )
        self.assertEqual(merge_exit, 0)
        self.assertIsNotNone(merge_payload)
        self.assertEqual(merge_payload["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_MANUAL)

        updated_root = self.fetch_document_row("root.eml")
        updated_reply = self.fetch_document_row("reply.eml")
        updated_target = self.fetch_document_row("target.eml")
        self.assertEqual(updated_root["conversation_id"], updated_target["conversation_id"])
        self.assertEqual(updated_reply["conversation_id"], updated_target["conversation_id"])
        self.assertEqual(updated_root["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_MANUAL)
        self.assertEqual(updated_reply["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)

    def test_ingest_mbox_groups_threaded_messages_into_one_conversation(self) -> None:
        root_message = self.build_fake_mbox_message(
            subject="Budget Review",
            body_text="Budget root body",
            message_id="<mbox-budget-root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        reply_message = self.build_fake_mbox_message(
            subject="Re: Budget Review",
            body_text="Budget reply body",
            message_id="<mbox-budget-reply@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )
        reply_message["In-Reply-To"] = "<mbox-budget-root@example.com>"
        reply_message["References"] = "<mbox-budget-root@example.com>"
        self.write_fake_mbox_file([root_message, reply_message])

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        root_row = self.fetch_document_row(
            retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-budget-root@example.com>")
        )
        reply_row = self.fetch_document_row(
            retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-budget-reply@example.com>")
        )
        self.assertIsNotNone(root_row["conversation_id"])
        self.assertEqual(root_row["conversation_id"], reply_row["conversation_id"])

    def test_ingest_mbox_same_subject_messages_without_thread_links_stay_separate(self) -> None:
        first_message = self.build_fake_mbox_message(
            subject="You've joined the Beagle group",
            body_text="Mailbox one welcome",
            message_id="<mbox-beagle-001@example.com>",
            author="Beagle <welcome@example.com>",
            recipients="Mailbox One <one@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        second_message = self.build_fake_mbox_message(
            subject="You've joined the Beagle group",
            body_text="Mailbox two welcome",
            message_id="<mbox-beagle-002@example.com>",
            author="Beagle <welcome@example.com>",
            recipients="Mailbox Two <two@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )
        self.write_fake_mbox_file([first_message, second_message])

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)

        first_row = self.fetch_document_row(
            retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-beagle-001@example.com>")
        )
        second_row = self.fetch_document_row(
            retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-beagle-002@example.com>")
        )
        self.assertIsNotNone(first_row["conversation_id"])
        self.assertIsNotNone(second_row["conversation_id"])
        self.assertNotEqual(first_row["conversation_id"], second_row["conversation_id"])

    def test_ingest_pst_email_groups_threaded_messages_from_transport_headers(self) -> None:
        self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-email-root",
                subject="Witness Prep",
                body_text="PST root body",
                date_created="2026-04-14T10:00:00Z",
                transport_headers="\n".join(
                    [
                        "Message-ID: <pst-root@example.com>",
                        "Conversation-Topic: Witness Prep",
                        "",
                    ]
                ),
                conversation_topic="Witness Prep",
            ),
            self.build_fake_pst_message(
                source_item_id="pst-email-reply",
                subject="Re: Witness Prep",
                body_text="PST reply body",
                date_created="2026-04-14T11:00:00Z",
                transport_headers="\n".join(
                    [
                        "Message-ID: <pst-reply@example.com>",
                        "In-Reply-To: <pst-root@example.com>",
                        "References: <pst-root@example.com>",
                        "Conversation-Topic: Witness Prep",
                        "",
                    ]
                ),
                conversation_topic="Witness Prep",
            ),
        ]

        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            retriever_tools.bootstrap(self.root)
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        root_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-email-root"))
        reply_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-email-reply"))
        self.assertIsNotNone(root_row["conversation_id"])
        self.assertEqual(root_row["conversation_id"], reply_row["conversation_id"])

    def test_ingest_pst_email_topic_only_matches_do_not_merge_unrelated_messages(self) -> None:
        self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-beagle-001",
                subject="You've joined the Beagle group",
                body_text="Mailbox one welcome",
                author="Beagle",
                recipients="Mailbox One <one@example.com>",
                date_created="2026-04-24T00:00:00Z",
                conversation_topic="You've joined the Beagle group",
            ),
            self.build_fake_pst_message(
                source_item_id="pst-beagle-002",
                subject="You've joined the Beagle group",
                body_text="Mailbox two welcome",
                author="Beagle",
                recipients="Mailbox Two <two@example.com>",
                date_created="2026-04-24T01:00:00Z",
                conversation_topic="You've joined the Beagle group",
            ),
            self.build_fake_pst_message(
                source_item_id="pst-beagle-003",
                subject="You've joined the Beagle group",
                body_text="Mailbox three welcome",
                author="Beagle",
                recipients="Mailbox Three <three@example.com>",
                date_created="2026-04-24T02:00:00Z",
                conversation_topic="You've joined the Beagle group",
            ),
        ]

        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            retriever_tools.bootstrap(self.root)
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)

        first_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-beagle-001"))
        second_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-beagle-002"))
        third_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-beagle-003"))
        self.assertIsNotNone(first_row["conversation_id"])
        self.assertIsNotNone(second_row["conversation_id"])
        self.assertIsNotNone(third_row["conversation_id"])
        self.assertNotEqual(first_row["conversation_id"], second_row["conversation_id"])
        self.assertNotEqual(first_row["conversation_id"], third_row["conversation_id"])
        self.assertNotEqual(second_row["conversation_id"], third_row["conversation_id"])

    def test_search_parser_accepts_alias_flags_for_paging_and_sorting(self) -> None:
        parser = retriever_tools.build_parser()
        args = parser.parse_args(
            [
                "search",
                str(self.root),
                "",
                "--limit",
                "10",
                "--sort-by",
                "created_date",
                "--sort-order",
                "desc",
            ]
        )

        self.assertEqual(args.command, "search")
        self.assertEqual(args.per_page, 10)
        self.assertEqual(args.sort, "created_date")
        self.assertEqual(args.order, "desc")

    def test_search_accepts_created_date_sort_alias(self) -> None:
        first_path = self.root / "chat-older.txt"
        first_path.write_text(
            "\n".join(
                [
                    "[2026-04-14 09:00] Alice Example: Earlier thread.",
                    "[2026-04-14 09:05] Bob Example: Copy that.",
                ]
            )
            + "\n",
            encoding="utf-8",
        )
        second_path = self.root / "chat-newer.txt"
        second_path.write_text(
            "\n".join(
                [
                    "[2026-04-15 09:00] Alice Example: Later thread.",
                    "[2026-04-15 09:05] Bob Example: Drafting now.",
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        search_result = retriever_tools.search(self.root, "", None, "created_date", "desc", 1, 10)
        self.assertEqual(search_result["sort"], "date_created")
        self.assertEqual(search_result["order"], "desc")
        self.assertEqual(search_result["results"][0]["file_name"], "chat-newer.txt")
        self.assertEqual(search_result["results"][1]["file_name"], "chat-older.txt")

    def test_search_cli_view_mode_defaults_thread_listing_queries_to_newest_first(self) -> None:
        self.write_email_message(
            self.root / "older.eml",
            subject="Beagle Feature -- Auto Tagging",
            body_text="Beagle Feature -- Auto Tagging thread update from the older message.",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            self.root / "middle.eml",
            subject="Beagle Feature -- Auto Tagging",
            body_text="Beagle Feature -- Auto Tagging thread update from the middle message.",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )
        self.write_email_message(
            self.root / "newer.eml",
            subject="Beagle Feature -- Auto Tagging",
            body_text="Beagle Feature -- Auto Tagging thread update from the newest message.",
            date_created="Tue, 14 Apr 2026 12:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 3)

        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "Beagle Feature -- Auto Tagging thread",
            "--mode",
            "view",
            "--per-page",
            "2",
        )
        next_exit, next_stdout, next_stderr = self.run_cli_raw("slash", str(self.root), "/next")

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        assert search_payload is not None
        self.assertEqual(search_payload["sort"], "date_created")
        self.assertEqual(search_payload["order"], "desc")
        self.assertEqual(
            [item["file_name"] for item in search_payload["results"]],
            ["newer.eml", "middle.eml"],
        )

        self.assertEqual(next_exit, 0)
        self.assertEqual(next_stderr, "")
        self.assertIn("Page: 2 of 2  (docs 3-3 of 3)", next_stdout)
        self.assertIn("older.eml", next_stdout)

    def test_search_cli_view_mode_keeps_relevance_for_plain_content_queries(self) -> None:
        (self.root / "alpha.txt").write_text("needle\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("needle needle needle\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "needle",
            "--mode",
            "view",
        )

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        assert search_payload is not None
        self.assertEqual(search_payload["sort"], "relevance")
        self.assertEqual(search_payload["order"], "asc")

    def test_dataset_cli_commands_manage_manual_dataset_membership(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("sample dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")

        exit_code, create_payload, _, _ = self.run_cli("create-dataset", str(self.root), "Review Set")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(create_payload)
        dataset_id = int(create_payload["dataset"]["id"])
        self.assertEqual(create_payload["dataset"]["dataset_name"], "Review Set")
        self.assertEqual(create_payload["dataset"]["source_kind"], retriever_tools.MANUAL_DATASET_SOURCE_KIND)

        exit_code, add_payload, _, _ = self.run_cli(
            "add-to-dataset",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
            "--doc-id",
            str(row["id"]),
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(add_payload)
        self.assertEqual(add_payload["added_document_ids"], [row["id"]])
        self.assertEqual(add_payload["already_present_document_ids"], [])

        dataset_filtered = retriever_tools.search(
            self.root,
            "",
            [["dataset_name", "eq", "Review Set"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(dataset_filtered["total_hits"], 1)
        self.assertEqual(dataset_filtered["results"][0]["id"], row["id"])

        exit_code, list_payload, _, _ = self.run_cli("list-datasets", str(self.root))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        dataset_names = [item["dataset_name"] for item in list_payload["datasets"]]
        self.assertIn("Review Set", dataset_names)
        self.assertIn(self.root.name, dataset_names)

        exit_code, remove_payload, _, _ = self.run_cli(
            "remove-from-dataset",
            str(self.root),
            "--dataset-name",
            "Review Set",
            "--doc-id",
            str(row["id"]),
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(remove_payload)
        self.assertEqual(remove_payload["removed_document_ids"], [row["id"]])
        self.assertEqual(remove_payload["documents_without_dataset_memberships"], [])

        dataset_filtered = retriever_tools.search(
            self.root,
            "",
            [["dataset_name", "eq", "Review Set"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(dataset_filtered["total_hits"], 0)

        exit_code, delete_payload, _, _ = self.run_cli(
            "delete-dataset",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(delete_payload)
        self.assertEqual(delete_payload["deleted_dataset"]["id"], dataset_id)

        exit_code, list_payload, _, _ = self.run_cli("list-datasets", str(self.root))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        dataset_names = [item["dataset_name"] for item in list_payload["datasets"]]
        self.assertNotIn("Review Set", dataset_names)

    def test_create_dataset_rejects_normalized_name_collision(self) -> None:
        retriever_tools.bootstrap(self.root)

        first_exit, first_payload, _, _ = self.run_cli("create-dataset", str(self.root), "Review Set")
        second_exit, second_payload, _, _ = self.run_cli("create-dataset", str(self.root), "  review   set  ")

        self.assertEqual(first_exit, 0)
        self.assertIsNotNone(first_payload)
        self.assertEqual(second_exit, 2)
        self.assertIsNotNone(second_payload)
        self.assertIn("already exists", second_payload["error"])

    def test_search_cli_accepts_sql_like_filter_boolean_groups(self) -> None:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("beta body\n", encoding="utf-8")
        (self.root / "gamma.md").write_text("gamma body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 3)

        exit_code, payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "--filter",
            "(file_name = 'alpha.txt' OR file_name = 'beta.txt') AND file_type = 'txt'",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["total_hits"], 2)
        self.assertEqual(sorted(item["file_name"] for item in payload["results"]), ["alpha.txt", "beta.txt"])

    def test_search_cli_sql_like_filter_errors_are_actionable(self) -> None:
        retriever_tools.bootstrap(self.root)

        unknown_exit, unknown_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "--filter",
            "authr = 'Alice Example'",
        )
        mismatch_exit, mismatch_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "--filter",
            "is_attachment > 5",
        )

        self.assertEqual(unknown_exit, 2)
        self.assertIsNotNone(unknown_payload)
        self.assertIn("Unknown field 'authr'", unknown_payload["error"])
        self.assertIn("author", unknown_payload["error"])

        self.assertEqual(mismatch_exit, 2)
        self.assertIsNotNone(mismatch_payload)
        self.assertIn("Field 'is_attachment' is boolean", mismatch_payload["error"])
        self.assertIn("IS NULL", mismatch_payload["error"])

    def test_sql_filter_randomized_semantics_match_sqlite_reference(self) -> None:
        corpus_rows = self.setup_randomized_sql_filter_corpus()
        rng = random.Random(RANDOMIZED_FILTER_TEST_SEED)

        for _ in range(40):
            expression = self.render_random_sql_filter_node(self.build_random_sql_filter_node(rng))
            expected_doc_ids = self.reference_doc_ids_for_sql_filter(corpus_rows, expression)

            with self.subTest(expression=expression):
                exit_code, payload, _, _ = self.run_cli(
                    "search",
                    str(self.root),
                    "--filter",
                    expression,
                )
                self.assertEqual(exit_code, 0)
                self.assertIsNotNone(payload)
                actual_doc_ids = sorted(int(item["id"]) for item in payload["results"])
                self.assertEqual(actual_doc_ids, expected_doc_ids)

    def test_sql_filter_in_list_boundary_is_enforced(self) -> None:
        retriever_tools.bootstrap(self.root)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            accepted_expression = "file_name IN (" + ", ".join("'a'" for _ in range(retriever_tools.MAX_FILTER_IN_LIST_ITEMS)) + ")"
            accepted_clause, accepted_params = retriever_tools.compile_sql_filter_expression(connection, accepted_expression)
            self.assertIn(" IN (", accepted_clause)
            self.assertEqual(len(accepted_params), retriever_tools.MAX_FILTER_IN_LIST_ITEMS)

            rejected_expression = "file_name IN (" + ", ".join("'a'" for _ in range(retriever_tools.MAX_FILTER_IN_LIST_ITEMS + 1)) + ")"
            with self.assertRaises(retriever_tools.RetrieverError) as excinfo:
                retriever_tools.compile_sql_filter_expression(connection, rejected_expression)
            self.assertIn(
                f"capped at {retriever_tools.MAX_FILTER_IN_LIST_ITEMS}",
                str(excinfo.exception),
            )
        finally:
            connection.close()

    def test_search_cli_columns_returns_display_values_and_rejects_filter_only_fields(self) -> None:
        (self.root / "sample.txt").write_text("sample body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        self.assertEqual(self.run_cli("add-field", str(self.root), "review_score", "integer")[0], 0)
        self.assertEqual(
            self.run_cli(
                "set-field",
                str(self.root),
                "--doc-id",
                str(row["id"]),
                "--field",
                "review_score",
                "--value",
                "7",
            )[0],
            0,
        )

        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "--columns",
            "title,review_score,control_number",
        )
        error_exit, error_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "--columns",
            "has_attachments",
        )

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        self.assertEqual(search_payload["display"]["columns"], ["title", "review_score", "control_number"])
        self.assertEqual(search_payload["results"][0]["display_values"]["review_score"], 7)
        self.assertIn("title", search_payload["results"][0]["display_values"])

        self.assertEqual(error_exit, 2)
        self.assertIsNotNone(error_payload)
        self.assertIn("filter-only", error_payload["error"])
        self.assertIn("has_attachments", error_payload["error"])

    def test_search_cli_keyword_query_paginates_with_explicit_sort(self) -> None:
        for index in range(25):
            (self.root / f"doc-{index:02d}.txt").write_text(f"needle document {index}\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 25)

        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "needle",
            "--sort",
            "file_name",
            "--order",
            "asc",
            "--per-page",
            "5",
            "--page",
            "2",
        )

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        assert search_payload is not None
        self.assertEqual(search_payload["total_hits"], 25)
        self.assertEqual(search_payload["total_pages"], 5)
        self.assertEqual(
            [item["file_name"] for item in search_payload["results"]],
            [f"doc-{index:02d}.txt" for index in range(5, 10)],
        )
        self.assertEqual(search_payload["sort"], "file_name")
        self.assertEqual(search_payload["order"], "asc")

    def test_search_cli_view_mode_returns_rendered_markdown_with_current_compact_shape(self) -> None:
        (self.root / "sample.txt").write_text("sample body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "sample",
            "--mode",
            "view",
            "--columns",
            "title,control_number",
        )

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        assert search_payload is not None
        self.assertIn("results", search_payload)
        self.assertNotIn("documents", search_payload)
        self.assertIn("rendered_markdown", search_payload)
        rendered = str(search_payload["rendered_markdown"])
        self.assertIn("Keyword: 'sample'", rendered)
        self.assertIn("| title | control_number |", rendered)
        self.assertIn("](computer://", rendered)
        self.assertIn("Documents 1–1 of 1.", rendered)
        self.assertNotIn("| # |", rendered)

    def test_search_cli_compose_mode_also_returns_rendered_markdown_for_standard_listing_schema(self) -> None:
        (self.root / "sample.txt").write_text("sample body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "sample",
            "--columns",
            "title,control_number",
        )

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        assert search_payload is not None
        self.assertIn("rendered_markdown", search_payload)
        rendered = str(search_payload["rendered_markdown"])
        self.assertIn("Keyword: 'sample'", rendered)
        self.assertIn("| title | control_number |", rendered)
        self.assertIn("](computer://", rendered)
        self.assertIn("Documents 1–1 of 1.", rendered)
        self.assertNotIn("| # |", rendered)

    def test_search_cli_view_mode_renders_attachment_rows_and_parent_context(self) -> None:
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Upgrade test",
            body_text="Hello team,\nThis is the email body.",
            attachment_name="notes.txt",
            attachment_text="confidential attachment detail",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        parent_exit, parent_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "Upgrade test",
            "--mode",
            "view",
        )
        self.assertEqual(parent_exit, 0)
        self.assertIsNotNone(parent_payload)
        assert parent_payload is not None
        parent_rendered = str(parent_payload["rendered_markdown"])
        self.assertIn("[Upgrade test](computer://", parent_rendered)
        self.assertIn("| E-Doc | [↳ notes.txt](computer://", parent_rendered)

        attachment_exit, attachment_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "confidential attachment detail",
            "--mode",
            "view",
        )
        self.assertEqual(attachment_exit, 0)
        self.assertIsNotNone(attachment_payload)
        assert attachment_payload is not None
        attachment_rendered = str(attachment_payload["rendered_markdown"])
        self.assertIn("| E-Doc | [↳ notes.txt (parent: Upgrade test)](computer://", attachment_rendered)

    def test_search_cli_view_mode_uses_unrecognized_for_unknown_attachment_types(self) -> None:
        email_path = self.root / "thread.eml"
        message = EmailMessage()
        message["From"] = "Alice Example <alice@example.com>"
        message["To"] = "Bob Example <bob@example.com>"
        message["Subject"] = "Binary attachment test"
        message["Date"] = "Tue, 14 Apr 2026 10:00:00 +0000"
        message.set_content("Hello team,\nThis email carries a binary attachment.")
        message.add_attachment(b"\x00\xff\x10\x80", maintype="application", subtype="octet-stream")
        email_path.write_bytes(message.as_bytes(policy=policy.default))

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "Binary attachment test",
            "--mode",
            "view",
        )
        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        assert search_payload is not None
        rendered = str(search_payload["rendered_markdown"])
        self.assertIn("| Unrecognized | [↳ attachment-001.bin](computer://", rendered)
        self.assertNotIn("| ↳ Unrecognized |", rendered)
        self.assertNotIn("| Attachment | [↳ attachment-001.bin](computer://", rendered)

    def test_search_cli_view_mode_renders_author_email_on_new_line(self) -> None:
        rendered = retriever_tools.render_search_markdown_cell(
            {
                "author": '"Sood, Udit" usood@cov.com',
                "display_values": {
                    "author": '"Sood, Udit" usood@cov.com',
                },
            },
            {"name": "author", "type": "text"},
        )

        self.assertEqual(rendered, "Sood, Udit<br>usood@cov.com")

    def test_ingest_eml_attachment_without_filename_uses_detected_extension(self) -> None:
        email_path = self.root / "thread.eml"
        attachment_path = self.root / "attachment-source.pdf"
        self.write_minimal_pdf(attachment_path, "Detected attachment")
        attachment_bytes = attachment_path.read_bytes()
        attachment_path.unlink()

        message = EmailMessage()
        message["From"] = "Alice Example <alice@example.com>"
        message["To"] = "Bob Example <bob@example.com>"
        message["Subject"] = "Detected attachment type"
        message["Date"] = "Tue, 14 Apr 2026 10:00:00 +0000"
        message.set_content("Hello team,\nThis carries a nameless PDF attachment.")
        message.add_attachment(attachment_bytes, maintype="application", subtype="pdf")
        email_path.write_bytes(message.as_bytes(policy=policy.default))

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        parent_row = self.fetch_document_row("thread.eml")
        child_rows = self.fetch_child_rows(parent_row["id"])
        self.assertEqual(len(child_rows), 1)
        self.assertEqual(child_rows[0]["file_name"], "attachment-001.pdf")
        self.assertEqual(child_rows[0]["file_type"], "pdf")

    def test_slash_search_persists_scope_and_search_within_keyword(self) -> None:
        (self.root / "alpha.txt").write_text("alpha beta body\n", encoding="utf-8")
        (self.root / "second.txt").write_text("alpha only body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        first_exit, first_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")

        self.assertEqual(first_exit, 0)
        self.assertIsNotNone(first_payload)
        self.assertEqual(first_payload["scope"]["keyword"], "alpha")
        self.assertEqual(first_payload["header"]["keyword"], "Keyword: 'alpha'")

        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(session_payload["schema_version"], retriever_tools.SESSION_SCHEMA_VERSION)
        self.assertEqual(session_payload["scope"]["keyword"], "alpha")

        second_exit, second_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "--within", "beta")

        self.assertEqual(second_exit, 0)
        self.assertIsNotNone(second_payload)
        self.assertEqual(second_payload["scope"]["keyword"], "(alpha) AND (beta)")
        self.assertEqual(second_payload["total_hits"], 1)
        self.assertEqual(second_payload["results"][0]["file_name"], "alpha.txt")

    def test_slash_scope_save_load_and_dataset_rename_refresh_saved_scope_labels(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("sample dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        create_exit, create_payload, _, _ = self.run_cli("create-dataset", str(self.root), "Review Set")
        self.assertEqual(create_exit, 0)
        self.assertIsNotNone(create_payload)
        dataset_id = int(create_payload["dataset"]["id"])
        add_exit, _, _, _ = self.run_cli(
            "add-to-dataset",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
            "--doc-id",
            str(row["id"]),
        )
        self.assertEqual(add_exit, 0)

        dataset_exit, dataset_payload, _, _ = self.run_cli("slash", str(self.root), "/dataset", "Review Set")
        save_exit, save_payload, _, _ = self.run_cli("slash", str(self.root), "/scope", "save", "review")
        rename_exit, rename_payload, _, _ = self.run_cli(
            "slash",
            str(self.root),
            '/dataset rename "Review Set" "Renamed Set"',
        )
        load_exit, load_payload, _, _ = self.run_cli("slash", str(self.root), "/scope", "load", "review")

        self.assertEqual(dataset_exit, 0)
        self.assertIsNotNone(dataset_payload)
        self.assertEqual(dataset_payload["scope"]["dataset"][0]["name"], "Review Set")

        self.assertEqual(save_exit, 0)
        self.assertIsNotNone(save_payload)
        self.assertEqual(save_payload["name"], "review")

        self.assertEqual(rename_exit, 0)
        self.assertIsNotNone(rename_payload)
        self.assertEqual(rename_payload["dataset"]["dataset_name"], "Renamed Set")

        self.assertEqual(load_exit, 0)
        self.assertIsNotNone(load_payload)
        self.assertEqual(load_payload["scope"]["dataset"][0]["id"], dataset_id)
        self.assertEqual(load_payload["scope"]["dataset"][0]["name"], "Renamed Set")
        self.assertEqual(load_payload["total_hits"], 1)

        saved_scopes_payload = json.loads(self.paths["saved_scopes_path"].read_text(encoding="utf-8"))
        self.assertEqual(saved_scopes_payload["schema_version"], retriever_tools.SESSION_SCHEMA_VERSION)
        self.assertEqual(saved_scopes_payload["scopes"]["review"]["dataset"][0]["id"], dataset_id)
        self.assertEqual(saved_scopes_payload["scopes"]["review"]["dataset"][0]["name"], "Renamed Set")

    def test_slash_scope_and_dataset_list_commands_show_current_and_available_state(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("alpha dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        create_exit, create_payload, _, _ = self.run_cli("create-dataset", str(self.root), "Review Set")
        self.assertEqual(create_exit, 0)
        self.assertIsNotNone(create_payload)
        dataset_id = int(create_payload["dataset"]["id"])
        add_exit, _, _, _ = self.run_cli(
            "add-to-dataset",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
            "--doc-id",
            str(row["id"]),
        )
        self.assertEqual(add_exit, 0)
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute(
                """
                UPDATE document_occurrences
                SET file_size = ?, custodian = ?, extracted_doc_authored_at = ?, extracted_doc_modified_at = ?, extracted_content_type = ?
                WHERE document_id = ?
                """,
                (
                    12,
                    "mailbox@example.com",
                    "2024-01-02T10:00:00Z",
                    "2024-01-03T11:00:00Z",
                    "E-Doc",
                    row["id"],
                ),
            )
            retriever_tools.refresh_document_from_occurrences(connection, row["id"])
            connection.commit()
        finally:
            connection.close()

        self.assertEqual(self.run_cli("slash", str(self.root), "/search", "alpha")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/dataset", "Review Set")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/scope", "save", "review")[0], 0)
        list_exit, list_payload, _, _ = self.run_cli("list-datasets", str(self.root))

        search_show_exit, search_show_stdout, search_show_stderr = self.run_cli_raw("slash", str(self.root), "/search")
        scope_show_exit, scope_show_stdout, scope_show_stderr = self.run_cli_raw("slash", str(self.root), "/scope")
        scope_list_exit, scope_list_stdout, scope_list_stderr = self.run_cli_raw("slash", str(self.root), "/scope", "list")
        dataset_show_exit, dataset_show_stdout, dataset_show_stderr = self.run_cli_raw("slash", str(self.root), "/dataset")
        dataset_list_exit, dataset_list_stdout, dataset_list_stderr = self.run_cli_raw("slash", str(self.root), "/dataset", "list")

        self.assertEqual(search_show_exit, 0)
        self.assertEqual(search_show_stderr, "")
        self.assertEqual(search_show_stdout.strip(), "Search: alpha")

        self.assertEqual(scope_show_exit, 0)
        self.assertEqual(scope_show_stderr, "")
        self.assertIn("Scope:", scope_show_stdout)
        self.assertIn("- keyword: alpha", scope_show_stdout)
        self.assertIn("- dataset: Review Set", scope_show_stdout)

        self.assertEqual(scope_list_exit, 0)
        self.assertEqual(scope_list_stderr, "")
        self.assertIn("Saved scopes:", scope_list_stdout)
        self.assertIn("- review: keyword=alpha; dataset=Review Set", scope_list_stdout)

        self.assertEqual(dataset_show_exit, 0)
        self.assertEqual(dataset_show_stderr, "")
        self.assertEqual(dataset_show_stdout, "Dataset: Review Set")

        self.assertEqual(list_exit, 0)
        self.assertIsNotNone(list_payload)
        review_dataset = next(item for item in list_payload["datasets"] if item["dataset_name"] == "Review Set")
        self.assertEqual(review_dataset["size_bytes"], 12)
        self.assertEqual(review_dataset["sized_document_count"], 1)
        self.assertEqual(review_dataset["custodians"], ["mailbox@example.com"])
        self.assertEqual(review_dataset["content_types"], [{"name": "E-Doc", "count": 1}])
        self.assertEqual(review_dataset["time_range_start"], "2024-01-02T10:00:00Z")
        self.assertEqual(review_dataset["time_range_end"], "2024-01-03T11:00:00Z")

        self.assertEqual(dataset_list_exit, 0)
        self.assertEqual(dataset_list_stderr, "")
        self.assertIn("Datasets:", dataset_list_stdout)
        self.assertIn("| Dataset | Docs | Size | Custodians |", dataset_list_stdout)
        self.assertIn(
            "| Review Set | 1 | 12 B | mailbox@example.com |",
            dataset_list_stdout,
        )
        self.assertIn(
            f"| {self.root.name} | 1 | 12 B | mailbox@example.com |",
            dataset_list_stdout,
        )

    def test_dataset_list_prefers_container_file_size_for_mbox_sources(self) -> None:
        mbox_path = self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="Container-backed dataset",
                    body_text="Message body for the mailbox dataset.",
                    message_id="<mbox-size@example.com>",
                    attachment_name="notes.txt",
                    attachment_text="attachment bytes",
                )
            ]
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["mbox_messages_created"], 1)

        list_exit, list_payload, _, _ = self.run_cli("list-datasets", str(self.root))
        dataset_list_exit, dataset_list_stdout, dataset_list_stderr = self.run_cli_raw(
            "slash",
            str(self.root),
            "/dataset",
            "list",
        )

        self.assertEqual(list_exit, 0)
        self.assertIsNotNone(list_payload)
        mailbox_dataset = next(item for item in list_payload["datasets"] if item["dataset_name"] == "mailbox.mbox")
        self.assertEqual(mailbox_dataset["size_bytes"], mbox_path.stat().st_size)
        self.assertEqual(mailbox_dataset["size_basis"], "container")
        self.assertEqual(mailbox_dataset["document_count"], 2)
        self.assertEqual(mailbox_dataset["sized_document_count"], 1)

        self.assertEqual(dataset_list_exit, 0)
        self.assertEqual(dataset_list_stderr, "")
        self.assertIn(
            f"| mailbox.mbox | 2 | {retriever_tools.format_dataset_size_summary(mailbox_dataset)} | mailbox |",
            dataset_list_stdout,
        )
        self.assertNotIn("(1/2 sized)", dataset_list_stdout)

    def test_slash_list_commands_auto_refresh_stale_workspace_runtime_metadata(self) -> None:
        retriever_tools.bootstrap(self.root)

        runtime = json.loads(self.paths["runtime_path"].read_text(encoding="utf-8"))
        runtime["template_sha256"] = "0" * 64
        runtime["tool_version"] = "0.0-test-stale"
        runtime["template_source"] = "skills/tool-template/retriever_tools.py"
        self.paths["runtime_path"].write_text(
            json.dumps(runtime, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )

        dataset_list_exit, dataset_list_stdout, dataset_list_stderr = self.run_cli_raw(
            "slash",
            str(self.root),
            "/dataset",
            "list",
        )

        self.assertEqual(dataset_list_exit, 0)
        self.assertIn("Datasets:", dataset_list_stdout)
        self.assertTrue(dataset_list_stderr.startswith("retriever-runtime-sync: "))
        runtime_sync_payload = json.loads(dataset_list_stderr.removeprefix("retriever-runtime-sync: "))
        self.assertEqual(runtime_sync_payload["status"], "updated-runtime")
        self.assertEqual(runtime_sync_payload["reason"], "auto-runtime-sync")

        refreshed_runtime = json.loads(self.paths["runtime_path"].read_text(encoding="utf-8"))
        self.assertEqual(refreshed_runtime["tool_version"], retriever_tools.TOOL_VERSION)
        self.assertEqual(refreshed_runtime["template_sha256"], hashlib.sha256(TOOL_BYTES).hexdigest())
        self.assertEqual(refreshed_runtime["template_source"], retriever_tools.TEMPLATE_SOURCE)

    def test_slash_bates_within_intersects_ranges_and_rejects_cross_slot_and_mixed_prefix(self) -> None:
        retriever_tools.bootstrap(self.root)

        first_exit, first_payload, _, _ = self.run_cli("slash", str(self.root), "/bates", "ABC0001-ABC0010")
        within_exit, within_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "--within", "ABC0005-ABC0007")
        cross_exit, cross_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "--within", "alpha")
        mixed_exit, mixed_payload, _, _ = self.run_cli("slash", str(self.root), "/bates", "ABC0001-XYZ0002")

        self.assertEqual(first_exit, 0)
        self.assertIsNotNone(first_payload)
        self.assertEqual(first_payload["scope"]["bates"], {"begin": "ABC0001", "end": "ABC0010"})

        self.assertEqual(within_exit, 0)
        self.assertIsNotNone(within_payload)
        self.assertEqual(within_payload["scope"]["bates"], {"begin": "ABC0005", "end": "ABC0007"})

        self.assertEqual(cross_exit, 2)
        self.assertIsNotNone(cross_payload)
        self.assertIn("only composes within the current slot", cross_payload["error"])

        self.assertEqual(mixed_exit, 2)
        self.assertIsNotNone(mixed_payload)
        self.assertIn("Mixed-prefix Bates ranges", mixed_payload["error"])

    def test_slash_search_errors_when_scope_dataset_reference_is_deleted(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("sample dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        create_exit, _, _, _ = self.run_cli("create-dataset", str(self.root), "Review Set")
        self.assertEqual(create_exit, 0)
        scope_exit, _, _, _ = self.run_cli("slash", str(self.root), "/dataset", "Review Set")
        self.assertEqual(scope_exit, 0)
        delete_exit, _, _, _ = self.run_cli("delete-dataset", str(self.root), "--dataset-name", "Review Set")
        self.assertEqual(delete_exit, 0)

        search_exit, search_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        self.assertEqual(search_exit, 2)
        self.assertIsNotNone(search_payload)
        self.assertIn("no longer exists", search_payload["error"])

    def test_slash_search_errors_when_scope_from_run_reference_is_missing(self) -> None:
        retriever_tools.bootstrap(self.root)
        self.paths["session_path"].write_text(
            json.dumps(
                {
                    "schema_version": retriever_tools.SESSION_SCHEMA_VERSION,
                    "scope": {"from_run_id": 999},
                    "browsing": {},
                    "display": {},
                },
                indent=2,
                sort_keys=True,
            )
            + "\n",
            encoding="utf-8",
        )

        search_exit, search_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        self.assertEqual(search_exit, 2)
        self.assertIsNotNone(search_payload)
        self.assertIn("scope.from_run_id no longer exists", search_payload["error"])

    def test_slash_search_filter_bates_and_from_run_show_current_scope_slots(self) -> None:
        retriever_tools.bootstrap(self.root)

        search_set_exit, _, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        filter_set_exit, _, _, _ = self.run_cli("slash", str(self.root), "/filter", "content_type = 'Email'")
        bates_set_exit, _, _, _ = self.run_cli("slash", str(self.root), "/bates", "ABC0001-ABC0010")
        self.assertEqual(search_set_exit, 0)
        self.assertEqual(filter_set_exit, 0)
        self.assertEqual(bates_set_exit, 0)

        self.paths["session_path"].write_text(
            json.dumps(
                {
                    "schema_version": retriever_tools.SESSION_SCHEMA_VERSION,
                    "scope": {
                        "keyword": "alpha",
                        "filter": "content_type = 'Email'",
                        "bates": {"begin": "ABC0001", "end": "ABC0010"},
                        "from_run_id": 42,
                    },
                    "browsing": {},
                    "display": {},
                },
                indent=2,
                sort_keys=True,
            )
            + "\n",
            encoding="utf-8",
        )

        search_show_exit, search_show_stdout, search_show_stderr = self.run_cli_raw("slash", str(self.root), "/search")
        filter_show_exit, filter_show_stdout, filter_show_stderr = self.run_cli_raw("slash", str(self.root), "/filter")
        bates_show_exit, bates_show_stdout, bates_show_stderr = self.run_cli_raw("slash", str(self.root), "/bates")
        from_run_show_exit, from_run_show_stdout, from_run_show_stderr = self.run_cli_raw("slash", str(self.root), "/from-run")

        self.assertEqual(search_show_exit, 0)
        self.assertEqual(search_show_stderr, "")
        self.assertEqual(search_show_stdout.strip(), "Search: alpha")

        self.assertEqual(filter_show_exit, 0)
        self.assertEqual(filter_show_stderr, "")
        self.assertEqual(filter_show_stdout.strip(), "Filter: content_type = 'Email'")

        self.assertEqual(bates_show_exit, 0)
        self.assertEqual(bates_show_stderr, "")
        self.assertEqual(bates_show_stdout.strip(), "Bates: ABC0001-ABC0010")

        self.assertEqual(from_run_show_exit, 0)
        self.assertEqual(from_run_show_stderr, "")
        self.assertEqual(from_run_show_stdout.strip(), "From run: 42")

    def test_slash_sort_and_paging_persist_browsing_state(self) -> None:
        for index in range(25):
            (self.root / f"doc-{index:02d}.txt").write_text(f"document {index}\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 25)

        sort_exit, sort_payload, _, _ = self.run_cli("slash", str(self.root), "/sort", "file_name asc")

        self.assertEqual(sort_exit, 0)
        self.assertIsNotNone(sort_payload)
        self.assertEqual(sort_payload["results"][0]["file_name"], "doc-00.txt")
        self.assertEqual(sort_payload["header"]["sort"], "Sort: file_name asc")

        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(
            session_payload["browsing"]["documents"]["sort"],
            [["file_name", "asc"]],
        )
        self.assertEqual(session_payload["browsing"]["documents"]["offset"], 0)

        next_exit, next_stdout, next_stderr = self.run_cli_raw("slash", str(self.root), "/next")
        page_exit, page_payload, _, _ = self.run_cli("slash", str(self.root), "/page", "last")
        previous_exit, previous_stdout, previous_stderr = self.run_cli_raw("slash", str(self.root), "/previous")
        default_exit, default_payload, _, _ = self.run_cli("slash", str(self.root), "/sort", "default")

        self.assertEqual(next_exit, 0)
        self.assertEqual(next_stderr, "")
        self.assertIn("Sort: file_name asc", next_stdout)
        self.assertIn("Page: 2 of 3  (docs 11-20 of 25)", next_stdout)
        self.assertIn("doc-10.txt", next_stdout)

        self.assertEqual(page_exit, 0)
        self.assertIsNotNone(page_payload)
        self.assertEqual(page_payload["page"], 3)
        self.assertEqual(page_payload["offset"], 20)

        self.assertEqual(previous_exit, 0)
        self.assertEqual(previous_stderr, "")
        self.assertIn("Sort: file_name asc", previous_stdout)
        self.assertIn("Page: 2 of 3  (docs 11-20 of 25)", previous_stdout)
        self.assertIn("doc-10.txt", previous_stdout)

        self.assertEqual(default_exit, 0)
        self.assertIsNotNone(default_payload)
        self.assertEqual(default_payload["offset"], 0)

        final_session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertNotIn("sort", final_session_payload["browsing"]["documents"])
        self.assertEqual(final_session_payload["browsing"]["documents"]["offset"], 0)

    def test_slash_sort_page_page_size_and_columns_show_current_and_available_state(self) -> None:
        for index in range(25):
            (self.root / f"doc-{index:02d}.txt").write_text(f"document {index}\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 25)

        sort_exit, _, _, _ = self.run_cli("slash", str(self.root), "/sort", "file_name asc")
        self.assertEqual(sort_exit, 0)

        sort_show_exit, sort_show_stdout, sort_show_stderr = self.run_cli_raw("slash", str(self.root), "/sort")
        sort_list_exit, sort_list_stdout, sort_list_stderr = self.run_cli_raw("slash", str(self.root), "/sort", "list")
        page_size_show_exit, page_size_show_stdout, page_size_show_stderr = self.run_cli_raw("slash", str(self.root), "/page-size")
        self.assertEqual(self.run_cli("slash", str(self.root), "/page-size 5")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/next")[0], 0)
        page_show_exit, page_show_stdout, page_show_stderr = self.run_cli_raw("slash", str(self.root), "/page")
        page_size_after_exit, page_size_after_stdout, page_size_after_stderr = self.run_cli_raw("slash", str(self.root), "/page-size")
        columns_show_exit, columns_show_stdout, columns_show_stderr = self.run_cli_raw("slash", str(self.root), "/columns")
        columns_list_exit, columns_list_stdout, columns_list_stderr = self.run_cli_raw("slash", str(self.root), "/columns", "list")

        self.assertEqual(sort_show_exit, 0)
        self.assertEqual(sort_show_stderr, "")
        self.assertEqual(sort_show_stdout, "Sort: file_name asc (override)")

        self.assertEqual(sort_list_exit, 0)
        self.assertEqual(sort_list_stderr, "")
        self.assertIn("Sortable fields:", sort_list_stdout)
        self.assertIn("- file_name", sort_list_stdout)
        self.assertIn("- title", sort_list_stdout)
        self.assertNotIn("- dataset_name", sort_list_stdout)

        self.assertEqual(page_size_show_exit, 0)
        self.assertEqual(page_size_show_stderr, "")
        self.assertEqual(page_size_show_stdout, f"Page size: {retriever_tools.DEFAULT_PAGE_SIZE}")

        self.assertEqual(page_show_exit, 0)
        self.assertEqual(page_show_stderr, "")
        self.assertEqual(page_show_stdout, "Page: 2 of 5 (docs 6-10 of 25)")

        self.assertEqual(page_size_after_exit, 0)
        self.assertEqual(page_size_after_stderr, "")
        self.assertEqual(page_size_after_stdout, "Page size: 5")

        self.assertEqual(columns_show_exit, 0)
        self.assertEqual(columns_show_stderr, "")
        self.assertIn(
            "Columns: content_type, title, author, date_created, control_number",
            columns_show_stdout,
        )
        self.assertIn("Page size: 5", columns_show_stdout)

        self.assertEqual(columns_list_exit, 0)
        self.assertEqual(columns_list_stderr, "")
        self.assertIn("Displayable columns:", columns_list_stdout)
        self.assertIn("- title", columns_list_stdout)
        self.assertIn("- dataset_name", columns_list_stdout)
        self.assertNotIn("- has_attachments", columns_list_stdout)

    def test_slash_conversations_uses_conversation_defaults_and_full_preview_target(self) -> None:
        self.write_email_message(
            self.root / "root.eml",
            subject="Status Update",
            body_text="Root message body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            cc=None,
            message_id="<root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            self.root / "reply.eml",
            subject="Re: Status Update",
            body_text="Reply message body",
            author="Bob Example <bob@example.com>",
            recipients="Alice Example <alice@example.com>",
            cc=None,
            message_id="<reply@example.com>",
            in_reply_to="<root@example.com>",
            references="<root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        payload = retriever_tools.run_slash_command(self.root, "/conversations")
        self.assertEqual(payload["browse_mode"], retriever_tools.BROWSE_MODE_CONVERSATIONS)
        self.assertEqual(
            payload["display"]["columns"],
            list(retriever_tools.DEFAULT_CONVERSATION_DISPLAY_COLUMNS),
        )
        self.assertEqual(payload["total_hits"], 1)
        result = payload["results"][0]
        self.assertEqual(result["conversation_type"], "email")
        self.assertEqual(result["title"], "Status Update")
        self.assertEqual(result["document_count"], 2)
        self.assertEqual(result["matching_document_count"], 2)
        self.assertIn("Alice Example", str(result["participants"]))
        self.assertTrue(result["preview_rel_path"].endswith("/conversation.html"))
        preview_path = Path(str(result["preview_abs_path"]).split("#", 1)[0])
        self.assertTrue(preview_path.exists())
        preview_html = preview_path.read_text(encoding="utf-8")
        self.assertIn("Root message body", preview_html)
        self.assertIn("Reply message body", preview_html)

        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(session_payload["browse_mode"], retriever_tools.BROWSE_MODE_CONVERSATIONS)
        self.assertEqual(session_payload["browsing"]["conversations"]["offset"], 0)
        self.assertEqual(session_payload["browsing"]["conversations"]["total_known"], 1)

    def test_list_conversations_paginates_sorts_and_seeds_slash_browse(self) -> None:
        for subject, email, created in [
            ("Alpha Thread", "alpha@example.com", "Tue, 14 Apr 2026 10:00:00 +0000"),
            ("Beta Thread", "beta@example.com", "Tue, 14 Apr 2026 11:00:00 +0000"),
            ("Gamma Thread", "gamma@example.com", "Tue, 14 Apr 2026 12:00:00 +0000"),
            ("Zeta Thread", "zeta@example.com", "Tue, 14 Apr 2026 13:00:00 +0000"),
        ]:
            self.write_email_message(
                self.root / f"{email}.eml",
                subject=subject,
                body_text=f"{subject} body",
                author=f"{subject} Author <{email}>",
                recipients="Reviewer <reviewer@example.com>",
                cc=None,
                message_id=f"<{email}>",
                date_created=created,
            )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 4)

        first_exit, first_payload, _, _ = self.run_cli(
            "list-conversations",
            str(self.root),
            "--limit",
            "2",
            "--sort",
            "title",
            "--order",
            "asc",
        )
        self.assertEqual(first_exit, 0)
        self.assertIsNotNone(first_payload)
        assert first_payload is not None
        self.assertEqual(first_payload["browse_mode"], retriever_tools.BROWSE_MODE_CONVERSATIONS)
        self.assertEqual(first_payload["offset"], 0)
        self.assertEqual(first_payload["limit"], 2)
        self.assertEqual(first_payload["total_hits"], 4)
        self.assertEqual(
            [conversation["title"] for conversation in first_payload["conversations"]],
            ["Alpha Thread", "Beta Thread"],
        )

        second_exit, second_payload, _, _ = self.run_cli(
            "list-conversations",
            str(self.root),
            "--limit",
            "2",
            "--offset",
            "2",
            "--sort",
            "title",
            "--order",
            "asc",
        )
        self.assertEqual(second_exit, 0)
        self.assertIsNotNone(second_payload)
        assert second_payload is not None
        self.assertEqual(second_payload["offset"], 2)
        self.assertEqual(
            [conversation["title"] for conversation in second_payload["conversations"]],
            ["Gamma Thread", "Zeta Thread"],
        )

        state = retriever_tools.read_session_state(self.paths)
        self.assertEqual(state["browse_mode"], retriever_tools.BROWSE_MODE_CONVERSATIONS)
        self.assertEqual(state["display"][retriever_tools.BROWSE_MODE_CONVERSATIONS]["page_size"], 2)
        self.assertEqual(
            state["browsing"][retriever_tools.BROWSE_MODE_CONVERSATIONS]["sort"],
            [["title", "asc"]],
        )

        page_exit, page_stdout, _ = self.run_cli_raw("slash", str(self.root), "/page first")
        self.assertEqual(page_exit, 0)
        self.assertIn("Alpha Thread", page_stdout)
        self.assertIn("Beta Thread", page_stdout)
        self.assertNotIn("Gamma Thread", page_stdout)

        next_exit, next_stdout, _ = self.run_cli_raw("slash", str(self.root), "/next")
        self.assertEqual(next_exit, 0)
        self.assertIn("Gamma Thread", next_stdout)
        self.assertIn("Zeta Thread", next_stdout)
        self.assertNotIn("Alpha Thread", next_stdout)

    def test_slash_conversation_columns_are_separate_from_document_columns(self) -> None:
        self.write_email_message(
            self.root / "message.eml",
            subject="Status Update",
            body_text="Root message body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            cc=None,
            message_id="<root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        (self.root / "sample.txt").write_text("sample display body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        self.assertEqual(self.run_cli("slash", str(self.root), "/conversations")[0], 0)
        set_exit, set_payload, _, _ = self.run_cli(
            "slash",
            str(self.root),
            "/columns set title, document_count",
        )
        self.assertEqual(set_exit, 0)
        self.assertIsNotNone(set_payload)
        assert set_payload is not None
        self.assertEqual(set_payload["display"]["columns"], ["title", "document_count"])

        self.assertEqual(self.run_cli("slash", str(self.root), "/documents")[0], 0)
        columns_exit, columns_stdout, columns_stderr = self.run_cli_raw("slash", str(self.root), "/columns")
        self.assertEqual(columns_exit, 0)
        self.assertEqual(columns_stderr, "")
        self.assertEqual(
            columns_stdout,
            "Columns: content_type, title, author, date_created, control_number\nPage size: 10",
        )

        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(
            session_payload["display"]["conversations"]["columns"],
            ["title", "document_count"],
        )
        self.assertNotIn("columns", session_payload["display"]["documents"])

    def test_slash_columns_commands_persist_display_preferences_and_render_custom_fields(self) -> None:
        (self.root / "sample.txt").write_text("sample display body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        self.assertEqual(self.run_cli("add-field", str(self.root), "review_score", "integer")[0], 0)
        self.assertEqual(
            self.run_cli(
                "set-field",
                str(self.root),
                "--doc-id",
                str(row["id"]),
                "--field",
                "review_score",
                "--value",
                "7",
            )[0],
            0,
        )

        show_exit, show_stdout, show_stderr = self.run_cli_raw("slash", str(self.root), "/columns")
        set_exit, set_payload, _, _ = self.run_cli(
            "slash",
            str(self.root),
            "/columns set title, review_score, control_number",
        )

        self.assertEqual(show_exit, 0)
        self.assertEqual(show_stderr, "")
        self.assertEqual(
            show_stdout,
            "Columns: content_type, title, author, date_created, control_number\nPage size: 10",
        )

        self.assertEqual(set_exit, 0)
        self.assertIsNotNone(set_payload)
        self.assertEqual(set_payload["display"]["columns"], ["title", "review_score", "control_number"])
        self.assertEqual(set_payload["results"][0]["display_values"]["review_score"], 7)
        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(
            session_payload["display"]["documents"]["columns"],
            ["title", "review_score", "control_number"],
        )

        add_exit, add_payload, _, _ = self.run_cli("slash", str(self.root), "/columns add dataset_name")
        self.assertEqual(add_exit, 0)
        self.assertIsNotNone(add_payload)
        self.assertEqual(add_payload["display"]["columns"], ["title", "review_score", "control_number", "dataset_name"])
        self.assertEqual(add_payload["results"][0]["display_values"]["dataset_name"], self.root.name)

        remove_exit, remove_payload, _, _ = self.run_cli("slash", str(self.root), "/columns remove review_score")
        self.assertEqual(remove_exit, 0)
        self.assertIsNotNone(remove_payload)
        self.assertEqual(remove_payload["display"]["columns"], ["title", "control_number", "dataset_name"])

        default_exit, default_payload, _, _ = self.run_cli("slash", str(self.root), "/columns default")
        self.assertEqual(default_exit, 0)
        self.assertIsNotNone(default_payload)
        self.assertEqual(
            default_payload["display"]["columns"],
            ["content_type", "title", "author", "date_created", "control_number"],
        )
        final_session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertNotIn("columns", final_session_payload["display"]["documents"])

    def test_passive_field_label_uses_friendly_aliases_and_custom_fallbacks(self) -> None:
        self.assertEqual(retriever_tools.passive_field_label("content_type"), "Type")
        self.assertEqual(retriever_tools.passive_field_label("conversation_type"), "Type")
        self.assertEqual(
            retriever_tools.passive_field_label("conversation_type", mixed_context=True),
            "Conversation Type",
        )
        self.assertEqual(retriever_tools.passive_field_label("first_activity"), "Started")
        self.assertEqual(retriever_tools.passive_field_label("last_activity"), "Last Activity")
        self.assertEqual(retriever_tools.passive_field_label("control_number"), "Control #")
        self.assertEqual(retriever_tools.passive_field_label("dataset"), "Dataset")
        self.assertEqual(retriever_tools.passive_field_label("review_score"), "Review Score")
        self.assertEqual(retriever_tools.passive_field_label("ocr_score"), "OCR Score")
        self.assertEqual(retriever_tools.passive_field_label("pst_owner"), "PST Owner")

    def test_slash_page_size_persists_display_preferences(self) -> None:
        for index in range(12):
            (self.root / f"doc-{index:02d}.txt").write_text(f"document {index}\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 12)

        size_exit, size_payload, _, _ = self.run_cli("slash", str(self.root), "/page-size 5")
        next_exit, next_stdout, next_stderr = self.run_cli_raw("slash", str(self.root), "/next")

        self.assertEqual(size_exit, 0)
        self.assertIsNotNone(size_payload)
        self.assertEqual(size_payload["per_page"], 5)
        self.assertEqual(size_payload["display"]["page_size"], 5)
        self.assertEqual(len(size_payload["results"]), 5)

        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(session_payload["display"]["documents"]["page_size"], 5)

        self.assertEqual(next_exit, 0)
        self.assertEqual(next_stderr, "")
        self.assertIn("Page: 2 of 3  (docs 6-10 of 12)", next_stdout)
        self.assertIn("doc-05.txt", next_stdout)
        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(session_payload["display"]["documents"]["page_size"], 5)
        self.assertEqual(session_payload["browsing"]["documents"]["offset"], 5)

    def test_view_search_persists_browse_page_size_for_followup_slash_navigation(self) -> None:
        for index in range(25):
            (self.root / f"doc-{index:02d}.txt").write_text(f"document {index}\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 25)

        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "document",
            "--mode",
            "view",
            "--per-page",
            "5",
        )
        next_exit, next_stdout, next_stderr = self.run_cli_raw("slash", str(self.root), "/next")

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        self.assertEqual(search_payload["page"], 1)
        self.assertEqual(search_payload["per_page"], 5)
        self.assertEqual(len(search_payload["results"]), 5)

        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(session_payload["display"]["documents"]["page_size"], 5)
        self.assertEqual(session_payload["scope"]["keyword"], "document")

        self.assertEqual(next_exit, 0)
        self.assertEqual(next_stderr, "")
        self.assertIn("Page: 2 of 5  (docs 6-10 of 25)", next_stdout)
        self.assertIn("doc-05.txt", next_stdout)
        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(session_payload["display"]["documents"]["page_size"], 5)
        self.assertEqual(session_payload["scope"]["keyword"], "document")
        self.assertEqual(session_payload["browsing"]["documents"]["offset"], 5)

    def test_view_search_uses_saved_page_size_when_per_page_is_omitted(self) -> None:
        for index in range(25):
            (self.root / f"doc-{index:02d}.txt").write_text(f"document {index}\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 25)

        self.assertEqual(self.run_cli("slash", str(self.root), "/page-size 5")[0], 0)
        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "document",
            "--mode",
            "view",
        )

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        assert search_payload is not None
        self.assertEqual(search_payload["page"], 1)
        self.assertEqual(search_payload["per_page"], 5)
        self.assertEqual(len(search_payload["results"]), 5)
        rendered = str(search_payload["rendered_markdown"])
        self.assertIn("Documents 1–5 of 25.", rendered)
        self.assertIn("Navigate: `/retriever:next` for the next page.", rendered)

    def test_compose_search_uses_saved_page_size_when_per_page_is_omitted(self) -> None:
        for index in range(12):
            (self.root / f"doc-{index:02d}.txt").write_text(f"document {index}\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 12)

        self.assertEqual(self.run_cli("slash", str(self.root), "/page-size 5")[0], 0)
        search_exit, search_payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "document",
        )

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        assert search_payload is not None
        self.assertEqual(search_payload["per_page"], 5)
        self.assertEqual(len(search_payload["results"]), 5)
        rendered = str(search_payload["rendered_markdown"])
        self.assertIn("Documents 1–5 of 12.", rendered)
        self.assertIn("Navigate: `/retriever:next` for the next page.", rendered)

    def test_slash_search_drops_stale_display_columns_with_warning(self) -> None:
        (self.root / "sample.txt").write_text("sample body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        self.paths["session_path"].write_text(
            json.dumps(
                {
                    "schema_version": retriever_tools.SESSION_SCHEMA_VERSION,
                    "scope": {},
                    "browsing": {},
                    "display": {"columns": ["title", "missing_column", "has_attachments"]},
                },
                indent=2,
                sort_keys=True,
            )
            + "\n",
            encoding="utf-8",
        )

        search_exit, search_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "sample")

        self.assertEqual(search_exit, 0)
        self.assertIsNotNone(search_payload)
        self.assertEqual(search_payload["display"]["columns"], ["title"])
        self.assertIn("warnings", search_payload)
        self.assertIn("missing_column", search_payload["warnings"][0])
        persisted_session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(
            persisted_session_payload["display"]["documents"]["columns"],
            ["title"],
        )

    def test_deleting_source_backed_dataset_hides_documents_until_reingest(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("sample dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        original_row = self.fetch_document_row("sample.txt")

        exit_code, delete_payload, _, _ = self.run_cli(
            "delete-dataset",
            str(self.root),
            "--dataset-name",
            self.root.name,
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(delete_payload)
        self.assertEqual(delete_payload["deleted_dataset"]["dataset_name"], self.root.name)
        self.assertEqual(delete_payload["documents_without_dataset_memberships"], [original_row["id"]])

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_result["total_hits"], 0)

        reingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(reingest_result["skipped"], 1)

        restored_row = self.fetch_document_row("sample.txt")
        self.assertEqual(restored_row["id"], original_row["id"])
        self.assertEqual(restored_row["control_number"], original_row["control_number"])

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_result["total_hits"], 1)
        self.assertEqual(browse_result["results"][0]["dataset_name"], self.root.name)

    def test_ingest_supports_xls_sheet_previews(self) -> None:
        xls_path = self.root / "ledger.xls"
        self.write_xls_fixture(xls_path)

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("ledger.xls")
        self.assertEqual(row["content_type"], "Spreadsheet / Table")
        self.assertEqual(row["page_count"], 2)

        search_result = retriever_tools.search(self.root, "Memo", None, None, None, 1, 20)
        self.assertEqual(search_result["results"][0]["file_name"], "ledger.xls")
        self.assertEqual(search_result["results"][0]["preview_targets"][0]["preview_type"], "csv")
        self.assertTrue(search_result["results"][0]["preview_rel_path"].endswith(".csv"))
        self.assertEqual(
            [target["label"] for target in search_result["results"][0]["preview_targets"]],
            ["Sheet1", "Notes"],
        )
        value_only_result = retriever_tools.search(self.root, "Budget approved", None, None, None, 1, 20)
        self.assertEqual(value_only_result["total_hits"], 0)

    def test_ingest_supports_xlsx_structural_summary_and_sheet_chunks(self) -> None:
        xlsx_path = self.root / "budget.xlsx"
        self.write_xlsx_fixture(xlsx_path)

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("budget.xlsx")
        self.assertEqual(row["content_type"], "Spreadsheet / Table")
        self.assertEqual(row["page_count"], 2)
        self.assertEqual(row["author"], "Rachel Green")
        self.assertEqual(row["title"], "Quarterly Budget Workbook")
        self.assertEqual(row["subject"], "Finance Planning")
        self.assertEqual(row["date_created"], "2026-04-20T10:00:00Z")
        self.assertIsNotNone(row["date_modified"])
        self.assertIn("Rachel Green", row["participants"])
        self.assertIn("Sergey", row["participants"])

        search_result = retriever_tools.search(self.root, "Budget Totals", None, None, None, 1, 20)
        self.assertEqual(search_result["results"][0]["file_name"], "budget.xlsx")
        self.assertEqual(search_result["results"][0]["preview_targets"][0]["preview_type"], "csv")
        self.assertEqual(
            [target["label"] for target in search_result["results"][0]["preview_targets"]],
            ["Budget", "Notes"],
        )

        self.assertEqual(
            retriever_tools.search(self.root, "Needs review", None, None, None, 1, 20)["results"][0]["file_name"],
            "budget.xlsx",
        )
        self.assertEqual(
            retriever_tools.search(self.root, "DeptList", None, None, None, 1, 20)["results"][0]["file_name"],
            "budget.xlsx",
        )
        self.assertEqual(retriever_tools.search(self.root, "Budget approved", None, None, None, 1, 20)["total_hits"], 0)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            chunk_rows = connection.execute(
                """
                SELECT text_content
                FROM document_chunks
                WHERE document_id = ?
                ORDER BY chunk_index ASC
                """,
                (row["id"],),
            ).fetchall()
        finally:
            connection.close()
        self.assertGreaterEqual(len(chunk_rows), 3)
        self.assertTrue(any("Sheet: Budget" in chunk_row["text_content"] for chunk_row in chunk_rows))
        self.assertTrue(any("Sheet: Notes" in chunk_row["text_content"] for chunk_row in chunk_rows))

    def test_ingest_supports_csv_structural_summary_without_indexing_values(self) -> None:
        csv_path = self.root / "pipeline.csv"
        csv_path.write_text(
            "sep=,\nCustomer,Amount,Status\nAcme Corp,100,Paid\nBeta LLC,250,Hold\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("pipeline.csv")
        self.assertEqual(row["content_type"], "Spreadsheet / Table")
        self.assertEqual(row["page_count"], 1)

        search_result = retriever_tools.search(self.root, "Status", None, None, None, 1, 20)
        self.assertEqual(search_result["results"][0]["file_name"], "pipeline.csv")
        self.assertEqual(search_result["results"][0]["preview_targets"][0]["preview_type"], "native")
        self.assertEqual(retriever_tools.search(self.root, "Acme Corp", None, None, None, 1, 20)["total_hits"], 0)

    def test_ingest_supports_pptx_deck_preview_images_and_notes(self) -> None:
        pptx_path = self.root / "deck.pptx"
        self.write_pptx_fixture(pptx_path)

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)

        row = self.fetch_document_row("deck.pptx")
        self.assertEqual(row["content_type"], "Presentation")
        self.assertEqual(row["author"], "Rachel Green")
        self.assertEqual(row["title"], "Quarterly Strategy Deck")
        self.assertEqual(row["subject"], "Board Update")
        self.assertEqual(row["date_created"], "2026-04-15T10:00:00Z")
        self.assertEqual(row["date_modified"], "2026-04-15T11:00:00Z")
        self.assertEqual(row["page_count"], 2)
        self.assertEqual(row["text_status"], "ok")

        search_result = retriever_tools.search(self.root, "retained enterprise customers", None, None, None, 1, 20)
        result = search_result["results"][0]
        self.assertEqual(result["file_name"], "deck.pptx")
        self.assertTrue(result["preview_rel_path"].endswith(".html"))
        self.assertEqual(result["preview_targets"][0]["preview_type"], "html")

        preview_path = Path(result["preview_targets"][0]["abs_path"])
        preview_html = preview_path.read_text(encoding="utf-8")
        self.assertTrue(preview_html.startswith("<!DOCTYPE html>"))
        self.assertIn("<head>", preview_html)
        self.assertIn('<meta charset="utf-8"/>', preview_html)
        self.assertIn("<title>Quarterly Strategy Deck</title>", preview_html)
        self.assertIn("<h1>Quarterly Strategy Deck</h1>", preview_html)
        self.assertIn("Slide 1", preview_html)
        self.assertIn("Slide 2", preview_html)
        self.assertIn("Q3 Revenue Review", preview_html)
        self.assertIn("Hiring Plan", preview_html)
        self.assertIn("Speaker Notes", preview_html)
        self.assertIn("Emphasize retained enterprise customers.", preview_html)
        self.assertIn("data:image/png;base64,", preview_html)
        self.assertIn("Stark Therapeutics logo", preview_html)

        second_search = retriever_tools.search(self.root, "Open 12 roles in engineering", None, None, None, 1, 20)
        self.assertEqual(second_search["results"][0]["file_name"], "deck.pptx")

    def test_normalize_datetime_prefers_iso_before_pdf_date_parsing(self) -> None:
        self.assertEqual(
            retriever_tools.normalize_datetime("2023-01-18T09:00:00Z"),
            "2023-01-18T09:00:00Z",
        )
        self.assertEqual(
            retriever_tools.normalize_datetime("2023-01-18T09:00:00+00:00"),
            "2023-01-18T09:00:00Z",
        )
        self.assertEqual(
            retriever_tools.normalize_datetime("2023-01-18"),
            "2023-01-18T00:00:00Z",
        )
        self.assertEqual(
            retriever_tools.normalize_datetime("2023-01-18T09:00:00.123Z"),
            "2023-01-18T09:00:00Z",
        )
        self.assertEqual(
            retriever_tools.normalize_datetime("D:20230118090000"),
            "2023-01-18T09:00:00Z",
        )

    def test_ingest_supports_curated_source_and_config_text_files(self) -> None:
        text_files = {
            "main.py": "source marker python\n",
            "app.js": "source marker javascript\n",
            "types.ts": "source marker typescript\n",
            "widget.jsx": "source marker jsx\n",
            "view.tsx": "source marker tsx\n",
            "service.java": "source marker java\n",
            "worker.go": "source marker go\n",
            "script.rb": "source marker ruby\n",
            "index.php": "source marker php\n",
            "main.c": "source marker c\n",
            "header.h": "source marker h\n",
            "engine.cpp": "source marker cpp\n",
            "engine.hpp": "source marker hpp\n",
            "app.cs": "source marker csharp\n",
            "lib.rs": "source marker rust\n",
            "mobile.swift": "source marker swift\n",
            "service.kt": "source marker kotlin\n",
            "run.sh": "source marker sh\n",
            "login.bash": "source marker bash\n",
            "setup.zsh": "source marker zsh\n",
            "deploy.ps1": "source marker powershell\n",
            "app.yaml": "source marker yaml\n",
            "config.yml": "source marker yml\n",
            "settings.toml": "source marker toml\n",
            "local.ini": "source marker ini\n",
            "build.cfg": "source marker cfg\n",
            "service.conf": "source marker conf\n",
            "app.properties": "source marker properties\n",
            "schema.xml": "<config>source marker xml</config>\n",
            "query.sql": "select 'source marker sql';\n",
            "style.css": "/* source marker css */\n",
            "theme.scss": "/* source marker scss */\n",
            "theme.less": "/* source marker less */\n",
        }
        for file_name, content in text_files.items():
            (self.root / file_name).write_text(content, encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], len(text_files))
        self.assertEqual(ingest_result["failed"], 0)

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 100)
        for file_name in text_files:
            row = self.fetch_document_row(file_name)
            self.assertEqual(row["text_status"], "ok")
            result = next(item for item in browse_result["results"] if item["file_name"] == Path(file_name).name)
            self.assertEqual(result["preview_rel_path"], file_name)
            self.assertEqual(result["preview_targets"][0]["preview_type"], "native")

        xml_row = self.fetch_document_row("schema.xml")
        self.assertEqual(xml_row["content_type"], "Web")

        search_result = retriever_tools.search(self.root, "source marker", None, None, None, 1, 100)
        self.assertGreaterEqual(search_result["total_hits"], len(text_files))

    def test_v5_to_v7_migration_then_touch_reingest_keeps_backfilled_control_number(self) -> None:
        self.create_legacy_documents_table(with_row=True)
        legacy_path = self.root / "legacy.txt"
        legacy_path.write_text("updated legacy body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        migrated_row = self.fetch_document_row("legacy.txt")
        self.assertEqual(migrated_row["control_number"], "DOC001.00000001")
        self.assertEqual(migrated_row["control_number_batch"], 1)
        self.assertEqual(migrated_row["control_number_family_sequence"], 1)

        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["updated"], 1)

        updated_row = self.fetch_document_row("legacy.txt")
        self.assertEqual(updated_row["control_number"], migrated_row["control_number"])
        self.assertEqual(updated_row["control_number_batch"], 1)
        self.assertEqual(updated_row["control_number_family_sequence"], 1)

    def test_bootstrap_renames_v6_display_columns_to_control_number(self) -> None:
        self.create_v6_documents_table(with_row=True)
        legacy_path = self.root / "legacy.txt"
        legacy_path.write_text("updated legacy body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            columns = retriever_tools.table_columns(connection, "documents")
            self.assertIn("control_number", columns)
            self.assertNotIn("display_id", columns)
            row = connection.execute(
                """
                SELECT control_number, control_number_batch, control_number_family_sequence
                FROM documents
                WHERE rel_path = 'legacy.txt'
                """
            ).fetchone()
            table_names = {
                item["name"]
                for item in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'table'"
                ).fetchall()
            }
        finally:
            connection.close()

        self.assertEqual(row["control_number"], "DOC001.00000001")
        self.assertEqual(row["control_number_batch"], 1)
        self.assertEqual(row["control_number_family_sequence"], 1)
        self.assertIn("control_number_batches", table_names)
        self.assertNotIn("display_id_batches", table_names)

        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["updated"], 1)

        updated_row = self.fetch_document_row("legacy.txt")
        self.assertEqual(updated_row["control_number"], "DOC001.00000001")

    def test_bootstrap_repairs_partial_control_number_identity_fields(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("hello\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["new"], 1)

        row = self.fetch_document_row("sample.txt")
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute(
                """
                UPDATE documents
                SET control_number_batch = NULL, control_number_family_sequence = NULL
                WHERE id = ?
                """,
                (row["id"],),
            )
            connection.commit()
        finally:
            connection.close()

        retriever_tools.bootstrap(self.root)
        repaired_row = self.fetch_document_row("sample.txt")
        self.assertEqual(repaired_row["control_number"], row["control_number"])
        self.assertEqual(repaired_row["control_number_batch"], 1)
        self.assertEqual(repaired_row["control_number_family_sequence"], 1)

    def test_reingest_preserves_attachment_control_number_and_manual_locks(self) -> None:
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Phase 3.5 test",
            body_text="Original parent body",
            attachment_name="notes.txt",
            attachment_text="stable attachment body",
        )

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["new"], 1)

        parent_row = self.fetch_document_row("thread.eml")
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        retriever_tools.set_field(self.root, child_row["id"], "title", "Manual Attachment Title")

        self.write_email_message(
            email_path,
            subject="Phase 3.5 test updated",
            body_text="Updated parent body only",
            attachment_name="notes.txt",
            attachment_text="stable attachment body",
        )
        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_ingest["updated"], 1)
        self.assertEqual(second_ingest["failed"], 0)

        updated_parent = self.fetch_document_row("thread.eml")
        updated_child = self.fetch_child_rows(updated_parent["id"])[0]
        self.assertEqual(updated_parent["control_number"], parent_row["control_number"])
        self.assertEqual(updated_child["id"], child_row["id"])
        self.assertEqual(updated_child["control_number"], child_row["control_number"])
        self.assertEqual(updated_child["title"], "Manual Attachment Title")

        self.write_email_message(
            email_path,
            subject="Phase 3.5 test no attachment",
            body_text="Updated parent body without attachment",
        )
        third_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(third_ingest["updated"], 1)

        removed_child = self.fetch_document_by_id(child_row["id"])
        self.assertEqual(removed_child["lifecycle_status"], "deleted")

        parent_search = retriever_tools.search(self.root, "Phase 3.5 test no attachment", None, None, None, 1, 20)
        parent_result = next(item for item in parent_search["results"] if item["id"] == updated_parent["id"])
        self.assertEqual(parent_result["attachment_count"], 0)

    def test_rerun_ingest_after_attachment_extraction_keeps_document_total_stable(self) -> None:
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Attachment rerun guard",
            body_text="Parent body",
            attachment_name="notes.txt",
            attachment_text="attachment body",
        )

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["new"], 1)
        self.assertEqual(first_ingest["failed"], 0)

        parent_row = self.fetch_document_row("thread.eml")
        child_rows = self.fetch_child_rows(parent_row["id"])
        self.assertEqual(len(child_rows), 1)

        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_ingest["skipped"], 1)
        self.assertEqual(second_ingest["new"], 0)
        self.assertEqual(second_ingest["failed"], 0)

        browse_results = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_results["total_hits"], 2)
        attachments_only = retriever_tools.search(
            self.root,
            "",
            [["is_attachment", "eq", "true"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(attachments_only["total_hits"], 1)
        self.assertTrue(
            all(
                not str(item["rel_path"]).startswith(
                    f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/previews/thread.eml/attachments/"
                )
                or item["parent_document_id"] is not None
                for item in browse_results["results"]
            )
        )

    def test_ingest_reuses_workspace_dataset_on_reingest(self) -> None:
        alpha_path = self.root / "alpha.txt"
        alpha_path.write_text("alpha dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["new"], 1)
        self.assertEqual(first_ingest["failed"], 0)

        alpha_row = self.fetch_document_row("alpha.txt")
        self.assertIsNotNone(alpha_row["dataset_id"])
        initial_dataset_id = int(alpha_row["dataset_id"])
        initial_dataset_row = self.fetch_dataset_row(initial_dataset_id)
        self.assertEqual(initial_dataset_row["source_kind"], retriever_tools.FILESYSTEM_SOURCE_KIND)
        self.assertEqual(initial_dataset_row["dataset_locator"], ".")
        self.assertEqual(initial_dataset_row["dataset_name"], self.root.name)

        beta_path = self.root / "beta.txt"
        beta_path.write_text("beta dataset body\n", encoding="utf-8")

        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_ingest["new"], 1)
        self.assertEqual(second_ingest["skipped"], 1)
        self.assertEqual(second_ingest["failed"], 0)

        updated_alpha_row = self.fetch_document_row("alpha.txt")
        beta_row = self.fetch_document_row("beta.txt")
        self.assertEqual(updated_alpha_row["dataset_id"], initial_dataset_id)
        self.assertEqual(beta_row["dataset_id"], initial_dataset_id)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            dataset_rows = connection.execute(
                """
                SELECT *
                FROM datasets
                ORDER BY id ASC
                """
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(len(dataset_rows), 1)
        dataset_filtered = retriever_tools.search(
            self.root,
            "",
            [["dataset", "eq", self.root.name]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(dataset_filtered["total_hits"], 2)
        self.assertTrue(all(item["dataset_name"] == self.root.name for item in dataset_filtered["results"]))

    def test_document_can_belong_to_multiple_datasets_without_duplicate_search_hits(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("sample dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            review_dataset_id = retriever_tools.create_dataset_row(connection, "Review Set")
            retriever_tools.ensure_dataset_document_membership(
                connection,
                dataset_id=review_dataset_id,
                document_id=int(row["id"]),
                dataset_source_id=None,
            )
            connection.commit()
        finally:
            connection.close()

        dataset_filtered = retriever_tools.search(
            self.root,
            "",
            [["dataset_name", "eq", "Review Set"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(dataset_filtered["total_hits"], 1)
        result = dataset_filtered["results"][0]
        self.assertEqual(result["id"], row["id"])
        self.assertEqual(sorted(result["dataset_names"]), ["Review Set", self.root.name])
        self.assertIsNone(result["dataset_id"])
        self.assertNotIn("dataset_name", result)

    def test_exact_duplicate_files_share_one_canonical_document_and_two_occurrences(self) -> None:
        left_dir = self.root / "custodian-a"
        right_dir = self.root / "custodian-b"
        left_dir.mkdir()
        right_dir.mkdir()
        duplicate_text = "duplicate witness line\nsecond line\n"
        (left_dir / "dup.txt").write_text(duplicate_text, encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["new"], 1)
        self.assertEqual(first_ingest["failed"], 0)

        (right_dir / "dup-copy.txt").write_text(duplicate_text, encoding="utf-8")
        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_ingest["failed"], 0)

        search_result = retriever_tools.search(self.root, "duplicate witness line", None, None, None, 1, 20)
        self.assertEqual(search_result["total_hits"], 1)
        document_id = int(search_result["results"][0]["id"])
        occurrence_rows = self.fetch_occurrence_rows(document_id)
        active_rel_paths = sorted(
            str(row["rel_path"])
            for row in occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
        )
        self.assertEqual(active_rel_paths, ["custodian-a/dup.txt", "custodian-b/dup-copy.txt"])
        self.assertEqual(self.count_rows("documents"), 1)
        self.assertEqual(self.count_rows("document_occurrences"), 2)

        chunk_count = retriever_tools.search_chunks(
            self.root,
            "duplicate witness line",
            None,
            None,
            None,
            10,
            3,
            count_only=True,
            distinct_docs=True,
        )
        self.assertEqual(chunk_count["documents_with_hits"], 1)

    def test_rel_path_filter_picks_matching_occurrence_for_duplicate_result_links(self) -> None:
        left_dir = self.root / "custodian-a"
        right_dir = self.root / "custodian-b"
        left_dir.mkdir()
        right_dir.mkdir()
        duplicate_text = "same logical document body\n"
        (left_dir / "dup.txt").write_text(duplicate_text, encoding="utf-8")
        (right_dir / "dup-copy.txt").write_text(duplicate_text, encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        filtered_result = retriever_tools.search(
            self.root,
            "",
            [["rel_path", "eq", "custodian-b/dup-copy.txt"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(filtered_result["total_hits"], 1)
        result = filtered_result["results"][0]
        self.assertEqual(result["rel_path"], "custodian-b/dup-copy.txt")
        self.assertEqual(result["abs_path"], str(self.root / "custodian-b" / "dup-copy.txt"))
        self.assertEqual(result["source_rel_path"], "custodian-b/dup-copy.txt")

    def test_delete_docs_requires_confirm_and_deletes_document_family(self) -> None:
        self.write_email_message(
            self.root / "thread.eml",
            subject="Delete family thread",
            body_text="family delete body\n",
            attachment_name="notes.txt",
            attachment_text="attached note\n",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        root_row = self.fetch_document_row("thread.eml")
        child_rows = self.fetch_child_rows(int(root_row["id"]))
        self.assertEqual(len(child_rows), 1)

        preview_exit, preview_payload, _, _ = self.run_cli(
            "delete-docs",
            str(self.root),
            "--doc-id",
            str(root_row["id"]),
        )
        self.assertEqual(preview_exit, 0)
        self.assertIsNotNone(preview_payload)
        self.assertEqual(preview_payload["status"], "confirm_required")
        self.assertEqual(preview_payload["matched_document_count"], 1)
        self.assertEqual(preview_payload["would_delete_occurrences"], 2)
        self.assertEqual(preview_payload["would_touch_documents"], 2)

        apply_exit, apply_payload, _, _ = self.run_cli(
            "delete-docs",
            str(self.root),
            "--doc-id",
            str(root_row["id"]),
            "--confirm",
        )
        self.assertEqual(apply_exit, 0)
        self.assertIsNotNone(apply_payload)
        self.assertEqual(apply_payload["status"], "ok")
        self.assertEqual(apply_payload["deleted_occurrences"], 2)
        self.assertEqual(apply_payload["deleted_documents"], 2)

        deleted_root_row = self.fetch_document_by_id(int(root_row["id"]))
        self.assertEqual(deleted_root_row["lifecycle_status"], "deleted")
        deleted_child_row = self.fetch_document_by_id(int(child_rows[0]["id"]))
        self.assertEqual(deleted_child_row["lifecycle_status"], "deleted")

        root_search = retriever_tools.search(self.root, "family delete body", None, None, None, 1, 20)
        self.assertEqual(root_search["total_hits"], 0)
        attachment_search = retriever_tools.search(self.root, "attached note", None, None, None, 1, 20)
        self.assertEqual(attachment_search["total_hits"], 0)

    def test_delete_docs_filter_deletes_only_matching_duplicate_occurrence(self) -> None:
        left_dir = self.root / "custodian-a"
        right_dir = self.root / "custodian-b"
        left_dir.mkdir()
        right_dir.mkdir()
        duplicate_text = "same logical document body\n"
        (left_dir / "dup.txt").write_text(duplicate_text, encoding="utf-8")
        (right_dir / "dup-copy.txt").write_text(duplicate_text, encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        search_result = retriever_tools.search(self.root, "same logical document body", None, None, None, 1, 20)
        self.assertEqual(search_result["total_hits"], 1)
        document_id = int(search_result["results"][0]["id"])

        preview_exit, preview_payload, _, _ = self.run_cli(
            "delete-docs",
            str(self.root),
            "--filter",
            "rel_path = 'custodian-b/dup-copy.txt'",
        )
        self.assertEqual(preview_exit, 0)
        self.assertIsNotNone(preview_payload)
        self.assertEqual(preview_payload["status"], "confirm_required")
        self.assertTrue(preview_payload["occurrence_scoped"])
        self.assertEqual(preview_payload["matched_document_count"], 1)
        self.assertEqual(preview_payload["would_delete_occurrences"], 1)
        self.assertEqual(preview_payload["would_touch_documents"], 1)

        apply_exit, apply_payload, _, _ = self.run_cli(
            "delete-docs",
            str(self.root),
            "--filter",
            "rel_path = 'custodian-b/dup-copy.txt'",
            "--confirm",
        )
        self.assertEqual(apply_exit, 0)
        self.assertIsNotNone(apply_payload)
        self.assertEqual(apply_payload["status"], "ok")
        self.assertEqual(apply_payload["deleted_occurrences"], 1)
        self.assertEqual(apply_payload["deleted_documents"], 0)
        self.assertEqual(apply_payload["retained_documents"], 1)

        occurrence_rows = self.fetch_occurrence_rows(document_id)
        active_rel_paths = sorted(
            str(row["rel_path"])
            for row in occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
        )
        self.assertEqual(active_rel_paths, ["custodian-a/dup.txt"])

        deleted_filter_result = retriever_tools.search(
            self.root,
            "",
            [["rel_path", "eq", "custodian-b/dup-copy.txt"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(deleted_filter_result["total_hits"], 0)
        surviving_filter_result = retriever_tools.search(
            self.root,
            "",
            [["rel_path", "eq", "custodian-a/dup.txt"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(surviving_filter_result["total_hits"], 1)

    def test_delete_docs_path_prefix_shortcut_removes_matching_rows_and_dataset_membership(self) -> None:
        raw_dir = self.root / "raw"
        keep_dir = self.root / "keep"
        raw_dir.mkdir()
        keep_dir.mkdir()
        (raw_dir / "a.txt").write_text("raw body\n", encoding="utf-8")
        (keep_dir / "b.txt").write_text("keep body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        preview_exit, preview_payload, _, _ = self.run_cli(
            "delete-docs",
            str(self.root),
            "--path",
            "raw",
        )
        self.assertEqual(preview_exit, 0)
        self.assertIsNotNone(preview_payload)
        self.assertEqual(preview_payload["status"], "confirm_required")
        self.assertTrue(preview_payload["occurrence_scoped"])
        self.assertEqual(preview_payload["matched_document_count"], 1)

        apply_exit, apply_payload, _, _ = self.run_cli(
            "delete-docs",
            str(self.root),
            "--path",
            "raw",
            "--confirm",
        )
        self.assertEqual(apply_exit, 0)
        self.assertIsNotNone(apply_payload)
        self.assertEqual(apply_payload["status"], "ok")
        self.assertEqual(apply_payload["deleted_documents"], 1)

        deleted_row = self.fetch_document_row("raw/a.txt")
        self.assertEqual(deleted_row["lifecycle_status"], "deleted")
        kept_row = self.fetch_document_row("keep/b.txt")
        self.assertEqual(kept_row["lifecycle_status"], "active")

        deleted_search = retriever_tools.search(self.root, "", [["rel_path", "eq", "raw/a.txt"]], None, None, 1, 20)
        self.assertEqual(deleted_search["total_hits"], 0)
        kept_search = retriever_tools.search(self.root, "", [["rel_path", "eq", "keep/b.txt"]], None, None, 1, 20)
        self.assertEqual(kept_search["total_hits"], 1)

        list_exit, list_payload, _, _ = self.run_cli("list-datasets", str(self.root))
        self.assertEqual(list_exit, 0)
        self.assertIsNotNone(list_payload)
        dataset_entry = {item["dataset_name"]: item for item in list_payload["datasets"]}[self.root.name]
        self.assertEqual(dataset_entry["document_count"], 1)

    def test_sql_like_custodian_filter_picks_matching_occurrence_for_duplicate_links(self) -> None:
        left_dir = self.root / "custodian-a"
        right_dir = self.root / "custodian-b"
        left_dir.mkdir()
        right_dir.mkdir()
        duplicate_text = "same logical document body\n"
        (left_dir / "dup.txt").write_text(duplicate_text, encoding="utf-8")
        (right_dir / "dup-copy.txt").write_text(duplicate_text, encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        search_result = retriever_tools.search(self.root, "same logical document body", None, None, None, 1, 20)
        self.assertEqual(search_result["total_hits"], 1)
        document_id = int(search_result["results"][0]["id"])
        occurrence_rows = self.fetch_occurrence_rows(document_id)
        occurrence_ids_by_path = {str(row["rel_path"]): int(row["id"]) for row in occurrence_rows}

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            now = retriever_tools.utc_now()
            connection.execute(
                """
                UPDATE document_occurrences
                SET custodian = ?, updated_at = ?
                WHERE id = ?
                """,
                ("mailbox-a", now, occurrence_ids_by_path["custodian-a/dup.txt"]),
            )
            connection.execute(
                """
                UPDATE document_occurrences
                SET custodian = ?, updated_at = ?
                WHERE id = ?
                """,
                ("mailbox-b", now, occurrence_ids_by_path["custodian-b/dup-copy.txt"]),
            )
            retriever_tools.refresh_document_from_occurrences(connection, document_id)
            connection.commit()
        finally:
            connection.close()

        filter_expression = f"custodian = 'mailbox-b' AND dataset_name = '{self.root.name}'"
        exit_code, payload, _, _ = self.run_cli(
            "search",
            str(self.root),
            "",
            "--filter",
            filter_expression,
            "--verbose",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["total_hits"], 1)
        result = payload["results"][0]
        self.assertEqual(result["rel_path"], "custodian-b/dup-copy.txt")
        self.assertEqual(
            Path(result["abs_path"]).resolve(),
            (self.root / "custodian-b" / "dup-copy.txt").resolve(),
        )
        self.assertEqual(result["source_rel_path"], "custodian-b/dup-copy.txt")
        self.assertEqual(set(result["custodians"]), {"mailbox-a", "mailbox-b"})

        chunk_exit, chunk_payload, _, _ = self.run_cli(
            "search-chunks",
            str(self.root),
            "same logical document body",
            "--filter",
            filter_expression,
            "--verbose",
        )
        self.assertEqual(chunk_exit, 0)
        self.assertIsNotNone(chunk_payload)
        self.assertEqual(len(chunk_payload["results"]), 1)
        chunk_result = chunk_payload["results"][0]
        self.assertEqual(chunk_result["document_id"], document_id)
        self.assertEqual(chunk_result["rel_path"], "custodian-b/dup-copy.txt")
        self.assertEqual(chunk_result["file_name"], "dup-copy.txt")

    def test_exact_duplicate_email_family_keeps_attachment_occurrences_for_both_custodians(self) -> None:
        left_dir = self.root / "custodian-a"
        right_dir = self.root / "custodian-b"
        left_dir.mkdir()
        right_dir.mkdir()

        left_path = left_dir / "message.eml"
        right_path = right_dir / "message-copy.eml"
        self.write_email_message(
            left_path,
            subject="Family duplicate",
            body_text="Shared family body.\n",
            attachment_name="notes.txt",
            attachment_text="shared attachment detail",
            message_id="<family-duplicate@example.com>",
        )
        shutil.copyfile(left_path, right_path)

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        root_search = retriever_tools.search(self.root, "Family duplicate", None, None, None, 1, 20)
        self.assertEqual(root_search["total_hits"], 1)
        root_document_id = int(root_search["results"][0]["id"])
        root_occurrence_rows = self.fetch_occurrence_rows(root_document_id)
        active_root_occurrence_rows = [
            row
            for row in root_occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
        ]
        self.assertEqual(
            sorted(str(row["rel_path"]) for row in active_root_occurrence_rows),
            ["custodian-a/message.eml", "custodian-b/message-copy.eml"],
        )

        child_rows = self.fetch_child_rows(root_document_id)
        self.assertEqual(len(child_rows), 1)
        child_row = child_rows[0]
        child_occurrence_rows = self.fetch_occurrence_rows(int(child_row["id"]))
        active_child_occurrence_rows = [
            row
            for row in child_occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
        ]
        self.assertEqual(len(active_child_occurrence_rows), 2)
        self.assertEqual(
            {int(row["parent_occurrence_id"]) for row in active_child_occurrence_rows},
            {int(row["id"]) for row in active_root_occurrence_rows},
        )
        self.assertTrue(
            any(
                str(row["rel_path"]).startswith(
                    f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/previews/custodian-a/message.eml/attachments/"
                )
                for row in active_child_occurrence_rows
            )
        )
        self.assertTrue(
            any(
                str(row["rel_path"]).startswith(
                    f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/previews/custodian-b/message-copy.eml/attachments/"
                )
                for row in active_child_occurrence_rows
            )
        )
        right_occurrence_rel_path = next(
            str(row["rel_path"])
            for row in active_child_occurrence_rows
            if str(row["rel_path"]).startswith(
                f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/previews/custodian-b/message-copy.eml/attachments/"
            )
        )

        filtered_attachment_search = retriever_tools.search(
            self.root,
            "shared attachment detail",
            [["rel_path", "eq", right_occurrence_rel_path]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(filtered_attachment_search["total_hits"], 1)
        self.assertEqual(filtered_attachment_search["results"][0]["id"], child_row["id"])
        self.assertTrue(
            filtered_attachment_search["results"][0]["rel_path"].startswith(
                f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/previews/custodian-b/message-copy.eml/attachments/"
            )
        )

    def test_reconcile_duplicates_merges_same_email_attachment_family(self) -> None:
        self.write_email_message(
            self.root / "alpha.eml",
            subject="Family merge ready",
            body_text="Shared family merge body.\n",
            attachment_name="notes.txt",
            attachment_text="family attachment detail",
            message_id="<alpha-family@example.com>",
        )
        self.write_email_message(
            self.root / "beta.eml",
            subject="Family merge ready",
            body_text="Shared family merge body.\n",
            attachment_name="notes.txt",
            attachment_text="family attachment detail",
            message_id="<beta-family@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        exit_code, dry_run_payload, _, _ = self.run_cli(
            "reconcile-duplicates",
            str(self.root),
            "--dry-run",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(dry_run_payload)
        self.assertEqual(dry_run_payload["candidate_group_count"], 1)
        candidate_group = dry_run_payload["candidate_groups"][0]
        self.assertEqual(candidate_group["status"], "ready")
        self.assertEqual(candidate_group["family_child_group_count"], 1)

        exit_code, apply_payload, _, _ = self.run_cli(
            "reconcile-duplicates",
            str(self.root),
            "--apply",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(apply_payload)
        self.assertEqual(apply_payload["merged_group_count"], 1)
        merged_group = apply_payload["applied_groups"][0]
        self.assertEqual(merged_group["status"], "merged")
        self.assertEqual(merged_group["family_child_group_count"], 1)
        self.assertEqual(len(merged_group["merge_event_ids"]), 2)

        survivor_document_id = int(merged_group["survivor_document_id"])
        root_occurrence_rows = self.fetch_occurrence_rows(survivor_document_id)
        active_root_rel_paths = sorted(
            str(row["rel_path"])
            for row in root_occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
        )
        self.assertEqual(active_root_rel_paths, ["alpha.eml", "beta.eml"])

        survivor_child_rows = self.fetch_child_rows(survivor_document_id)
        self.assertEqual(len(survivor_child_rows), 1)
        child_occurrence_rows = self.fetch_occurrence_rows(int(survivor_child_rows[0]["id"]))
        active_child_rel_paths = sorted(
            str(row["rel_path"])
            for row in child_occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
        )
        self.assertEqual(len(active_child_rel_paths), 2)

        subject_search = retriever_tools.search(self.root, "Family merge ready", None, None, None, 1, 20)
        self.assertEqual(subject_search["total_hits"], 1)
        self.assertEqual(subject_search["results"][0]["id"], survivor_document_id)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            merged_document_count = connection.execute(
                """
                SELECT COUNT(*) AS row_count
                FROM documents
                WHERE canonical_status = ?
                """,
                (retriever_tools.CANONICAL_STATUS_MERGED,),
            ).fetchone()
            merge_event_count = connection.execute(
                "SELECT COUNT(*) AS row_count FROM document_merge_events",
            ).fetchone()
        finally:
            connection.close()
        self.assertEqual(int(merged_document_count["row_count"] or 0), 2)
        self.assertEqual(int(merge_event_count["row_count"] or 0), 2)

    def test_reconcile_duplicates_skips_same_email_body_with_different_attachment_family(self) -> None:
        self.write_email_message(
            self.root / "alpha.eml",
            subject="Family mismatch",
            body_text="Shared body but different family.\n",
            attachment_name="notes.txt",
            attachment_text="alpha attachment detail",
            message_id="<alpha-family-mismatch@example.com>",
        )
        self.write_email_message(
            self.root / "beta.eml",
            subject="Family mismatch",
            body_text="Shared body but different family.\n",
            attachment_name="notes.txt",
            attachment_text="beta attachment detail",
            message_id="<beta-family-mismatch@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        exit_code, dry_run_payload, _, _ = self.run_cli(
            "reconcile-duplicates",
            str(self.root),
            "--dry-run",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(dry_run_payload)
        self.assertEqual(dry_run_payload["candidate_group_count"], 0)
        self.assertEqual(dry_run_payload["mergeable_group_count"], 0)
        self.assertEqual(dry_run_payload["blocked_group_count"], 0)

        exit_code, apply_payload, _, _ = self.run_cli(
            "reconcile-duplicates",
            str(self.root),
            "--apply",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(apply_payload)
        self.assertEqual(apply_payload["merged_group_count"], 0)
        self.assertEqual(apply_payload["blocked_group_count"], 0)

        subject_search = retriever_tools.search(self.root, "Family mismatch", None, None, None, 1, 20)
        self.assertEqual(subject_search["total_hits"], 2)

    def test_reconcile_duplicates_dry_run_and_apply_merge_ready_content_hash_group(self) -> None:
        (self.root / "alpha.txt").write_text("shared clause alpha version\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("shared clause beta version\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        alpha_row = self.fetch_document_row("alpha.txt")
        beta_row = self.fetch_document_row("beta.txt")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            review_dataset_id = retriever_tools.create_dataset_row(connection, "Review Set")
            retriever_tools.ensure_dataset_document_membership(
                connection,
                dataset_id=review_dataset_id,
                document_id=int(beta_row["id"]),
                dataset_source_id=None,
            )
            retriever_tools.refresh_document_dataset_cache(connection, int(beta_row["id"]))
            connection.execute(
                """
                UPDATE documents
                SET content_hash = ?, updated_at = ?
                WHERE id = ?
                """,
                (alpha_row["content_hash"], retriever_tools.utc_now(), int(beta_row["id"])),
            )
            connection.commit()
        finally:
            connection.close()

        exit_code, dry_run_payload, _, _ = self.run_cli(
            "reconcile-duplicates",
            str(self.root),
            "--dry-run",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(dry_run_payload)
        self.assertEqual(dry_run_payload["mode"], "dry-run")
        self.assertEqual(dry_run_payload["candidate_group_count"], 1)
        self.assertEqual(dry_run_payload["mergeable_group_count"], 1)
        candidate_group = dry_run_payload["candidate_groups"][0]
        self.assertEqual(candidate_group["status"], "ready")
        self.assertEqual(
            sorted(candidate_group["document_ids"]),
            sorted([int(alpha_row["id"]), int(beta_row["id"])]),
        )

        exit_code, apply_payload, _, _ = self.run_cli(
            "reconcile-duplicates",
            str(self.root),
            "--apply",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(apply_payload)
        self.assertEqual(apply_payload["mode"], "apply")
        self.assertEqual(apply_payload["merged_group_count"], 1)

        merged_group = apply_payload["applied_groups"][0]
        survivor_document_id = int(merged_group["survivor_document_id"])
        loser_document_id = int(merged_group["loser_document_ids"][0])
        self.assertEqual(merged_group["status"], "merged")
        self.assertEqual(len(merged_group["merge_event_ids"]), 1)

        loser_row = self.fetch_document_by_id(loser_document_id)
        self.assertEqual(loser_row["canonical_status"], retriever_tools.CANONICAL_STATUS_MERGED)
        self.assertEqual(loser_row["merged_into_document_id"], survivor_document_id)
        self.assertEqual(loser_row["lifecycle_status"], "deleted")

        occurrence_rows = self.fetch_occurrence_rows(survivor_document_id)
        active_rel_paths = sorted(
            str(row["rel_path"])
            for row in occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
        )
        self.assertEqual(active_rel_paths, ["alpha.txt", "beta.txt"])

        search_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(search_result["total_hits"], 1)
        self.assertEqual(search_result["results"][0]["id"], survivor_document_id)

        review_set_result = retriever_tools.search(
            self.root,
            "",
            [["dataset_name", "eq", "Review Set"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(review_set_result["total_hits"], 1)
        self.assertEqual(review_set_result["results"][0]["id"], survivor_document_id)

        rel_path_result = retriever_tools.search(
            self.root,
            "",
            [["rel_path", "eq", "beta.txt"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(rel_path_result["total_hits"], 1)
        self.assertEqual(rel_path_result["results"][0]["id"], survivor_document_id)
        self.assertEqual(rel_path_result["results"][0]["rel_path"], "beta.txt")

        chunk_count = retriever_tools.search_chunks(
            self.root,
            "shared",
            None,
            None,
            None,
            10,
            3,
            count_only=True,
            distinct_docs=True,
        )
        self.assertEqual(chunk_count["documents_with_hits"], 1)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            loser_chunk_count = connection.execute(
                "SELECT COUNT(*) AS row_count FROM document_chunks WHERE document_id = ?",
                (loser_document_id,),
            ).fetchone()
            loser_chunk_fts_count = connection.execute(
                "SELECT COUNT(*) AS row_count FROM chunks_fts WHERE document_id = ?",
                (loser_document_id,),
            ).fetchone()
            loser_doc_fts_count = connection.execute(
                "SELECT COUNT(*) AS row_count FROM documents_fts WHERE document_id = ?",
                (loser_document_id,),
            ).fetchone()
            merge_event_count = connection.execute(
                "SELECT COUNT(*) AS row_count FROM document_merge_events",
            ).fetchone()
            title_conflict_row = connection.execute(
                """
                SELECT field_name, resolution
                FROM document_field_conflicts
                WHERE document_id = ?
                  AND field_name = 'title'
                ORDER BY id ASC
                LIMIT 1
                """,
                (survivor_document_id,),
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(int(loser_chunk_count["row_count"] or 0), 0)
        self.assertEqual(int(loser_chunk_fts_count["row_count"] or 0), 0)
        self.assertEqual(int(loser_doc_fts_count["row_count"] or 0), 0)
        self.assertEqual(int(merge_event_count["row_count"] or 0), 1)
        self.assertIsNotNone(title_conflict_row)
        self.assertEqual(title_conflict_row["resolution"], "kept_survivor")

    def test_reconcile_duplicates_blocks_conflicting_custom_fields(self) -> None:
        (self.root / "alpha.txt").write_text("shared clause alpha version\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("shared clause beta version\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)
        self.assertEqual(ingest_result["failed"], 0)

        alpha_row = self.fetch_document_row("alpha.txt")
        beta_row = self.fetch_document_row("beta.txt")

        self.assertEqual(self.run_cli("add-field", str(self.root), "effective_date", "text")[0], 0)
        self.assertEqual(
            self.run_cli(
                "set-field",
                str(self.root),
                "--doc-id",
                str(alpha_row["id"]),
                "--field",
                "effective_date",
                "--value",
                "2026-04-01",
            )[0],
            0,
        )
        self.assertEqual(
            self.run_cli(
                "set-field",
                str(self.root),
                "--doc-id",
                str(beta_row["id"]),
                "--field",
                "effective_date",
                "--value",
                "2026-04-15",
            )[0],
            0,
        )

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute(
                """
                UPDATE documents
                SET content_hash = ?, updated_at = ?
                WHERE id = ?
                """,
                (alpha_row["content_hash"], retriever_tools.utc_now(), int(beta_row["id"])),
            )
            connection.commit()
        finally:
            connection.close()

        exit_code, dry_run_payload, _, _ = self.run_cli(
            "reconcile-duplicates",
            str(self.root),
            "--dry-run",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(dry_run_payload)
        self.assertEqual(dry_run_payload["candidate_group_count"], 1)
        self.assertEqual(dry_run_payload["blocked_group_count"], 1)
        candidate_group = dry_run_payload["candidate_groups"][0]
        self.assertEqual(candidate_group["status"], "blocked")
        effective_date_conflict = next(
            conflict
            for conflict in candidate_group["blocking_conflicts"]
            if conflict["type"] == "custom_field_conflict" and conflict["field_name"] == "effective_date"
        )
        self.assertEqual(
            sorted(item["value"] for item in effective_date_conflict["values"]),
            ["2026-04-01", "2026-04-15"],
        )

        exit_code, apply_payload, _, _ = self.run_cli(
            "reconcile-duplicates",
            str(self.root),
            "--apply",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(apply_payload)
        self.assertEqual(apply_payload["merged_group_count"], 0)
        self.assertEqual(apply_payload["blocked_group_count"], 1)
        self.assertEqual(apply_payload["blocked_groups"][0]["status"], "blocked")

        search_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(search_result["total_hits"], 2)
        self.assertEqual(len(self.fetch_occurrence_rows(int(alpha_row["id"]))), 1)
        self.assertEqual(len(self.fetch_occurrence_rows(int(beta_row["id"]))), 1)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            merge_event_count = connection.execute(
                "SELECT COUNT(*) AS row_count FROM document_merge_events",
            ).fetchone()
        finally:
            connection.close()
        self.assertEqual(int(merge_event_count["row_count"] or 0), 0)

    def test_search_omits_documents_without_dataset_memberships(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("sample dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute("DELETE FROM dataset_documents WHERE document_id = ?", (row["id"],))
            connection.commit()
        finally:
            connection.close()

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_result["total_hits"], 0)

    def test_ingest_mbox_creates_message_rows_with_source_context_and_attachment_children(self) -> None:
        mbox_path = self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="MBOX Parent",
                    body_text="Parent message body",
                    message_id="<mbox-msg-001@example.com>",
                    attachment_name="notes.txt",
                    attachment_text="mbox attachment body",
                ),
                self.build_fake_mbox_message(
                    subject="MBOX Sibling",
                    body_text="Sibling body text",
                    message_id="<mbox-msg-002@example.com>",
                    date_created="Tue, 14 Apr 2026 10:05:00 +0000",
                ),
            ]
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["mbox_messages_created"], 2)
        self.assertEqual(ingest_result["workspace_parent_documents"], 2)
        self.assertEqual(ingest_result["workspace_attachment_children"], 1)
        self.assertEqual(ingest_result["workspace_documents_total"], 3)

        parent_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-msg-001@example.com>")
        sibling_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-msg-002@example.com>")
        parent_row = self.fetch_document_row(parent_rel_path)
        sibling_row = self.fetch_document_row(sibling_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        dataset_row = self.fetch_dataset_row(int(parent_row["dataset_id"]))

        self.assertEqual(parent_row["source_kind"], retriever_tools.MBOX_SOURCE_KIND)
        self.assertEqual(parent_row["source_rel_path"], "mailbox.mbox")
        self.assertEqual(parent_row["source_item_id"], "<mbox-msg-001@example.com>")
        self.assertIsNone(parent_row["source_folder_path"])
        self.assertEqual(parent_row["file_type"], "mbox")
        self.assertIsNone(parent_row["file_size"])
        self.assertEqual(parent_row["content_type"], "Email")
        self.assertEqual(parent_row["custodian"], "mailbox")
        self.assertEqual(sibling_row["custodian"], "mailbox")
        self.assertEqual(child_row["custodian"], "mailbox")
        self.assertEqual(parent_row["dataset_id"], sibling_row["dataset_id"])
        self.assertEqual(parent_row["dataset_id"], child_row["dataset_id"])
        self.assertTrue(
            str(child_row["rel_path"]).startswith(
                f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/previews/mailbox.mbox/attachments/"
            )
        )

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            container_row = connection.execute(
                "SELECT * FROM container_sources WHERE source_kind = ? AND source_rel_path = ?",
                (retriever_tools.MBOX_SOURCE_KIND, "mailbox.mbox"),
            ).fetchone()
            dataset_rows = connection.execute(
                """
                SELECT *
                FROM datasets
                ORDER BY id ASC
                """
            ).fetchall()
        finally:
            connection.close()

        self.assertIsNotNone(container_row)
        self.assertEqual(container_row["dataset_id"], parent_row["dataset_id"])
        self.assertEqual(container_row["message_count"], 2)
        self.assertEqual(container_row["file_size"], mbox_path.stat().st_size)
        self.assertIsNotNone(container_row["last_scan_completed_at"])
        self.assertEqual(dataset_row["source_kind"], retriever_tools.MBOX_SOURCE_KIND)
        self.assertEqual(dataset_row["dataset_locator"], "mailbox.mbox")
        self.assertEqual(dataset_row["dataset_name"], "mailbox.mbox")
        self.assertEqual(len(dataset_rows), 1)
        self.assertTrue(all(row["source_kind"] == retriever_tools.MBOX_SOURCE_KIND for row in dataset_rows))

        parent_search = retriever_tools.search(self.root, "MBOX Parent", None, None, None, 1, 20)
        parent_result = next(item for item in parent_search["results"] if item["id"] == parent_row["id"])
        self.assertEqual(parent_result["attachment_count"], 1)
        self.assertEqual(parent_result["source_rel_path"], "mailbox.mbox")
        self.assertEqual(parent_result["dataset_name"], "mailbox.mbox")
        self.assertEqual(
            parent_result["preview_rel_path"],
            self.preview_target_by_label(parent_result["preview_targets"], "message")["rel_path"],
        )
        self.assertEqual([target.get("label") for target in parent_result["preview_targets"]], ["message"])
        parent_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(parent_result["preview_targets"], "message")
        ).read_text(encoding="utf-8")
        self.assertIn("Apr 14, 2026 10:00 AM UTC", parent_preview_html)
        self.assertIn("MBOX Parent", parent_preview_html)

        mbox_only = retriever_tools.search(
            self.root,
            "",
            [["source_kind", "eq", retriever_tools.MBOX_SOURCE_KIND]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(mbox_only["total_hits"], 2)
        dataset_filtered = retriever_tools.search(
            self.root,
            "",
            [["dataset_name", "eq", "mailbox.mbox"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(dataset_filtered["total_hits"], 3)
        self.assertTrue(all(item["dataset_name"] == "mailbox.mbox" for item in dataset_filtered["results"]))
        attachment_search = retriever_tools.search(self.root, "mbox attachment body", None, None, None, 1, 20)
        attachment_result = next(item for item in attachment_search["results"] if item["id"] == child_row["id"])
        self.assertEqual(attachment_result["parent"]["control_number"], parent_row["control_number"])

    def test_ingest_mbox_applies_archive_basename_custodian_rules(self) -> None:
        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="Email basename custodian",
                    body_text="Email basename custodian body",
                    message_id="<mbox-custodian-email@example.com>",
                )
            ],
            name="jane@example.com.mbox",
        )
        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="Generic basename custodian",
                    body_text="Generic basename custodian body",
                    message_id="<mbox-custodian-generic@example.com>",
                )
            ],
            name="noreply@example.com.mbox",
        )
        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="Name basename custodian",
                    body_text="Name basename custodian body",
                    message_id="<mbox-custodian-name@example.com>",
                )
            ],
            name="Jane Doe Mailbox.mbox",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["mbox_messages_created"], 3)

        email_rel_path = retriever_tools.mbox_message_rel_path(
            "jane@example.com.mbox",
            "<mbox-custodian-email@example.com>",
        )
        generic_rel_path = retriever_tools.mbox_message_rel_path(
            "noreply@example.com.mbox",
            "<mbox-custodian-generic@example.com>",
        )
        name_rel_path = retriever_tools.mbox_message_rel_path(
            "Jane Doe Mailbox.mbox",
            "<mbox-custodian-name@example.com>",
        )
        email_row = self.fetch_document_row(email_rel_path)
        generic_row = self.fetch_document_row(generic_rel_path)
        name_row = self.fetch_document_row(name_rel_path)

        self.assertEqual(email_row["custodian"], "jane@example.com")
        self.assertIsNone(generic_row["custodian"])
        self.assertEqual(name_row["custodian"], "Jane Doe")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            custodian_rows = connection.execute(
                """
                SELECT d.rel_path, e.entity_type, e.primary_email, e.display_name, e.entity_origin
                FROM document_entities de
                JOIN documents d ON d.id = de.document_id
                JOIN entities e ON e.id = de.entity_id
                WHERE de.role = 'custodian'
                  AND d.id IN (?, ?, ?)
                ORDER BY d.rel_path ASC
                """,
                (email_row["id"], generic_row["id"], name_row["id"]),
            ).fetchall()
            name_resolution_key_count = connection.execute(
                """
                SELECT COUNT(*) AS row_count
                FROM entity_resolution_keys
                WHERE key_type = 'name' AND normalized_value = 'jane doe'
                """
            ).fetchone()
        finally:
            connection.close()

        custodian_entities_by_path = {row["rel_path"]: row for row in custodian_rows}
        self.assertEqual(set(custodian_entities_by_path), {email_rel_path, name_rel_path})
        self.assertEqual(
            custodian_entities_by_path[email_rel_path]["entity_type"],
            retriever_tools.ENTITY_TYPE_PERSON,
        )
        self.assertEqual(custodian_entities_by_path[email_rel_path]["primary_email"], "jane@example.com")
        self.assertEqual(
            custodian_entities_by_path[name_rel_path]["entity_type"],
            retriever_tools.ENTITY_TYPE_PERSON,
        )
        self.assertEqual(custodian_entities_by_path[name_rel_path]["display_name"], "Jane Doe")
        self.assertEqual(
            custodian_entities_by_path[name_rel_path]["entity_origin"],
            retriever_tools.ENTITY_ORIGIN_OBSERVED,
        )
        self.assertEqual(int(name_resolution_key_count["row_count"] or 0), 0)

    def test_ingest_mbox_fixture_file_is_searchable(self) -> None:
        fixture_source = REGRESSION_CORPUS_ROOT / "sample_utf8.mbox"
        if not fixture_source.exists():
            self.skipTest(f"local regression corpus fixture unavailable: {fixture_source}")
        fixture_target = self.root / fixture_source.name
        fixture_target.write_bytes(fixture_source.read_bytes())

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["mbox_messages_created"], 2)

        fixture_search = retriever_tools.search(self.root, "retriever regression fixture", None, None, None, 1, 20)
        self.assertEqual(fixture_search["total_hits"], 1)
        result = fixture_search["results"][0]
        self.assertEqual(result["source_kind"], retriever_tools.MBOX_SOURCE_KIND)
        self.assertEqual(result["dataset_name"], "sample_utf8.mbox")

    def test_ingest_gmail_export_enriches_messages_and_skips_auxiliary_files(self) -> None:
        export_root = self.root / "gmail-export"
        export_root.mkdir()

        mbox_path = export_root / "Export.mbox"
        archive = mailbox.mbox(str(mbox_path), create=True)
        try:
            archive.add(
                self.build_fake_mbox_message(
                    subject="Drive sharing update",
                    body_text="Email body that references a linked Drive document.",
                    message_id="<gmail-msg-001@example.com>",
                    author="Sender Example <sender@example.com>",
                    recipients="Receiver Example <receiver@example.com>",
                )
            )
            archive.add(
                self.build_fake_mbox_message(
                    subject="Plain Gmail export message",
                    body_text="Regular mailbox content.",
                    message_id="<gmail-msg-002@example.com>",
                    author="Sender Example <sender@example.com>",
                    recipients="Receiver Example <receiver@example.com>",
                )
            )
            archive.flush()
        finally:
            archive.close()

        metadata_headers = [
            "Rfc822MessageId",
            "GmailMessageId",
            "FileName",
            "Account",
            "Labels",
            "From",
            "Subject",
            "To",
            "CC",
            "BCC",
            "DateSent",
            "DateReceived",
            "SubjectAtStart",
            "SubjectAtEnd",
            "DateFirstMessageSent",
            "DateLastMessageSent",
            "DateFirstMessageReceived",
            "DateLastMessageReceived",
            "ThreadedMessageCount",
        ]
        with (export_root / "Export-metadata.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=metadata_headers)
            writer.writeheader()
            writer.writerow(
                {
                    "Rfc822MessageId": "gmail-msg-001@example.com",
                    "GmailMessageId": "1767000000000000001",
                    "FileName": "1767000000000000001-1685123311000",
                    "Account": "owner@example.com",
                    "Labels": "^INBOX,projectalpha",
                    "From": "sender@example.com Sender Example",
                    "Subject": "Drive sharing update",
                    "To": "receiver@example.com Receiver Example",
                    "CC": "",
                    "BCC": "",
                    "DateSent": "2026-04-14T10:00:00Z",
                    "DateReceived": "2026-04-14T10:00:05Z",
                    "SubjectAtStart": "",
                    "SubjectAtEnd": "",
                    "DateFirstMessageSent": "",
                    "DateLastMessageSent": "",
                    "DateFirstMessageReceived": "",
                    "DateLastMessageReceived": "",
                    "ThreadedMessageCount": "1",
                }
            )
            writer.writerow(
                {
                    "Rfc822MessageId": "gmail-msg-002@example.com",
                    "GmailMessageId": "1767000000000000002",
                    "FileName": "1767000000000000002-1685123312000",
                    "Account": "owner@example.com",
                    "Labels": "^INBOX",
                    "From": "sender@example.com Sender Example",
                    "Subject": "Plain Gmail export message",
                    "To": "receiver@example.com Receiver Example",
                    "CC": "",
                    "BCC": "",
                    "DateSent": "2026-04-14T11:00:00Z",
                    "DateReceived": "2026-04-14T11:00:03Z",
                    "SubjectAtStart": "",
                    "SubjectAtEnd": "",
                    "DateFirstMessageSent": "",
                    "DateLastMessageSent": "",
                    "DateFirstMessageReceived": "",
                    "DateLastMessageReceived": "",
                    "ThreadedMessageCount": "1",
                }
            )

        with (export_root / "Export-drive-links.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=["Account", "Rfc822MessageId", "GmailMessageId", "DriveUrl", "DriveItemId"],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "Account": "owner@example.com",
                    "Rfc822MessageId": "gmail-msg-001@example.com",
                    "GmailMessageId": "1767000000000000001",
                    "DriveUrl": "https://docs.google.com/document/d/drive-doc-001/edit",
                    "DriveItemId": "drive-doc-001",
                }
            )

        drive_export_dir = export_root / "Export_Drive_Link_Export_0"
        drive_export_dir.mkdir()
        drive_file = drive_export_dir / "Linked notes_drive-doc-001.txt"
        drive_file.write_text("Drive export body text.\n", encoding="utf-8")

        (export_root / "archive_browser.html").write_text("<html><body>Archive browser</body></html>\n", encoding="utf-8")
        (export_root / "Export-metadata.xml").write_text("<Root><Documents/></Root>\n", encoding="utf-8")
        (export_root / "Export-errors.xml").write_text("<Errors/>\n", encoding="utf-8")
        (export_root / "Export-result-counts.csv").write_text(
            "Email,AccountStatus,SuccessCount,MessageErrorCount,ChatErrorCount\nTotals,,2,0,0\n",
            encoding="utf-8",
        )
        (export_root / "Export.md5").write_text("deadbeef\n", encoding="utf-8")
        (export_root / "Export Drive Link Export.md5").write_text("deadbeef\n", encoding="utf-8")
        (export_root / "Export_Drive_Link_Export-errors.csv").write_text(
            "Document ID,Document type,File type,Title,Size,Creator,Collaborators,Viewers,Others,Creation time,Last modified time,Error Description,Drive Document ID\n"
            "missing-doc,,,,,,,,,,,Document not found,missing-doc\n",
            encoding="utf-8",
        )
        (export_root / "Export_Drive_Link_Export-metadata.xml").write_text(
            """<?xml version='1.0' encoding='UTF-8' standalone='yes'?>
<Root DataInterchangeType='Update' Description='Test'>
  <Batch name='New Batch'>
    <Documents>
      <Document DocID='drive-doc-001'>
        <Tags>
          <Tag TagName='#Author' TagDataType='Text' TagValue='owner@example.com'/>
          <Tag TagName='Collaborators' TagDataType='Text' TagValue='reviewer@example.com'/>
          <Tag TagName='Viewers' TagDataType='Text' TagValue='observer@example.com'/>
          <Tag TagName='Others' TagDataType='Text' TagValue=''/>
          <Tag TagName='#DateCreated' TagDataType='DateTime' TagValue='2026-04-13T08:30:00Z'/>
          <Tag TagName='#DateModified' TagDataType='DateTime' TagValue='2026-04-14T09:45:00Z'/>
          <Tag TagName='#Title' TagDataType='Text' TagValue='Linked notes from metadata'/>
          <Tag TagName='DocumentType' TagDataType='Text' TagValue='DOCUMENT'/>
        </Tags>
        <Files>
          <File FileType='Native'>
            <ExternalFile FileName='Linked notes_drive-doc-001.txt' FileSize='24' Hash='abc123'/>
          </File>
        </Files>
      </Document>
    </Documents>
  </Batch>
</Root>
""",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["gmail_exports_detected"], 1)
        self.assertEqual(ingest_result["mbox_messages_created"], 2)
        self.assertEqual(ingest_result["gmail_linked_documents_created"], 0)
        self.assertEqual(ingest_result["workspace_parent_documents"], 2)
        self.assertEqual(ingest_result["workspace_attachment_children"], 1)
        self.assertEqual(ingest_result["workspace_documents_total"], 3)

        email_rel_path = retriever_tools.mbox_message_rel_path(
            "gmail-export/Export.mbox",
            "<gmail-msg-001@example.com>",
        )
        email_row = self.fetch_document_row(email_rel_path)
        child_rows = self.fetch_child_rows(int(email_row["id"]))
        self.assertEqual(len(child_rows), 1)
        drive_attachment_row = child_rows[0]

        self.assertEqual(drive_attachment_row["parent_document_id"], email_row["id"])
        self.assertEqual(email_row["dataset_id"], drive_attachment_row["dataset_id"])
        self.assertEqual(drive_attachment_row["dataset_id"], self.fetch_dataset_row(int(drive_attachment_row["dataset_id"]))["id"])
        self.assertEqual(drive_attachment_row["source_kind"], retriever_tools.EMAIL_ATTACHMENT_SOURCE_KIND)
        self.assertEqual(drive_attachment_row["title"], "Linked notes from metadata")
        self.assertEqual(drive_attachment_row["author"], "owner@example.com")
        self.assertIn("reviewer@example.com", str(drive_attachment_row["participants"]))
        self.assertEqual(drive_attachment_row["date_created"], "2026-04-13T08:30:00Z")
        self.assertEqual(drive_attachment_row["date_modified"], "2026-04-14T09:45:00Z")

        label_search = retriever_tools.search(self.root, "projectalpha", None, None, None, 1, 20)
        self.assertEqual(label_search["total_hits"], 1)
        self.assertEqual(label_search["results"][0]["id"], email_row["id"])
        self.assertEqual(label_search["results"][0]["attachment_count"], 1)

        linked_title_search = retriever_tools.search(self.root, "Linked notes from metadata", None, None, None, 1, 20)
        returned_ids = {item["id"] for item in linked_title_search["results"]}
        self.assertIn(email_row["id"], returned_ids)
        self.assertIn(drive_attachment_row["id"], returned_ids)
        drive_result = next(item for item in linked_title_search["results"] if item["id"] == drive_attachment_row["id"])
        self.assertEqual(drive_result["dataset_name"], "Export.mbox")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            parent_rel_paths = {
                str(row["rel_path"])
                for row in connection.execute(
                    """
                    SELECT rel_path
                    FROM documents
                    WHERE parent_document_id IS NULL
                    ORDER BY rel_path ASC
                    """
                ).fetchall()
            }
        finally:
            connection.close()

        self.assertNotIn("gmail-export/archive_browser.html", parent_rel_paths)
        self.assertNotIn("gmail-export/Export-metadata.csv", parent_rel_paths)
        self.assertNotIn("gmail-export/Export-metadata.xml", parent_rel_paths)
        self.assertNotIn("gmail-export/Export-drive-links.csv", parent_rel_paths)
        self.assertNotIn("gmail-export/Export-errors.xml", parent_rel_paths)
        self.assertNotIn("gmail-export/Export-result-counts.csv", parent_rel_paths)
        self.assertNotIn("gmail-export/Export_Drive_Link_Export_0/Linked notes_drive-doc-001.txt", parent_rel_paths)
        self.assertNotIn("gmail-export/Export_Drive_Link_Export-errors.csv", parent_rel_paths)
        self.assertNotIn("gmail-export/Export_Drive_Link_Export-metadata.xml", parent_rel_paths)

    def test_ingest_gmail_export_with_mbox_file_filter_preserves_sidecar_enrichment(self) -> None:
        export_root = self.root / "gmail-filtered"
        export_root.mkdir()

        mbox_path = export_root / "Filtered.mbox"
        archive = mailbox.mbox(str(mbox_path), create=True)
        try:
            archive.add(
                self.build_fake_mbox_message(
                    subject="Drive sharing update",
                    body_text="Email body that references a linked Drive document.",
                    message_id="<gmail-filter-001@example.com>",
                    author="Sender Example <sender@example.com>",
                    recipients="Receiver Example <receiver@example.com>",
                )
            )
            archive.flush()
        finally:
            archive.close()

        with (export_root / "Filtered-metadata.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "Rfc822MessageId",
                    "GmailMessageId",
                    "Account",
                    "Labels",
                    "Subject",
                    "From",
                    "To",
                    "DateSent",
                    "DateReceived",
                    "ThreadedMessageCount",
                ],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "Rfc822MessageId": "gmail-filter-001@example.com",
                    "GmailMessageId": "1767000000000000003",
                    "Account": "owner@example.com",
                    "Labels": "^INBOX,projectalpha",
                    "Subject": "Drive sharing update",
                    "From": "sender@example.com Sender Example",
                    "To": "receiver@example.com Receiver Example",
                    "DateSent": "2026-04-14T10:00:00Z",
                    "DateReceived": "2026-04-14T10:00:05Z",
                    "ThreadedMessageCount": "1",
                }
            )

        with (export_root / "Filtered-drive-links.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=["Account", "Rfc822MessageId", "GmailMessageId", "DriveUrl", "DriveItemId"],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "Account": "owner@example.com",
                    "Rfc822MessageId": "gmail-filter-001@example.com",
                    "GmailMessageId": "1767000000000000003",
                    "DriveUrl": "https://docs.google.com/document/d/drive-doc-001/edit",
                    "DriveItemId": "drive-doc-001",
                }
            )

        drive_export_dir = export_root / "Filtered_Drive_Link_Export_0"
        drive_export_dir.mkdir()
        drive_file = drive_export_dir / "Linked notes_drive-doc-001.txt"
        drive_file.write_text("Drive export body text.\n", encoding="utf-8")
        (export_root / "Filtered_Drive_Link_Export-metadata.xml").write_text(
            """<?xml version='1.0' encoding='UTF-8' standalone='yes'?>
<Root DataInterchangeType='Update' Description='Test'>
  <Batch name='New Batch'>
    <Documents>
      <Document DocID='drive-doc-001'>
        <Tags>
          <Tag TagName='#Author' TagDataType='Text' TagValue='owner@example.com'/>
          <Tag TagName='#Title' TagDataType='Text' TagValue='Linked notes from metadata'/>
        </Tags>
        <Files>
          <File FileType='Native'>
            <ExternalFile FileName='Linked notes_drive-doc-001.txt' FileSize='24' Hash='abc123'/>
          </File>
        </Files>
      </Document>
    </Documents>
  </Batch>
</Root>
""",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types="mbox")

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["gmail_exports_detected"], 1)
        self.assertEqual(ingest_result["mbox_messages_created"], 1)
        self.assertEqual(ingest_result["gmail_linked_documents_created"], 0)
        self.assertEqual(ingest_result["workspace_parent_documents"], 1)
        self.assertEqual(ingest_result["workspace_attachment_children"], 1)
        self.assertEqual(ingest_result["workspace_documents_total"], 2)

        email_rel_path = retriever_tools.mbox_message_rel_path(
            "gmail-filtered/Filtered.mbox",
            "<gmail-filter-001@example.com>",
        )
        email_row = self.fetch_document_row(email_rel_path)
        child_rows = self.fetch_child_rows(int(email_row["id"]))
        self.assertEqual(len(child_rows), 1)
        self.assertEqual(child_rows[0]["title"], "Linked notes from metadata")
        self.assertEqual(child_rows[0]["source_kind"], retriever_tools.EMAIL_ATTACHMENT_SOURCE_KIND)

        label_search = retriever_tools.search(self.root, "projectalpha", None, None, None, 1, 20)
        self.assertEqual(label_search["total_hits"], 1)
        self.assertEqual(label_search["results"][0]["id"], email_row["id"])
        self.assertEqual(label_search["results"][0]["attachment_count"], 1)

        linked_title_search = retriever_tools.search(self.root, "Linked notes from metadata", None, None, None, 1, 20)
        self.assertEqual(linked_title_search["total_hits"], 2)
        returned_ids = {item["id"] for item in linked_title_search["results"]}
        self.assertIn(email_row["id"], returned_ids)
        self.assertIn(child_rows[0]["id"], returned_ids)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            parent_rel_paths = {
                str(row["rel_path"])
                for row in connection.execute(
                    """
                    SELECT rel_path
                    FROM documents
                    WHERE parent_document_id IS NULL
                    ORDER BY rel_path ASC
                    """
                ).fetchall()
            }
        finally:
            connection.close()

        self.assertNotIn(
            "gmail-filtered/Filtered_Drive_Link_Export_0/Linked notes_drive-doc-001.txt",
            parent_rel_paths,
        )
        self.assertNotIn("gmail-filtered/Filtered-metadata.csv", parent_rel_paths)
        self.assertNotIn("gmail-filtered/Filtered-drive-links.csv", parent_rel_paths)
        self.assertNotIn("gmail-filtered/Filtered_Drive_Link_Export-metadata.xml", parent_rel_paths)

    def test_ingest_mbox_without_message_id_uses_stable_fallback_source_item_id(self) -> None:
        first_messages = [
            self.build_fake_mbox_message(
                subject="Missing ID Parent",
                body_text="Fallback source id body",
                message_id=None,
            )
        ]
        self.write_fake_mbox_file(first_messages)

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["new"], 1)
        self.assertEqual(first_ingest["mbox_messages_created"], 1)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            first_row = connection.execute(
                """
                SELECT *
                FROM documents
                WHERE parent_document_id IS NULL
                  AND source_kind = ?
                  AND source_rel_path = ?
                ORDER BY id ASC
                """,
                (retriever_tools.MBOX_SOURCE_KIND, "mailbox.mbox"),
            ).fetchone()
        finally:
            connection.close()

        self.assertIsNotNone(first_row)
        fallback_source_item_id = str(first_row["source_item_id"])
        self.assertTrue(fallback_source_item_id.startswith("mbox-hash:"))
        original_control_number = str(first_row["control_number"])

        second_messages = [
            self.build_fake_mbox_message(
                subject="Missing ID Parent",
                body_text="Fallback source id body",
                message_id=None,
            ),
            self.build_fake_mbox_message(
                subject="Sibling",
                body_text="Sibling body text",
                message_id="<mbox-msg-002@example.com>",
                date_created="Tue, 14 Apr 2026 10:05:00 +0000",
            ),
        ]
        self.write_fake_mbox_file(second_messages)
        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["updated"], 1)
        self.assertEqual(second_ingest["mbox_messages_created"], 1)
        self.assertEqual(second_ingest["mbox_messages_updated"], 1)

        updated_row = self.fetch_document_row(
            retriever_tools.mbox_message_rel_path("mailbox.mbox", fallback_source_item_id)
        )
        self.assertEqual(updated_row["control_number"], original_control_number)

    def test_unchanged_mbox_source_skips_without_reparsing(self) -> None:
        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="MBOX Parent",
                    body_text="Parent message body",
                    message_id="<mbox-msg-001@example.com>",
                )
            ]
        )

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["new"], 1)
        with mock.patch.object(
            retriever_tools,
            "iter_mbox_messages",
            side_effect=AssertionError("MBOX iterator should not run on unchanged source"),
        ):
            second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["skipped"], 1)
        self.assertEqual(second_ingest["mbox_sources_skipped"], 1)
        self.assertEqual(second_ingest["failed"], 0)
        browse_results = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_results["total_hits"], 1)

    def test_refresh_generated_previews_regenerates_mbox_message_previews_by_dataset_name(self) -> None:
        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="Gmail Max Refresh",
                    body_text="Refresh this generated preview body",
                    message_id="<mbox-refresh-001@example.com>",
                )
            ],
            name="gmail-max.mbox",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 1)
        message_rel_path = retriever_tools.mbox_message_rel_path("gmail-max.mbox", "<mbox-refresh-001@example.com>")
        search_result = retriever_tools.search(self.root, "Refresh this generated preview body", None, None, None, 1, 20)
        message_result = next(item for item in search_result["results"] if item["rel_path"] == message_rel_path)
        message_preview_path = self.preview_target_file_path(
            self.preview_target_by_label(message_result["preview_targets"], "message")
        )
        message_preview_path.write_text("stale preview html", encoding="utf-8")

        refresh_result = retriever_tools.refresh_generated_previews(self.root, dataset_name="gmail-max.mbox")

        self.assertEqual(refresh_result["status"], "ok")
        self.assertEqual(refresh_result["dataset"]["dataset_name"], "gmail-max.mbox")
        self.assertEqual(refresh_result["refreshed_conversations"], 1)
        refreshed_html = message_preview_path.read_text(encoding="utf-8")
        self.assertNotIn("stale preview html", refreshed_html)
        self.assertIn("Refresh this generated preview body", refreshed_html)
        self.assertIn('class="gmail-thread-title"', refreshed_html)

    def test_changed_mbox_reingest_preserves_control_numbers_and_retires_removed_messages(self) -> None:
        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="Original MBOX Parent",
                    body_text="Parent v1",
                    message_id="<mbox-msg-001@example.com>",
                    attachment_name="notes.txt",
                    attachment_text="stable attachment body",
                ),
                self.build_fake_mbox_message(
                    subject="Removed later",
                    body_text="Remove me",
                    message_id="<mbox-msg-002@example.com>",
                    date_created="Tue, 14 Apr 2026 10:05:00 +0000",
                ),
            ]
        )

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["mbox_messages_created"], 2)
        parent_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-msg-001@example.com>")
        removed_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-msg-002@example.com>")
        parent_row = self.fetch_document_row(parent_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        retriever_tools.set_field(self.root, parent_row["id"], "title", "Manual MBOX Title")

        self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="Updated MBOX Parent",
                    body_text="Parent v2",
                    message_id="<mbox-msg-001@example.com>",
                    attachment_name="notes.txt",
                    attachment_text="stable attachment body",
                )
            ]
        )
        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["updated"], 1)
        self.assertEqual(second_ingest["mbox_messages_updated"], 1)
        self.assertEqual(second_ingest["mbox_messages_deleted"], 1)

        updated_parent = self.fetch_document_row(parent_rel_path)
        updated_child = self.fetch_child_rows(updated_parent["id"])[0]
        retired_row = self.fetch_document_row(removed_rel_path)
        self.assertEqual(updated_parent["control_number"], parent_row["control_number"])
        self.assertEqual(updated_parent["title"], "Manual MBOX Title")
        self.assertEqual(updated_child["id"], child_row["id"])
        self.assertEqual(updated_child["control_number"], child_row["control_number"])
        self.assertEqual(retired_row["lifecycle_status"], "deleted")

    def test_missing_mbox_source_marks_messages_and_children_missing(self) -> None:
        mbox_path = self.write_fake_mbox_file(
            [
                self.build_fake_mbox_message(
                    subject="MBOX Parent",
                    body_text="Parent body",
                    message_id="<mbox-msg-001@example.com>",
                    attachment_name="notes.txt",
                    attachment_text="attachment body",
                )
            ]
        )

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["mbox_messages_created"], 1)
        parent_rel_path = retriever_tools.mbox_message_rel_path("mailbox.mbox", "<mbox-msg-001@example.com>")
        parent_row = self.fetch_document_row(parent_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]

        mbox_path.unlink()
        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["missing"], 1)
        self.assertEqual(second_ingest["mbox_sources_missing"], 1)
        self.assertEqual(second_ingest["mbox_documents_missing"], 2)
        missing_parent = self.fetch_document_row(parent_rel_path)
        missing_child = self.fetch_document_by_id(child_row["id"])
        self.assertEqual(missing_parent["lifecycle_status"], "missing")
        self.assertEqual(missing_child["lifecycle_status"], "missing")

    def test_ingest_pst_creates_message_rows_with_source_context_and_attachment_children(self) -> None:
        pst_path = self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-msg-001",
                subject="PST Parent",
                body_text="Parent message body",
                folder_path="Inbox",
                recipients=None,
                transport_headers="\n".join(
                    [
                        "To: Bob Example <bob@example.com>",
                        "Cc: Carol Example <carol@example.com>",
                        "",
                    ]
                ),
                attachment_name="notes.txt",
                attachment_text="pst attachment body",
            ),
            self.build_fake_pst_message(
                source_item_id="pst-msg-002",
                subject="PST Sibling",
                body_text="Sibling body text",
                folder_path="Sent Items",
            ),
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 2)
        self.assertEqual(ingest_result["workspace_parent_documents"], 2)
        self.assertEqual(ingest_result["workspace_attachment_children"], 1)
        self.assertEqual(ingest_result["workspace_documents_total"], 3)

        parent_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "pst-msg-001")
        sibling_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "pst-msg-002")
        parent_row = self.fetch_document_row(parent_rel_path)
        sibling_row = self.fetch_document_row(sibling_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]

        self.assertEqual(parent_row["source_kind"], retriever_tools.PST_SOURCE_KIND)
        self.assertEqual(parent_row["source_rel_path"], "mailbox.pst")
        self.assertEqual(parent_row["source_item_id"], "pst-msg-001")
        self.assertEqual(parent_row["source_folder_path"], "Inbox")
        self.assertEqual(parent_row["file_type"], "pst")
        self.assertIsNone(parent_row["file_size"])
        self.assertEqual(parent_row["content_type"], "Email")
        self.assertEqual(parent_row["custodian"], "mailbox")
        self.assertEqual(
            parent_row["recipients"],
            "Bob Example <bob@example.com>, Carol Example <carol@example.com>",
        )
        self.assertEqual(sibling_row["source_folder_path"], "Sent Items")
        self.assertEqual(sibling_row["custodian"], "mailbox")
        self.assertEqual(child_row["custodian"], "mailbox")
        self.assertIsNotNone(parent_row["dataset_id"])
        self.assertEqual(parent_row["dataset_id"], sibling_row["dataset_id"])
        self.assertEqual(parent_row["dataset_id"], child_row["dataset_id"])
        self.assertTrue(
            str(child_row["rel_path"]).startswith(
                f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/previews/mailbox.pst/attachments/"
            )
        )

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            container_row = connection.execute(
                "SELECT * FROM container_sources WHERE source_kind = ? AND source_rel_path = ?",
                (retriever_tools.PST_SOURCE_KIND, "mailbox.pst"),
            ).fetchone()
            dataset_row = connection.execute(
                "SELECT * FROM datasets WHERE id = ?",
                (parent_row["dataset_id"],),
            ).fetchone()
            dataset_rows = connection.execute(
                """
                SELECT *
                FROM datasets
                ORDER BY id ASC
                """
            ).fetchall()
        finally:
            connection.close()

        self.assertIsNotNone(container_row)
        self.assertEqual(container_row["dataset_id"], parent_row["dataset_id"])
        self.assertEqual(container_row["message_count"], 2)
        self.assertEqual(container_row["file_size"], pst_path.stat().st_size)
        self.assertIsNotNone(container_row["last_scan_completed_at"])
        self.assertIsNotNone(dataset_row)
        self.assertEqual(dataset_row["source_kind"], retriever_tools.PST_SOURCE_KIND)
        self.assertEqual(dataset_row["dataset_locator"], "mailbox.pst")
        self.assertEqual(dataset_row["dataset_name"], "mailbox.pst")
        self.assertEqual(len(dataset_rows), 1)
        self.assertTrue(all(row["source_kind"] == retriever_tools.PST_SOURCE_KIND for row in dataset_rows))

        parent_search = retriever_tools.search(self.root, "PST Parent", None, None, None, 1, 20)
        parent_result = next(item for item in parent_search["results"] if item["id"] == parent_row["id"])
        self.assertEqual(parent_result["attachment_count"], 1)
        self.assertEqual(parent_result["source_rel_path"], "mailbox.pst")
        self.assertEqual(parent_result["source_folder_path"], "Inbox")
        self.assertEqual(parent_result["dataset_name"], "mailbox.pst")
        self.assertEqual(
            parent_result["preview_rel_path"],
            self.preview_target_by_label(parent_result["preview_targets"], "message")["rel_path"],
        )
        message_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(parent_result["preview_targets"], "message")
        ).read_text(encoding="utf-8")
        self.assertIn('class="gmail-thread-title">PST Parent</h1>', message_preview_html)
        self.assertIn("to Bob Example &lt;bob@example.com&gt;, Carol Example &lt;carol@example.com&gt;", message_preview_html)
        self.assertIn("Parent message body", message_preview_html)
        self.assertNotIn("Viewing message", message_preview_html)
        self.assertNotIn("2 messages", message_preview_html)
        self.assertEqual([target.get("label") for target in parent_result["preview_targets"]], ["message"])
        parent_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(parent_result["preview_targets"], "message")
        ).read_text(encoding="utf-8")
        self.assertIn("Apr 14, 2026 10:00 AM UTC", parent_preview_html)
        self.assertIn("PST Parent", parent_preview_html)
        self.assertIn("to Bob Example &lt;bob@example.com&gt;, Carol Example &lt;carol@example.com&gt;", parent_preview_html)

        pst_only = retriever_tools.search(
            self.root,
            "",
            [["source_kind", "eq", retriever_tools.PST_SOURCE_KIND]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(pst_only["total_hits"], 2)
        custodian_filtered = retriever_tools.search(
            self.root,
            "",
            [["custodian", "eq", "mailbox"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(custodian_filtered["total_hits"], 3)
        dataset_filtered = retriever_tools.search(
            self.root,
            "",
            [["dataset_name", "eq", "mailbox.pst"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(dataset_filtered["total_hits"], 3)
        self.assertTrue(all(item["dataset_name"] == "mailbox.pst" for item in dataset_filtered["results"]))
        custodian_query = retriever_tools.search(self.root, "mailbox", None, None, None, 1, 20)
        self.assertEqual(custodian_query["total_hits"], 3)

        attachment_search = retriever_tools.search(self.root, "pst attachment body", None, None, None, 1, 20)
        attachment_result = next(item for item in attachment_search["results"] if item["id"] == child_row["id"])
        self.assertEqual(attachment_result["parent"]["control_number"], parent_row["control_number"])

    def test_ingest_pst_classifies_archive_basename_custodians_without_generic_mailbox_assignment(self) -> None:
        self.write_fake_pst_file(name="Acme Corp.pst")
        self.write_fake_pst_file(name="support@example.com.pst")
        messages_by_name = {
            "Acme Corp.pst": [
                self.build_fake_pst_message(
                    source_item_id="pst-org-custodian-001",
                    subject="Organization custodian",
                    body_text="Organization custodian body",
                )
            ],
            "support@example.com.pst": [
                self.build_fake_pst_message(
                    source_item_id="pst-shared-custodian-001",
                    subject="Shared mailbox basename",
                    body_text="Shared mailbox basename body",
                )
            ],
        }

        def fake_iter_pst_messages(path: Path):
            return iter(messages_by_name[path.name])

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", side_effect=fake_iter_pst_messages):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 2)

        organization_rel_path = retriever_tools.pst_message_rel_path("Acme Corp.pst", "pst-org-custodian-001")
        shared_rel_path = retriever_tools.pst_message_rel_path(
            "support@example.com.pst",
            "pst-shared-custodian-001",
        )
        organization_row = self.fetch_document_row(organization_rel_path)
        shared_row = self.fetch_document_row(shared_rel_path)

        self.assertEqual(organization_row["custodian"], "Acme Corp")
        self.assertIsNone(shared_row["custodian"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            organization_entity_row = connection.execute(
                """
                SELECT e.entity_type, e.display_name, e.primary_email, e.entity_origin
                FROM document_entities de
                JOIN entities e ON e.id = de.entity_id
                WHERE de.document_id = ? AND de.role = 'custodian'
                """,
                (organization_row["id"],),
            ).fetchone()
            shared_custodian_count = connection.execute(
                """
                SELECT COUNT(*) AS row_count
                FROM document_entities
                WHERE document_id = ? AND role = 'custodian'
                """,
                (shared_row["id"],),
            ).fetchone()
        finally:
            connection.close()

        self.assertIsNotNone(organization_entity_row)
        assert organization_entity_row is not None
        self.assertEqual(organization_entity_row["entity_type"], retriever_tools.ENTITY_TYPE_ORGANIZATION)
        self.assertEqual(organization_entity_row["display_name"], "Acme Corp")
        self.assertIsNone(organization_entity_row["primary_email"])
        self.assertEqual(organization_entity_row["entity_origin"], retriever_tools.ENTITY_ORIGIN_OBSERVED)
        self.assertEqual(int(shared_custodian_count["row_count"] or 0), 0)

    def test_ingest_pst_export_sidecars_augment_messages_and_stay_out_of_filesystem_docs(self) -> None:
        (self.root / "Exchange").mkdir(parents=True, exist_ok=True)
        self.write_fake_pst_file(name="Exchange/mailbox.pst")
        self.write_csv_rows(
            self.root / "Results.csv",
            fieldnames=[
                "Item Identity",
                "Document ID",
                "Target Path",
                "Original Path",
                "Location",
                "Location Name",
                "Subject or Title",
                "Sender or Created by",
                "Recipients in To line",
                "Sent",
            ],
            rows=[
                {
                    "Item Identity": "pst-export-001",
                    "Document ID": "101",
                    "Target Path": r"Exchange\mailbox.pst\Top of Information Store\Inbox\Sidecar Welcome",
                    "Original Path": (
                        r"mailbox@example.com, Primary, abc-123\mailbox@example.com (Primary)"
                        r"\Top of Information Store\Inbox"
                    ),
                    "Location": "mailbox@example.com, Primary, abc-123",
                    "Location Name": "mailbox@example.com",
                    "Subject or Title": "Sidecar Welcome",
                    "Sender or Created by": "Mailbox Sender <sender@example.com>",
                    "Recipients in To line": "Recipient Example <recipient@example.com>",
                    "Sent": "4/24/2026 11:59:00 PM",
                }
            ],
        )
        (self.root / "Manifest.xml").write_text(
            (
                "<?xml version='1.0' encoding='utf-8'?>"
                "<Root><Batch><Documents>"
                "<Document DocID='pst-export-001'>"
                "<Tags>"
                "<Tag TagName='TargetPath' TagDataType='Text' "
                "TagValue='Exchange\\mailbox.pst\\Top of Information Store\\Inbox\\Sidecar Welcome' />"
                "<Tag TagName='#OriginalUrl' TagDataType='Text' TagValue='pst-export-original://abc-123' />"
                "</Tags>"
                "</Document>"
                "</Documents></Batch></Root>"
            ),
            encoding="utf-8",
        )
        self.write_csv_rows(
            self.root / "Export Summary 05.24.2024-1357PM.csv",
            fieldnames=["Export Name", "Value"],
            rows=[{"Export Name": "Example Export", "Value": "done"}],
        )
        (self.root / "trace.log").write_text("trace entry\n", encoding="utf-8")

        messages = [
            self.build_fake_pst_message(
                source_item_id="2097188",
                subject=None,
                body_text="Body from PST",
                folder_path="Top of Personal Folders/Top-of-Information-Store/Inbox",
                author=None,
                recipients=None,
                date_created=None,
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 1)
        rel_path = retriever_tools.pst_message_rel_path("Exchange/mailbox.pst", "2097188")
        row = self.fetch_document_row(rel_path)
        self.assertEqual(row["title"], "Sidecar Welcome")
        self.assertEqual(row["subject"], "Sidecar Welcome")
        self.assertEqual(row["author"], "Mailbox Sender <sender@example.com>")
        self.assertEqual(row["recipients"], "Recipient Example <recipient@example.com>")
        self.assertEqual(row["date_created"], "2026-04-24T23:59:00Z")
        self.assertEqual(row["custodian"], "mailbox@example.com")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            sidecar_count = int(
                connection.execute(
                    """
                    SELECT COUNT(*) AS count
                    FROM documents
                    WHERE rel_path IN (?, ?, ?, ?)
                    """,
                    (
                        "Results.csv",
                        "Manifest.xml",
                        "Export Summary 05.24.2024-1357PM.csv",
                        "trace.log",
                    ),
                ).fetchone()["count"]
            )
            chunk_count = int(
                connection.execute(
                    """
                    SELECT COUNT(*) AS count
                    FROM document_chunks
                    WHERE document_id = ?
                      AND text_content LIKE ?
                    """,
                    (row["id"], "%abc-123%"),
                ).fetchone()["count"]
            )
        finally:
            connection.close()

        self.assertEqual(sidecar_count, 0)
        self.assertGreaterEqual(chunk_count, 1)

    def test_pst_export_location_external_id_hints_follow_dataset_policy_on_rebuild(self) -> None:
        (self.root / "Exchange").mkdir(parents=True, exist_ok=True)
        self.write_fake_pst_file(name="Exchange/mailbox.pst")
        shared_location = "PST Location, Primary, location-guid-123"
        self.write_csv_rows(
            self.root / "Results.csv",
            fieldnames=[
                "Item Identity",
                "Document ID",
                "Target Path",
                "Location",
                "Location Name",
                "Subject or Title",
            ],
            rows=[
                {
                    "Item Identity": "pst-location-001",
                    "Document ID": "101",
                    "Target Path": r"Exchange\mailbox.pst\Top of Information Store\Inbox\Location One",
                    "Location": shared_location,
                    "Location Name": "Alpha Owner One",
                    "Subject or Title": "Location One",
                },
                {
                    "Item Identity": "pst-location-002",
                    "Document ID": "102",
                    "Target Path": r"Exchange\mailbox.pst\Top of Information Store\Inbox\Location Two",
                    "Location": shared_location,
                    "Location Name": "Alpha Owner Alias",
                    "Subject or Title": "Location Two",
                },
            ],
        )
        self.write_csv_rows(
            self.root / "Export Summary 05.24.2024-1357PM.csv",
            fieldnames=["Export Name", "Value"],
            rows=[{"Export Name": "Example Export", "Value": "done"}],
        )
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-location-001",
                subject="Location One",
                body_text="First location body",
                folder_path="Top of Information Store/Inbox",
                author=None,
                recipients=None,
            ),
            self.build_fake_pst_message(
                source_item_id="pst-location-002",
                subject="Location Two",
                body_text="Second location body",
                folder_path="Top of Information Store/Inbox",
                author=None,
                recipients=None,
            ),
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 2)

        first_row = self.fetch_document_row(
            retriever_tools.pst_message_rel_path("Exchange/mailbox.pst", "pst-location-001")
        )
        second_row = self.fetch_document_row(
            retriever_tools.pst_message_rel_path("Exchange/mailbox.pst", "pst-location-002")
        )
        self.assertEqual(first_row["custodian"], "Alpha Owner One")
        self.assertEqual(second_row["custodian"], "Alpha Owner Alias")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            custodian_rows = connection.execute(
                """
                SELECT d.rel_path, de.entity_id, e.display_name, e.entity_origin
                FROM document_entities de
                JOIN documents d ON d.id = de.document_id
                JOIN entities e ON e.id = de.entity_id
                WHERE de.role = 'custodian'
                  AND d.id IN (?, ?)
                ORDER BY d.rel_path ASC
                """,
                (first_row["id"], second_row["id"]),
            ).fetchall()
            hint_row = connection.execute(
                """
                SELECT entity_hints_json
                FROM document_occurrences
                WHERE document_id = ?
                """,
                (first_row["id"],),
            ).fetchone()
            location_identifier_rows = connection.execute(
                """
                SELECT entity_id, identifier_name, identifier_scope, normalized_value
                FROM entity_identifiers
                WHERE identifier_type = 'external_id'
                  AND identifier_name = 'pst_location'
                  AND identifier_scope = 'exchange/mailbox.pst'
                  AND normalized_value = 'pst location primary location-guid-123'
                ORDER BY entity_id ASC
                """
            ).fetchall()
            location_key_count = connection.execute(
                """
                SELECT COUNT(*) AS row_count
                FROM entity_resolution_keys
                WHERE key_type = 'external_id'
                  AND identifier_name = 'pst_location'
                  AND identifier_scope = 'exchange/mailbox.pst'
                  AND normalized_value = 'pst location primary location-guid-123'
                """
            ).fetchone()
            document_id_identifier_count = connection.execute(
                """
                SELECT COUNT(*) AS row_count
                FROM entity_identifiers
                WHERE identifier_type = 'external_id'
                  AND identifier_name IN ('document_id', 'item_identity', 'export_item_id')
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(len(custodian_rows), 2)
        self.assertNotEqual(custodian_rows[0]["entity_id"], custodian_rows[1]["entity_id"])
        self.assertEqual({row["entity_origin"] for row in custodian_rows}, {retriever_tools.ENTITY_ORIGIN_OBSERVED})
        self.assertIsNotNone(hint_row)
        hint_payload = json.loads(hint_row["entity_hints_json"])
        self.assertEqual(hint_payload["custodian"][0]["identifiers"][0]["identifier_name"], "pst_location")
        self.assertEqual(hint_payload["custodian"][0]["identifiers"][0]["identifier_scope"], "exchange/mailbox.pst")
        self.assertEqual(len(location_identifier_rows), 2)
        self.assertEqual(int(location_key_count["row_count"] or 0), 0)
        self.assertEqual(int(document_id_identifier_count["row_count"] or 0), 0)

        exit_code, set_payload, _, _ = self.run_cli(
            "set-dataset-policy",
            str(self.root),
            "--dataset-id",
            str(first_row["dataset_id"]),
            "--external-id-auto-merge-name",
            "pst_location",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(set_payload)
        self.assertEqual(set_payload["merge_policy"]["external_id_auto_merge_names"], ["pst_location"])

        exit_code, rebuild_payload, _, _ = self.run_cli("rebuild-entities", str(self.root))
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(rebuild_payload)
        self.assertEqual(rebuild_payload["documents_synced"], 2)

        rebuilt_first = self.fetch_document_by_id(first_row["id"])
        rebuilt_second = self.fetch_document_by_id(second_row["id"])
        self.assertEqual(rebuilt_first["custodian"], "Alpha Owner One")
        self.assertEqual(rebuilt_second["custodian"], "Alpha Owner One")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            rebuilt_custodian_rows = connection.execute(
                """
                SELECT d.rel_path, de.entity_id, e.display_name, e.entity_origin
                FROM document_entities de
                JOIN documents d ON d.id = de.document_id
                JOIN entities e ON e.id = de.entity_id
                WHERE de.role = 'custodian'
                  AND d.id IN (?, ?)
                ORDER BY d.rel_path ASC
                """,
                (first_row["id"], second_row["id"]),
            ).fetchall()
            rebuilt_location_key_count = connection.execute(
                """
                SELECT COUNT(*) AS row_count
                FROM entity_resolution_keys
                WHERE key_type = 'external_id'
                  AND identifier_name = 'pst_location'
                  AND identifier_scope = 'exchange/mailbox.pst'
                  AND normalized_value = 'pst location primary location-guid-123'
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(len(rebuilt_custodian_rows), 2)
        self.assertEqual(
            {row["entity_id"] for row in rebuilt_custodian_rows},
            {rebuilt_custodian_rows[0]["entity_id"]},
        )
        self.assertEqual(
            {row["entity_origin"] for row in rebuilt_custodian_rows},
            {retriever_tools.ENTITY_ORIGIN_IDENTIFIED},
        )
        self.assertEqual(int(rebuilt_location_key_count["row_count"] or 0), 1)

    def test_changed_pst_export_results_sidecar_reingests_unchanged_pst_source(self) -> None:
        (self.root / "Exchange").mkdir(parents=True, exist_ok=True)
        self.write_fake_pst_file(name="Exchange/mailbox.pst", content=b"pst-static")
        export_results_path = self.root / "Export_Results_1_1.csv"
        fieldnames = [
            "Export Item Id",
            "Export Item Path",
            "Document ID",
            "Location",
            "Location Name",
            "Target Path",
            "Original Path",
            "Subject or Title",
            "Sender or Created by",
            "Recipients in To line",
            "Sent",
            "Internet Message Id",
        ]

        def write_export_results(subject: str) -> None:
            self.write_csv_rows(
                export_results_path,
                fieldnames=fieldnames,
                rows=[
                    {
                        "Export Item Id": "pst-export-002",
                        "Export Item Path": (
                            r"First export_1.zip\mailbox.pst\Top-of-Information-Store\Inbox\Welcome"
                        ),
                        "Document ID": "17",
                        "Location": "mailbox@example.com",
                        "Location Name": "mailbox@example.com",
                        "Target Path": "/Top of Information Store/Inbox",
                        "Original Path": "/Top of Information Store/Inbox",
                        "Subject or Title": subject,
                        "Sender or Created by": "Sidecar Sender <sender@example.com>",
                        "Recipients in To line": "Recipient Example <recipient@example.com>",
                        "Sent": "4/24/2026 11:59:00 PM",
                        "Internet Message Id": "<sidecar-message@example.com>",
                    }
                ],
            )

        write_export_results("First export subject")
        messages = [
            self.build_fake_pst_message(
                source_item_id="2097220",
                subject=None,
                body_text="Body from PST",
                folder_path="Top of Personal Folders/Top-of-Information-Store/Inbox",
                author=None,
                recipients=None,
                date_created=None,
                transport_headers=(
                    "Message-ID: <sidecar-message@example.com>\n"
                    "From: Sidecar Sender <sender@example.com>\n"
                    "To: Recipient Example <recipient@example.com>\n"
                ),
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["failed"], 0)
        rel_path = retriever_tools.pst_message_rel_path("Exchange/mailbox.pst", "2097220")
        first_row = self.fetch_document_row(rel_path)
        first_threading_row = self.fetch_email_threading_row(first_row["id"])
        self.assertEqual(first_row["title"], "First export subject")
        self.assertEqual(first_row["custodian"], "mailbox@example.com")
        self.assertIsNotNone(first_threading_row)
        self.assertEqual(
            first_threading_row["message_id"],
            retriever_tools.normalize_email_message_id("<sidecar-message@example.com>"),
        )

        write_export_results("Second export subject")
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["updated"], 1)
        self.assertEqual(second_ingest["pst_messages_updated"], 1)
        updated_row = self.fetch_document_row(rel_path)
        self.assertEqual(updated_row["title"], "Second export subject")
        self.assertEqual(updated_row["subject"], "Second export subject")

    def test_pst_only_ingest_prunes_stale_empty_filesystem_dataset(self) -> None:
        self.write_fake_pst_file()
        retriever_tools.bootstrap(self.root)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            stale_dataset_id, stale_dataset_source_id = retriever_tools.ensure_source_backed_dataset(
                connection,
                source_kind=retriever_tools.FILESYSTEM_SOURCE_KIND,
                source_locator=retriever_tools.filesystem_dataset_locator(),
                dataset_name=retriever_tools.filesystem_dataset_name(self.root),
            )
            connection.commit()
        finally:
            connection.close()

        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-msg-001",
                subject="PST Parent",
                body_text="Parent message body",
                folder_path="Inbox",
            )
        ]

        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pruned_unused_filesystem_dataset"], 1)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            dataset_rows = connection.execute(
                """
                SELECT *
                FROM datasets
                ORDER BY id ASC
                """
            ).fetchall()
            stale_dataset_row = connection.execute(
                "SELECT * FROM datasets WHERE id = ?",
                (stale_dataset_id,),
            ).fetchone()
            stale_dataset_source_row = connection.execute(
                "SELECT * FROM dataset_sources WHERE id = ?",
                (stale_dataset_source_id,),
            ).fetchone()
        finally:
            connection.close()

        self.assertIsNone(stale_dataset_row)
        self.assertIsNone(stale_dataset_source_row)
        self.assertEqual(len(dataset_rows), 1)
        self.assertEqual(dataset_rows[0]["source_kind"], retriever_tools.PST_SOURCE_KIND)

    def test_ingest_pst_chat_like_message_uses_chat_document_metadata(self) -> None:
        self.write_fake_pst_file()
        transcript = "\n".join(
            [
                "[2026-04-15 09:00] Alice Example: Kickoff thread for launch planning.",
                "[2026-04-15 09:05] Bob Example: I'll draft the update.",
                "[2026-04-15 09:07] Alice Example: Great, thanks.",
            ]
        )
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-chat-001",
                subject=None,
                body_text=transcript,
                folder_path="Conversation History",
                author="Conversation History",
                recipients=None,
                date_created="2026-04-16T00:00:00Z",
                message_class="IPM.Note.Microsoft.Conversation",
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 1)
        self.assertEqual(ingest_result["pst_chat_conversations"], 1)

        row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-chat-001"))
        self.assertEqual(row["content_type"], "Chat")
        self.assertIsNone(row["author"])
        self.assertEqual(row["custodian"], "mailbox")
        self.assertEqual(row["participants"], "Alice Example, Bob Example")
        self.assertEqual(row["date_created"], "2026-04-15T09:00:00Z")
        self.assertEqual(row["date_modified"], "2026-04-15T09:07:00Z")
        self.assertEqual(row["title"], "Kickoff thread for launch planning.")
        self.assertIsNone(row["subject"])
        self.assertIsNone(row["recipients"])
        self.assertEqual(row["source_folder_path"], "Conversation History")
        self.assertIsNotNone(row["conversation_id"])

        search_result = retriever_tools.search(self.root, "draft the update", None, None, None, 1, 20)
        result = search_result["results"][0]
        self.assertEqual(
            [target.get("label") for target in result["preview_targets"]],
            ["message", "conversation"],
        )
        self.assertEqual(
            result["preview_rel_path"],
            self.preview_target_by_label(result["preview_targets"], "message")["rel_path"],
        )
        preview_html = Path(str(result["preview_abs_path"]).split("#", 1)[0]).read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', preview_html)
        self.assertIn("Alice Example", preview_html)
        self.assertIn("Bob Example", preview_html)
        self.assertIn("Kickoff thread for launch planning.", preview_html)
        conversation_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(result["preview_targets"], "conversation")
        ).read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', conversation_preview_html)
        self.assertIn("draft the update.", conversation_preview_html)

    def test_ingest_groups_pst_chat_messages_by_thread_id_over_shared_folder_path(self) -> None:
        self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-chat-001",
                subject=None,
                body_text="Kickoff thread for launch planning.",
                folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                author="Alice Example",
                recipients=None,
                date_created="2026-04-15T09:00:00Z",
                chat_threading={
                    "thread_id": "19:launch-thread@unq.gbl.spaces",
                    "message_id": "1713882000000",
                    "thread_type": "chat",
                    "participants": ["Alice Example", "Bob Example"],
                },
            ),
            self.build_fake_pst_message(
                source_item_id="pst-chat-002",
                subject=None,
                body_text="Follow-up from the same Teams space.",
                folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                author="Bob Example",
                recipients=None,
                date_created="2026-05-16T09:00:00Z",
                chat_threading={
                    "thread_id": "19:launch-thread@unq.gbl.spaces",
                    "message_id": "1713968400000",
                    "thread_type": "chat",
                    "participants": ["Bob Example", "Alice Example"],
                },
            ),
            self.build_fake_pst_message(
                source_item_id="pst-chat-003",
                subject=None,
                body_text="Separate one-off Teams thread.",
                folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                author="Alice Example",
                recipients=None,
                date_created="2026-04-16T11:00:00Z",
                chat_threading={
                    "thread_id": "19:separate-thread@unq.gbl.spaces",
                    "message_id": "1713975600000",
                    "thread_type": "chat",
                    "participants": ["Alice Example", "Carol Example"],
                },
            ),
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 3)
        self.assertEqual(ingest_result["pst_chat_conversations"], 2)

        first_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-chat-001"))
        second_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-chat-002"))
        third_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-chat-003"))
        self.assertIsNotNone(first_row["conversation_id"])
        self.assertEqual(first_row["conversation_id"], second_row["conversation_id"])
        self.assertNotEqual(first_row["conversation_id"], third_row["conversation_id"])
        self.assertEqual(first_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)
        self.assertEqual(second_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)
        self.assertEqual(third_row["conversation_assignment_mode"], retriever_tools.CONVERSATION_ASSIGNMENT_MODE_AUTO)
        self.assertEqual(first_row["title"], "Alice Example / Bob Example")
        self.assertEqual(second_row["title"], "Alice Example / Bob Example")
        self.assertEqual(third_row["title"], "Alice Example / Carol Example")
        self.assertEqual(first_row["participants"], "Alice Example, Bob Example")
        self.assertEqual(second_row["participants"], "Alice Example, Bob Example")
        self.assertEqual(third_row["participants"], "Alice Example, Carol Example")
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            threading_rows = connection.execute(
                """
                SELECT document_id, thread_id, thread_type
                FROM document_chat_threading
                ORDER BY document_id ASC
                """
            ).fetchall()
        finally:
            connection.close()
        self.assertEqual(
            [(row["document_id"], row["thread_id"], row["thread_type"]) for row in threading_rows],
            [
                (first_row["id"], "19:launch-thread@unq.gbl.spaces", "chat"),
                (second_row["id"], "19:launch-thread@unq.gbl.spaces", "chat"),
                (third_row["id"], "19:separate-thread@unq.gbl.spaces", "chat"),
            ],
        )

        search_result = retriever_tools.search(self.root, "Follow-up from the same Teams space", None, None, None, 1, 20)
        result = search_result["results"][0]
        self.assertEqual(
            [target.get("label") for target in result["preview_targets"]],
            ["message", "conversation"],
        )
        self.assertEqual(
            result["preview_rel_path"],
            self.preview_target_by_label(result["preview_targets"], "message")["rel_path"],
        )
        self.assertEqual(
            self.preview_target_file_path(
                self.preview_target_by_label(result["preview_targets"], "conversation")
            ).name,
            "conversation.html",
        )
        message_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(result["preview_targets"], "message")
        ).read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', message_preview_html)
        self.assertIn("Bob Example", message_preview_html)
        self.assertIn("Follow-up from the same Teams space.", message_preview_html)
        preview_html = self.preview_target_file_path(
            self.preview_target_by_label(result["preview_targets"], "conversation")
        ).read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', preview_html)
        self.assertIn("Alice Example", preview_html)
        self.assertIn("Bob Example", preview_html)
        self.assertIn("Kickoff thread for launch planning.", preview_html)
        self.assertIn("Follow-up from the same Teams space.", preview_html)

    def test_ingest_deduplicates_exact_pst_chat_messages_across_custodians(self) -> None:
        self.write_fake_pst_file(name="max.pst", content=b"pst-max-v1")
        self.write_fake_pst_file(name="sergey.pst", content=b"pst-sergey-v1")
        thread_id = "19:launch-thread@unq.gbl.spaces"
        shared_body = "Kickoff thread for launch planning."
        messages_by_name = {
            "max.pst": [
                self.build_fake_pst_message(
                    source_item_id="pst-chat-max-001",
                    subject=None,
                    body_text=shared_body,
                    folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                    author="max",
                    recipients=None,
                    date_created="2026-04-15T09:00:00Z",
                    chat_threading={
                        "thread_id": thread_id,
                        "message_id": "1713882000000",
                        "thread_type": "chat",
                        "participants": ["max", "sergey"],
                    },
                )
            ],
            "sergey.pst": [
                self.build_fake_pst_message(
                    source_item_id="pst-chat-sergey-001",
                    subject=None,
                    body_text=shared_body,
                    folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                    author="max",
                    recipients=None,
                    date_created="2026-04-15T09:00:00Z",
                    chat_threading={
                        "thread_id": thread_id,
                        "message_id": "1713882000000",
                        "thread_type": "chat",
                        "participants": ["max", "sergey"],
                    },
                )
            ],
        }

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(
            retriever_tools,
            "iter_pst_messages",
            side_effect=lambda path: iter(messages_by_name[path.name]),
        ):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        search_result = retriever_tools.search(self.root, shared_body, None, None, None, 1, 20)
        self.assertEqual(search_result["total_hits"], 1)
        root_row = self.fetch_document_by_id(int(search_result["results"][0]["id"]))
        self.assertEqual(root_row["content_type"], "Chat")
        self.assertEqual(root_row["title"], "max / sergey")
        self.assertEqual(set(root_row["custodians"]), {"max", "sergey"})
        self.assertIsNotNone(root_row["conversation_id"])

        occurrence_rows = self.fetch_occurrence_rows(int(root_row["id"]))
        active_root_occurrence_rows = [
            row
            for row in occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
            and row["parent_occurrence_id"] is None
        ]
        self.assertEqual(
            sorted((str(row["source_rel_path"]), str(row["source_item_id"])) for row in active_root_occurrence_rows),
            [("max.pst", "pst-chat-max-001"), ("sergey.pst", "pst-chat-sergey-001")],
        )
        self.assertEqual(self.count_rows("documents"), 1)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            conversation_rows = connection.execute(
                """
                SELECT source_locator, conversation_key
                FROM conversations
                ORDER BY id ASC
                """
            ).fetchall()
        finally:
            connection.close()
        self.assertEqual(len(conversation_rows), 1)
        self.assertEqual(conversation_rows[0]["source_locator"], retriever_tools.filesystem_dataset_locator())
        self.assertEqual(conversation_rows[0]["conversation_key"], f"thread:{thread_id}")

    def test_reingest_splits_deduplicated_pst_chat_when_one_source_copy_changes(self) -> None:
        self.write_fake_pst_file(name="max.pst", content=b"pst-max-v1")
        sergey_path = self.write_fake_pst_file(name="sergey.pst", content=b"pst-sergey-v1")
        thread_id = "19:launch-thread@unq.gbl.spaces"
        shared_body = "Kickoff thread for launch planning."
        diverged_body = "Diverged copy from Sergey's PST."
        messages_by_name = {
            "max.pst": [
                self.build_fake_pst_message(
                    source_item_id="pst-chat-max-001",
                    subject=None,
                    body_text=shared_body,
                    folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                    author="max",
                    recipients=None,
                    date_created="2026-04-15T09:00:00Z",
                    chat_threading={
                        "thread_id": thread_id,
                        "message_id": "1713882000000",
                        "thread_type": "chat",
                        "participants": ["max", "sergey"],
                    },
                )
            ],
            "sergey.pst": [
                self.build_fake_pst_message(
                    source_item_id="pst-chat-sergey-001",
                    subject=None,
                    body_text=shared_body,
                    folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                    author="max",
                    recipients=None,
                    date_created="2026-04-15T09:00:00Z",
                    chat_threading={
                        "thread_id": thread_id,
                        "message_id": "1713882000000",
                        "thread_type": "chat",
                        "participants": ["max", "sergey"],
                    },
                )
            ],
        }

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(
            retriever_tools,
            "iter_pst_messages",
            side_effect=lambda path: iter(messages_by_name[path.name]),
        ):
            first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["failed"], 0)
        self.assertEqual(
            retriever_tools.search(self.root, shared_body, None, None, None, 1, 20)["total_hits"],
            1,
        )

        sergey_path.write_bytes(b"pst-sergey-v2")
        messages_by_name["sergey.pst"] = [
            self.build_fake_pst_message(
                source_item_id="pst-chat-sergey-001",
                subject=None,
                body_text=diverged_body,
                folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                author="max",
                recipients=None,
                date_created="2026-04-15T09:00:00Z",
                chat_threading={
                    "thread_id": thread_id,
                    "message_id": "1713882000000",
                    "thread_type": "chat",
                    "participants": ["max", "sergey"],
                },
            )
        ]
        with mock.patch.object(
            retriever_tools,
            "iter_pst_messages",
            side_effect=lambda path: iter(messages_by_name[path.name]),
        ):
            second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["failed"], 0)
        self.assertEqual(self.count_rows("documents"), 2)

        shared_search = retriever_tools.search(self.root, shared_body, None, None, None, 1, 20)
        diverged_search = retriever_tools.search(self.root, diverged_body, None, None, None, 1, 20)
        self.assertEqual(shared_search["total_hits"], 1)
        self.assertEqual(diverged_search["total_hits"], 1)

        shared_row = self.fetch_document_by_id(int(shared_search["results"][0]["id"]))
        diverged_row = self.fetch_document_by_id(int(diverged_search["results"][0]["id"]))
        self.assertNotEqual(shared_row["id"], diverged_row["id"])
        self.assertEqual(set(shared_row["custodians"]), {"max"})
        self.assertEqual(set(diverged_row["custodians"]), {"sergey"})
        self.assertEqual(shared_row["conversation_id"], diverged_row["conversation_id"])

        shared_occurrence_rows = self.fetch_occurrence_rows(int(shared_row["id"]))
        diverged_occurrence_rows = self.fetch_occurrence_rows(int(diverged_row["id"]))
        active_shared_occurrence_rows = [
            row
            for row in shared_occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
            and row["parent_occurrence_id"] is None
        ]
        active_diverged_occurrence_rows = [
            row
            for row in diverged_occurrence_rows
            if row["lifecycle_status"] == retriever_tools.ACTIVE_OCCURRENCE_STATUS
            and row["parent_occurrence_id"] is None
        ]
        self.assertEqual(
            [(str(row["source_rel_path"]), str(row["source_item_id"])) for row in active_shared_occurrence_rows],
            [("max.pst", "pst-chat-max-001")],
        )
        self.assertEqual(
            [(str(row["source_rel_path"]), str(row["source_item_id"])) for row in active_diverged_occurrence_rows],
            [("sergey.pst", "pst-chat-sergey-001")],
        )

    def test_ingest_pst_teams_space_sender_hint_resolves_sender_by_email(self) -> None:
        self.write_fake_pst_file()
        aad_object_id = "45895f37-3f55-4093-9978-87581df21d97"
        compact_aad_object_id = aad_object_id.replace("-", "")
        thread_id = "19:channel-thread@thread.tacv2"

        def text_entry(entry_type: int, value: str) -> object:
            return types.SimpleNamespace(
                entry_type=entry_type,
                value_type=0x001F,
                data=(value + "\x00").encode("utf-16-le"),
            )

        raw_message = types.SimpleNamespace(
            sender_name="Denys",
            sender_email_address="",
            record_sets=[
                types.SimpleNamespace(
                    entries=[
                        text_entry(0x0C1A, "Denys"),
                        text_entry(0x0C1F, "denys@discoverbeagle.com"),
                        text_entry(
                            0x8043,
                            json.dumps(
                                {
                                    "SenderId": compact_aad_object_id,
                                    "ThreadId": thread_id,
                                    "MessageId": 1713916215808,
                                    "ParentMessageId": 1713916215808,
                                    "ThreadType": "space",
                                },
                                ensure_ascii=True,
                            ),
                        ),
                        text_entry(
                            0x804F,
                            json.dumps(
                                {
                                    "SchemaVersion": 1,
                                    "ItemData": json.dumps(
                                        {
                                            "messageFrom": f"8:orgid:{aad_object_id}",
                                            "from": {"internalId": f"8:orgid:{aad_object_id}"},
                                        },
                                        ensure_ascii=True,
                                    ),
                                },
                                ensure_ascii=True,
                            ),
                        ),
                    ]
                )
            ],
        )
        chat_threading = retriever_tools.extract_pst_chat_threading(raw_message)
        self.assertIsNotNone(chat_threading)
        assert chat_threading is not None
        self.assertEqual(chat_threading["thread_type"], "space")
        self.assertEqual(chat_threading["participants"], ["Denys"])
        participant_hints = chat_threading["participant_entity_hints"]
        self.assertEqual(len(participant_hints), 1)
        self.assertEqual(participant_hints[0]["display_value"], "Denys <denys@discoverbeagle.com>")
        self.assertEqual(participant_hints[0]["identifiers"][0]["identifier_name"], "aad_oid")
        self.assertEqual(participant_hints[0]["identifiers"][0]["normalized_value"], aad_object_id)

        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-space-001",
                subject="HI",
                body_text="HI",
                folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                author="Denys",
                recipients=None,
                date_created="2026-04-15T09:00:00Z",
                chat_threading=chat_threading,
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 1)
        row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-space-001"))
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            participant_rows = connection.execute(
                """
                SELECT e.display_name, e.primary_email, e.entity_origin, de.evidence_json
                FROM document_entities de
                JOIN entities e ON e.id = de.entity_id
                WHERE de.document_id = ?
                  AND de.role = 'participant'
                """,
                (row["id"],),
            ).fetchall()
            hint_row = connection.execute(
                """
                SELECT entity_hints_json
                FROM document_occurrences
                WHERE document_id = ?
                """,
                (row["id"],),
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual(len(participant_rows), 1)
        self.assertEqual(participant_rows[0]["primary_email"], "denys@discoverbeagle.com")
        self.assertEqual(participant_rows[0]["entity_origin"], retriever_tools.ENTITY_ORIGIN_IDENTIFIED)
        participant_evidence = json.loads(participant_rows[0]["evidence_json"])
        self.assertEqual(participant_evidence["raw_value"], "Denys <denys@discoverbeagle.com>")
        self.assertIsNotNone(hint_row)
        hint_payload = json.loads(hint_row["entity_hints_json"])
        self.assertEqual(hint_payload["participants"][0]["display_value"], "Denys <denys@discoverbeagle.com>")

    def test_ingest_pst_routes_teams_messages_to_chat_and_skips_system_folders(self) -> None:
        self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-team-001",
                subject=None,
                body_text="hey there",
                folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                author="Sergey Demyanov",
                recipients=None,
                date_created="2026-04-15T09:00:00Z",
            ),
            self.build_fake_pst_message(
                source_item_id="pst-spool-001",
                subject=None,
                body_text="",
                folder_path="Top of Personal Folders/user (Primary)/SubstrateFiles/SPOOLS",
                author=None,
                recipients=None,
                date_created="2026-04-15T09:01:00Z",
            ),
            self.build_fake_pst_message(
                source_item_id="pst-meeting-system-001",
                subject=None,
                body_text="<addmember><eventtime>1713916258178</eventtime></addmember>",
                folder_path="Top of Personal Folders/user (Primary)/SkypeSpacesData/TeamsMeetings",
                author="19:meeting@unq.gbl.spaces",
                recipients=None,
                date_created="2026-04-15T09:02:00Z",
            ),
            self.build_fake_pst_message(
                source_item_id="pst-email-001",
                subject="Ordinary inbox message",
                body_text="Parent message body",
                folder_path="Top of Information Store/Inbox",
                author="Alice Example <alice@example.com>",
                recipients="Bob Example <bob@example.com>",
                date_created="2026-04-15T09:03:00Z",
            ),
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 2)
        self.assertEqual(ingest_result["workspace_parent_documents"], 2)

        chat_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "pst-team-001")
        email_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "pst-email-001")
        chat_row = self.fetch_document_row(chat_rel_path)
        email_row = self.fetch_document_row(email_rel_path)

        self.assertEqual(chat_row["content_type"], "Chat")
        self.assertIsNone(chat_row["author"])
        self.assertEqual(chat_row["custodian"], "mailbox")
        self.assertEqual(chat_row["participants"], "Sergey Demyanov")
        self.assertEqual(chat_row["title"], "hey there")
        self.assertEqual(chat_row["source_folder_path"], "Top of Personal Folders/user (Primary)/TeamsMessagesData")
        self.assertIsNotNone(chat_row["conversation_id"])
        self.assertEqual(email_row["content_type"], "Email")
        self.assertEqual(email_row["custodian"], "mailbox")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            skipped_count = int(
                connection.execute(
                    """
                    SELECT COUNT(*) AS count
                    FROM documents
                    WHERE rel_path IN (?, ?)
                    """,
                    (
                        retriever_tools.pst_message_rel_path("mailbox.pst", "pst-spool-001"),
                        retriever_tools.pst_message_rel_path("mailbox.pst", "pst-meeting-system-001"),
                    ),
                ).fetchone()["count"]
            )
        finally:
            connection.close()
        self.assertEqual(skipped_count, 0)

        search_result = retriever_tools.search(self.root, "hey there", None, None, None, 1, 20)
        result = next(item for item in search_result["results"] if item["id"] == chat_row["id"])
        self.assertEqual(
            [target.get("label") for target in result["preview_targets"]],
            ["message", "conversation"],
        )
        self.assertEqual(
            result["preview_rel_path"],
            self.preview_target_by_label(result["preview_targets"], "message")["rel_path"],
        )
        preview_html = Path(str(result["preview_abs_path"]).split("#", 1)[0]).read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', preview_html)
        self.assertIn("Sergey Demyanov", preview_html)
        self.assertIn("hey there", preview_html)
        conversation_preview_html = self.preview_target_file_path(
            self.preview_target_by_label(result["preview_targets"], "conversation")
        ).read_text(encoding="utf-8")
        self.assertIn('class="chat-message"', conversation_preview_html)
        self.assertIn("Sergey Demyanov", conversation_preview_html)
        self.assertIn("hey there", conversation_preview_html)

    def test_ingest_pst_calendar_folder_uses_calendar_content_type(self) -> None:
        self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-calendar-001",
                subject="Labor Day",
                body_text="",
                folder_path="Top of Information Store/Calendar/United States holidays",
                author="sergey",
                recipients=None,
                date_created="2026-09-07T00:00:00Z",
                message_class="IPM.Appointment",
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 1)

        row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-calendar-001"))
        self.assertEqual(row["content_type"], "Calendar")
        self.assertEqual(row["custodian"], "mailbox")
        self.assertEqual(row["title"], "Labor Day")
        self.assertEqual(row["subject"], "Labor Day")
        self.assertEqual(row["author"], "sergey")

        search_result = retriever_tools.search(
            self.root,
            "",
            [["content_type", "eq", "Calendar"], ["source_kind", "eq", retriever_tools.PST_SOURCE_KIND]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(search_result["total_hits"], 1)
        preview_html = Path(search_result["results"][0]["preview_targets"][0]["abs_path"]).read_text(encoding="utf-8")
        self.assertIn("<title>Labor Day</title>", preview_html)
        self.assertIn("<h1>Labor Day</h1>", preview_html)
        self.assertIn("Labor Day", preview_html)
        self.assertIn("Sep 7, 2026 12:00 AM UTC", preview_html)

    def test_ingest_pst_calendar_uses_mapi_smtp_identity_hints(self) -> None:
        self.write_fake_pst_file()
        aad_object_id = "ed697494-fafd-46b5-aaec-222f396a4264"

        def text_entry(entry_type: int, value: str) -> object:
            return types.SimpleNamespace(
                entry_type=entry_type,
                value_type=0x001F,
                data=(value + "\x00").encode("utf-16-le"),
            )

        raw_message = types.SimpleNamespace(
            sender_name="sergey",
            sender_email_address="",
            record_sets=[
                types.SimpleNamespace(
                    entries=[
                        text_entry(0x0042, "sergey"),
                        text_entry(0x0C1A, "sergey"),
                        text_entry(0x5D01, "sergey@discoverbeagle.com"),
                        text_entry(0x5D02, "sergey@discoverbeagle.com"),
                        text_entry(0x5D0A, "sergey@discoverbeagle.com"),
                        text_entry(0x5D0B, "sergey@discoverbeagle.com"),
                        text_entry(
                            0x0C1F,
                            (
                                "/O=EXCHANGELABS/OU=EXCHANGE ADMINISTRATIVE GROUP/"
                                f"CN=RECIPIENTS/CN=FC6C02487FCA4662A66CD5AE66748DBD-{aad_object_id}"
                            ),
                        ),
                    ]
                )
            ],
        )
        entity_hints = retriever_tools.pst_message_mapi_entity_hints(raw_message)
        self.assertEqual(entity_hints["author"][0]["display_value"], "sergey <sergey@discoverbeagle.com>")
        self.assertEqual(entity_hints["participants"][0]["display_value"], "sergey <sergey@discoverbeagle.com>")
        self.assertEqual(entity_hints["author"][0]["identifiers"][0]["identifier_name"], "aad_oid")
        self.assertEqual(entity_hints["author"][0]["identifiers"][0]["normalized_value"], aad_object_id)

        message = self.build_fake_pst_message(
            source_item_id="pst-calendar-identity-001",
            subject="Administrative Professionals Day",
            body_text="United States holiday",
            folder_path="Top of Information Store/Calendar/United States holidays",
            author="sergey",
            recipients=None,
            date_created="2026-04-22T00:00:00Z",
            message_class="IPM.Appointment",
        )
        message["entity_hints"] = entity_hints

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter([message])):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["pst_messages_created"], 1)

        row = self.fetch_document_row(
            retriever_tools.pst_message_rel_path("mailbox.pst", "pst-calendar-identity-001")
        )
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            entity_rows = connection.execute(
                """
                SELECT de.role, de.entity_id, e.display_name, e.primary_email, e.entity_origin, de.evidence_json
                FROM document_entities de
                JOIN entities e ON e.id = de.entity_id
                WHERE de.document_id = ?
                  AND de.role IN ('author', 'participant')
                ORDER BY de.role ASC
                """,
                (row["id"],),
            ).fetchall()
            hint_row = connection.execute(
                """
                SELECT entity_hints_json
                FROM document_occurrences
                WHERE document_id = ?
                """,
                (row["id"],),
            ).fetchone()
            aad_identifier_row = connection.execute(
                """
                SELECT entity_id, identifier_name, normalized_value
                FROM entity_identifiers
                WHERE identifier_type = 'external_id'
                  AND identifier_name = 'aad_oid'
                  AND normalized_value = ?
                """,
                (aad_object_id,),
            ).fetchone()
        finally:
            connection.close()

        self.assertEqual([item["role"] for item in entity_rows], ["author", "participant"])
        self.assertEqual(len({int(item["entity_id"]) for item in entity_rows}), 1)
        self.assertEqual({item["primary_email"] for item in entity_rows}, {"sergey@discoverbeagle.com"})
        self.assertEqual({item["entity_origin"] for item in entity_rows}, {retriever_tools.ENTITY_ORIGIN_IDENTIFIED})
        evidence_values = [json.loads(item["evidence_json"])["raw_value"] for item in entity_rows]
        self.assertEqual(evidence_values, ["sergey <sergey@discoverbeagle.com>", "sergey <sergey@discoverbeagle.com>"])
        self.assertIsNotNone(hint_row)
        hint_payload = json.loads(hint_row["entity_hints_json"])
        self.assertEqual(hint_payload["author"][0]["display_value"], "sergey <sergey@discoverbeagle.com>")
        self.assertEqual(hint_payload["participants"][0]["display_value"], "sergey <sergey@discoverbeagle.com>")
        self.assertIsNotNone(aad_identifier_row)
        self.assertEqual(int(aad_identifier_row["entity_id"]), int(entity_rows[0]["entity_id"]))

    def test_unchanged_pst_source_skips_without_reparsing(self) -> None:
        self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-msg-001",
                subject="PST Parent",
                body_text="Parent message body",
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["new"], 1)
        with mock.patch.object(retriever_tools, "iter_pst_messages", side_effect=AssertionError("PST iterator should not run on unchanged source")):
            second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["skipped"], 1)
        self.assertEqual(second_ingest["pst_sources_skipped"], 1)
        self.assertEqual(second_ingest["failed"], 0)
        browse_results = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_results["total_hits"], 1)

    def test_unchanged_pst_source_recreates_missing_dataset_source_without_nested_transaction(self) -> None:
        self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-msg-001",
                subject="PST Parent",
                body_text="Parent message body",
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["new"], 1)
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute(
                """
                DELETE FROM dataset_sources
                WHERE source_kind = ? AND source_locator = ?
                """,
                (retriever_tools.PST_SOURCE_KIND, "mailbox.pst"),
            )
            connection.commit()
        finally:
            connection.close()

        with mock.patch.object(
            retriever_tools,
            "iter_pst_messages",
            side_effect=AssertionError("PST iterator should not run on unchanged source"),
        ):
            second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["skipped"], 1)
        self.assertEqual(second_ingest["pst_sources_skipped"], 1)
        self.assertEqual(second_ingest["failed"], 0)
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            dataset_source_row = connection.execute(
                """
                SELECT id, dataset_id
                FROM dataset_sources
                WHERE source_kind = ? AND source_locator = ?
                """,
                (retriever_tools.PST_SOURCE_KIND, "mailbox.pst"),
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(dataset_source_row)

    def test_ingest_rolls_back_failed_pst_source_and_continues(self) -> None:
        self.write_fake_pst_file()
        (self.root / "note.txt").write_text("plain text body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)

        def broken_ingest_pst_source(connection: sqlite3.Connection, *args, **kwargs):
            connection.execute(
                """
                UPDATE workspace_meta
                SET updated_at = updated_at
                WHERE id = 1
                """
            )
            raise sqlite3.OperationalError("synthetic nested transaction trigger")

        with mock.patch.object(retriever_tools, "ingest_pst_source", side_effect=broken_ingest_pst_source):
            ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["failed"], 1)
        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(len(ingest_result["failures"]), 1)
        self.assertEqual(ingest_result["failures"][0]["rel_path"], "mailbox.pst")
        note_row = self.fetch_document_row("note.txt")
        self.assertIsNotNone(note_row)

    def test_unchanged_pst_chat_source_reparses_when_chat_threading_rows_are_missing(self) -> None:
        self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-chat-001",
                subject=None,
                body_text="Kickoff thread for launch planning.",
                folder_path="Top of Personal Folders/user (Primary)/TeamsMessagesData",
                author="Alice Example",
                recipients=None,
                date_created="2026-04-15T09:00:00Z",
                chat_threading={
                    "thread_id": "19:launch-thread@unq.gbl.spaces",
                    "message_id": "1713882000000",
                    "thread_type": "chat",
                    "participants": ["Alice Example", "Bob Example"],
                },
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["pst_messages_created"], 1)
        parent_row = self.fetch_document_row(retriever_tools.pst_message_rel_path("mailbox.pst", "pst-chat-001"))
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            connection.execute(
                "DELETE FROM document_chat_threading WHERE document_id = ?",
                (int(parent_row["id"]),),
            )
            connection.commit()
        finally:
            connection.close()

        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["failed"], 0)
        self.assertEqual(second_ingest["pst_sources_skipped"], 0)
        self.assertEqual(second_ingest["pst_messages_updated"], 1)
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            threading_row = connection.execute(
                """
                SELECT thread_id, thread_type
                FROM document_chat_threading
                WHERE document_id = ?
                """,
                (int(parent_row["id"]),),
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(threading_row)
        assert threading_row is not None
        self.assertEqual(threading_row["thread_id"], "19:launch-thread@unq.gbl.spaces")
        self.assertEqual(threading_row["thread_type"], "chat")

    def test_changed_pst_reingest_preserves_control_numbers_and_retires_removed_messages(self) -> None:
        pst_path = self.write_fake_pst_file(content=b"pst-v1")
        first_messages = [
            self.build_fake_pst_message(
                source_item_id="pst-msg-001",
                subject="Original PST Parent",
                body_text="Parent v1",
                attachment_name="notes.txt",
                attachment_text="stable attachment body",
            ),
            self.build_fake_pst_message(
                source_item_id="pst-msg-002",
                subject="Removed later",
                body_text="Remove me",
            ),
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(first_messages)):
            first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["pst_messages_created"], 2)
        parent_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "pst-msg-001")
        removed_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "pst-msg-002")
        parent_row = self.fetch_document_row(parent_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        removed_row = self.fetch_document_row(removed_rel_path)
        retriever_tools.set_field(self.root, parent_row["id"], "title", "Manual PST Title")

        pst_path.write_bytes(b"pst-v2-expanded")
        second_messages = [
            self.build_fake_pst_message(
                source_item_id="pst-msg-001",
                subject="Updated PST Parent",
                body_text="Parent v2",
                attachment_name="notes.txt",
                attachment_text="stable attachment body",
            )
        ]
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(second_messages)):
            second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["updated"], 1)
        self.assertEqual(second_ingest["pst_messages_updated"], 1)
        self.assertEqual(second_ingest["pst_messages_deleted"], 1)

        updated_parent = self.fetch_document_row(parent_rel_path)
        updated_child = self.fetch_child_rows(updated_parent["id"])[0]
        retired_row = self.fetch_document_row(removed_rel_path)
        self.assertEqual(updated_parent["control_number"], parent_row["control_number"])
        self.assertEqual(updated_parent["title"], "Manual PST Title")
        self.assertEqual(updated_child["id"], child_row["id"])
        self.assertEqual(updated_child["control_number"], child_row["control_number"])
        self.assertEqual(retired_row["lifecycle_status"], "deleted")

    def test_missing_pst_source_marks_messages_and_children_missing(self) -> None:
        pst_path = self.write_fake_pst_file()
        messages = [
            self.build_fake_pst_message(
                source_item_id="pst-msg-001",
                subject="PST Parent",
                body_text="Parent body",
                attachment_name="notes.txt",
                attachment_text="attachment body",
            )
        ]

        retriever_tools.bootstrap(self.root)
        with mock.patch.object(retriever_tools, "iter_pst_messages", return_value=iter(messages)):
            first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(first_ingest["pst_messages_created"], 1)
        parent_rel_path = retriever_tools.pst_message_rel_path("mailbox.pst", "pst-msg-001")
        parent_row = self.fetch_document_row(parent_rel_path)
        child_row = self.fetch_child_rows(parent_row["id"])[0]

        pst_path.unlink()
        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(second_ingest["missing"], 1)
        self.assertEqual(second_ingest["pst_sources_missing"], 1)
        self.assertEqual(second_ingest["pst_documents_missing"], 2)
        missing_parent = self.fetch_document_row(parent_rel_path)
        missing_child = self.fetch_document_by_id(child_row["id"])
        self.assertEqual(missing_parent["lifecycle_status"], "missing")
        self.assertEqual(missing_child["lifecycle_status"], "missing")

    def test_plain_ingest_auto_routes_detected_production_roots(self) -> None:
        self.write_production_fixture()
        loose_file = self.root / "notes.txt"
        loose_file.write_text("loose workspace note\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 5)
        self.assertEqual(ingest_result["updated"], 0)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["ingested_production_roots"], ["Synthetic_Production"])
        self.assertEqual(ingest_result["skipped_production_roots"], [])
        self.assertEqual(ingest_result["production_documents_created"], 4)
        self.assertEqual(ingest_result["production_documents_updated"], 0)
        self.assertEqual(ingest_result["production_documents_unchanged"], 0)
        self.assertEqual(ingest_result["production_documents_retired"], 0)
        self.assertEqual(ingest_result["production_families_reconstructed"], 1)
        self.assertEqual(ingest_result["production_docs_missing_linked_text"], 0)
        self.assertEqual(ingest_result["production_docs_missing_linked_images"], 1)
        self.assertEqual(ingest_result["production_docs_missing_linked_natives"], 0)
        self.assertNotIn("warnings", ingest_result)

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_result["total_hits"], 5)
        self.assertIn("notes.txt", [item["file_name"] for item in browse_result["results"]])

        production_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000001.logical"
        )
        self.assertEqual(production_row["source_kind"], retriever_tools.PRODUCTION_SOURCE_KIND)

    def test_plain_ingest_auto_routes_multiple_production_roots_without_double_indexing_loose_artifacts(self) -> None:
        self.write_production_fixture()
        self.write_production_fixture(production_name="Second_Production", control_prefix="QDX")
        loose_file = self.root / "summary.txt"
        loose_file.write_text("workspace summary\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 9)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(
            ingest_result["ingested_production_roots"],
            ["Second_Production", "Synthetic_Production"],
        )
        self.assertEqual(ingest_result["skipped_production_roots"], [])
        self.assertEqual(ingest_result["production_documents_created"], 8)
        self.assertEqual(ingest_result["production_documents_updated"], 0)
        self.assertEqual(ingest_result["production_documents_unchanged"], 0)
        self.assertEqual(ingest_result["production_documents_retired"], 0)
        self.assertEqual(ingest_result["production_families_reconstructed"], 2)

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            production_doc_count = connection.execute(
                "SELECT COUNT(*) AS count FROM documents WHERE source_kind = ?",
                (retriever_tools.PRODUCTION_SOURCE_KIND,),
            ).fetchone()["count"]
            loose_artifact_rows = connection.execute(
                """
                SELECT rel_path
                FROM documents
                WHERE rel_path LIKE 'Synthetic_Production/TEXT/%'
                   OR rel_path LIKE 'Synthetic_Production/IMAGES/%'
                   OR rel_path LIKE 'Synthetic_Production/NATIVES/%'
                   OR rel_path LIKE 'Second_Production/TEXT/%'
                   OR rel_path LIKE 'Second_Production/IMAGES/%'
                   OR rel_path LIKE 'Second_Production/NATIVES/%'
                """
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(production_doc_count, 8)
        self.assertEqual(loose_artifact_rows, [])

    def test_plain_ingest_with_file_type_filter_still_skips_detected_production_roots(self) -> None:
        self.write_production_fixture()
        loose_file = self.root / "notes.txt"
        loose_file.write_text("loose workspace note\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types="txt")

        self.assertEqual(ingest_result["new"], 1)
        self.assertEqual(ingest_result["failed"], 0)
        self.assertEqual(ingest_result["ingested_production_roots"], [])
        self.assertEqual(ingest_result["skipped_production_roots"], ["Synthetic_Production"])
        self.assertEqual(ingest_result["production_documents_created"], 0)
        self.assertIn("use ingest-production", ingest_result["warnings"][0])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            production_doc_count = connection.execute(
                "SELECT COUNT(*) AS count FROM documents WHERE source_kind = ?",
                (retriever_tools.PRODUCTION_SOURCE_KIND,),
            ).fetchone()["count"]
        finally:
            connection.close()
        self.assertEqual(production_doc_count, 0)

    def test_plain_ingest_auto_routed_production_matches_direct_ingest_production(self) -> None:
        production_root = self.write_production_fixture()

        def production_snapshot(workspace_root: Path) -> tuple[list[dict[str, object]], dict[str, list[tuple[str, str, int]]]]:
            connection = retriever_tools.connect_db(retriever_tools.workspace_paths(workspace_root)["db_path"])
            try:
                rows = connection.execute(
                    """
                    SELECT id, control_number, begin_bates, end_bates, begin_attachment, end_attachment,
                           parent_document_id, rel_path, file_name, content_type, text_status
                    FROM documents
                    WHERE source_kind = ?
                    ORDER BY control_number ASC, id ASC
                    """,
                    (retriever_tools.PRODUCTION_SOURCE_KIND,),
                ).fetchall()
                control_by_id = {int(row["id"]): str(row["control_number"]) for row in rows}
                documents: list[dict[str, object]] = []
                source_parts: dict[str, list[tuple[str, str, int]]] = {}
                for row in rows:
                    documents.append(
                        {
                            "control_number": row["control_number"],
                            "begin_bates": row["begin_bates"],
                            "end_bates": row["end_bates"],
                            "begin_attachment": row["begin_attachment"],
                            "end_attachment": row["end_attachment"],
                            "parent_control_number": (
                                control_by_id[int(row["parent_document_id"])]
                                if row["parent_document_id"] is not None
                                else None
                            ),
                            "rel_path": row["rel_path"],
                            "file_name": row["file_name"],
                            "content_type": row["content_type"],
                            "text_status": row["text_status"],
                        }
                    )
                    part_rows = connection.execute(
                        """
                        SELECT part_kind, rel_source_path, ordinal
                        FROM document_source_parts
                        WHERE document_id = ?
                        ORDER BY part_kind ASC, ordinal ASC, id ASC
                        """,
                        (row["id"],),
                    ).fetchall()
                    source_parts[str(row["control_number"])] = [
                        (str(part["part_kind"]), str(part["rel_source_path"]), int(part["ordinal"]))
                        for part in part_rows
                    ]
                return documents, source_parts
            finally:
                connection.close()

        retriever_tools.bootstrap(self.root)
        direct_result = retriever_tools.ingest_production(self.root, production_root)
        direct_documents, direct_source_parts = production_snapshot(self.root)

        with tempfile.TemporaryDirectory(prefix="retriever-auto-production-") as auto_tempdir:
            auto_root = Path(auto_tempdir)
            shutil.copytree(production_root, auto_root / production_root.name)
            auto_paths = retriever_tools.workspace_paths(auto_root)
            retriever_tools.ensure_layout(auto_paths)
            retriever_tools.bootstrap(auto_root)
            auto_result = retriever_tools.ingest(auto_root, recursive=True, raw_file_types=None)
            auto_documents, auto_source_parts = production_snapshot(auto_root)

        self.assertEqual(direct_result["created"], 4)
        self.assertEqual(auto_result["production_documents_created"], 4)
        self.assertEqual(auto_result["production_documents_updated"], 0)
        self.assertEqual(auto_result["production_documents_retired"], 0)
        self.assertEqual(auto_result["production_families_reconstructed"], direct_result["families_reconstructed"])
        self.assertEqual(auto_result["ingested_production_roots"], ["Synthetic_Production"])
        self.assertEqual(auto_documents, direct_documents)
        self.assertEqual(auto_source_parts, direct_source_parts)

    def test_plain_ingest_continues_after_detected_production_root_failure(self) -> None:
        self.write_production_fixture(production_name="Valid_Production", control_prefix="VDX")
        broken_root = self.root / "Broken_Production"
        (broken_root / "DATA").mkdir(parents=True, exist_ok=True)
        broken_text_dir = broken_root / "TEXT" / "TEXT001"
        broken_text_dir.mkdir(parents=True, exist_ok=True)
        (broken_root / "IMAGES" / "IMG001").mkdir(parents=True, exist_ok=True)
        (broken_text_dir / "BDX000001.txt").write_text("broken production text should stay unindexed\n", encoding="utf-8")
        loose_file = self.root / "notes.txt"
        loose_file.write_text("loose workspace note\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            retriever_tools.upsert_production_row(
                connection,
                dataset_id=None,
                rel_root="Broken_Production",
                production_name="Broken_Production",
                metadata_load_rel_path="Broken_Production/DATA/Broken_Production.dat",
                image_load_rel_path="Broken_Production/DATA/Broken_Production.opt",
                source_type="concordance-dat-opt",
            )
            connection.commit()
        finally:
            connection.close()

        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)

        self.assertEqual(ingest_result["new"], 5)
        self.assertEqual(ingest_result["failed"], 1)
        self.assertEqual(ingest_result["ingested_production_roots"], ["Valid_Production"])
        self.assertEqual(ingest_result["production_documents_created"], 4)
        self.assertTrue(any(item["rel_path"] == "Broken_Production" for item in ingest_result["failures"]))

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            broken_doc_count = connection.execute(
                "SELECT COUNT(*) AS count FROM documents WHERE rel_path LIKE 'Broken_Production/%'",
            ).fetchone()["count"]
            valid_doc_count = connection.execute(
                """
                SELECT COUNT(*) AS count
                FROM documents
                WHERE source_kind = ?
                  AND production_id = (SELECT id FROM productions WHERE rel_root = ?)
                """,
                (retriever_tools.PRODUCTION_SOURCE_KIND, "Valid_Production"),
            ).fetchone()["count"]
        finally:
            connection.close()

        self.assertEqual(broken_doc_count, 0)
        self.assertEqual(valid_doc_count, 4)

    def test_ingest_production_creates_logical_documents_bates_lookup_and_preview_precedence(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)

        self.assertEqual(ingest_result["created"], 4)
        self.assertEqual(ingest_result["retired"], 0)
        self.assertEqual(ingest_result["docs_missing_linked_text"], 0)
        self.assertEqual(ingest_result["families_reconstructed"], 1)
        self.assertEqual(ingest_result["failures"], [])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            production_row = connection.execute("SELECT * FROM productions WHERE rel_root = ?", ("Synthetic_Production",)).fetchone()
            source_parts = connection.execute(
                """
                SELECT part_kind, rel_source_path
                FROM document_source_parts
                WHERE document_id = (SELECT id FROM documents WHERE control_number = 'PDX000004')
                ORDER BY part_kind ASC, ordinal ASC
                """
            ).fetchall()
            dataset_row = connection.execute(
                """
                SELECT *
                FROM datasets
                WHERE id = (SELECT dataset_id FROM productions WHERE rel_root = ?)
                """,
                ("Synthetic_Production",),
            ).fetchone()
        finally:
            connection.close()

        self.assertIsNotNone(production_row)
        self.assertEqual(production_row["production_name"], "Synthetic_Production")
        self.assertIsNotNone(production_row["dataset_id"])
        self.assertIsNotNone(dataset_row)
        self.assertEqual(dataset_row["source_kind"], retriever_tools.PRODUCTION_SOURCE_KIND)
        self.assertEqual(dataset_row["dataset_locator"], "Synthetic_Production")
        self.assertEqual(dataset_row["dataset_name"], "Synthetic_Production")
        self.assertTrue(any(row["part_kind"] == "native" for row in source_parts))

        parent_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000001.logical"
        )
        child_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000003.logical"
        )
        native_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000004.logical"
        )
        image_only_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )

        self.assertEqual(parent_row["control_number"], "PDX000001")
        self.assertEqual(parent_row["begin_bates"], "PDX000001")
        self.assertEqual(parent_row["end_bates"], "PDX000002")
        self.assertEqual(parent_row["source_kind"], retriever_tools.PRODUCTION_SOURCE_KIND)
        self.assertEqual(parent_row["dataset_id"], production_row["dataset_id"])
        self.assertEqual(parent_row["content_type"], "Email")
        self.assertEqual(parent_row["author"], "Elena Steven <elena@example.com>")
        self.assertEqual(parent_row["date_created"], "2026-04-14T10:32:00Z")
        self.assertEqual(parent_row["title"], "Attachment Handling")
        self.assertEqual(parent_row["subject"], "Attachment Handling")
        self.assertEqual(parent_row["recipients"], "Harry Montoro <harry@example.com>")
        self.assertEqual(
            parent_row["participants"],
            "Elena Steven <elena@example.com>, Harry Montoro <harry@example.com>",
        )
        self.assertEqual(child_row["parent_document_id"], parent_row["id"])
        self.assertEqual(child_row["dataset_id"], production_row["dataset_id"])
        self.assertEqual(child_row["content_type"], "Email")
        self.assertEqual(child_row["author"], "Review Team")
        self.assertEqual(child_row["date_created"], "2026-04-14T09:00:00Z")
        self.assertEqual(child_row["title"], "Case status update")
        self.assertEqual(native_row["dataset_id"], production_row["dataset_id"])
        self.assertEqual(native_row["file_name"], "PDX000004.pdf")
        self.assertEqual(native_row["content_type"], "E-Doc")
        self.assertEqual(image_only_row["dataset_id"], production_row["dataset_id"])
        self.assertEqual(image_only_row["text_status"], "empty")
        self.assertEqual(image_only_row["content_type"], "Image")
        self.assertEqual(image_only_row["page_count"], 2)

        exact_search = retriever_tools.search(self.root, "PDX000001", None, None, None, 1, 20)
        self.assertEqual(exact_search["total_hits"], 1)
        self.assertEqual(exact_search["results"][0]["control_number"], "PDX000001")
        self.assertEqual(exact_search["results"][0]["attachment_count"], 1)
        self.assertEqual(exact_search["results"][0]["attachments"][0]["control_number"], "PDX000003")

        containing_search = retriever_tools.search(self.root, "PDX000002", None, None, None, 1, 20)
        self.assertEqual(containing_search["total_hits"], 1)
        self.assertEqual(containing_search["results"][0]["control_number"], "PDX000001")

        range_search = retriever_tools.search(self.root, "PDX000002-PDX000005", None, None, None, 1, 20)
        self.assertEqual(
            [item["control_number"] for item in range_search["results"]],
            ["PDX000001", "PDX000003", "PDX000004", "PDX000005"],
        )

        production_filtered = retriever_tools.search(
            self.root,
            "",
            [["source_kind", "eq", "production"], ["production_name", "contains", "Synthetic"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(production_filtered["total_hits"], 4)
        dataset_filtered = retriever_tools.search(
            self.root,
            "",
            [["dataset", "eq", "Synthetic_Production"]],
            None,
            None,
            1,
            20,
        )
        self.assertEqual(dataset_filtered["total_hits"], 4)
        self.assertTrue(all(item["dataset_name"] == "Synthetic_Production" for item in dataset_filtered["results"]))

        native_search = retriever_tools.search(self.root, "Native-backed production doc", None, None, None, 1, 20)
        native_result = next(item for item in native_search["results"] if item["control_number"] == "PDX000004")
        self.assertEqual(native_result["preview_rel_path"], "Synthetic_Production/NATIVES/NAT001/PDX000004.pdf")
        self.assertEqual(native_result["preview_targets"][0]["preview_type"], "native")
        self.assertEqual(native_result["production_name"], "Synthetic_Production")

        html_search = retriever_tools.search(self.root, "Discuss attachment handling", None, None, None, 1, 20)
        html_result = next(item for item in html_search["results"] if item["control_number"] == "PDX000001")
        self.assertTrue(html_result["preview_rel_path"].endswith(".html"))
        preview_html = Path(html_result["preview_targets"][0]["abs_path"]).read_text(encoding="utf-8")
        self.assertIn("<title>Attachment Handling</title>", preview_html)
        self.assertIn("<h1>Attachment Handling</h1>", preview_html)
        self.assertIn("<th>Control #</th>", preview_html)
        self.assertNotIn("<th>Control Number</th>", preview_html)
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            child_preview_target = retriever_tools.default_preview_target(self.paths, child_row, connection)
            child_preview_abs_path = str(
                child_preview_target.get("file_abs_path") or child_preview_target.get("abs_path") or ""
            ).split("#", 1)[0]
        finally:
            connection.close()
        expected_href = retriever_tools.urllib_request.pathname2url(
            os.path.relpath(
                child_preview_abs_path,
                start=str(Path(html_result["preview_targets"][0]["abs_path"]).parent),
            )
        )
        self.assertIn("<h2>Attachments</h2>", preview_html)
        self.assertIn(f'href="{expected_href}"', preview_html)
        self.assertIn("PDX000001", preview_html)
        self.assertIn("Discuss attachment handling.", preview_html)
        self.assertIn("data:image/png;base64,", preview_html)

    def test_ingest_production_falls_back_when_loadfile_paths_include_missing_volume_prefix(self) -> None:
        production_root = self.write_production_fixture(loadfile_volume_prefix="Sunrise_Production_01")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)

        self.assertEqual(ingest_result["created"], 4)
        self.assertEqual(ingest_result["docs_missing_linked_text"], 0)
        self.assertEqual(ingest_result["docs_missing_linked_images"], 1)
        self.assertEqual(ingest_result["docs_missing_linked_natives"], 0)
        self.assertEqual(ingest_result["page_images_linked"], 5)
        self.assertEqual(ingest_result["failures"], [])

        parent_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000001.logical"
        )
        native_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000004.logical"
        )
        self.assertEqual(parent_row["content_type"], "Email")
        self.assertEqual(parent_row["author"], "Elena Steven <elena@example.com>")
        self.assertEqual(parent_row["date_created"], "2026-04-14T10:32:00Z")
        self.assertEqual(parent_row["title"], "Attachment Handling")
        self.assertEqual(parent_row["text_status"], "ok")
        self.assertEqual(parent_row["page_count"], 2)
        self.assertEqual(native_row["file_name"], "PDX000004.pdf")

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            parent_source_parts = connection.execute(
                """
                SELECT part_kind, rel_source_path
                FROM document_source_parts
                WHERE document_id = ?
                ORDER BY part_kind ASC, ordinal ASC
                """,
                (parent_row["id"],),
            ).fetchall()
        finally:
            connection.close()

        self.assertEqual(
            [row["rel_source_path"] for row in parent_source_parts if row["part_kind"] == "text"],
            ["Synthetic_Production/TEXT/TEXT001/PDX000001.txt"],
        )
        self.assertEqual(
            [row["rel_source_path"] for row in parent_source_parts if row["part_kind"] == "image"],
            [
                "Synthetic_Production/IMAGES/IMG001/PDX000001.tif",
                "Synthetic_Production/IMAGES/IMG001/PDX000002.tif",
            ],
        )

    def test_ingest_production_rerun_retires_missing_rows(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        first_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(first_result["created"], 4)

        data_path = production_root / "DATA" / "Synthetic_Production.dat"
        lines = data_path.read_bytes().splitlines()
        filtered = [line for line in lines if b"PDX000005" not in line]
        data_path.write_bytes(b"\r\n".join(filtered) + b"\r\n")

        second_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(second_result["retired"], 1)

        retired_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000005.logical"
        )
        self.assertEqual(retired_row["lifecycle_status"], "deleted")

    def test_search_docs_cli_alias_matches_search(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("termination notice appears here\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        search_exit, search_payload, _, _ = self.run_cli("search", str(self.root), "termination")
        search_docs_exit, search_docs_payload, _, _ = self.run_cli("search-docs", str(self.root), "termination")

        self.assertEqual(search_exit, 0)
        self.assertEqual(search_docs_exit, 0)
        self.assertIsNotNone(search_payload)
        self.assertIsNotNone(search_docs_payload)
        self.assertEqual(search_docs_payload["query"], search_payload["query"])
        self.assertEqual(search_docs_payload["sort"], search_payload["sort"])
        self.assertEqual(search_docs_payload["total_hits"], search_payload["total_hits"])
        self.assertEqual(
            [item["id"] for item in search_docs_payload["results"]],
            [item["id"] for item in search_payload["results"]],
        )

    def test_search_cli_defaults_to_compact_payload_with_verbose_escape_hatch(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("termination notice appears here\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        compact_exit, compact_payload, _, _ = self.run_cli("search", str(self.root), "termination")
        verbose_exit, verbose_payload, _, _ = self.run_cli("search", str(self.root), "termination", "--verbose")

        self.assertEqual(compact_exit, 0)
        self.assertEqual(verbose_exit, 0)
        self.assertIsNotNone(compact_payload)
        self.assertIsNotNone(verbose_payload)

        compact_result = compact_payload["results"][0]
        verbose_result = verbose_payload["results"][0]

        self.assertNotIn("preview_targets", compact_result)
        self.assertNotIn("manual_field_locks", compact_result)
        self.assertIn("content_type", compact_result["metadata"])
        self.assertIn("updated_at", compact_result["metadata"])
        self.assertNotIn("page_count", compact_result["metadata"])
        self.assertIn("preview_abs_path", compact_result)
        self.assertNotIn("preview_rel_path", compact_result)
        self.assertIn("preview_targets", verbose_result)
        self.assertIn("manual_field_locks", verbose_result)
        self.assertIn("page_count", verbose_result["metadata"])

    def test_compact_search_keeps_attachment_counts_without_verbose_children(self) -> None:
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Upgrade test",
            body_text="Please review the attached notes.",
            attachment_name="notes.txt",
            attachment_text="confidential attachment detail",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        compact_exit, compact_payload, _, _ = self.run_cli("search", str(self.root), "Upgrade test")
        verbose_exit, verbose_payload, _, _ = self.run_cli("search", str(self.root), "Upgrade test", "--verbose")

        self.assertEqual(compact_exit, 0)
        self.assertEqual(verbose_exit, 0)
        self.assertIsNotNone(compact_payload)
        self.assertIsNotNone(verbose_payload)

        compact_result = compact_payload["results"][0]
        verbose_result = verbose_payload["results"][0]

        self.assertEqual(compact_result["attachment_count"], 1)
        self.assertNotIn("attachments", compact_result)
        self.assertNotIn("child_documents", compact_result)
        self.assertNotIn("preview_targets", compact_result)
        self.assertEqual(verbose_result["attachment_count"], 1)
        self.assertEqual(verbose_result["attachments"][0]["file_name"], "notes.txt")

    def test_catalog_lists_dataset_name_and_date_granularities(self) -> None:
        retriever_tools.bootstrap(self.root)
        retriever_tools.add_field(self.root, "effective_date", "date", "Contract effective date")

        exit_code, payload, _, _ = self.run_cli("catalog", str(self.root))

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        intrinsic = {item["name"]: item for item in payload["intrinsic"]}
        custom = {item["name"]: item for item in payload["custom"]}
        virtual = {item["name"]: item for item in payload["virtual"]}

        self.assertIn("date_created", intrinsic)
        self.assertEqual(intrinsic["date_created"]["date_granularities"], ["year", "quarter", "month", "week"])
        self.assertEqual(custom["effective_date"]["type"], "date")
        self.assertEqual(custom["effective_date"]["date_granularities"], ["year", "quarter", "month", "week"])
        self.assertTrue(virtual["dataset_name"]["aggregatable"])
        self.assertEqual(virtual["dataset_name"]["description"], retriever_tools.VIRTUAL_FIELD_DESCRIPTIONS["dataset_name"])

    def test_export_csv_cli_writes_requested_fields_for_filtered_collection(self) -> None:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("beta body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        alpha_row = self.fetch_document_row("alpha.txt")
        self.assertEqual(self.run_cli("add-field", str(self.root), "effective_date", "date")[0], 0)
        self.assertEqual(
            self.run_cli(
                "set-field",
                str(self.root),
                "--doc-id",
                str(alpha_row["id"]),
                "--field",
                "effective_date",
                "--value",
                "2026-04-16",
            )[0],
            0,
        )

        export_path = self.root / ".retriever" / "exports" / "review.csv"
        exit_code, payload, _, _ = self.run_cli(
            "export-csv",
            str(self.root),
            "review.csv",
            "--field",
            "dataset_name",
            "--field",
            "control_number",
            "--field",
            "effective_date",
            "--field",
            "file_name",
            "--filter",
            "file_name",
            "eq",
            "alpha.txt",
            "--sort",
            "file_name",
            "--order",
            "asc",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["document_count"], 1)
        self.assertEqual(payload["output_rel_path"], ".retriever/exports/review.csv")
        self.assertEqual(payload["selector"]["mode"], "search")
        self.assertEqual([field["field_name"] for field in payload["fields"]], ["dataset_name", "control_number", "effective_date", "file_name"])

        with export_path.open("r", encoding="utf-8", newline="") as handle:
            rows = list(csv.reader(handle))

        self.assertEqual(rows[0], ["dataset_name", "control_number", "effective_date", "file_name"])
        self.assertEqual(rows[1], [self.root.name, alpha_row["control_number"], "2026-04-16", "alpha.txt"])
        self.assertEqual(len(rows), 2)

    def test_export_csv_select_from_scope_exports_current_scope(self) -> None:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("beta body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        scope_exit, scope_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        self.assertEqual(scope_exit, 0)
        self.assertIsNotNone(scope_payload)

        export_path = self.root / ".retriever" / "exports" / "scope.csv"
        exit_code, payload, _, _ = self.run_cli(
            "export-csv",
            str(self.root),
            "scope.csv",
            "--field",
            "file_name",
            "--select-from-scope",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["document_count"], 1)
        self.assertEqual(payload["selector"]["mode"], "scope_search")
        self.assertTrue(payload["selector"]["selected_from_scope"])
        self.assertEqual(payload["selector"]["scope"]["keyword"], "alpha")

        with export_path.open("r", encoding="utf-8", newline="") as handle:
            rows = list(csv.reader(handle))

        self.assertEqual(rows[0], ["file_name"])
        self.assertEqual(rows[1], ["alpha.txt"])
        self.assertEqual(len(rows), 2)

    def test_export_csv_select_from_scope_and_narrows_with_explicit_filter(self) -> None:
        (self.root / "alpha-one.txt").write_text("alpha body one\n", encoding="utf-8")
        (self.root / "alpha-two.txt").write_text("alpha body two\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        scope_exit, scope_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        self.assertEqual(scope_exit, 0)
        self.assertIsNotNone(scope_payload)
        self.assertEqual(scope_payload["total_hits"], 2)

        export_path = self.root / ".retriever" / "exports" / "scope-narrowed.csv"
        exit_code, payload, _, _ = self.run_cli(
            "export-csv",
            str(self.root),
            "scope-narrowed.csv",
            "--field",
            "file_name",
            "--select-from-scope",
            "--filter",
            "file_name = 'alpha-two.txt'",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["document_count"], 1)
        self.assertEqual(payload["selector"]["mode"], "scope_search")
        self.assertEqual(payload["selector"]["scope"]["keyword"], "alpha")
        self.assertIn("alpha-two.txt", payload["selector"]["scope"]["filter"])

        with export_path.open("r", encoding="utf-8", newline="") as handle:
            rows = list(csv.reader(handle))

        self.assertEqual(rows[0], ["file_name"])
        self.assertEqual(rows[1], ["alpha-two.txt"])
        self.assertEqual(len(rows), 2)

    def test_export_csv_rejects_combining_doc_ids_with_scope_selector(self) -> None:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("alpha.txt")
        scope_exit, _, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        self.assertEqual(scope_exit, 0)

        exit_code, payload, _, _ = self.run_cli(
            "export-csv",
            str(self.root),
            "scope-doc-id.csv",
            "--field",
            "file_name",
            "--select-from-scope",
            "--doc-id",
            str(row["id"]),
        )

        self.assertEqual(exit_code, 2)
        self.assertIsNotNone(payload)
        self.assertIn("query/filter/scope selectors", payload["error"])

    def test_export_csv_cli_preserves_explicit_document_order(self) -> None:
        (self.root / "first.txt").write_text("first body\n", encoding="utf-8")
        (self.root / "second.txt").write_text("second body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        first_row = self.fetch_document_row("first.txt")
        second_row = self.fetch_document_row("second.txt")

        export_path = self.root / ".retriever" / "exports" / "ordered.csv"
        exit_code, payload, _, _ = self.run_cli(
            "export-csv",
            str(self.root),
            "ordered.csv",
            "--field",
            "file_name",
            "--field",
            "control_number",
            "--doc-id",
            str(second_row["id"]),
            "--doc-id",
            str(first_row["id"]),
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["document_count"], 2)
        self.assertEqual(payload["selector"]["mode"], "document_ids")
        self.assertEqual(payload["selector"]["document_ids"], [second_row["id"], first_row["id"]])

        with export_path.open("r", encoding="utf-8", newline="") as handle:
            rows = list(csv.reader(handle))

        self.assertEqual(rows[0], ["file_name", "control_number"])
        self.assertEqual(rows[1], ["second.txt", second_row["control_number"]])
        self.assertEqual(rows[2], ["first.txt", first_row["control_number"]])

    def test_export_csv_relative_output_does_not_get_reingested(self) -> None:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("beta body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        first_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(first_ingest["new"], 2)

        exit_code, payload, _, _ = self.run_cli(
            "export-csv",
            str(self.root),
            "nested/review.csv",
            "--field",
            "file_name",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["output_rel_path"], ".retriever/exports/nested/review.csv")

        second_ingest = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(second_ingest["new"], 0)
        self.assertEqual(second_ingest["skipped"], 2)

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_result["total_hits"], 2)
        self.assertEqual(sorted(item["file_name"] for item in browse_result["results"]), ["alpha.txt", "beta.txt"])

    def test_export_csv_rejects_workspace_output_path_outside_retriever_state_dir(self) -> None:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        unsafe_path = self.root / "exports" / "review.csv"
        exit_code, payload, _, _ = self.run_cli(
            "export-csv",
            str(self.root),
            str(unsafe_path),
            "--field",
            "file_name",
        )

        self.assertEqual(exit_code, 2)
        self.assertIsNotNone(payload)
        self.assertIn("must live under", payload["error"])

    def test_export_csv_doc_id_mode_rejects_rows_without_dataset_membership(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("sample dataset body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        delete_exit, delete_payload, _, _ = self.run_cli(
            "delete-dataset",
            str(self.root),
            "--dataset-name",
            self.root.name,
        )
        self.assertEqual(delete_exit, 0)
        self.assertIsNotNone(delete_payload)
        self.assertEqual(delete_payload["documents_without_dataset_memberships"], [row["id"]])

        browse_result = retriever_tools.search(self.root, "", None, None, None, 1, 20)
        self.assertEqual(browse_result["total_hits"], 0)

        exit_code, payload, _, _ = self.run_cli(
            "export-csv",
            str(self.root),
            "doc-id.csv",
            "--field",
            "file_name",
            "--doc-id",
            str(row["id"]),
        )

        self.assertEqual(exit_code, 2)
        self.assertIsNotNone(payload)
        self.assertIn("not visible because they have no dataset memberships", payload["error"])

    def test_export_previews_expands_reply_selection_to_full_email_conversation(self) -> None:
        root_path = self.root / "root.eml"
        reply_path = self.root / "reply.eml"
        self.write_email_message(
            root_path,
            subject="Status Update",
            body_text="Root message body",
            message_id="<root@example.com>",
            date_created="Tue, 14 Apr 2026 10:00:00 +0000",
        )
        self.write_email_message(
            reply_path,
            subject="Re: Status Update",
            body_text="Reply message body",
            message_id="<reply@example.com>",
            in_reply_to="<root@example.com>",
            references="<root@example.com>",
            date_created="Tue, 14 Apr 2026 11:00:00 +0000",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        root_row = self.fetch_document_row("root.eml")
        reply_row = self.fetch_document_row("reply.eml")

        exit_code, payload, _, _ = self.run_cli(
            "export-previews",
            str(self.root),
            "email-preview",
            "--doc-id",
            str(reply_row["id"]),
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["selected_document_count"], 1)
        self.assertEqual(payload["unit_count"], 1)
        self.assertEqual(payload["output_rel_path"], ".retriever/exports/email-preview")
        self.assertEqual(len(payload["document_targets"]), 1)
        self.assertEqual(payload["document_targets"][0]["document_id"], reply_row["id"])
        self.assertEqual(payload["document_targets"][0]["target_fragment"], f"doc-{reply_row['id']}")
        self.assertEqual(payload["document_targets"][0]["unit_output_rel_path"], payload["units"][0]["output_rel_path"])

        unit = payload["units"][0]
        self.assertEqual(unit["unit_kind"], "email_conversation")
        self.assertEqual(unit["selected_document_ids"], [reply_row["id"]])
        self.assertEqual(unit["document_ids"], [root_row["id"], reply_row["id"]])
        self.assertNotEqual(unit["output_rel_path"], payload["document_targets"][0]["output_rel_path"])

        unit_path = Path(unit["output_path"])
        document_path = Path(payload["document_targets"][0]["output_path"])
        index_path = Path(payload["index_path"])
        manifest_path = Path(payload["manifest_path"])
        self.assertTrue(unit_path.exists())
        self.assertTrue(document_path.exists())
        self.assertTrue(index_path.exists())
        self.assertTrue(manifest_path.exists())

        unit_html = unit_path.read_text(encoding="utf-8")
        document_html = document_path.read_text(encoding="utf-8")
        index_html = index_path.read_text(encoding="utf-8")
        self.assertIn("<th>Conversation Type</th>", unit_html)
        self.assertNotIn("<th>Conversation type</th>", unit_html)
        self.assertIn("Root message body", unit_html)
        self.assertIn("Reply message body", unit_html)
        self.assertIn(f'id="doc-{root_row["id"]}"', unit_html)
        self.assertIn(f'id="doc-{reply_row["id"]}"', unit_html)
        self.assertIn("Root message body", document_html)
        self.assertIn("Reply message body", document_html)
        self.assertEqual(document_html.count('class="gmail-message-card'), 2)
        self.assertIn('class="gmail-thread-title-link"', document_html)
        self.assertIn(f'href="../{unit["output_rel_path"]}"', document_html)
        self.assertIn("units/", index_html)
        self.assertIn(payload["document_targets"][0]["output_rel_path"], index_html)
        self.assertNotIn(f"#doc-{reply_row['id']}", index_html)

        units_dir = unit_path.parent
        documents_dir = document_path.parent
        self.assertEqual(sorted(path.name for path in units_dir.glob("*.html")), [unit_path.name])
        self.assertEqual(sorted(path.name for path in documents_dir.glob("*.html")), [document_path.name])

    def test_export_previews_merges_contiguous_slack_documents_and_splits_gaps(self) -> None:
        export_root = self.root / "data" / "slack"
        export_root.mkdir(parents=True)
        (export_root / "users.json").write_text(
            json.dumps(
                [
                    {
                        "id": "U04SERGEY1",
                        "name": "sergey",
                        "profile": {
                            "real_name": "Sergey Demyanov",
                            "display_name": "Sergey",
                        },
                    },
                    {
                        "id": "U04MAX0001",
                        "name": "maksim",
                        "profile": {
                            "real_name": "Maksim Faleev",
                            "display_name": "Maksim",
                        },
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (export_root / "channels.json").write_text(
            json.dumps(
                [
                    {
                        "id": "C04GENERAL1",
                        "name": "general",
                        "is_channel": True,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        channel_dir = export_root / "general"
        channel_dir.mkdir()
        thread_ts = "1671235434.237949"
        (channel_dir / "2022-12-16.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Kickoff thread",
                        "user": "U04SERGEY1",
                        "ts": thread_ts,
                        "thread_ts": thread_ts,
                        "reply_count": 1,
                    },
                    {
                        "type": "message",
                        "text": "Standalone channel update",
                        "user": "U04MAX0001",
                        "ts": "1671235834.237949",
                    },
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        (channel_dir / "2022-12-17.json").write_text(
            json.dumps(
                [
                    {
                        "type": "message",
                        "text": "Following up on kickoff",
                        "user": "U04MAX0001",
                        "ts": "1671321834.237949",
                        "thread_ts": thread_ts,
                    }
                ],
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["failed"], 0)

        day_one_row = self.fetch_document_row("data/slack/general/2022-12-16.json")
        day_two_row = self.fetch_document_row("data/slack/general/2022-12-17.json")
        child_row = self.fetch_document_row(
            retriever_tools.slack_reply_thread_rel_path("C04GENERAL1", thread_ts)
        )

        contiguous_exit, contiguous_payload, _, _ = self.run_cli(
            "export-previews",
            str(self.root),
            "slack-contiguous",
            "--doc-id",
            str(day_one_row["id"]),
            "--doc-id",
            str(child_row["id"]),
        )

        self.assertEqual(contiguous_exit, 0)
        self.assertIsNotNone(contiguous_payload)
        self.assertEqual(contiguous_payload["unit_count"], 1)
        self.assertEqual(len(contiguous_payload["document_targets"]), 2)
        contiguous_unit = contiguous_payload["units"][0]
        self.assertEqual(contiguous_unit["unit_kind"], "conversation_run")
        self.assertEqual(contiguous_unit["selected_document_ids"], [day_one_row["id"], child_row["id"]])
        self.assertEqual(contiguous_unit["document_ids"], [day_one_row["id"], child_row["id"]])
        self.assertEqual(
            {
                target["output_rel_path"]
                for target in contiguous_payload["document_targets"]
            },
            {contiguous_unit["output_rel_path"]},
        )
        contiguous_html = Path(contiguous_unit["output_path"]).read_text(encoding="utf-8")
        self.assertIn("Kickoff thread", contiguous_html)
        self.assertIn("Following up on kickoff", contiguous_html)
        self.assertIn(f'id="doc-{day_one_row["id"]}"', contiguous_html)
        self.assertIn(f'id="doc-{child_row["id"]}"', contiguous_html)
        self.assertNotIn(f'id="doc-{day_two_row["id"]}"', contiguous_html)

        split_exit, split_payload, _, _ = self.run_cli(
            "export-previews",
            str(self.root),
            "slack-split",
            "--doc-id",
            str(day_two_row["id"]),
            "--doc-id",
            str(child_row["id"]),
        )

        self.assertEqual(split_exit, 0)
        self.assertIsNotNone(split_payload)
        self.assertEqual(split_payload["unit_count"], 2)
        self.assertEqual(
            [unit["selected_document_ids"] for unit in split_payload["units"]],
            [[day_two_row["id"]], [child_row["id"]]],
        )
        self.assertEqual(
            [unit["document_ids"] for unit in split_payload["units"]],
            [[day_two_row["id"]], [child_row["id"]]],
        )
        self.assertEqual(
            len({target["output_rel_path"] for target in split_payload["document_targets"]}),
            2,
        )
        first_split_html = Path(split_payload["units"][0]["output_path"]).read_text(encoding="utf-8")
        second_split_html = Path(split_payload["units"][1]["output_path"]).read_text(encoding="utf-8")
        self.assertIn(f'id="doc-{day_two_row["id"]}"', first_split_html)
        self.assertNotIn(f'id="doc-{day_one_row["id"]}"', first_split_html)
        self.assertIn(f'id="doc-{child_row["id"]}"', second_split_html)
        self.assertNotIn(f'id="doc-{day_one_row["id"]}"', second_split_html)

    def test_export_archive_cli_includes_previews_and_attachment_family(self) -> None:
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Archive export",
            body_text="Parent email body text.",
            attachment_name="notes.txt",
            attachment_text="confidential attachment detail",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        parent_row = self.fetch_document_row("thread.eml")
        child_row = self.fetch_child_rows(parent_row["id"])[0]
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            parent_preview_rows = connection.execute(
                """
                SELECT rel_preview_path
                FROM document_previews
                WHERE document_id = ?
                ORDER BY ordinal ASC, id ASC
                """,
                (parent_row["id"],),
            ).fetchall()
        finally:
            connection.close()
        self.assertGreaterEqual(len(parent_preview_rows), 1)
        parent_preview_archive_path = str(Path(".retriever") / str(parent_preview_rows[0]["rel_preview_path"]))

        export_path = self.root / ".retriever" / "exports" / "family.zip"
        exit_code, payload, _, _ = self.run_cli(
            "export-archive",
            str(self.root),
            "family.zip",
            "--keyword",
            "confidential",
            "--family-mode",
            "with_family",
            "--limit",
            "1",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["document_count"], 2)
        self.assertEqual(payload["output_rel_path"], ".retriever/exports/family.zip")
        self.assertEqual(payload["family_mode"], "with_family")
        self.assertFalse(payload["portable_workspace"])
        self.assertEqual(payload["selector"]["keyword"], "confidential")

        with zipfile.ZipFile(export_path, "r") as archive:
            names = set(archive.namelist())
            manifest = json.loads(archive.read(".retriever/export-manifest.json").decode("utf-8"))

        self.assertIn("thread.eml", names)
        self.assertIn(child_row["rel_path"], names)
        self.assertIn(parent_preview_archive_path, names)
        self.assertEqual(manifest["document_count"], 2)
        self.assertEqual(manifest["family_mode"], "with_family")
        self.assertEqual(manifest["warnings"], [])

        manifest_by_document_id = {
            int(item["document_id"]): item
            for item in manifest["documents"]
        }
        self.assertEqual(set(manifest_by_document_id), {int(parent_row["id"]), int(child_row["id"])})
        self.assertEqual(
            manifest_by_document_id[int(child_row["id"])]["inclusion_reason"]["direct_reasons"][0]["type"],
            "keyword",
        )
        self.assertEqual(
            manifest_by_document_id[int(parent_row["id"])]["inclusion_reason"]["family_seed_document_ids"],
            [int(child_row["id"])],
        )

    def test_export_archive_select_from_scope_exports_current_scope(self) -> None:
        (self.root / "alpha.txt").write_text("alpha body\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("beta body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        scope_exit, scope_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        self.assertEqual(scope_exit, 0)
        self.assertIsNotNone(scope_payload)

        export_path = self.root / ".retriever" / "exports" / "scope.zip"
        exit_code, payload, _, _ = self.run_cli(
            "export-archive",
            str(self.root),
            "scope.zip",
            "--select-from-scope",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["document_count"], 1)
        self.assertEqual(payload["selector"]["keyword"], "alpha")

        with zipfile.ZipFile(export_path, "r") as archive:
            names = set(archive.namelist())
            manifest = json.loads(archive.read(".retriever/export-manifest.json").decode("utf-8"))

        self.assertIn("alpha.txt", names)
        self.assertNotIn("beta.txt", names)
        self.assertEqual(manifest["document_count"], 1)
        self.assertEqual(manifest["selector"]["keyword"], "alpha")

    def test_export_archive_portable_workspace_supports_selected_child_with_parent_context_stub(self) -> None:
        email_path = self.root / "thread.eml"
        self.write_email_message(
            email_path,
            subject="Portable archive",
            body_text="Parent email body text.",
            attachment_name="notes.txt",
            attachment_text="confidential attachment detail",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        parent_row = self.fetch_document_row("thread.eml")
        child_row = self.fetch_child_rows(parent_row["id"])[0]

        export_path = self.root / ".retriever" / "exports" / "portable-child.zip"
        exit_code, payload, _, _ = self.run_cli(
            "export-archive",
            str(self.root),
            "portable-child.zip",
            "--filter",
            f"id = {child_row['id']}",
            "--portable-workspace",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["document_count"], 1)
        self.assertTrue(payload["portable_workspace"])
        self.assertEqual(payload["selector"]["filter"], f"(id = {child_row['id']})")

        with zipfile.ZipFile(export_path, "r") as archive:
            names = set(archive.namelist())
            manifest = json.loads(archive.read(".retriever/export-manifest.json").decode("utf-8"))
            self.assertIn(".retriever/retriever.db", names)
            self.assertIn(child_row["rel_path"], names)

            with tempfile.TemporaryDirectory(prefix="retriever-portable-archive-") as extract_dir:
                extracted_root = Path(extract_dir) / "workspace"
                archive.extractall(extracted_root)

                portable_search = retriever_tools.search(extracted_root, "confidential", None, None, None, 1, 20)
                self.assertEqual(portable_search["total_hits"], 1)
                self.assertEqual(portable_search["results"][0]["id"], child_row["id"])
                self.assertEqual(portable_search["results"][0]["parent"]["id"], parent_row["id"])
                self.assertEqual(
                    portable_search["results"][0]["parent"]["control_number"],
                    parent_row["control_number"],
                )

                parent_search = retriever_tools.search(extracted_root, "Portable archive", None, None, None, 1, 20)
                self.assertEqual(parent_search["total_hits"], 0)

                portable_connection = retriever_tools.connect_db(extracted_root / ".retriever" / "retriever.db")
                try:
                    parent_stub_row = portable_connection.execute(
                        """
                        SELECT source_text_revision_id, active_search_text_revision_id
                        FROM documents
                        WHERE id = ?
                        """,
                        (parent_row["id"],),
                    ).fetchone()
                    self.assertIsNotNone(parent_stub_row)
                    self.assertIsNone(parent_stub_row["source_text_revision_id"])
                    self.assertIsNone(parent_stub_row["active_search_text_revision_id"])

                    dataset_document_count = portable_connection.execute(
                        "SELECT COUNT(*) AS count FROM dataset_documents"
                    ).fetchone()["count"]
                    self.assertEqual(dataset_document_count, 1)
                finally:
                    portable_connection.close()

        self.assertEqual(manifest["portable_workspace_document_ids"], [child_row["id"]])
        self.assertEqual(manifest["portable_workspace_stub_document_ids"], [parent_row["id"]])

    def test_export_archive_includes_production_source_parts_and_synthetic_logical_entry(self) -> None:
        production_root = self.write_production_fixture()

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest_production(self.root, production_root)
        self.assertEqual(ingest_result["created"], 4)

        native_row = self.fetch_document_row(
            f"{retriever_tools.INTERNAL_REL_PATH_PREFIX}/productions/Synthetic_Production/documents/PDX000004.logical"
        )
        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            source_part_rows = connection.execute(
                """
                SELECT part_kind, rel_source_path
                FROM document_source_parts
                WHERE document_id = ?
                ORDER BY part_kind ASC, ordinal ASC, id ASC
                """,
                (native_row["id"],),
            ).fetchall()
        finally:
            connection.close()
        self.assertTrue(any(row["part_kind"] == "native" for row in source_part_rows))

        export_path = self.root / ".retriever" / "exports" / "production.zip"
        exit_code, payload, _, _ = self.run_cli(
            "export-archive",
            str(self.root),
            "production.zip",
            "--filter",
            f"id = {native_row['id']}",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["document_count"], 1)
        self.assertEqual(payload["warnings"], [])
        self.assertEqual(payload["selector"]["filter"], f"(id = {native_row['id']})")

        with zipfile.ZipFile(export_path, "r") as archive:
            names = set(archive.namelist())
            manifest = json.loads(archive.read(".retriever/export-manifest.json").decode("utf-8"))
            descriptor_payload = json.loads(archive.read(native_row["rel_path"]).decode("utf-8"))

        self.assertIn(native_row["rel_path"], names)
        self.assertTrue(all(row["rel_source_path"] in names for row in source_part_rows))
        self.assertEqual(manifest["warnings"], [])
        self.assertEqual(manifest["documents"][0]["document_entry_kind"], "synthetic")
        self.assertTrue(any(part["part_kind"] == "native" for part in manifest["documents"][0]["source_part_entries"]))
        self.assertEqual(descriptor_payload["document_id"], native_row["id"])
        self.assertEqual(descriptor_payload["control_number"], native_row["control_number"])

    def test_get_doc_and_list_chunks_return_summary_and_exact_chunk_text(self) -> None:
        paragraph = "Termination notice requires careful review and supporting detail. "
        body = "\n".join(f"Section {index}: {paragraph * 70}" for index in range(1, 8)) + "\n"
        document_path = self.root / "long.txt"
        document_path.write_text(body, encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("long.txt")
        expected_chunks = retriever_tools.chunk_text(body)
        self.assertGreater(len(expected_chunks), 2)

        exit_code, get_payload, _, _ = self.run_cli(
            "get-doc",
            str(self.root),
            "--doc-id",
            str(row["id"]),
            "--include-text",
            "summary",
            "--chunk",
            "0",
            "--chunk",
            "1",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(get_payload)
        self.assertEqual(get_payload["chunk_count"], len(expected_chunks))
        self.assertEqual(
            get_payload["text_summary"],
            retriever_tools.normalize_whitespace(body)[: retriever_tools.GET_DOC_SUMMARY_CHARS],
        )
        self.assertEqual(get_payload["chunks"][0]["chunk_index"], 0)
        self.assertEqual(get_payload["chunks"][0]["text"], expected_chunks[0]["text_content"])
        self.assertEqual(get_payload["chunks"][1]["chunk_index"], 1)
        self.assertEqual(get_payload["chunks"][1]["text"], expected_chunks[1]["text_content"])

        exit_code, list_payload, _, _ = self.run_cli(
            "list-chunks",
            str(self.root),
            "--doc-id",
            str(row["id"]),
            "--page",
            "1",
            "--per-page",
            "1",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(list_payload)
        self.assertEqual(list_payload["total_chunks"], len(expected_chunks))
        self.assertEqual(list_payload["total_pages"], len(expected_chunks))
        self.assertEqual(list_payload["chunks"][0]["chunk_index"], 0)
        self.assertEqual(
            list_payload["chunks"][0]["snippet"],
            retriever_tools.chunk_preview_text(expected_chunks[0]["text_content"]),
        )

    def test_get_doc_and_search_chunks_cli_default_to_compact_payloads(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("Termination notice appears here.\nSupporting detail follows.\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")

        get_exit, get_payload, _, _ = self.run_cli("get-doc", str(self.root), "--doc-id", str(row["id"]))
        get_verbose_exit, get_verbose_payload, _, _ = self.run_cli(
            "get-doc",
            str(self.root),
            "--doc-id",
            str(row["id"]),
            "--verbose",
        )

        self.assertEqual(get_exit, 0)
        self.assertEqual(get_verbose_exit, 0)
        self.assertIsNotNone(get_payload)
        self.assertIsNotNone(get_verbose_payload)
        self.assertNotIn("preview_targets", get_payload["document"])
        self.assertNotIn("manual_field_locks", get_payload["document"])
        self.assertIn("preview_targets", get_verbose_payload["document"])
        self.assertIn("manual_field_locks", get_verbose_payload["document"])

        chunk_exit, chunk_payload, _, _ = self.run_cli("search-chunks", str(self.root), "termination")
        chunk_verbose_exit, chunk_verbose_payload, _, _ = self.run_cli(
            "search-chunks",
            str(self.root),
            "termination",
            "--verbose",
        )

        self.assertEqual(chunk_exit, 0)
        self.assertEqual(chunk_verbose_exit, 0)
        self.assertIsNotNone(chunk_payload)
        self.assertIsNotNone(chunk_verbose_payload)

        compact_result = chunk_payload["results"][0]
        verbose_result = chunk_verbose_payload["results"][0]
        self.assertNotIn("text", compact_result)
        self.assertNotIn("preview_targets", compact_result)
        self.assertIn("citation", compact_result)
        self.assertIn("text", verbose_result)
        self.assertIn("preview_targets", verbose_result)

    def test_search_chunks_supports_citations_and_distinct_doc_count_mode(self) -> None:
        (self.root / "nda-one.txt").write_text("Termination notice must be delivered within thirty days.\n", encoding="utf-8")
        (self.root / "nda-two.txt").write_text("The agreement has a termination notice period of sixty days.\n", encoding="utf-8")
        (self.root / "other.txt").write_text("Unrelated payment processing clause.\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 3)

        exit_code, payload, _, _ = self.run_cli(
            "search-chunks",
            str(self.root),
            "termination",
            "--top-k",
            "5",
            "--per-doc-cap",
            "1",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(len(payload["results"]), 2)
        for item in payload["results"]:
            self.assertIn("citation", item)
            self.assertEqual(item["citation"]["document_id"], item["document_id"])
            self.assertEqual(item["citation"]["chunk_index"], item["chunk_index"])
            self.assertTrue(item["citation"]["snippet"])

        exit_code, count_payload, _, _ = self.run_cli(
            "search-chunks",
            str(self.root),
            "termination",
            "--count-only",
            "--distinct-docs",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(count_payload)
        self.assertEqual(count_payload["documents_with_hits"], 2)
        self.assertEqual(count_payload["total_docs_filtered"], 3)
        self.assertEqual(count_payload["count_mode"], "distinct-documents")

    def test_search_chunks_select_from_scope_narrows_to_scoped_documents(self) -> None:
        (self.root / "alpha-scope.txt").write_text("Alpha matter termination notice.\n", encoding="utf-8")
        (self.root / "beta-scope.txt").write_text("Beta project termination notice.\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        scope_exit, scope_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        self.assertEqual(scope_exit, 0)
        self.assertIsNotNone(scope_payload)
        self.assertEqual(scope_payload["total_hits"], 1)

        exit_code, payload, _, _ = self.run_cli(
            "search-chunks",
            str(self.root),
            "termination",
            "--select-from-scope",
            "--top-k",
            "10",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertTrue(payload["selected_from_scope"])
        self.assertEqual(payload["scope"]["keyword"], "alpha")
        self.assertEqual([item["file_name"] for item in payload["results"]], ["alpha-scope.txt"])

    def test_list_fields_and_describe_field_report_custom_field_metadata(self) -> None:
        (self.root / "sample.txt").write_text("field metadata body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        add_exit, add_payload, _, _ = self.run_cli(
            "add-field",
            str(self.root),
            "review_status",
            "text",
            "--instruction",
            "Initial review label",
        )
        self.assertEqual(add_exit, 0)
        self.assertIsNotNone(add_payload)

        fill_exit, fill_payload, _, _ = self.run_cli(
            "fill-field",
            str(self.root),
            "--field",
            "review_status",
            "--value",
            "hot",
            "--doc-id",
            str(row["id"]),
        )
        self.assertEqual(fill_exit, 0)
        self.assertIsNotNone(fill_payload)
        self.assertEqual(fill_payload["written"], 1)

        list_exit, list_payload, _, _ = self.run_cli("list-fields", str(self.root))
        self.assertEqual(list_exit, 0)
        self.assertIsNotNone(list_payload)
        review_status_entry = {
            item["field_name"]: item for item in list_payload["fields"]
        }["review_status"]
        self.assertEqual(review_status_entry["field_type"], "text")
        self.assertEqual(review_status_entry["documents_with_values"], 1)
        self.assertEqual(review_status_entry["instruction"], "Initial review label")

        describe_exit, describe_payload, _, _ = self.run_cli(
            "describe-field",
            str(self.root),
            "review_status",
            "--text",
            "Normalized review label",
        )
        self.assertEqual(describe_exit, 0)
        self.assertIsNotNone(describe_payload)
        self.assertEqual(describe_payload["instruction"], "Normalized review label")
        self.assertEqual(
            self.fetch_custom_field_registry_row("review_status")["instruction"],
            "Normalized review label",
        )

        table_exit, table_stdout, table_stderr = self.run_cli_raw(
            "list-fields",
            str(self.root),
            "--format",
            "table",
        )
        self.assertEqual(table_exit, 0)
        self.assertEqual(table_stderr, "")
        self.assertIn("review_status | text | 1 | Normalized review label", table_stdout)

    def test_fill_field_bulk_scope_requires_confirm_and_locks_documents(self) -> None:
        (self.root / "alpha-one.txt").write_text("alpha body one\n", encoding="utf-8")
        (self.root / "alpha-two.txt").write_text("alpha body two\n", encoding="utf-8")
        (self.root / "beta.txt").write_text("beta body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 3)

        self.assertEqual(self.run_cli("add-field", str(self.root), "review_status", "text")[0], 0)
        scope_payload = retriever_tools.run_slash_command(self.root, "/search alpha")
        self.assertEqual(scope_payload["total_hits"], 2)

        preview_payload = retriever_tools.run_slash_command(self.root, "/fill review_status responsive")
        self.assertEqual(preview_payload["status"], "confirm_required")
        self.assertEqual(preview_payload["would_write"], 2)

        apply_exit, apply_payload, _, _ = self.run_cli(
            "slash",
            str(self.root),
            "/fill",
            "review_status",
            "responsive",
            "--confirm",
        )
        self.assertEqual(apply_exit, 0)
        self.assertIsNotNone(apply_payload)
        self.assertEqual(apply_payload["written"], 2)
        self.assertEqual(apply_payload["skipped"], 0)

        for rel_path in ("alpha-one.txt", "alpha-two.txt"):
            row = self.fetch_document_row(rel_path)
            self.assertEqual(row["review_status"], "responsive")
            self.assertIn("review_status", retriever_tools.normalize_string_list(row["manual_field_locks_json"]))
        self.assertIsNone(self.fetch_document_row("beta.txt").get("review_status"))

    def test_slash_fill_accepts_control_number_for_single_document(self) -> None:
        (self.root / "single.txt").write_text("single fill body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("single.txt")
        self.assertEqual(self.run_cli("add-field", str(self.root), "review_status", "text")[0], 0)

        payload = retriever_tools.run_slash_command(
            self.root,
            f"/fill review_status hot on {row['control_number']}",
        )
        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["written"], 1)
        self.assertEqual(payload["document_ids"], [row["id"]])

        updated_row = self.fetch_document_row("single.txt")
        self.assertEqual(updated_row["review_status"], "hot")
        self.assertIn("review_status", retriever_tools.normalize_string_list(updated_row["manual_field_locks_json"]))

    def test_rename_field_updates_session_and_saved_scope_refs(self) -> None:
        (self.root / "sample.txt").write_text("rename field body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        self.assertEqual(self.run_cli("add-field", str(self.root), "review_score", "integer")[0], 0)
        self.assertEqual(
            self.run_cli(
                "fill-field",
                str(self.root),
                "--field",
                "review_score",
                "--value",
                "7",
                "--doc-id",
                str(row["id"]),
            )[0],
            0,
        )
        self.assertEqual(self.run_cli("slash", str(self.root), "/columns set title, review_score, control_number")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/filter review_score >= 5")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/sort review_score asc")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/scope save review")[0], 0)

        rename_exit, rename_payload, _, _ = self.run_cli(
            "rename-field",
            str(self.root),
            "review_score",
            "responsiveness_score",
        )
        self.assertEqual(rename_exit, 0)
        self.assertIsNotNone(rename_payload)
        self.assertEqual(rename_payload["status"], "ok")
        self.assertEqual(rename_payload["field_name"], "responsiveness_score")

        updated_row = self.fetch_document_row("sample.txt")
        self.assertEqual(updated_row["responsiveness_score"], 7)
        self.assertNotIn("review_score", updated_row)

        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(
            session_payload["display"]["documents"]["columns"],
            ["title", "responsiveness_score", "control_number"],
        )
        self.assertEqual(
            session_payload["browsing"]["documents"]["sort"],
            [["responsiveness_score", "asc"]],
        )
        self.assertIn("responsiveness_score", session_payload["scope"]["filter"])
        self.assertNotIn("review_score", session_payload["scope"]["filter"])

        saved_scopes_payload = json.loads(self.paths["saved_scopes_path"].read_text(encoding="utf-8"))
        self.assertIn("responsiveness_score", saved_scopes_payload["scopes"]["review"]["filter"])
        self.assertNotIn("review_score", saved_scopes_payload["scopes"]["review"]["filter"])

        load_exit, load_payload, _, _ = self.run_cli("slash", str(self.root), "/scope", "load", "review")
        self.assertEqual(load_exit, 0)
        self.assertIsNotNone(load_payload)
        self.assertIn("responsiveness_score", load_payload["scope"]["filter"])
        self.assertEqual(load_payload["total_hits"], 1)

    def test_delete_field_blocks_on_scope_filters_and_scrubs_session_state(self) -> None:
        (self.root / "sample.txt").write_text("delete field body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        self.assertEqual(self.run_cli("add-field", str(self.root), "review_score", "integer")[0], 0)
        self.assertEqual(
            self.run_cli(
                "fill-field",
                str(self.root),
                "--field",
                "review_score",
                "--value",
                "7",
                "--doc-id",
                str(row["id"]),
            )[0],
            0,
        )
        self.assertEqual(self.run_cli("slash", str(self.root), "/columns set title, review_score, control_number")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/sort review_score asc")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/filter review_score >= 5")[0], 0)
        self.assertEqual(self.run_cli("slash", str(self.root), "/scope save review")[0], 0)

        blocked_exit, blocked_payload, _, _ = self.run_cli(
            "delete-field",
            str(self.root),
            "review_score",
            "--confirm",
        )
        self.assertEqual(blocked_exit, 0)
        self.assertIsNotNone(blocked_payload)
        self.assertEqual(blocked_payload["status"], "blocked")
        self.assertEqual(len(blocked_payload["blockers"]), 2)

        self.assertEqual(self.run_cli("slash", str(self.root), "/filter clear")[0], 0)
        saved_scopes_payload = json.loads(self.paths["saved_scopes_path"].read_text(encoding="utf-8"))
        saved_scopes_payload["scopes"]["review"].pop("filter", None)
        self.paths["saved_scopes_path"].write_text(
            json.dumps(saved_scopes_payload, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )

        preview_exit, preview_payload, _, _ = self.run_cli("delete-field", str(self.root), "review_score")
        self.assertEqual(preview_exit, 0)
        self.assertIsNotNone(preview_payload)
        self.assertEqual(preview_payload["status"], "confirm_required")

        delete_exit, delete_payload, _, _ = self.run_cli(
            "delete-field",
            str(self.root),
            "review_score",
            "--confirm",
        )
        self.assertEqual(delete_exit, 0)
        self.assertIsNotNone(delete_payload)
        self.assertEqual(delete_payload["status"], "ok")
        self.assertEqual(delete_payload["deleted"], "review_score")

        session_payload = json.loads(self.paths["session_path"].read_text(encoding="utf-8"))
        self.assertEqual(session_payload["display"]["documents"]["columns"], ["title", "control_number"])
        self.assertNotIn("sort", session_payload["browsing"]["documents"])

        connection = retriever_tools.connect_db(self.paths["db_path"])
        try:
            columns = {
                row["name"]
                for row in connection.execute("PRAGMA table_info(documents)")
            }
        finally:
            connection.close()
        self.assertNotIn("review_score", columns)
        list_fields_payload = retriever_tools.list_fields(self.root)
        self.assertEqual(list_fields_payload["fields"], [])

    def test_change_field_type_supports_integer_to_boolean(self) -> None:
        (self.root / "flag-a.txt").write_text("flag a\n", encoding="utf-8")
        (self.root / "flag-b.txt").write_text("flag b\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        first_row = self.fetch_document_row("flag-a.txt")
        second_row = self.fetch_document_row("flag-b.txt")
        self.assertEqual(self.run_cli("add-field", str(self.root), "review_flag", "integer")[0], 0)
        self.assertEqual(
            self.run_cli(
                "fill-field",
                str(self.root),
                "--field",
                "review_flag",
                "--value",
                "1",
                "--doc-id",
                str(first_row["id"]),
            )[0],
            0,
        )
        self.assertEqual(
            self.run_cli(
                "fill-field",
                str(self.root),
                "--field",
                "review_flag",
                "--value",
                "0",
                "--doc-id",
                str(second_row["id"]),
            )[0],
            0,
        )

        exit_code, payload, _, _ = self.run_cli(
            "change-field-type",
            str(self.root),
            "review_flag",
            "boolean",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["status"], "ok")
        self.assertTrue(payload["conversion_applied"])
        self.assertEqual(self.fetch_custom_field_registry_row("review_flag")["field_type"], "boolean")
        self.assertEqual(self.fetch_document_row("flag-a.txt")["review_flag"], 1)
        self.assertEqual(self.fetch_document_row("flag-b.txt")["review_flag"], 0)

    def test_set_field_validates_custom_date_fields(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("hello\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")

        exit_code, payload, _, _ = self.run_cli("add-field", str(self.root), "effective_date", "date")
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)

        exit_code, payload, _, _ = self.run_cli(
            "set-field",
            str(self.root),
            "--doc-id",
            str(row["id"]),
            "--field",
            "effective_date",
            "--value",
            "2026-04-16",
        )
        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)

        updated_row = self.fetch_document_row("sample.txt")
        self.assertEqual(updated_row["effective_date"], "2026-04-16")

        exit_code, error_payload, _, _ = self.run_cli(
            "set-field",
            str(self.root),
            "--doc-id",
            str(row["id"]),
            "--field",
            "effective_date",
            "--value",
            "next Tuesday",
        )
        self.assertEqual(exit_code, 2)
        self.assertIsNotNone(error_payload)
        self.assertIn("Expected ISO date value", error_payload["error"])

    def test_promote_field_type_supports_week_aggregate_without_reingest(self) -> None:
        (self.root / "contract-a.txt").write_text("Contract A\n", encoding="utf-8")
        (self.root / "contract-b.txt").write_text("Contract B\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        first_row = self.fetch_document_row("contract-a.txt")
        second_row = self.fetch_document_row("contract-b.txt")

        self.assertEqual(self.run_cli("add-field", str(self.root), "effective_date", "text")[0], 0)
        self.assertEqual(
            self.run_cli(
                "set-field",
                str(self.root),
                "--doc-id",
                str(first_row["id"]),
                "--field",
                "effective_date",
                "--value",
                "2026-04-01",
            )[0],
            0,
        )
        self.assertEqual(
            self.run_cli(
                "set-field",
                str(self.root),
                "--doc-id",
                str(second_row["id"]),
                "--field",
                "effective_date",
                "--value",
                "2026-04-15",
            )[0],
            0,
        )

        exit_code, promote_payload, _, _ = self.run_cli(
            "promote-field-type",
            str(self.root),
            "effective_date",
            "date",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(promote_payload)
        self.assertEqual(promote_payload["status"], "ok")
        self.assertTrue(promote_payload["promotion_applied"])
        self.assertEqual(promote_payload["documents_with_values"], 2)

        aggregate_payload = retriever_tools.aggregate(
            self.root,
            None,
            ["week:effective_date"],
            "count",
            None,
            None,
            20,
            False,
        )
        self.assertEqual(aggregate_payload["graph"]["type"], "line")
        self.assertEqual(len(aggregate_payload["buckets"]), 2)
        self.assertTrue(all(str(bucket["week"]).startswith("2026-W") for bucket in aggregate_payload["buckets"]))

    def test_promote_field_type_blocks_invalid_existing_values(self) -> None:
        document_path = self.root / "contract.txt"
        document_path.write_text("Contract text\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("contract.txt")
        self.assertEqual(self.run_cli("add-field", str(self.root), "effective_date", "text")[0], 0)
        self.assertEqual(
            self.run_cli(
                "set-field",
                str(self.root),
                "--doc-id",
                str(row["id"]),
                "--field",
                "effective_date",
                "--value",
                "tomorrow",
            )[0],
            0,
        )

        exit_code, payload, _, _ = self.run_cli(
            "promote-field-type",
            str(self.root),
            "effective_date",
            "date",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["status"], "blocked")
        self.assertFalse(payload["promotion_applied"])
        self.assertEqual(payload["invalid_value_samples"][0]["value"], "tomorrow")

    def test_aggregate_groups_by_dataset_name_and_explain_flag(self) -> None:
        (self.root / "doc-one.txt").write_text("alpha\n", encoding="utf-8")
        (self.root / "doc-two.txt").write_text("beta\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        first_row = self.fetch_document_row("doc-one.txt")
        create_exit, create_payload, _, _ = self.run_cli("create-dataset", str(self.root), "Review Set")
        self.assertEqual(create_exit, 0)
        self.assertIsNotNone(create_payload)
        dataset_id = int(create_payload["dataset"]["id"])

        add_exit, _, _, _ = self.run_cli(
            "add-to-dataset",
            str(self.root),
            "--dataset-id",
            str(dataset_id),
            "--doc-id",
            str(first_row["id"]),
        )
        self.assertEqual(add_exit, 0)

        exit_code, payload, _, _ = self.run_cli(
            "aggregate",
            str(self.root),
            "--group-by",
            "dataset_name",
            "--explain",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        bucket_map = {item["dataset_name"]: item["count"] for item in payload["buckets"]}
        self.assertEqual(bucket_map[self.root.name], 2)
        self.assertEqual(bucket_map["Review Set"], 1)
        self.assertEqual(payload["graph"]["description"], "Count by Dataset")
        self.assertIn("sql", payload)
        self.assertIn("COUNT(DISTINCT d.id)", payload["sql"])
        self.assertIn("JOIN datasets ds", payload["sql"])

    def test_aggregate_groups_entities_by_type_origin_and_role(self) -> None:
        self.write_email_message(
            self.root / "entity-source.eml",
            subject="Entity aggregate",
            body_text="Entity aggregate body",
            author="Alice Example <alice@example.com>",
            recipients="Bob Example <bob@example.com>",
            cc=None,
            message_id="<entity-aggregate@example.com>",
        )

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        create_exit, create_payload, _, _ = self.run_cli(
            "create-entity",
            str(self.root),
            "--entity-type",
            "shared_mailbox",
            "--display-name",
            "Review Team",
            "--email",
            "review@example.com",
        )
        self.assertEqual(create_exit, 0)
        self.assertIsNotNone(create_payload)

        type_exit, type_payload, _, _ = self.run_cli(
            "aggregate",
            str(self.root),
            "--group-by",
            "entity_type",
            "--order-by",
            "entity_type",
        )
        self.assertEqual(type_exit, 0)
        self.assertIsNotNone(type_payload)
        assert type_payload is not None
        self.assertEqual(type_payload["aggregate_scope"], "entities")
        type_counts = {bucket["entity_type"]: bucket["count"] for bucket in type_payload["buckets"]}
        self.assertGreaterEqual(type_counts["person"], 2)
        self.assertEqual(type_counts["shared_mailbox"], 1)

        origin_exit, origin_payload, _, _ = self.run_cli(
            "aggregate",
            str(self.root),
            "--group-by",
            "entity_origin",
        )
        self.assertEqual(origin_exit, 0)
        self.assertIsNotNone(origin_payload)
        assert origin_payload is not None
        origin_counts = {bucket["entity_origin"]: bucket["count"] for bucket in origin_payload["buckets"]}
        self.assertGreaterEqual(origin_counts["manual"], 1)

        role_exit, role_payload, _, _ = self.run_cli(
            "aggregate",
            str(self.root),
            "--group-by",
            "entity_role",
            "--order-by",
            "entity_role",
        )
        self.assertEqual(role_exit, 0)
        self.assertIsNotNone(role_payload)
        assert role_payload is not None
        role_counts = {bucket["entity_role"]: bucket["count"] for bucket in role_payload["buckets"]}
        self.assertEqual(role_counts["author"], 1)
        self.assertEqual(role_counts["recipient"], 1)

        status_exit, status_payload, _, _ = self.run_cli(
            "aggregate",
            str(self.root),
            "--group-by",
            "entity_status",
        )
        self.assertEqual(status_exit, 0)
        self.assertIsNotNone(status_payload)
        assert status_payload is not None
        status_counts = {bucket["entity_status"]: bucket["count"] for bucket in status_payload["buckets"]}
        self.assertGreaterEqual(status_counts["active"], 3)

        mixed_exit, mixed_payload, _, _ = self.run_cli(
            "aggregate",
            str(self.root),
            "--group-by",
            "entity_type",
            "--group-by",
            "content_type",
        )
        self.assertEqual(mixed_exit, 2)
        self.assertIsNotNone(mixed_payload)
        assert mixed_payload is not None
        self.assertIn("cannot be mixed", mixed_payload["error"])

    def test_claude_routing_ladder_lists_entity_subcommands(self) -> None:
        routing_text = (REPO_ROOT / "CLAUDE.md").read_text(encoding="utf-8")

        self.assertIn("### Entities", routing_text)
        self.assertIn("`list-entities`", routing_text)
        self.assertIn("`show-entity`", routing_text)
        self.assertIn("`list-entity-role-inventory`", routing_text)
        self.assertIn("`list-conversations`", routing_text)
        self.assertIn("`/entities`", routing_text)
        self.assertIn("entities by type", routing_text)

    def test_aggregate_select_from_scope_narrows_bucket_population(self) -> None:
        (self.root / "alpha-scope.txt").write_text("alpha body\n", encoding="utf-8")
        (self.root / "beta-scope.txt").write_text("beta body\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 2)

        scope_exit, scope_payload, _, _ = self.run_cli("slash", str(self.root), "/search", "alpha")
        self.assertEqual(scope_exit, 0)
        self.assertIsNotNone(scope_payload)
        self.assertEqual(scope_payload["total_hits"], 1)

        exit_code, payload, _, _ = self.run_cli(
            "aggregate",
            str(self.root),
            "--group-by",
            "file_name",
            "--select-from-scope",
        )

        self.assertEqual(exit_code, 0)
        self.assertIsNotNone(payload)
        self.assertTrue(payload["selected_from_scope"])
        self.assertEqual(payload["scope"]["keyword"], "alpha")
        self.assertEqual(payload["buckets"], [{"file_name": "alpha-scope.txt", "count": 1}])

    def test_set_field_rejects_system_managed_fields(self) -> None:
        document_path = self.root / "sample.txt"
        document_path.write_text("hello\n", encoding="utf-8")

        retriever_tools.bootstrap(self.root)
        ingest_result = retriever_tools.ingest(self.root, recursive=True, raw_file_types=None)
        self.assertEqual(ingest_result["new"], 1)

        row = self.fetch_document_row("sample.txt")
        blocked_fields = [
            "control_number",
            "dataset_id",
            "ingested_at",
            "last_seen_at",
            "parent_document_id",
            "updated_at",
            retriever_tools.MANUAL_FIELD_LOCKS_COLUMN,
        ]

        for field_name in blocked_fields:
            with self.subTest(field_name=field_name):
                with self.assertRaises(retriever_tools.RetrieverError) as context:
                    retriever_tools.set_field(self.root, row["id"], field_name, "override")
                self.assertIn("system-managed", str(context.exception))


class CidInliningTests(unittest.TestCase):
    PNG_PIXEL = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+a9mQAAAAASUVORK5CYII="
    )

    def test_sniff_image_mime_type_detects_common_formats(self) -> None:
        self.assertEqual(retriever_tools.sniff_image_mime_type(self.PNG_PIXEL), "image/png")
        self.assertEqual(retriever_tools.sniff_image_mime_type(b"\xff\xd8\xff\xe0..."), "image/jpeg")
        self.assertEqual(retriever_tools.sniff_image_mime_type(b"GIF89a...."), "image/gif")
        self.assertIsNone(retriever_tools.sniff_image_mime_type(b"not-an-image"))

    def test_normalize_content_id_strips_brackets_and_whitespace(self) -> None:
        self.assertEqual(retriever_tools.normalize_content_id("  <foo@bar>  "), "foo@bar")
        self.assertEqual(retriever_tools.normalize_content_id("plain"), "plain")
        self.assertIsNone(retriever_tools.normalize_content_id(""))
        self.assertIsNone(retriever_tools.normalize_content_id(None))
        self.assertEqual(retriever_tools.normalize_content_id(b"<inline>"), "inline")

    def test_inline_cid_references_swaps_src_with_data_uri(self) -> None:
        html_body = (
            '<p>hi</p>'
            '<img src="cid:logo_icon">'
            "<img src='cid:LOGO_ICON'>"
            '<td background="cid:logo_icon">x</td>'
        )
        attachments = [
            {
                "file_name": "logo.png",
                "payload": self.PNG_PIXEL,
                "content_id": "<logo_icon>",
            },
        ]
        result = retriever_tools.inline_cid_references_in_html(html_body, attachments)
        self.assertIsNotNone(result)
        self.assertNotIn("cid:logo_icon", result)
        self.assertNotIn("cid:LOGO_ICON", result)
        self.assertEqual(result.count("data:image/png;base64,"), 3)

    def test_inline_cid_references_leaves_unknown_cids_intact(self) -> None:
        html_body = '<img src="cid:present"><img src="cid:missing">'
        attachments = [
            {
                "file_name": "present.png",
                "payload": self.PNG_PIXEL,
                "content_id": "present",
            },
        ]
        result = retriever_tools.inline_cid_references_in_html(html_body, attachments)
        self.assertIn("data:image/png;base64,", result)
        self.assertIn('src="cid:missing"', result)

    def test_inline_cid_references_noop_without_attachments_or_html(self) -> None:
        self.assertEqual(retriever_tools.inline_cid_references_in_html("<p>x</p>", []), "<p>x</p>")
        self.assertEqual(retriever_tools.inline_cid_references_in_html("<p>x</p>", None), "<p>x</p>")
        self.assertIsNone(retriever_tools.inline_cid_references_in_html(None, [{"content_id": "x", "payload": b""}]))
        html_body = '<img src="cid:logo">'
        attachments_missing_cid = [{"file_name": "logo.png", "payload": self.PNG_PIXEL}]
        self.assertEqual(
            retriever_tools.inline_cid_references_in_html(html_body, attachments_missing_cid),
            html_body,
        )

    def test_inline_cid_references_sniffs_mime_when_extension_missing(self) -> None:
        html_body = '<img src="cid:bareref">'
        attachments = [
            {
                "file_name": "",
                "payload": self.PNG_PIXEL,
                "content_id": "bareref",
            },
        ]
        result = retriever_tools.inline_cid_references_in_html(html_body, attachments)
        self.assertIn("data:image/png;base64,", result)

    def test_extract_eml_attachments_captures_content_id(self) -> None:
        boundary = "boundary-inline-test"
        encoded_png = base64.b64encode(self.PNG_PIXEL).decode("ascii")
        raw_message = (
            "From: sender@example.com\r\n"
            "To: recipient@example.com\r\n"
            "Subject: Inline icon test\r\n"
            "MIME-Version: 1.0\r\n"
            f'Content-Type: multipart/related; boundary="{boundary}"\r\n'
            "\r\n"
            f"--{boundary}\r\n"
            "Content-Type: text/html; charset=utf-8\r\n"
            "\r\n"
            '<html><body><img src="cid:logo-icon" alt="logo"/></body></html>\r\n'
            f"--{boundary}\r\n"
            "Content-Type: image/png\r\n"
            "Content-Transfer-Encoding: base64\r\n"
            "Content-ID: <logo-icon>\r\n"
            'Content-Disposition: inline; filename="logo.png"\r\n'
            "\r\n"
            f"{encoded_png}\r\n"
            f"--{boundary}--\r\n"
        ).encode("ascii")

        parsed = BytesParser(policy=policy.default).parsebytes(raw_message)
        attachments = retriever_tools.extract_eml_attachments(parsed)
        self.assertTrue(any(a.get("content_id") == "logo-icon" for a in attachments))
        self.assertTrue(any(a.get("is_inline") is True for a in attachments))

    def test_build_email_extracted_payload_inlines_cid_images_in_preview(self) -> None:
        html_body = '<html><body><img src="cid:icon"/></body></html>'
        attachments = [
            {
                "file_name": "icon.png",
                "ordinal": 1,
                "payload": self.PNG_PIXEL,
                "file_hash": "deadbeef",
                "content_id": "icon",
                "content_type": "image/png",
            },
            {
                "file_name": "notes.txt",
                "ordinal": 2,
                "payload": b"notes",
                "file_hash": "feedbead",
                "content_type": "text/plain",
            },
        ]
        payload = retriever_tools.build_email_extracted_payload(
            subject="Test",
            author="a@example.com",
            recipients="b@example.com",
            date_created="2026-04-16T00:00:00Z",
            text_body="See icon.",
            html_body=html_body,
            attachments=attachments,
            preview_file_name="msg.html",
        )
        preview_content = payload["preview_artifacts"][0]["content"]
        self.assertIn("data:image/png;base64,", preview_content)
        self.assertNotIn('src="cid:icon"', preview_content)
        self.assertEqual([attachment["file_name"] for attachment in payload["attachments"]], ["notes.txt"])

    def test_build_email_extracted_payload_uses_subject_for_preview_title_and_heading(self) -> None:
        payload = retriever_tools.build_email_extracted_payload(
            subject="Legalweek 2023 Mobile App Now Available Attachments: agenda.pdf; deck.pdf",
            author="events@example.com",
            recipients="sergey@example.com",
            date_created="2026-04-17T15:31:00Z",
            text_body="Email body.",
            html_body=None,
            attachments=[],
            preview_file_name="msg.html",
        )

        preview_content = payload["preview_artifacts"][0]["content"]
        self.assertIn("<title>Legalweek 2023 Mobile App Now Available</title>", preview_content)
        self.assertIn('class="gmail-thread-title">Legalweek 2023 Mobile App Now Available</h1>', preview_content)
        self.assertNotIn("Attachments: agenda.pdf", preview_content)

    def test_build_email_extracted_payload_promotes_calendar_invites_into_parent_text(self) -> None:
        payload = retriever_tools.build_email_extracted_payload(
            subject="Discuss Relativity",
            author="Sergey Demyanov <sergey@discoverbeagle.com>",
            recipients="Max Faleev <max@discoverbeagle.com>",
            date_created="2023-06-01T00:31:26Z",
            text_body="Invitation from Google Calendar",
            html_body=None,
            attachments=[
                {
                    "file_name": "invite.ics",
                    "ordinal": 1,
                    "payload": "\r\n".join(
                        [
                            "BEGIN:VCALENDAR",
                            "VERSION:2.0",
                            "METHOD:REQUEST",
                            "BEGIN:VEVENT",
                            "SUMMARY:Discuss Relativity",
                            "DTSTART;TZID=America/New_York:20230601T150000",
                            "DTEND;TZID=America/New_York:20230601T153000",
                            "ORGANIZER;CN=Sergey Demyanov:mailto:sergey@discoverbeagle.com",
                            "ATTENDEE;CN=Max Faleev:mailto:max@discoverbeagle.com",
                            "STATUS:CONFIRMED",
                            "X-GOOGLE-CONFERENCE:https://meet.google.com/fps-qara-aie",
                            "END:VEVENT",
                            "END:VCALENDAR",
                            "",
                        ]
                    ).encode("utf-8"),
                    "file_hash": "invite-hash",
                    "content_type": "text/calendar",
                },
                {
                    "file_name": "notes.txt",
                    "ordinal": 2,
                    "payload": b"Agenda note",
                    "file_hash": "notes-hash",
                    "content_type": "text/plain",
                },
            ],
            preview_file_name="msg.html",
        )

        self.assertEqual([attachment["file_name"] for attachment in payload["attachments"]], ["notes.txt"])
        self.assertIn("[[RETRIEVER_CALENDAR_INVITE]]", payload["text_content"])
        self.assertIn("Join: https://meet.google.com/fps-qara-aie", payload["text_content"])
        self.assertIn("Sergey Demyanov <sergey@discoverbeagle.com>", str(payload["participants"]))

        preview_content = payload["preview_artifacts"][0]["content"]
        self.assertIn("Calendar invite", preview_content)
        self.assertIn("Discuss Relativity", preview_content)
        self.assertNotIn(">invite.ics<", preview_content)


class CalendarInviteParsingTests(unittest.TestCase):
    def test_parse_icalendar_event_metadata_extracts_invite_fields(self) -> None:
        metadata = retriever_tools.parse_icalendar_event_metadata(
            "\r\n".join(
                [
                    "BEGIN:VCALENDAR",
                    "VERSION:2.0",
                    "METHOD:REQUEST",
                    "BEGIN:VEVENT",
                    "SUMMARY:Discuss Relativity",
                    "DTSTART;TZID=America/New_York:20230601T150000",
                    "DTEND;TZID=America/New_York:20230601T153000",
                    "ORGANIZER;CN=Sergey Demyanov:mailto:sergey@discoverbeagle.com",
                    "ATTENDEE;CN=Max Faleev:mailto:max@discoverbeagle.com",
                    "STATUS:CONFIRMED",
                    "X-GOOGLE-CONFERENCE:https://meet.google.com/fps-qara-aie",
                    "UID:4742uqv7uab4rpn1fmps1bqvs5@google.com",
                    "END:VEVENT",
                    "END:VCALENDAR",
                    "",
                ]
            )
        )

        self.assertIsNotNone(metadata)
        self.assertEqual(metadata["summary"], "Discuss Relativity")
        self.assertEqual(metadata["when"], "Jun 1, 2023 3:00 PM - 3:30 PM EDT")
        self.assertEqual(metadata["organizer"], "Sergey Demyanov <sergey@discoverbeagle.com>")
        self.assertEqual(metadata["attendees_display"], "Max Faleev <max@discoverbeagle.com>")
        self.assertEqual(metadata["conference_url"], "https://meet.google.com/fps-qara-aie")
        self.assertEqual(metadata["method"], "REQUEST")
        self.assertEqual(metadata["status"], "CONFIRMED")
        self.assertEqual(
            retriever_tools.summarize_icalendar_invite_status(metadata),
            "Request · Confirmed",
        )


class AttachmentResolutionTests(unittest.TestCase):
    def test_sniff_attachment_file_type_detects_pdf_and_docx(self) -> None:
        self.assertEqual(retriever_tools.sniff_attachment_file_type(b"%PDF-1.4\n1 0 obj\n"), "pdf")

        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                "[Content_Types].xml",
                """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="xml" ContentType="application/xml"/>
</Types>
""",
            )
            archive.writestr(
                "word/document.xml",
                """<?xml version="1.0" encoding="UTF-8"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"/>
""",
            )
        self.assertEqual(retriever_tools.sniff_attachment_file_type(buffer.getvalue()), "docx")

    def test_pst_attachment_file_name_recovers_record_set_metadata(self) -> None:
        class FakeRecordEntry:
            def __init__(self, *, entry_type: int, value_type: int, data: bytes):
                self.entry_type = entry_type
                self.value_type = value_type
                self.data = data

        class FakeRecordSet:
            def __init__(self, entries: list[object]):
                self.entries = entries

        class FakeAttachment:
            def __init__(self, *, payload: bytes, record_sets: list[object]):
                self.data = payload
                self.record_sets = record_sets
                self.name = None
                self.filename = None
                self.long_filename = None
                self.display_name = None
                self.mime_tag = None

        payload = b"%PDF-1.7\n1 0 obj\n"
        attachment = FakeAttachment(
            payload=payload,
            record_sets=[
                FakeRecordSet(
                    [
                        FakeRecordEntry(
                            entry_type=retriever_tools.PST_PROP_ATTACH_EXTENSION,
                            value_type=0x001F,
                            data=".pdf\x00".encode("utf-16-le"),
                        ),
                        FakeRecordEntry(
                            entry_type=retriever_tools.PST_PROP_ATTACH_FILENAME,
                            value_type=0x001F,
                            data="Mcneil~1.pdf\x00".encode("utf-16-le"),
                        ),
                        FakeRecordEntry(
                            entry_type=retriever_tools.PST_PROP_ATTACH_LONG_FILENAME,
                            value_type=0x001F,
                            data="Mcneill, Walter.pdf\x00".encode("utf-16-le"),
                        ),
                        FakeRecordEntry(
                            entry_type=retriever_tools.PST_PROP_ATTACH_MIME_TAG,
                            value_type=0x001F,
                            data="application/pdf\x00".encode("utf-16-le"),
                        ),
                    ]
                )
            ],
        )

        self.assertEqual(
            retriever_tools.pst_attachment_file_name(attachment, 1, payload=payload),
            "Mcneill, Walter.pdf",
        )
        self.assertEqual(retriever_tools.pst_attachment_content_type(attachment), "application/pdf")


if __name__ == "__main__":
    unittest.main()
